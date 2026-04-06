import collections
import json
import logging
import os
import uuid
import base64
from decimal import Decimal
from html import escape
from django.utils.html import strip_tags
from random import choice, shuffle
import time
import datetime
from datetime import datetime as dt
from io import BytesIO
import boto3
import requests
from boto3 import client
from dateutil.relativedelta import relativedelta
from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.staticfiles import finders
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db.models import Q
from django.core.serializers.json import DjangoJSONEncoder
from .tasks import upload_video_to_s3
from .utils import process_uploaded_image, send_telegram_new_entity_notification, get_planet_image_url, get_fallback_planet_url, upload_file_to_s3_sync, handle_entity_file_uploads, upload_entity_logo, upload_entity_catalog_card, upload_entity_files
from django.db import (
    models,
    transaction,
)
from django.db.models import (
    Avg,
    Case,
    Count,
    DecimalField,
    Exists,
    ExpressionWrapper,
    F,
    FloatField,
    IntegerField,
    Max,
    Min,
    OuterRef,
    Q,
    Subquery,
    Sum,
    Value,
    When,
)
from django.db.models.functions import (
    Coalesce,
    TruncMonth,
    Floor,
)
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.text import slugify
from django.views.decorators.cache import cache_page
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.vary import vary_on_headers
from django.contrib.messages import get_messages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .forms import (
    CommentForm,
    FranchiseCommentForm,
    AgencyCommentForm,
    SpecialistCommentForm,
    ContactForm,
    LoginForm,
    MessageForm,
    ModeratorTicketForm,
    ProfileEditForm,
    RegisterForm,
    StartupForm,
    StartupEditForm,
    FranchiseForm,
    FranchiseEditForm,
    AgencyForm,
    AgencyEditForm,
    SpecialistForm,
    SpecialistEditForm,
    SupportTicketForm,
    UserSearchForm,
)
from .models import (
    ChatConversations,
    ChatParticipants,
    City,
    Comments,
    Directions,
    FranchiseDirections,
    EntityTypes,
    FileStorage,
    FileTypes,
    Franchises,
    InvestmentTransactions,
    Messages,
    MessageStatuses,
    ModerationLog,
    NewsArticles,
    NewsCategories,
    NewsComments,
    NewsDislikes,
    NewsLikes,
    NewsViews,
    PaymentMethods,
    ReviewStatuses,
    Roles,
    Startups,
    StartupTimeline,
    SupportTicket,
    TransactionTypes,
    Users,
    UserVotes,
    FranchiseVotes,
    Agencies,
    AgencyComments,
    AgencyVotes,
    Specialists,
    SpecialistComments,
    SpecialistVotes,
)
from .utils import send_telegram_support_message, send_telegram_contact_form_message
logger = logging.getLogger(__name__)


def is_moderator(user):
    """Централизованная проверка прав модератора."""
    if not user or not user.is_authenticated:
        return False
    role = getattr(user, "role", None)
    if not role:
        return False
    return (getattr(role, "role_name", "") or "").lower() == "moderator"


def invalidate_catalog_cache():
    """Сбрасываем кэш каталогов и главной страницы при изменении данных."""
    from django.core.cache import cache as django_cache
    django_cache.delete('home_page_anonymous_v1')
    # cache_page использует ключи на основе URL, сбрасываем всё с нашим префиксом
    try:
        django_cache.delete_pattern('*views.decorators.cache*')
    except (AttributeError, NotImplementedError):
        # Если бэкенд не поддерживает delete_pattern, очищаем весь кэш
        django_cache.clear()


RATE_WINDOW_SECONDS = 60
RATE_MAX_ATTEMPTS = 15
BLOCK_SECONDS = 30
CAPTCHA_FAILS_THRESHOLD = 3
FREQUENT_ATTEMPTS_THRESHOLD = 3
CAPTCHA_INVALID_MESSAGE = "Неверный ответ на капчу."

def _session_key(prefix: str, suffix: str) -> str:
    return f"{prefix}_{suffix}"

def _now_ts() -> int:
    return int(time.time())

def _get_attempts_in_window(session, prefix: str) -> int:
    key = _session_key(prefix, "attempt_times")
    times = session.get(key, [])
    now_ts = _now_ts()
    recent = [t for t in times if now_ts - t <= RATE_WINDOW_SECONDS]
    session[key] = recent
    return len(recent)

def _record_attempt(session, prefix: str) -> None:
    key = _session_key(prefix, "attempt_times")
    times = session.get(key, [])
    times.append(_now_ts())
    session[key] = times
    if _get_attempts_in_window(session, prefix) > RATE_MAX_ATTEMPTS:
        session[_session_key(prefix, "block_until")] = _now_ts() + BLOCK_SECONDS

def _is_blocked(session, prefix: str) -> int:
    block_until = session.get(_session_key(prefix, "block_until"))
    if not block_until:
        return 0
    remaining = block_until - _now_ts()
    if remaining <= 0:
        session.pop(_session_key(prefix, "block_until"), None)
        return 0
    return remaining

def _should_require_captcha(session, prefix: str) -> bool:
    if session.get(_session_key(prefix, "captcha_required")):
        logger.debug(f"[{prefix}] Captcha required flag set in session")
        return True
    fail_count = _get_fail_count(session, prefix)
    attempts = _get_attempts_in_window(session, prefix)
    require = fail_count >= CAPTCHA_FAILS_THRESHOLD or attempts >= FREQUENT_ATTEMPTS_THRESHOLD
    logger.debug(f"[{prefix}] should_require_captcha? fail_count={fail_count}, attempts_in_window={attempts} => {require}")
    return require

def _generate_captcha(session, prefix: str) -> None:
    import random
    a = random.choice([3, 5, 6])
    b = random.choice([3, 5, 6])
    session[_session_key(prefix, "captcha_question")] = f"Сколько будет {a} + {b}?"
    session[_session_key(prefix, "captcha_expected")] = str(a + b)
    session[_session_key(prefix, "captcha_required")] = True
    session[_session_key(prefix, "captcha_set_at")] = _now_ts()
    logger.debug(f"[{prefix}] Generated captcha: question={session.get(_session_key(prefix, 'captcha_question'))}, expected={session.get(_session_key(prefix, 'captcha_expected'))}")

def _clear_captcha(session, prefix: str) -> None:
    session.pop(_session_key(prefix, "captcha_question"), None)
    session.pop(_session_key(prefix, "captcha_expected"), None)
    session.pop(_session_key(prefix, "captcha_required"), None)
    session.pop(_session_key(prefix, "captcha_set_at"), None)
    logger.debug(f"[{prefix}] Cleared captcha requirement")

def _reset_limits(session, prefix: str) -> None:
    session.pop(_session_key(prefix, "attempt_times"), None)
    session.pop(_session_key(prefix, "fail_count"), None)
    session.pop(_session_key(prefix, "fail_last_ts"), None)
    session.pop(_session_key(prefix, "block_until"), None)
    _clear_captcha(session, prefix)
    logger.debug(f"[{prefix}] Reset limits")

def _get_fail_count(session, prefix: str) -> int:
    last_ts = session.get(_session_key(prefix, "fail_last_ts"))
    if last_ts and (_now_ts() - last_ts) > RATE_WINDOW_SECONDS:
        session[_session_key(prefix, "fail_count")] = 0
        session.pop(_session_key(prefix, "fail_last_ts"), None)
        logger.debug(f"[{prefix}] Fail count expired window -> reset to 0")
    return session.get(_session_key(prefix, "fail_count"), 0)

def _inc_fail_count(session, prefix: str) -> int:
    count = _get_fail_count(session, prefix) + 1
    session[_session_key(prefix, "fail_count")] = count
    session[_session_key(prefix, "fail_last_ts")] = _now_ts()
    logger.debug(f"[{prefix}] Increased fail_count -> {count}")
    return count

def _expire_captcha_if_old(session, prefix: str) -> None:
    set_at = session.get(_session_key(prefix, "captcha_set_at"))
    if set_at and (_now_ts() - set_at) > RATE_WINDOW_SECONDS:
        logger.debug(f"[{prefix}] Captcha expired by time window")
        _clear_captcha(session, prefix)

def _clear_captcha_messages(request):
    storage = get_messages(request)
    kept = []
    for m in storage:
        if str(m) != CAPTCHA_INVALID_MESSAGE:
            kept.append((m.level, str(m)))
    for level, msg in kept:
        messages.add_message(request, level, msg)
def safe_create_file_storage(entity_type, entity_id, file_type, file_url, uploaded_at, startup, original_file_name):
    """
    Безопасно создает объект FileStorage, учитывая наличие/отсутствие поля original_file_name
    Проверяет на дублирование по file_url перед созданием
    """
    # Проверяем, не существует ли уже файл с таким file_url
    existing_file = FileStorage.objects.filter(file_url=file_url).first()
    if existing_file:
        logger.warning(f"Файл с ID {file_url} уже существует, пропускаем создание")
        return existing_file
    
    if hasattr(FileStorage, 'original_file_name'):
        try:
            return FileStorage.objects.create(
                entity_type=entity_type,
                entity_id=entity_id,
                file_type=file_type,
                file_url=file_url,
                uploaded_at=uploaded_at,
                startup=startup,
                original_file_name=original_file_name,
            )
        except Exception:
            return FileStorage.objects.create(
                entity_type=entity_type,
                entity_id=entity_id,
                file_type=file_type,
                file_url=file_url,
                uploaded_at=uploaded_at,
                startup=startup,
            )
    else:
        return FileStorage.objects.create(
            entity_type=entity_type,
            entity_id=entity_id,
            file_type=file_type,
            file_url=file_url,
            uploaded_at=uploaded_at,
            startup=startup,
        )
def get_synced_files(entity, file_type_name, file_urls_field):
    """
    Синхронизирует файлы между редактором и деталями
    Возвращает только те файлы, которые есть и в поле сущности, и в FileStorage
    """
    if not hasattr(entity, file_urls_field):
        return FileStorage.objects.none()
    
    file_urls = getattr(entity, file_urls_field) or []
    if not file_urls:
        return FileStorage.objects.none()
    
    try:
        file_type = FileTypes.objects.get(type_name=file_type_name)
        
        # Исправляем получение entity_type - убираем 's' в конце
        entity_class_name = entity.__class__.__name__.lower()
        if entity_class_name.endswith('s'):
            entity_type_name = entity_class_name[:-1]  # убираем 's'
        else:
            entity_type_name = entity_class_name
        
        # Специальная обработка для агентств
        if entity_class_name == 'agencies':
            entity_type_name = 'agency'
        
        entity_type = EntityTypes.objects.get(type_name=entity_type_name)
        
        # Исправляем получение entity_id - убираем 's' в конце
        if entity_class_name.endswith('s'):
            entity_id_field = f"{entity_class_name[:-1]}_id"  # убираем 's'
        else:
            entity_id_field = f"{entity_class_name}_id"
        
        # Специальная обработка для агентств
        if entity_class_name == 'agencies':
            entity_id = getattr(entity, 'agency_id')
        else:
            entity_id = getattr(entity, entity_id_field)
        
        # Получаем файлы, которые есть и в поле сущности, и в FileStorage
        synced_files = FileStorage.objects.filter(
            file_url__in=file_urls,
            file_type=file_type,
            entity_type=entity_type,
            entity_id=entity_id
        ).order_by("-uploaded_at")
        
        return synced_files
    except (FileTypes.DoesNotExist, EntityTypes.DoesNotExist, AttributeError):
        return FileStorage.objects.none()

def delete_file_from_s3(file_path):
    """
    Удаляет файл из S3 хранилища
    """
    try:
        from storages.backends.s3boto3 import S3Boto3Storage
        storage = S3Boto3Storage()
        
        # Проверяем, существует ли файл перед удалением
        if storage.exists(file_path):
            storage.delete(file_path)
            logger.info(f"Файл успешно удален из S3: {file_path}")
            return True
        else:
            logger.warning(f"Файл не найден в S3: {file_path}")
            return False
    except Exception as e:
        logger.error(f"Ошибка удаления файла из S3 {file_path}: {e}")
        return False

def get_unique_filename(original_name, startup_id, file_type_name):
    """
    Генерирует уникальное имя файла, добавляя (2), (3) и т.д. если файл с таким именем уже существует
    """
    name, ext = os.path.splitext(original_name)
    try:
        file_type = FileTypes.objects.get(type_name=file_type_name)
        if not hasattr(FileStorage, 'original_file_name'):
            return original_name
        try:
            existing_files = FileStorage.objects.filter(
                startup_id=startup_id,
                file_type=file_type,
                original_file_name=original_name
            )
            if not existing_files.exists():
                return original_name
            counter = 2
            while True:
                new_name = f"{name} ({counter}){ext}"
                existing_duplicate = FileStorage.objects.filter(
                    startup_id=startup_id,
                    file_type=file_type,
                    original_file_name=new_name
                )
                if not existing_duplicate.exists():
                    return new_name
                counter += 1
        except Exception:
            return original_name
    except FileTypes.DoesNotExist:
        logger.error(f"FileType '{file_type_name}' не найден")
        return original_name
DIRECTION_TRANSLATIONS = {
    'Beauty': 'Красота', 'Technology': 'Технологии', 'Healthcare': 'Здравоохранение', 'Health': 'Здоровье',
    'Finance': 'Финансы', 'Cafe': 'Кафе/рестораны', 'Restaurant': 'Кафе/рестораны', 'Delivery': 'Доставка',
    'Fastfood': 'Фастфуд', 'Sport': 'Спорт', 'Transport': 'Транспорт', 'Psychology': 'Психология',
    'AI': 'ИИ', 'Auto': 'Авто',
    'Education': 'Образование', 'Entertainment': 'Развлечения',
    'Fashion': 'Мода', 'Food': 'Еда', 'Gaming': 'Игры', 'Real Estate': 'Недвижимость', 'Travel': 'Путешествия',
    'Agriculture': 'Сельское хозяйство', 'Energy': 'Энергетика', 'Environment': 'Экология', 'Social': 'Социальные проекты', 'Media': 'Медиа',
    'E-commerce': 'Электронная коммерция', 'Biotech': 'Биотехнологии', 'Construction': 'Строительство',
    'Logistics': 'Логистика', 'Manufacturing': 'Производство', 'Retail': 'Розничная торговля', 'Security': 'Безопасность', 'Insurance': 'Страхование',
    'Legal': 'Юридические услуги', 'Consulting': 'Консалтинг', 'Marketing': 'Маркетинг', 'IT': 'ИТ', 'Software': 'Программное обеспечение',
    'Hardware': 'Аппаратное обеспечение', 'Mobile': 'Мобильные приложения', 'Web': 'Веб-разработка', 'Blockchain': 'Блокчейн',
    'Cryptocurrency': 'Криптовалюты', 'VR': 'Виртуальная реальность', 'AR': 'Дополненная реальность', 'IoT': 'Интернет вещей',
    'Robotics': 'Робототехника', 'Space': 'Космические технологии', 'Science': 'Наука', 'Research': 'Исследования',     'Other': 'Другое',
}
FIXED_CATEGORIES = [
    {'original_name': 'AI', 'direction_name': 'ИИ', 'icon': 'ai.webp'},
    {'original_name': 'Auto', 'direction_name': 'Авто', 'icon': 'auto.webp'},
    {'original_name': 'Beauty', 'direction_name': 'Красота', 'icon': 'beauty.webp'},
    {'original_name': 'Cafe', 'direction_name': 'Кафе/рестораны', 'icon': 'cafe.webp'},
    {'original_name': 'Delivery', 'direction_name': 'Доставка', 'icon': 'delivery-b562f7.webp'},
    {'original_name': 'Fastfood', 'direction_name': 'Фастфуд', 'icon': 'fastfood.webp'},
    {'original_name': 'Health', 'direction_name': 'Здоровье', 'icon': 'healthcare.webp', 'match_names': ['Health', 'Healthcare', 'Medicine']},
    {'original_name': 'Finance', 'direction_name': 'Финансы', 'icon': 'finance.webp'},
    {'original_name': 'Psychology', 'direction_name': 'Психология', 'icon': 'psychology-16bdc1.webp'},
    {'original_name': 'Technology', 'direction_name': 'Технологии', 'icon': 'technology.webp'},
    {'original_name': 'Sport', 'direction_name': 'Спорт', 'icon': 'sport.webp'},
    {'original_name': 'Transport', 'direction_name': 'Транспорт', 'icon': 'transport.webp'},
]
def home(request):
    if not request.user.is_authenticated:
        from django.core.cache import cache as django_cache
        cache_key = 'home_page_anonymous_v1'
        cached_response = django_cache.get(cache_key)
        if cached_response:
            return cached_response
        import random
        from django.db.models import Avg, Count, Sum, F, Case, When, Value, FloatField, DecimalField
        from django.db.models.functions import Coalesce
        from django.templatetags.static import static
        startups_query = Startups.objects.filter(status="approved").select_related('owner', 'direction', 'stage').annotate(
            rating_avg=Coalesce(Avg("uservotes__rating"), 0.0, output_field=FloatField()),
            voters_count=Count("uservotes", distinct=True),
            total_investors=Count("investmenttransactions__investor", distinct=True),
            current_funding=Coalesce(
                Sum("investmenttransactions__amount"), 0, output_field=DecimalField()
            ),
            comment_count=Count("comments", distinct=True),
            progress=Case(
                When(funding_goal__gt=0, then=(F("amount_raised") * 100.0 / F("funding_goal"))),
                default=Value(0),
                output_field=FloatField(),
            )
        )
        all_startups = list(startups_query)
        demo_startups = []
        if all_startups:
            num_startups = min(6, len(all_startups))
            demo_startups = random.sample(all_startups, num_startups)
        startups_data = []
        for startup in demo_startups:
            planet_image_url = get_planet_image_url(startup.planet_image) if startup.planet_image else get_fallback_planet_url(startup.startup_id)
            startups_data.append({
                "id": startup.startup_id,
                "name": startup.title,
                "description": strip_tags(startup.short_description or startup.description[:200]) if (startup.short_description or startup.description) else "",
                "image": planet_image_url,
                "rating": round(startup.rating_avg, 2),
                "voters_count": startup.voters_count,
                "comment_count": startup.comment_count,
                "direction": startup.direction.direction_name if startup.direction else "Не указано",
                "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не указано",
                "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указано",
                "investors": startup.total_investors,
                "progress": round(startup.progress, 2) if startup.progress is not None else 0,
                "investment_type": "Выкуп+инвестирование" if startup.both_mode else ("Только выкуп" if startup.only_buy else "Только инвестирование")
            })

        directions_data = FIXED_CATEGORIES.copy()
        selected_startups = []
        if len(all_startups) > 0:
            selected_startups = all_startups[:6]
        else:
            selected_startups = []
        planets_data = []
        for i, startup in enumerate(selected_startups):
            planet_image_url = None

            if startup.planet_image:
                planet_image_url = get_planet_image_url(startup.planet_image)

            if not planet_image_url:
                planet_image_url = get_fallback_planet_url(startup.startup_id)

            direction_original = 'Не указано'
            if startup.direction:
                for cat in directions_data:
                    if cat['direction_name'] == startup.direction.direction_name or cat['original_name'] == getattr(startup.direction, 'original_name', None):
                        direction_original = cat['original_name']
                        break
            planets_data.append({
                "id": startup.startup_id,
                "startup_id": startup.startup_id,
                "name": startup.title,
                "description": strip_tags(startup.short_description or startup.description[:200]) if (startup.short_description or startup.description) else "",
                "image": planet_image_url,
                "rating": round(startup.rating_avg, 2) if hasattr(startup, 'rating_avg') else startup.get_average_rating(),
                "voters_count": startup.voters_count if hasattr(startup, 'voters_count') else startup.total_voters,
                "comment_count": startup.comment_count if hasattr(startup, 'comment_count') else 0,
                "direction": direction_original,
                "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не указано",
                "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указано",
                "investors": startup.total_investors if hasattr(startup, 'total_investors') else startup.get_investors_count(),
                "progress": round(startup.progress, 2) if hasattr(startup, 'progress') and startup.progress is not None else startup.get_progress_percentage(),
                "investment_type": "Выкуп+инвестирование" if startup.both_mode else ("Только выкуп" if startup.only_buy else "Только инвестирование"),
                "logo": startup.get_logo_url(),
            })


        random_startupers = []
        try:

            startuper_users = Users.objects.filter(
                role__role_name__iexact='startuper',
                rating__isnull=False
            ).select_related('role').exclude(rating=0).order_by('?')[:3]

            for user in startuper_users:

                rating = float(user.rating or 0)
                full_stars = int(rating)
                has_half_star = rating % 1 >= 0.5

                stars_html = "★" * full_stars
                if has_half_star:
                    stars_html += "☆"
                else:
                    stars_html += "☆" * (5 - full_stars)

                random_startupers.append({
                    'name': user.get_full_name() or user.username or f"Стартапер {user.user_id}",
                    'rating': rating,
                    'stars_html': stars_html,
                    'avatar_url': user.get_profile_picture_url() or static('accounts/images/avatars/default_avatar_ufo.png')
                })


            if len(random_startupers) < 3:

                additional_startupers = Users.objects.filter(
                    role__role_name__iexact='startuper'
                ).exclude(user_id__in=[s.get('user_id', 0) for s in random_startupers]).order_by('?')[:3-len(random_startupers)]

                for user in additional_startupers:

                    import random
                    rating = round(random.uniform(3.5, 5.0), 1)
                    full_stars = int(rating)
                    has_half_star = rating % 1 >= 0.5

                    stars_html = "★" * full_stars
                    if has_half_star:
                        stars_html += "☆"
                    else:
                        stars_html += "☆" * (5 - full_stars)

                    random_startupers.append({
                        'name': user.get_full_name() or user.username or f"Стартапер {user.user_id}",
                        'rating': rating,
                        'stars_html': stars_html,
                        'avatar_url': user.get_profile_picture_url() or static('accounts/images/avatars/default_avatar_ufo.png')
                    })


            if len(random_startupers) == 0:
                random_startupers = [
                    {
                        'name': 'Виктор Смирнов',
                        'rating': 4.5,
                        'stars_html': '★★★★☆',
                        'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                    },
                    {
                        'name': 'Анна Кузнецова',
                        'rating': 4.9,
                        'stars_html': '★★★★★',
                        'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                    },
                    {
                        'name': 'Дмитрий Иванов',
                        'rating': 4.3,
                        'stars_html': '★★★★☆',
                        'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                    }
                ]

        except Exception as e:
            logger.error(f"Error getting random startupers: {e}")

            random_startupers = [
                {
                    'name': 'Виктор Смирнов',
                    'rating': 4.5,
                    'stars_html': '★★★★☆',
                    'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                },
                {
                    'name': 'Анна Кузнецова',
                    'rating': 4.9,
                    'stars_html': '★★★★★',
                    'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                },
                {
                    'name': 'Дмитрий Иванов',
                    'rating': 4.3,
                    'stars_html': '★★★★☆',
                    'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                }
            ]
            logger.info("Using fallback startuper data")


        random_startups = []
        try:

            featured_startups = Startups.objects.filter(
                status="approved"
            ).select_related('owner', 'direction').order_by('?')[:3]


            if len(featured_startups) == 0:
                featured_startups = Startups.objects.all().select_related('owner', 'direction').order_by('?')[:3]

            for startup in featured_startups:

                try:
                    rating = startup.get_average_rating() or 0
                except Exception as e:
                    logger.warning(f"Could not get rating for startup {getattr(startup, 'title', 'Unknown')}: {e}")
                    rating = 4.5

                rating_formatted = str(round(rating))

                # Получаем логотип стартапа
                startup_logo = None
                if hasattr(startup, 'get_logo_url') and startup.get_logo_url():
                    startup_logo = startup.get_logo_url()
                
                # Получаем планету для декоративного отображения
                planet_image = None
                if hasattr(startup, 'planet_image') and startup.planet_image:
                    planet_image = get_planet_image_url(startup.planet_image)
                else:
                    planet_image = get_fallback_planet_url(startup.startup_id)
                
                # Основное изображение - логотип если есть, иначе планета
                startup_image = startup_logo if startup_logo else planet_image


                owner_avatar = static('accounts/images/avatars/default_avatar_ufo.png')
                try:
                    if hasattr(startup, 'owner') and startup.owner and hasattr(startup.owner, 'get_profile_picture_url'):
                        owner_avatar = startup.owner.get_profile_picture_url() or owner_avatar
                except Exception as e:
                    logger.warning(f"Could not get owner avatar for startup {getattr(startup, 'startup_id', 'Unknown')}: {e}")
                    owner_avatar = static('accounts/images/avatars/default_avatar_ufo.png')


                description = getattr(startup, 'short_description', None) or getattr(startup, 'description', None) or "Описание стартапа"
                description = strip_tags(description)
                if len(description) > 100:
                    description = description[:97] + "..."


                startup_id = getattr(startup, 'startup_id', None)
                if startup_id and str(startup_id).isdigit():
                    startup_url = f"/startups/{startup_id}/"
                else:
                    startup_url = "/startups_list/"

                startup_data = {
                    'id': startup_id or 'Unknown',
                    'name': getattr(startup, 'title', 'Unknown'),
                    'rating': rating_formatted,
                    'description': description,
                    'image': startup_image,
                    'planet_image': planet_image,
                    'has_logo': bool(startup_logo),
                    'owner_avatar': owner_avatar,
                    'url': startup_url
                }

                random_startups.append(startup_data)


            if len(random_startups) == 0:
                random_startups = [
                    {
                        'id': 1,
                        'name': 'VoltForge Dynamics',
                        'rating': '4',
                        'description': 'VoltForge разрабатывает твердотельные батареи с графеновыми наноструктурами, которые заряжаются...',
                        'image': static('accounts/images/main_page/volt_forge.webp'),
                        'planet_image': static('accounts/images/planetary_system/textures/planet_1.webp'),
                        'has_logo': False,
                        'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                        'url': '/startups_list/'
                    },
                    {
                        'id': 2,
                        'name': 'NeuroBloom',
                        'rating': '5',
                        'description': 'NeuroBloom предлагает носимый гаджет с ИИ, который анализирует нейронные паттерны...',
                        'image': static('accounts/images/main_page/neuro_bloom.webp'),
                        'planet_image': static('accounts/images/planetary_system/textures/planet_2.webp'),
                        'has_logo': False,
                        'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                        'url': '/startups_list/'
                    },
                    {
                        'id': 3,
                        'name': 'BioCrop Nexus',
                        'rating': '4',
                        'description': 'BioCrop Nexus создает генетически оптимизированные семена, устойчивые к климату...',
                        'image': static('accounts/images/main_page/biocrop_nexus.webp'),
                        'planet_image': static('accounts/images/planetary_system/textures/planet_3.webp'),
                        'has_logo': False,
                        'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                        'url': '/startups_list/'
                    }
                ]

        except Exception as e:
            logger.error(f"Error getting random startups: {e}")

            random_startups = [
                {
                    'id': 1,
                    'name': 'VoltForge Dynamics',
                    'rating': '4',
                    'description': 'VoltForge разрабатывает твердотельные батареи с графеновыми наноструктурами, которые заряжаются...',
                    'image': static('accounts/images/main_page/volt_forge.webp'),
                    'planet_image': static('accounts/images/planetary_system/textures/planet_1.webp'),
                    'has_logo': False,
                    'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                    'url': '/startups_list/'
                },
                {
                    'id': 2,
                    'name': 'NeuroBloom',
                    'rating': '5',
                    'description': 'NeuroBloom предлагает носимый гаджет с ИИ, который анализирует нейронные паттерны...',
                    'image': static('accounts/images/main_page/neuro_bloom.webp'),
                    'planet_image': static('accounts/images/planetary_system/textures/planet_2.webp'),
                    'has_logo': False,
                    'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                    'url': '/startups_list/'
                },
                {
                    'id': 3,
                    'name': 'BioCrop Nexus',
                    'rating': '4',
                    'description': 'BioCrop Nexus создает генетически оптимизированные семена, устойчивые к климату...',
                    'image': static('accounts/images/main_page/biocrop_nexus.webp'),
                    'planet_image': static('accounts/images/planetary_system/textures/planet_3.webp'),
                    'has_logo': False,
                    'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                    'url': '/startups_list/'
                }
            ]
            logger.info("Using fallback startup data")

        try:
            latest_news = list(NewsArticles.objects.filter(
                status="published"
            ).select_related("author").order_by("-published_at")[:5])
        except Exception:
            latest_news = []

        from accounts.promotions import get_active_ads
        context = {
            "demo_startups_data": json.dumps(startups_data, cls=DjangoJSONEncoder),
            "planets_data_json": json.dumps(planets_data, ensure_ascii=False),
            "directions_data_json": json.dumps(directions_data, ensure_ascii=False),
            "directions": directions_data,
            "random_startupers": random_startupers,
            "random_startups": random_startups,
            "latest_news": latest_news,
            "main_ads": get_active_ads("main_under_sidebar"),
        }

        response = render(request, "accounts/main.html", context)
        django_cache.set(cache_key, response, 60 * 5)  # Кэш 5 минут
        return response
    if hasattr(request.user, "role") and request.user.role:
        role_name = request.user.role.role_name.lower()
        if role_name == "investor":
            return redirect("investor_main")
        elif role_name == "startuper":
            return redirect("startuper_main")
        elif role_name == "moderator":
            return redirect("main_page_moderator")
    return redirect("profile")
def faq_page_view(request):
    return render(request, "accounts/faq.html")
@login_required
def contacts_page_view(request):
    prefix = "contacts"
    _expire_captcha_if_old(request.session, prefix)
    captcha_q = None

    if request.method == "POST":
        form = ContactForm(request.POST)


        if True:
            _expire_captcha_if_old(request.session, prefix)
            expected = request.session.get(_session_key(prefix, "captcha_expected"))
            answer_raw = (form.data.get("captcha_answer") or "").strip()
            try:
                answer_normalized = str(int(answer_raw))
            except ValueError:
                answer_normalized = ""
            if not expected or answer_normalized != expected:
                _clear_captcha_messages(request)
                messages.error(request, CAPTCHA_INVALID_MESSAGE)
                _record_attempt(request.session, prefix)
                _inc_fail_count(request.session, prefix)
                _generate_captcha(request.session, prefix)
                captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
                return render(request, "accounts/contacts.html", {"form": form, "captcha_question": captcha_q})
            else:
                logger.debug("[contacts] captcha ok, clearing requirement")
                _clear_captcha(request.session, prefix)

        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            subject = form.cleaned_data['subject']
            message = form.cleaned_data['message']

            try:

                logger.info(f"Sending contact form message from {email} to Telegram")
                sent_ok = send_telegram_contact_form_message(name, email, subject, message)
                logger.info(f"Telegram dispatch result for contact form from {email}: {sent_ok}")

                if sent_ok:
                    messages.success(request, "Спасибо за ваше сообщение! Мы свяжемся с вами в ближайшее время.")
                else:
                    messages.warning(request, "Сообщение отправлено, но возникли проблемы с уведомлением. Мы все равно получим ваше обращение.")


                return redirect('contacts')
            except Exception as e:
                logger.error(f"Unexpected error during Telegram dispatch for contact form from {email}: {e}", exc_info=True)
                messages.success(request, "Спасибо за ваше сообщение! Мы свяжемся с вами в ближайшее время.")

                return redirect('contacts')
        else:

            _generate_captcha(request.session, prefix)
            captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
    else:
        form = ContactForm()

        _generate_captcha(request.session, prefix)
        captcha_q = request.session.get(_session_key(prefix, "captcha_question"))

    return render(request, "accounts/contacts.html", {"form": form, "captcha_question": captcha_q})
def register(request):
    next_url = request.GET.get("next") or request.POST.get("next")
    prefix = "register"
    block_left = _is_blocked(request.session, prefix)
    _expire_captcha_if_old(request.session, prefix)
    captcha_q = None

    if request.method == "POST":
        if block_left:
            _clear_captcha_messages(request)
            messages.error(request, f"Слишком много попыток. Попробуйте через {block_left} сек.")
            form = RegisterForm(request.POST)
            return render(request, "accounts/register.html", {"form": form, "next": next_url})

        form = RegisterForm(request.POST)
        if _should_require_captcha(request.session, prefix):
            _expire_captcha_if_old(request.session, prefix)
            expected = request.session.get(_session_key(prefix, "captcha_expected"))
            answer_raw = (form.data.get("captcha_answer") or "").strip()
            try:
                answer_normalized = str(int(answer_raw))
            except ValueError:
                answer_normalized = ""
            if not expected or answer_normalized != expected:
                _clear_captcha_messages(request)
                messages.error(request, CAPTCHA_INVALID_MESSAGE)
                _record_attempt(request.session, prefix)
                _inc_fail_count(request.session, prefix)
                _generate_captcha(request.session, prefix)
                captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
                logger.debug(f"[register] captcha invalid: provided={answer_normalized!r}, expected={expected!r}")
                return render(request, "accounts/register.html", {"form": form, "next": next_url, "captcha_question": captcha_q})
            else:
                logger.debug("[register] captcha ok, clearing requirement")
                _clear_captcha(request.session, prefix)

        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data["password"])
            # Назначаем роль "startuper" по умолчанию (pk=1) - отображается как "Пользователь"
            try:
                default_role = Roles.objects.get(pk=1)
                user.role = default_role
            except Roles.DoesNotExist:
                logger.warning("Role with pk=1 not found, skipping default role assignment")
            try:
                user.save()
            except Exception as e:
                logger.error(f"Failed to save new user (email={form.cleaned_data.get('email')}): {e}", exc_info=True)
                messages.error(request, "Произошла ошибка при регистрации. Попробуйте позже.")
                return render(request, "accounts/register.html", {"form": form, "next": next_url})
            _reset_limits(request.session, prefix)
            messages.success(
                request, "Регистрация прошла успешно! Теперь вы можете войти."
            )
            if next_url:
                login_url = reverse("login") + f"?next={next_url}"
                return redirect(login_url)
            return redirect("login")
        else:
            _record_attempt(request.session, prefix)
            _inc_fail_count(request.session, prefix)
            if _should_require_captcha(request.session, prefix):
                _generate_captcha(request.session, prefix)
                captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
                return render(request, "accounts/register.html", {"form": form, "next": next_url, "captcha_question": captcha_q})
            return render(request, "accounts/register.html", {"form": form, "next": next_url})
    else:
        form = RegisterForm()
    _clear_captcha_messages(request)
    return render(request, "accounts/register.html", {"form": form, "next": next_url})
def user_login(request):
    logger.debug("Entering user_login view")
    next_url = request.GET.get("next") or request.POST.get("next")
    prefix = "login"
    block_left = _is_blocked(request.session, prefix)
    _expire_captcha_if_old(request.session, prefix)
    captcha_q = None

    if request.method == "POST":
        logger.debug("Processing POST request in user_login")
        if block_left:
            _clear_captcha_messages(request)
            messages.error(request, f"Слишком много попыток. Попробуйте через {block_left} сек.")
            form = LoginForm(request.POST)
            return render(request, "accounts/login.html", {"form": form, "next": next_url})

        form = LoginForm(request.POST)
        if _should_require_captcha(request.session, prefix):
            _expire_captcha_if_old(request.session, prefix)
            expected = request.session.get(_session_key(prefix, "captcha_expected"))
            answer_raw = (form.data.get("captcha_answer") or "").strip()
            try:
                answer_normalized = str(int(answer_raw))
            except ValueError:
                answer_normalized = ""
            if not expected or answer_normalized != expected:
                _clear_captcha_messages(request)
                messages.error(request, CAPTCHA_INVALID_MESSAGE)
                _record_attempt(request.session, prefix)
                _inc_fail_count(request.session, prefix)
                _generate_captcha(request.session, prefix)
                captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
                logger.debug(f"[login] captcha invalid: provided={answer_normalized!r}, expected={expected!r}")
                return render(request, "accounts/login.html", {"form": form, "next": next_url, "captcha_question": captcha_q})
            else:
                logger.debug("[login] captcha ok, clearing requirement")
                _clear_captcha(request.session, prefix)

        if form.is_valid():
            logger.debug(f"Form is valid. Email: {form.cleaned_data['email']}")
            user = authenticate(
                request,
                username=form.cleaned_data["email"],
                password=form.cleaned_data["password"],
            )
            if user is not None:
                logger.info(f"User authenticated: {user.email}")
                _reset_limits(request.session, prefix)
                login(request, user)
                messages.success(
                    request, f"Добро пожаловать, {user.first_name or user.email}!"
                )
                if next_url == reverse("create_startup"):
                    role_name = user.role.role_name.lower() if hasattr(user, "role") and user.role else None
                    if role_name in ["startuper", "moderator"]:
                        return redirect(next_url)
                    else:
                        if role_name == "investor":
                            return redirect("investor_main")
                        elif role_name == "moderator":
                            return redirect("main_page_moderator")
                        else:
                            return redirect("home")
                if hasattr(user, "role") and user.role:
                    role_name = user.role.role_name.lower()
                    if role_name == "investor":
                        return redirect("investor_main")
                    elif role_name == "startuper":
                        return redirect("startuper_main")
                    elif role_name == "moderator":
                        return redirect("main_page_moderator")
                return redirect("home")
            else:
                logger.warning("Authentication failed for email")
                messages.error(request, "Неверный email или пароль.")
                _record_attempt(request.session, prefix)
                _inc_fail_count(request.session, prefix)
                if _should_require_captcha(request.session, prefix):
                    _generate_captcha(request.session, prefix)
                    captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
        else:
            logger.warning(f"Form invalid: {form.errors}")
            _record_attempt(request.session, prefix)
            _inc_fail_count(request.session, prefix)
            if _should_require_captcha(request.session, prefix):
                _generate_captcha(request.session, prefix)
                captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
        return render(request, "accounts/login.html", {"form": form, "next": next_url, "captcha_question": captcha_q})
    else:
        logger.debug("Rendering login form")
        form = LoginForm()
    _clear_captcha_messages(request)
    return render(request, "accounts/login.html", {"form": form, "next": next_url})
def user_logout(request):
    logout(request)
    messages.success(request, "Вы успешно вышли из системы.")
    return redirect("home")

@vary_on_headers('X-Requested-With')
def startups_list(request):
    # Формируем список направлений для сайдбара каталога: объединяем три категории здоровья в одну визуальную «Health»
    health_group = ['Health', 'Healthcare', 'Medicine']
    base_directions = [
        'Technology', 'Finance', 'Education', 'Entertainment', 'Fashion', 'Food', 'Gaming',
        'Real Estate', 'Travel', 'Agriculture', 'Energy', 'Environment', 'Social', 'Auto',
        'Delivery', 'Cafe', 'Fastfood', 'Beauty', 'Transport', 'Sport', 'Psychology', 'AI', 'IT', 'Retail'
    ]
    # Загружаем все доступные направления из базы
    existing = set(Directions.objects.values_list('direction_name', flat=True))
    # Собираем итоговый отображаемый список с учетом наличия
    display_names = [name for name in base_directions if name in existing]
    if any(h in existing for h in health_group):
        display_names.append('Health')
    startup_directions = Directions.objects.filter(direction_name__in=display_names).order_by('direction_name')

    startups_qs = Startups.objects.filter(status="approved").select_related("owner", "direction", "stage")
    selected_categories = request.GET.getlist("category")
    micro_investment_str = request.GET.get("micro_investment", "0")
    min_goal_str = request.GET.get("min_goal", "0")
    max_goal_str = request.GET.get("max_goal", "10000000")
    min_micro_str = request.GET.get("min_micro", "0")
    max_micro_str = request.GET.get("max_micro", "1000000")
    search_query = request.GET.get("search", "").strip()
    min_rating_str = request.GET.get("min_rating", "0")
    max_rating_str = request.GET.get("max_rating", "5")
    sort_order = request.GET.get("sort_order", "newest")
    page_number = request.GET.get("page", 1)

    startups_qs = startups_qs.annotate(
        total_voters_agg=Count("uservotes", distinct=True),
        rating_agg=ExpressionWrapper(
            Coalesce(Avg("uservotes__rating"), 0.0), output_field=FloatField()
        ),
        total_investors_agg=Count("investmenttransactions__investor", distinct=True),
        rating_bucket=Floor(Coalesce(Avg("uservotes__rating"), 0.0)),
    )

    # Категории для JSON: заменяем группу здоровья одной записью
    categories = []
    for d in startup_directions:
        name = d.direction_name
        if name in health_group:
            if not any(c['name'] == 'Health' for c in categories):
                categories.append({'id': d.direction_id, 'name': 'Health'})
        else:
            categories.append({'id': d.direction_id, 'name': name})

    if selected_categories:
        expanded = []
        for cat in selected_categories:
            if cat in ['Health', 'Healthcare', 'Medicine', 'Здоровье']:
                expanded.extend(health_group)
            else:
                expanded.append(cat)
        startups_qs = startups_qs.filter(direction__direction_name__in=list(set(expanded)))

    micro_investment = micro_investment_str == "1"
    if micro_investment:
        startups_qs = startups_qs.filter(micro_investment_available=True)

    if search_query:
        startups_qs = startups_qs.filter(title__icontains=search_query)

    try:
        min_goal = int(min_goal_str)
        max_goal = int(max_goal_str)
        if min_goal > 0:
            startups_qs = startups_qs.filter(funding_goal__gte=min_goal)
        if max_goal < 10000000:
            startups_qs = startups_qs.filter(funding_goal__lte=max_goal)
    except ValueError:
        min_goal = 0
        max_goal = 10000000

    try:
        min_micro = int(min_micro_str)
        max_micro = int(max_micro_str)
        if min_micro > 0:
            startups_qs = startups_qs.filter(percent_amount__gte=min_micro)
        if max_micro < 1000000:
            startups_qs = startups_qs.filter(percent_amount__lte=max_micro)
    except ValueError:
        min_micro = 0
        max_micro = 1000000

    try:
        min_rating = float(min_rating_str)
        max_rating = float(max_rating_str)
        if min_rating > 0:
            startups_qs = startups_qs.filter(rating_agg__gte=min_rating)
        if max_rating < 5:
            startups_qs = startups_qs.filter(rating_agg__lte=max_rating)
    except ValueError:
        min_rating = 0
        max_rating = 5

    filters_active = (
        bool(selected_categories) or
        (search_query != "") or
        (min_goal > 0) or (max_goal < 10000000) or
        (min_micro > 0) or (max_micro < 1000000) or
        (min_rating > 0) or (max_rating < 5) or
        micro_investment
    )
    rating_active = (min_rating > 0 or max_rating < 5)
    goal_active = (min_goal > 0 or max_goal < 10000000)
    micro_active = (min_micro > 0 or max_micro < 1000000)
    if goal_active:
        startups_qs = startups_qs.order_by("funding_goal", "rating_bucket", "rating_agg", "-created_at")
    elif micro_active:
        startups_qs = startups_qs.order_by("percent_amount", "rating_bucket", "rating_agg", "-created_at")
    elif rating_active:
        startups_qs = startups_qs.order_by("rating_bucket", "rating_agg", "-created_at")
    elif filters_active:
        startups_qs = startups_qs.order_by("-created_at")
    else:
        if sort_order == "newest":
            startups_qs = startups_qs.order_by("-created_at")
        elif sort_order == "oldest":
            startups_qs = startups_qs.order_by("created_at")

    # Pinned cards on page 1 (no filters)
    from accounts.promotions import get_pinned_entities, get_active_ads
    pinned_startups = []
    page_num_int = int(page_number) if str(page_number).isdigit() else 1
    if page_num_int == 1 and not filters_active:
        pinned_startups = get_pinned_entities("startup", Startups, "startup_id")
        pinned_ids = [s.startup_id for s in pinned_startups]
        if pinned_ids:
            startups_qs = startups_qs.exclude(startup_id__in=pinned_ids)

    paginator = Paginator(startups_qs, 6)
    page_obj = paginator.get_page(page_number)

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string(
            "accounts/partials/_startup_cards.html", {"page_obj": page_obj}
        )
        return JsonResponse(
            {
                "html": html,
                "has_next": page_obj.has_next(),
                "page_number": page_obj.number,
                "num_pages": paginator.num_pages,
                "count": paginator.count,
            }
        )
    else:
        context = {
            "page_obj": page_obj,
            "paginator": paginator,
            "initial_has_next": page_obj.has_next(),
            "selected_categories": selected_categories,
            "search_query": search_query,
            "min_rating": min_rating,
            "max_rating": max_rating,
            "min_goal": min_goal,
            "max_goal": max_goal,
            "min_micro": min_micro,
            "max_micro": max_micro,
            "micro_investment": micro_investment,
            "sort_order": sort_order,
            "directions": startup_directions,
            "pinned_items": pinned_startups,
            "sidebar_ads": get_active_ads("catalog_sidebar"),
        }
        return render(request, "accounts/startups_list.html", context)

@vary_on_headers('X-Requested-With')
def franchises_list(request):

    existing_dir_ids = (
        Franchises.objects.filter(status="approved", direction__isnull=False)
        .values_list("direction_id", flat=True)
        .distinct()
    )
    franchise_directions = Directions.objects.filter(direction_id__in=existing_dir_ids).order_by("direction_name")

    franchises_qs = Franchises.objects.filter(status="approved").select_related("owner", "direction", "stage")
    selected_categories = request.GET.getlist("category")
    min_payback_str = request.GET.get("min_payback", "0")
    max_payback_str = request.GET.get("max_payback", "60")
    min_investment_str = request.GET.get("min_investment", "0")
    max_investment_str = request.GET.get("max_investment", "10000000")
    search_query = request.GET.get("search", "").strip()
    min_rating_str = request.GET.get("min_rating", "0")
    max_rating_str = request.GET.get("max_rating", "5")
    sort_order = request.GET.get("sort_order", "newest")
    page_number = request.GET.get("page", 1)

    franchises_qs = franchises_qs.annotate(
        rating_agg=ExpressionWrapper(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            ),
            output_field=FloatField()
        ),
        rating_bucket=Floor(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            )
        ),
    )


    if selected_categories:
        try:
            selected_ids = [int(cid) for cid in selected_categories]
        except Exception:
            selected_ids = []
        if selected_ids:
            franchises_qs = franchises_qs.filter(direction_id__in=selected_ids)

    if search_query:
        franchises_qs = franchises_qs.filter(title__icontains=search_query)

    try:
        min_payback = int(min_payback_str)
        max_payback = int(max_payback_str)
        if min_payback > 0:
            franchises_qs = franchises_qs.filter(payback_period__gte=min_payback)
        if max_payback < 60:
            franchises_qs = franchises_qs.filter(payback_period__lte=max_payback)
    except ValueError:
        min_payback = 0
        max_payback = 60

    try:
        min_investment = int(min_investment_str)
        max_investment = int(max_investment_str)
        if max_investment <= 0 or max_investment < min_investment:
            max_investment = 10000000
        if min_investment < 0:
            min_investment = 0
        if min_investment > 0:
            franchises_qs = franchises_qs.filter(investment_size__gte=min_investment)
        if max_investment < 10000000:
            franchises_qs = franchises_qs.filter(investment_size__lte=max_investment)
    except ValueError:
        min_investment = 0
        max_investment = 10000000

    try:
        min_rating = float(min_rating_str)
        max_rating = float(max_rating_str)
        if min_rating > 0:
            franchises_qs = franchises_qs.filter(rating_agg__gte=min_rating)
        if max_rating < 5:
            franchises_qs = franchises_qs.filter(rating_agg__lte=max_rating)
    except ValueError:
        min_rating = 0
        max_rating = 5

    # City filter
    from .models import FranchiseLocation, City
    selected_city = request.GET.get("city", "").strip()
    if selected_city:
        franchise_ids_in_city = FranchiseLocation.objects.filter(
            city__slug=selected_city, status="active"
        ).values_list("franchise_id", flat=True)
        franchises_qs = franchises_qs.filter(franchise_id__in=franchise_ids_in_city)

    filters_active = (
        bool(selected_categories) or
        (search_query != "") or
        (min_payback > 0) or (max_payback < 60) or
        (min_investment > 0) or (max_investment < 10000000) or
        (min_rating > 0) or (max_rating < 5) or
        bool(selected_city)
    )
    rating_active = (min_rating > 0 or max_rating < 5)
    payback_active = (min_payback > 0 or max_payback < 60)
    investment_active = (min_investment > 0 or max_investment < 10000000)
    if payback_active:
        franchises_qs = franchises_qs.order_by("payback_period", "investment_size", "rating_bucket", "rating_agg", "-created_at")
    elif investment_active:
        franchises_qs = franchises_qs.order_by("investment_size", "rating_bucket", "rating_agg", "-created_at")
    elif rating_active:
        franchises_qs = franchises_qs.order_by("rating_bucket", "rating_agg", "-created_at")
    elif filters_active:
        franchises_qs = franchises_qs.order_by("-created_at")
    else:
        if sort_order == "newest":
            franchises_qs = franchises_qs.order_by("-created_at")
        elif sort_order == "oldest":
            franchises_qs = franchises_qs.order_by("created_at")
        elif sort_order == "popular":
            from .models import AnalyticsDailyStat
            from django.db.models import Sum, Subquery, OuterRef
            views_subq = (
                AnalyticsDailyStat.objects
                .filter(entity_type="franchise", entity_id=OuterRef("franchise_id"))
                .values("entity_id")
                .annotate(total=Sum("total_views"))
                .values("total")[:1]
            )
            franchises_qs = franchises_qs.annotate(
                popularity=Coalesce(Subquery(views_subq, output_field=IntegerField()), Value(0))
            ).order_by("-popularity", "-created_at")
        elif sort_order == "rating":
            franchises_qs = franchises_qs.order_by("-rating_agg", "-total_voters", "-created_at")

    

    from accounts.promotions import get_pinned_entities, get_active_ads
    pinned_franchises = []
    page_num_int = int(page_number) if str(page_number).isdigit() else 1
    if page_num_int == 1 and not filters_active:
        pinned_franchises = get_pinned_entities("franchise", Franchises, "franchise_id")
        pinned_ids = [f.franchise_id for f in pinned_franchises]
        if pinned_ids:
            franchises_qs = franchises_qs.exclude(franchise_id__in=pinned_ids)

    paginator = Paginator(franchises_qs, 6)
    page_obj = paginator.get_page(page_number)

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string(
            "accounts/partials/_franchise_cards.html", {"page_obj": page_obj}
        )
        return JsonResponse(
            {
                "html": html,
                "has_next": page_obj.has_next(),
                "page_number": page_obj.number,
                "num_pages": paginator.num_pages,
                "count": paginator.count,
            }
        )
    else:
        context = {
            "page_obj": page_obj,
            "paginator": paginator,
            "initial_has_next": page_obj.has_next(),
            "selected_categories": selected_categories,
            "search_query": search_query,
            "min_rating": min_rating,
            "max_rating": max_rating,
            "min_payback": min_payback,
            "max_payback": max_payback,
            "min_investment": min_investment,
            "max_investment": max_investment,
            "sort_order": sort_order,
            "franchise_directions": franchise_directions,
            "pinned_items": pinned_franchises,
            "sidebar_ads": get_active_ads("catalog_sidebar"),
            "selected_city": selected_city,
            "available_cities": City.objects.filter(
                franchise_locations__status="active"
            ).distinct().order_by("name"),
        }
        return render(request, "accounts/franchises_list.html", context)
@vary_on_headers('X-Requested-With')
def agencies_list(request):
    # Используем distinct() для предотвращения дубликатов (проблема с PRIMARY KEY в таблице agencies)
    agencies_qs = Agencies.objects.filter(status="approved").select_related("owner", "direction").distinct()
    agency_categories = [
        "Веб-разработка",
        "Мобильная разработка",
        "Дизайн",
        "Маркетинг",
        "ИИ",
        "Брендинг",
        "Видео и мультимедиа",
        "Перевод",
    ]

    selected_categories = request.GET.getlist("category")
    search_query = request.GET.get("search", "").strip()
    min_rating_str = request.GET.get("min_rating", "0")
    max_rating_str = request.GET.get("max_rating", "5")
    sort_order = request.GET.get("sort_order", "newest")
    page_number = request.GET.get("page", 1)

    agencies_qs = agencies_qs.annotate(
        rating_agg=ExpressionWrapper(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            ),
            output_field=FloatField()
        ),
        rating_bucket=Floor(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            )
        ),
    )

    if selected_categories:
        agencies_qs = agencies_qs.filter(
            Q(customization_data__agency_category__in=selected_categories)
        )

    if search_query:
        agencies_qs = agencies_qs.filter(title__icontains=search_query)

    try:
        min_rating = float(min_rating_str)
        max_rating = float(max_rating_str)
        if min_rating > 0:
            agencies_qs = agencies_qs.filter(rating_agg__gte=min_rating)
        if max_rating < 5:
            agencies_qs = agencies_qs.filter(rating_agg__lte=max_rating)
    except ValueError:
        min_rating = 0
        max_rating = 5

    filters_active = (
        bool(selected_categories) or
        (search_query != "") or
        (min_rating > 0) or (max_rating < 5)
    )
    rating_active = (min_rating > 0 or max_rating < 5)
    if rating_active:
        agencies_qs = agencies_qs.order_by("rating_bucket", "rating_agg", "-created_at")
    elif filters_active:
        agencies_qs = agencies_qs.order_by("-created_at")
    else:
        if sort_order == "newest":
            agencies_qs = agencies_qs.order_by("-created_at")
        elif sort_order == "oldest":
            agencies_qs = agencies_qs.order_by("created_at")

    from accounts.promotions import get_pinned_entities, get_active_ads
    pinned_agencies = []
    page_num_int = int(page_number) if str(page_number).isdigit() else 1
    if page_num_int == 1 and not filters_active:
        pinned_agencies = get_pinned_entities("agency", Agencies, "agency_id")
        pinned_ids = [a.agency_id for a in pinned_agencies]
        if pinned_ids:
            agencies_qs = agencies_qs.exclude(agency_id__in=pinned_ids)

    paginator = Paginator(agencies_qs, 6)
    page_obj = paginator.get_page(page_number)

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string(
            "accounts/partials/_agency_cards.html", {"page_obj": page_obj}
        )
        return JsonResponse(
            {
                "html": html,
                "has_next": page_obj.has_next(),
                "page_number": page_obj.number,
                "num_pages": paginator.num_pages,
                "count": paginator.count,
            }
        )
    else:
        context = {
            "page_obj": page_obj,
            "paginator": paginator,
            "initial_has_next": page_obj.has_next(),
            "selected_categories": selected_categories,
            "search_query": search_query,
            "min_rating": min_rating,
            "max_rating": max_rating,
            "sort_order": sort_order,
            "agency_categories": agency_categories,
            "pinned_items": pinned_agencies,
            "sidebar_ads": get_active_ads("catalog_sidebar"),
        }
        return render(request, "accounts/agencies_list.html", context)

@vary_on_headers('X-Requested-With')
def specialists_list(request):
    specialists_qs = Specialists.objects.filter(status="approved").select_related("owner", "direction")
    specialist_categories = [
        "Веб-разработка",
        "Мобильная разработка",
        "Дизайн",
        "Маркетинг",
        "ИИ",
        "Брендинг",
        "Видео и мультимедиа",
        "Перевод",
    ]

    selected_categories = request.GET.getlist("category")
    search_query = request.GET.get("search", "").strip()
    min_rating_str = request.GET.get("min_rating", "0")
    max_rating_str = request.GET.get("max_rating", "5")
    sort_order = request.GET.get("sort_order", "newest")
    page_number = request.GET.get("page", 1)

    specialists_qs = specialists_qs.annotate(
        rating_agg=ExpressionWrapper(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            ),
            output_field=FloatField()
        ),
        rating_bucket=Floor(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            )
        ),
    )

    if selected_categories:
        specialists_qs = specialists_qs.filter(
            Q(customization_data__specialist_category__in=selected_categories)
        )

    if search_query:
        specialists_qs = specialists_qs.filter(title__icontains=search_query)

    try:
        min_rating = float(min_rating_str)
        max_rating = float(max_rating_str)
        if min_rating > 0:
            specialists_qs = specialists_qs.filter(rating_agg__gte=min_rating)
        if max_rating < 5:
            specialists_qs = specialists_qs.filter(rating_agg__lte=max_rating)
    except ValueError:
        min_rating = 0
        max_rating = 5

    filters_active = (
        bool(selected_categories) or
        (search_query != "") or
        (min_rating > 0) or (max_rating < 5)
    )
    rating_active = (min_rating > 0 or max_rating < 5)
    if rating_active:
        specialists_qs = specialists_qs.order_by("rating_bucket", "rating_agg", "-created_at")
    elif filters_active:
        specialists_qs = specialists_qs.order_by("-created_at")
    else:
        if sort_order == "newest":
            specialists_qs = specialists_qs.order_by("-created_at")
        elif sort_order == "oldest":
            specialists_qs = specialists_qs.order_by("created_at")

    from accounts.promotions import get_pinned_entities, get_active_ads
    pinned_specialists = []
    page_num_int = int(page_number) if str(page_number).isdigit() else 1
    if page_num_int == 1 and not filters_active:
        pinned_specialists = get_pinned_entities("specialist", Specialists, "specialist_id")
        pinned_ids = [s.specialist_id for s in pinned_specialists]
        if pinned_ids:
            specialists_qs = specialists_qs.exclude(specialist_id__in=pinned_ids)

    paginator = Paginator(specialists_qs, 6)
    page_obj = paginator.get_page(page_number)

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string(
            "accounts/partials/_specialist_cards.html", {"page_obj": page_obj}
        )
        return JsonResponse(
            {
                "html": html,
                "has_next": page_obj.has_next(),
                "page_number": page_obj.number,
                "num_pages": paginator.num_pages,
                "count": paginator.count,
            }
        )
    else:
        context = {
            "page_obj": page_obj,
            "paginator": paginator,
            "initial_has_next": page_obj.has_next(),
            "selected_categories": selected_categories,
            "search_query": search_query,
            "min_rating": min_rating,
            "max_rating": max_rating,
            "sort_order": sort_order,
            "specialist_categories": specialist_categories,
            "pinned_items": pinned_specialists,
            "sidebar_ads": get_active_ads("catalog_sidebar"),
        }
        return render(request, "accounts/specialists_list.html", context)
def agency_detail_by_id(request, agency_id):
    agency = get_object_or_404(Agencies, agency_id=agency_id)
    if agency.slug:
        return redirect("agency_detail", slug=agency.slug, permanent=True)
    return agency_detail(request, slug=str(agency_id))

def agency_detail(request, slug):
    agency = Agencies.objects.filter(slug=slug).first()
    if not agency:
        try:
            agency = Agencies.objects.filter(agency_id=slug).first()
            if agency and agency.slug:
                return redirect("agency_detail", slug=agency.slug, permanent=True)
        except ValueError:
            pass
    if not agency:
        return render(request, "accounts/404.html", status=404)

    if request.method == "POST":
        if "status" in request.POST:
            if not is_moderator(request.user):
                messages.error(request, "У вас нет прав для этого действия.")
                return redirect("agency_detail", slug=agency.slug or agency.agency_id)
            new_status = (request.POST.get("status", "") or "").strip().lower()
            allowed_statuses = {"approved", "blocked", "closed", "pending", "rejected"}
            if new_status in allowed_statuses:
                agency.status = new_status
                agency.save(update_fields=["status"])
                messages.success(request, "Статус агентства обновлён.")
            else:
                messages.error(request, "Недопустимый статус.")
            return redirect("agency_detail", slug=agency.slug or agency.agency_id)
        if not request.user.is_authenticated:
            return redirect("login")
        form = AgencyCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.agency = agency
            comment.user = request.user

            try:
                new_rating = int(form.cleaned_data.get("user_rating") or 0)
            except (TypeError, ValueError):
                new_rating = 0
            user_vote = AgencyVotes.objects.filter(user=request.user, agency=agency).first()
            if 1 <= new_rating <= 5:
                comment.user_rating = new_rating
                with transaction.atomic():
                    if user_vote:
                        if user_vote.rating != new_rating:
                            old_rating = int(user_vote.rating or 0)
                            user_vote.rating = new_rating
                            user_vote.save(update_fields=["rating"])
                            Agencies.objects.filter(agency_id=agency.agency_id).update(
                                sum_votes=models.F("sum_votes") + (new_rating - old_rating)
                            )
                            agency.refresh_from_db(fields=["sum_votes"])
                    else:
                        AgencyVotes.objects.create(user=request.user, agency=agency, rating=new_rating)
                        Agencies.objects.filter(agency_id=agency.agency_id).update(
                            total_voters=models.F("total_voters") + 1,
                            sum_votes=models.F("sum_votes") + new_rating,
                        )
                        agency.refresh_from_db(fields=["total_voters", "sum_votes"])
            else:
                if user_vote:
                    comment.user_rating = user_vote.rating

            comment.save()
            messages.success(request, "Ваш комментарий был добавлен.")
            return redirect("agency_detail", slug=agency.slug or agency.agency_id)
        else:
            messages.error(request, "Ошибка при добавлении комментария.")
    else:
        form = AgencyCommentForm()

    agency_category = None
    try:
        agency_category = (agency.customization_data or {}).get("agency_category")
    except Exception:
        agency_category = None

    if agency_category:
        candidates_qs = Agencies.objects.filter(
            customization_data__agency_category=agency_category,
            status="approved",
        ).exclude(agency_id=agency.agency_id)
    else:
        candidates_qs = Agencies.objects.filter(
            status="approved",
        ).exclude(agency_id=agency.agency_id)
    similar_agencies = candidates_qs.order_by("-created_at")[:4]

    comments_with_rating = (
        AgencyComments.objects.filter(agency=agency, parent_comment__isnull=True)
        .annotate(
            user_vote_rating=models.Subquery(
                AgencyVotes.objects.filter(
                    agency=agency, user=models.OuterRef("user_id")
                ).values("rating")[:1]
            )
        )
        .order_by("-created_at")
    )
    average_rating = agency.get_average_rating()
    total_votes = agency.total_voters
    total_comments_count = AgencyComments.objects.filter(agency=agency, parent_comment__isnull=True).count()
    user_has_voted = False
    if request.user.is_authenticated:
        user_has_voted = AgencyVotes.objects.filter(
            user=request.user, agency=agency
        ).exists()
    rating_distribution_query = (
        AgencyVotes.objects.filter(agency=agency)
        .values("rating")
        .annotate(count=Count("rating"))
        .order_by("-rating")
    )
    rating_distribution = {item["rating"]: item["count"] for item in rating_distribution_query}
    for i in range(1, 6):
        rating_distribution.setdefault(i, 0)

    logo_urls = agency.logo_urls if isinstance(agency.logo_urls, list) else []
    creatives_urls = (
        agency.creatives_urls if isinstance(agency.creatives_urls, list) else []
    )
    slider_images = agency.slider_images if isinstance(agency.slider_images, list) else []
    # Используем slider_images напрямую, если пусто - берём первые 4 creatives
    if not slider_images:
        slider_images = creatives_urls[:4]
    all_creatives = creatives_urls
    video_urls = agency.video_urls if isinstance(agency.video_urls, list) else []
    proofs_urls = agency.proofs_urls if isinstance(agency.proofs_urls, list) else []

    # Получаем синхронизированные документы (только те, что есть в редакторе)
    agency_documents = get_synced_files(agency, "proof", "proofs_urls")

    # Batch-resolve file URLs
    from .utils import batch_resolve_file_urls
    file_url_map = {}
    file_url_map.update(batch_resolve_file_urls(logo_urls, agency.agency_id, "logo", "agency"))
    file_url_map.update(batch_resolve_file_urls(all_creatives, agency.agency_id, "creative", "agency"))
    file_url_map.update(batch_resolve_file_urls(video_urls, agency.agency_id, "video", "agency"))
    proof_ids = [d.file_url for d in agency_documents] if agency_documents else []
    file_url_map.update(batch_resolve_file_urls(proof_ids, agency.agency_id, "proof", "agency"))

    ai_rating = None
    if agency.customization_data and isinstance(agency.customization_data, dict):
        ai_rating = agency.customization_data.get('ai_rating', 5)
    else:
        ai_rating = 5

    context = {
        "agency": agency,
        "similar_agencies": similar_agencies,
        "has_similar": candidates_qs.exists(),
        "comments": comments_with_rating,
        "form": form,
        "average_rating": average_rating,
        "total_votes_count": total_votes,
        "total_comments_count": total_comments_count,
        "user_has_voted": user_has_voted,
        "rating_distribution": rating_distribution,
        "logo_urls": logo_urls,
        "creatives_urls": creatives_urls,
        "slider_images": slider_images,
        "all_creatives": all_creatives,
        "video_urls": video_urls,
        "proofs_urls": proofs_urls,
        "agency_documents": agency_documents,
        "ai_rating": ai_rating,
        "canonical_url": request.build_absolute_uri(),
        "file_url_map": file_url_map,
    }
    return render(request, "accounts/agency_detail.html", context)

def specialist_detail_by_id(request, specialist_id):
    specialist = get_object_or_404(Specialists, specialist_id=specialist_id)
    if specialist.slug:
        return redirect("specialist_detail", slug=specialist.slug, permanent=True)
    return specialist_detail(request, slug=str(specialist_id))

def specialist_detail(request, slug):
    specialist = Specialists.objects.filter(slug=slug).first()
    if not specialist:
        try:
            specialist = Specialists.objects.filter(specialist_id=slug).first()
            if specialist and specialist.slug:
                return redirect("specialist_detail", slug=specialist.slug, permanent=True)
        except ValueError:
            pass
    if not specialist:
        return render(request, "accounts/404.html", status=404)

    if request.method == "POST":
        if "status" in request.POST:
            if not is_moderator(request.user):
                messages.error(request, "У вас нет прав для этого действия.")
                return redirect("specialist_detail", slug=specialist.slug or specialist.specialist_id)
            new_status = (request.POST.get("status", "") or "").strip().lower()
            allowed_statuses = {"approved", "blocked", "closed", "pending", "rejected"}
            if new_status in allowed_statuses:
                specialist.status = new_status
                specialist.save(update_fields=["status"])
                messages.success(request, "Статус специалиста обновлён.")
            else:
                messages.error(request, "Недопустимый статус.")
            return redirect("specialist_detail", slug=specialist.slug or specialist.specialist_id)
        if not request.user.is_authenticated:
            return redirect("login")
        form = SpecialistCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.specialist = specialist
            comment.user = request.user

            try:
                new_rating = int(form.cleaned_data.get("user_rating") or 0)
            except (TypeError, ValueError):
                new_rating = 0
            user_vote = SpecialistVotes.objects.filter(user=request.user, specialist=specialist).first()
            if 1 <= new_rating <= 5:
                comment.user_rating = new_rating
                with transaction.atomic():
                    if user_vote:
                        if user_vote.rating != new_rating:
                            old_rating = int(user_vote.rating or 0)
                            user_vote.rating = new_rating
                            user_vote.save(update_fields=["rating"])
                            Specialists.objects.filter(specialist_id=specialist.specialist_id).update(
                                sum_votes=models.F("sum_votes") + (new_rating - old_rating)
                            )
                            specialist.refresh_from_db(fields=["sum_votes"])
                    else:
                        SpecialistVotes.objects.create(user=request.user, specialist=specialist, rating=new_rating)
                        Specialists.objects.filter(specialist_id=specialist.specialist_id).update(
                            total_voters=models.F("total_voters") + 1,
                            sum_votes=models.F("sum_votes") + new_rating,
                        )
                        specialist.refresh_from_db(fields=["total_voters", "sum_votes"])
            else:
                if user_vote:
                    comment.user_rating = user_vote.rating

            comment.save()
            messages.success(request, "Ваш комментарий был добавлен.")
            return redirect("specialist_detail", slug=specialist.slug or specialist.specialist_id)
        else:
            messages.error(request, "Ошибка при добавлении комментария.")
    else:
        form = SpecialistCommentForm()

    specialist_category = None
    try:
        specialist_category = (specialist.customization_data or {}).get("specialist_category")
    except Exception:
        specialist_category = None

    if specialist_category:
        candidates_qs = Specialists.objects.filter(
            customization_data__specialist_category=specialist_category,
            status="approved",
        ).exclude(specialist_id=specialist.specialist_id)
    else:
        candidates_qs = Specialists.objects.filter(
            status="approved",
        ).exclude(specialist_id=specialist.specialist_id)
    similar_specialists = candidates_qs.order_by("-created_at")[:4]

    comments_with_rating = (
        SpecialistComments.objects.filter(specialist=specialist, parent_comment__isnull=True)
        .annotate(
            user_vote_rating=models.Subquery(
                SpecialistVotes.objects.filter(
                    specialist=specialist, user=models.OuterRef("user_id")
                ).values("rating")[:1]
            )
        )
        .order_by("-created_at")
    )
    average_rating = specialist.get_average_rating()
    total_votes = specialist.total_voters
    total_comments_count = SpecialistComments.objects.filter(specialist=specialist, parent_comment__isnull=True).count()
    user_has_voted = False
    if request.user.is_authenticated:
        user_has_voted = SpecialistVotes.objects.filter(
            user=request.user, specialist=specialist
        ).exists()
    rating_distribution_query = (
        SpecialistVotes.objects.filter(specialist=specialist)
        .values("rating")
        .annotate(count=Count("rating"))
        .order_by("-rating")
    )
    rating_distribution = {item["rating"]: item["count"] for item in rating_distribution_query}
    for i in range(1, 6):
        rating_distribution.setdefault(i, 0)

    logo_urls = specialist.logo_urls if isinstance(specialist.logo_urls, list) else []
    creatives_urls = (
        specialist.creatives_urls if isinstance(specialist.creatives_urls, list) else []
    )
    slider_images = specialist.slider_images if isinstance(specialist.slider_images, list) else []
    all_creatives = creatives_urls
    video_urls = specialist.video_urls if isinstance(specialist.video_urls, list) else []
    proofs_urls = specialist.proofs_urls if isinstance(specialist.proofs_urls, list) else []

    # Получаем синхронизированные документы (только те, что есть в редакторе)
    specialist_documents = get_synced_files(specialist, "proof", "proofs_urls")

    # Batch-resolve file URLs
    from .utils import batch_resolve_file_urls
    file_url_map = {}
    file_url_map.update(batch_resolve_file_urls(logo_urls, specialist.specialist_id, "logo", "specialist"))
    file_url_map.update(batch_resolve_file_urls(all_creatives, specialist.specialist_id, "creative", "specialist"))
    file_url_map.update(batch_resolve_file_urls(video_urls, specialist.specialist_id, "video", "specialist"))
    proof_ids = [d.file_url for d in specialist_documents] if specialist_documents else []
    file_url_map.update(batch_resolve_file_urls(proof_ids, specialist.specialist_id, "proof", "specialist"))

    ai_rating = None
    if specialist.customization_data and isinstance(specialist.customization_data, dict):
        ai_rating = specialist.customization_data.get('ai_rating', 5)
    else:
        ai_rating = 5

    context = {
        "specialist": specialist,
        "similar_specialists": similar_specialists,
        "has_similar": candidates_qs.exists(),
        "comments": comments_with_rating,
        "form": form,
        "average_rating": average_rating,
        "total_votes_count": total_votes,
        "total_comments_count": total_comments_count,
        "user_has_voted": user_has_voted,
        "rating_distribution": rating_distribution,
        "logo_urls": logo_urls,
        "creatives_urls": creatives_urls,
        "slider_images": slider_images,
        "all_creatives": all_creatives,
        "video_urls": video_urls,
        "proofs_urls": proofs_urls,
        "specialist_documents": specialist_documents,
        "ai_rating": ai_rating,
        "canonical_url": request.build_absolute_uri(),
        "file_url_map": file_url_map,
    }
    return render(request, "accounts/specialist_detail.html", context)
def franchise_detail_by_id(request, franchise_id):
    franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
    if franchise.slug:
        return redirect("franchise_detail", slug=franchise.slug, permanent=True)
    return franchise_detail(request, slug=str(franchise_id))

def franchise_detail(request, slug):
    try:
        franchise = Franchises.objects.get(slug=slug)
    except Franchises.DoesNotExist:
        try:
            franchise = Franchises.objects.get(franchise_id=slug)
            if franchise.slug:
                return redirect("franchise_detail", slug=franchise.slug, permanent=True)
        except (Franchises.DoesNotExist, ValueError):
            return render(request, "accounts/404.html", status=404)

    if request.method == "POST":
        if "status" in request.POST:
            if not is_moderator(request.user):
                messages.error(request, "У вас нет прав для этого действия.")
                return redirect("franchise_detail", slug=franchise.slug or franchise.franchise_id)
            new_status = (request.POST.get("status", "") or "").strip().lower()
            allowed_statuses = {"approved", "blocked", "closed", "pending", "rejected"}
            if new_status in allowed_statuses:
                franchise.status = new_status
                try:
                    status_map = {
                        "approved": "Approved",
                        "blocked": "Blocked",
                        "closed": "Closed",
                        "pending": "Pending",
                        "rejected": "Rejected",
                    }
                    franchise.status_id = ReviewStatuses.objects.get(status_name=status_map[new_status])
                except ReviewStatuses.DoesNotExist:
                    pass
                franchise.save(update_fields=["status", "status_id"])
                messages.success(request, "Статус франшизы обновлён.")
            else:
                messages.error(request, "Недопустимый статус.")
            return redirect("franchise_detail", slug=franchise.slug or franchise.franchise_id)
        if not request.user.is_authenticated:
            return redirect("login")
        form = FranchiseCommentForm(request.POST)
        if form.is_valid():
            from .models import FranchiseComments
            comment = form.save(commit=False)
            comment.franchise = franchise
            comment.user = request.user
            try:
                new_rating = int(form.cleaned_data.get("user_rating") or 0)
            except (TypeError, ValueError):
                new_rating = 0
            user_vote = FranchiseVotes.objects.filter(user=request.user, franchise=franchise).first()
            if 1 <= new_rating <= 5:
                comment.user_rating = new_rating
                with transaction.atomic():
                    if user_vote:
                        if user_vote.rating != new_rating:
                            old_rating = int(user_vote.rating or 0)
                            user_vote.rating = new_rating
                            user_vote.save(update_fields=["rating"])
                            Franchises.objects.filter(franchise_id=franchise.franchise_id).update(
                                sum_votes=models.F("sum_votes") + (new_rating - old_rating)
                            )
                            franchise.refresh_from_db(fields=["sum_votes"])
                    else:
                        FranchiseVotes.objects.create(user=request.user, franchise=franchise, rating=new_rating)
                        Franchises.objects.filter(franchise_id=franchise.franchise_id).update(
                            total_voters=models.F("total_voters") + 1,
                            sum_votes=models.F("sum_votes") + new_rating,
                        )
                        franchise.refresh_from_db(fields=["total_voters", "sum_votes"])
            else:
                if user_vote:
                    comment.user_rating = user_vote.rating
            comment.save()
            messages.success(request, "Ваш комментарий был добавлен.")
            return redirect("franchise_detail", slug=franchise.slug or franchise.franchise_id)
        else:
            messages.error(request, "Ошибка при добавлении комментария.")
    else:
        form = FranchiseCommentForm()

    candidates_qs = Franchises.objects.filter(
        direction=franchise.direction,
        status="approved",
    ).exclude(franchise_id=franchise.franchise_id)
    similar_franchises = candidates_qs.order_by("-created_at")[:4]

    from .models import FranchiseComments
    comments_with_rating = (
        FranchiseComments.objects.filter(franchise=franchise, parent_comment__isnull=True)
        .annotate(
            user_vote_rating=models.Subquery(
                FranchiseVotes.objects.filter(
                    franchise=franchise, user=models.OuterRef("user_id")
                ).values("rating")[:1]
            )
        )
        .order_by("-created_at")
    )
    average_rating = franchise.get_average_rating()
    total_votes = franchise.total_voters
    total_comments_count = FranchiseComments.objects.filter(franchise=franchise, parent_comment__isnull=True).count()
    user_has_voted = False
    if request.user.is_authenticated:
        user_has_voted = FranchiseVotes.objects.filter(
            user=request.user, franchise=franchise
        ).exists()
    rating_distribution_query = (
        FranchiseVotes.objects.filter(franchise=franchise)
        .values("rating")
        .annotate(count=Count("rating"))
        .order_by("-rating")
    )
    rating_distribution = {item["rating"]: item["count"] for item in rating_distribution_query}
    for i in range(1, 6):
        rating_distribution.setdefault(i, 0)

    logo_urls = franchise.logo_urls if isinstance(franchise.logo_urls, list) else []
    creatives_urls = (
        franchise.creatives_urls if isinstance(franchise.creatives_urls, list) else []
    )
    slider_images = franchise.slider_images if isinstance(franchise.slider_images, list) else []
    all_creatives = creatives_urls
    video_urls = franchise.video_urls if isinstance(franchise.video_urls, list) else []
    proofs_urls = franchise.proofs_urls if isinstance(franchise.proofs_urls, list) else []

    # Получаем синхронизированные документы (только те, что есть в редакторе)
    franchise_documents = get_synced_files(franchise, "proof", "proofs_urls")

    # Batch-resolve file URLs
    from .utils import batch_resolve_file_urls
    file_url_map = {}
    file_url_map.update(batch_resolve_file_urls(logo_urls, franchise.franchise_id, "logo", "franchise"))
    file_url_map.update(batch_resolve_file_urls(all_creatives, franchise.franchise_id, "creative", "franchise"))
    file_url_map.update(batch_resolve_file_urls(video_urls, franchise.franchise_id, "video", "franchise"))
    proof_ids = [d.file_url for d in franchise_documents] if franchise_documents else []
    file_url_map.update(batch_resolve_file_urls(proof_ids, franchise.franchise_id, "proof", "franchise"))

    ai_rating = None
    if franchise.customization_data and isinstance(franchise.customization_data, dict):
        ai_rating = franchise.customization_data.get('ai_rating', 5)
    else:
        ai_rating = 5

    progress_percentage = 0
    if franchise.investment_size and franchise.investment_size > 0:
        from django.db.models import Sum as _Sum
        try:
            current_total = InvestmentTransactions.objects.filter(franchise=franchise).aggregate(total=_Sum("amount")).get("total") or Decimal("0")
        except Exception:
            current_total = franchise.total_invested or Decimal("0")
        goal = Decimal(franchise.investment_size) if franchise.investment_size is not None else Decimal("0")
        if goal > 0:
            progress_percentage = (current_total * Decimal("100")) / goal
        else:
            progress_percentage = Decimal("0")
        if progress_percentage < 0:
            progress_percentage = 0
        if progress_percentage > 100:
            progress_percentage = 100
        progress_percentage = float(progress_percentage)
    investors_count = franchise.get_investors_count()

    # Geography — franchise locations
    from .models import FranchiseLocation
    franchise_locations = (
        FranchiseLocation.objects.filter(franchise=franchise)
        .select_related("city")
        .order_by("city__name")
    )
    locations_by_region = {}
    for loc in franchise_locations:
        region = loc.city.get_region_display() if loc.city.region else "Другой"
        locations_by_region.setdefault(region, []).append(loc)

    # Aggregate stats for calculator
    active_locations = [l for l in franchise_locations if l.status == 'active']
    avg_monthly_profit = None
    avg_payback = None
    if active_locations:
        profits = [float(l.monthly_profit) for l in active_locations if l.monthly_profit]
        if profits:
            avg_monthly_profit = int(sum(profits) / len(profits))
        paybacks = [l.get_payback_months() for l in active_locations]
        paybacks = [p for p in paybacks if p]
        if paybacks:
            avg_payback = int(sum(paybacks) / len(paybacks))

    context = {
        "franchise": franchise,
        "similar_franchises": similar_franchises,
        "has_similar": candidates_qs.exists(),
        "comments": comments_with_rating,
        "form": form,
        "average_rating": average_rating,
        "total_votes_count": total_votes,
        "total_comments_count": total_comments_count,
        "user_has_voted": user_has_voted,
        "rating_distribution": rating_distribution,
        "logo_urls": logo_urls,
        "creatives_urls": creatives_urls,
        "slider_images": slider_images,
        "all_creatives": all_creatives,
        "video_urls": video_urls,
        "proofs_urls": proofs_urls,
        "franchise_documents": franchise_documents,
        "ai_rating": ai_rating,
        "progress_percentage": progress_percentage,
        "investors_count": investors_count,
        "canonical_url": request.build_absolute_uri(),
        "file_url_map": file_url_map,
        "franchise_locations": franchise_locations,
        "locations_by_region": locations_by_region,
        "locations_count": len(active_locations),
        "avg_monthly_profit": avg_monthly_profit,
        "avg_payback": avg_payback,
        "cities": City.objects.all().order_by("name"),
        # Перелинковка: статьи по теме
        "related_articles": NewsArticles.objects.filter(
            status="published", entity_focus="franchise"
        ).order_by("-created_at")[:4],
        # Перелинковка: города присутствия этой франшизы (для ссылок на городские страницы)
        "franchise_cities": City.objects.filter(
            franchise_locations__franchise=franchise,
            franchise_locations__status="active",
        ).distinct().order_by("name")[:12],
    }
    return render(request, "accounts/franchise_detail.html", context)

def search_suggestions(request):
    query = request.GET.get("q", "").strip()
    users = []
    if len(query) >= 2:
        search_results = Users.objects.filter(
            Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(email__icontains=query)
        ).distinct()[:10]
        users = [
            {
                "id": user.user_id,
                "name": f"{user.first_name or ''} {user.last_name or ''} ({user.email})".strip(),
            }
            for user in search_results
        ]
    return JsonResponse({"suggestions": users})

def global_search(request):
    """Глобальный поиск по всем типам карточек"""
    try:
        query = request.GET.get("q", "").strip()

        if len(query) < 2:
            return JsonResponse({
                "users": [],
                "startups": [],
                "franchises": [],
                "agencies": [],
                "specialists": []
            })

        results = {
            "users": [],
            "startups": [],
            "franchises": [],
            "agencies": [],
            "specialists": []
        }


        try:
            users = Users.objects.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(email__icontains=query)
            ).distinct()[:5]

            for user in users:
                try:
                    results["users"].append({
                        "id": user.user_id,
                        "name": f"{user.first_name or ''} {user.last_name or ''}".strip() or user.email,
                        "type": "user",
                        "url": reverse('profile', kwargs={'user_id': user.user_id})
                    })
                except Exception as e:
                    logger.error(f"Ошибка при обработке пользователя {user.user_id}: {e}")
                    continue
        except Exception as e:
            logger.error(f"Ошибка при поиске пользователей: {e}")


        try:

            try:

                approved_status = ReviewStatuses.objects.get(status_name="Approved")
                startups = Startups.objects.filter(
                    Q(title__icontains=query) |
                    Q(short_description__icontains=query)
                ).filter(status_id=approved_status).distinct()[:5]
            except ReviewStatuses.DoesNotExist:

                startups = Startups.objects.filter(
                    Q(title__icontains=query) |
                    Q(short_description__icontains=query)
                ).filter(status="approved").distinct()[:5]

            for startup in startups:
                try:

                    if not startup.title:
                        continue

                    results["startups"].append({
                        "id": startup.startup_id,
                        "name": startup.title,
                        "type": "startup",
                        "url": reverse('startup_detail', kwargs={'slug': startup.slug or startup.startup_id})
                    })
                except Exception as e:
                    logger.error(f"Ошибка при обработке стартапа {startup.startup_id}: {e}")
                    continue
        except Exception as e:
            logger.error(f"Ошибка при поиске стартапов: {e}")


        try:

            try:

                approved_status = ReviewStatuses.objects.get(status_name="Approved")
                franchises = Franchises.objects.filter(
                    Q(title__icontains=query) |
                    Q(short_description__icontains=query)
                ).filter(status_id=approved_status).distinct()[:5]
            except ReviewStatuses.DoesNotExist:

                franchises = Franchises.objects.filter(
                    Q(title__icontains=query) |
                    Q(short_description__icontains=query)
                ).filter(status="approved").distinct()[:5]

            for franchise in franchises:
                try:

                    if not franchise.title:
                        continue

                    results["franchises"].append({
                        "id": franchise.franchise_id,
                        "name": franchise.title,
                        "type": "franchise",
                        "url": reverse('franchise_detail', kwargs={'slug': franchise.slug or franchise.franchise_id})
                    })
                except Exception as e:
                    logger.error(f"Ошибка при обработке франшизы {franchise.franchise_id}: {e}")
                    continue
        except Exception as e:
            logger.error(f"Ошибка при поиске франшиз: {e}")


        try:

            agencies = Agencies.objects.filter(
                Q(title__icontains=query) |
                Q(short_description__icontains=query)
            ).filter(status="approved").distinct()[:5]

            for agency in agencies:
                try:

                    if not agency.title:
                        continue

                    results["agencies"].append({
                        "id": agency.agency_id,
                        "name": agency.title,
                        "type": "agency",
                        "url": reverse('agency_detail', kwargs={'slug': agency.slug or agency.agency_id})
                    })
                except Exception as e:
                    logger.error(f"Ошибка при обработке агентства {agency.agency_id}: {e}")
                    continue
        except Exception as e:
            logger.error(f"Ошибка при поиске агентств: {e}")


        try:

            specialists = Specialists.objects.filter(
                Q(title__icontains=query) |
                Q(short_description__icontains=query)
            ).filter(status="approved").distinct()[:5]

            for specialist in specialists:
                try:

                    if not specialist.title:
                        continue

                    results["specialists"].append({
                        "id": specialist.specialist_id,
                        "name": specialist.title,
                        "type": "specialist",
                        "url": reverse('specialist_detail', kwargs={'slug': specialist.slug or specialist.specialist_id})
                    })
                except Exception as e:
                    logger.error(f"Ошибка при обработке специалиста {specialist.specialist_id}: {e}")
                    continue
        except Exception as e:
            logger.error(f"Ошибка при поиске специалистов: {e}")

        return JsonResponse(results)

    except Exception as e:
        logger.error(f"Критическая ошибка в global_search: {e}")
        return JsonResponse({
            "error": "Ошибка при выполнении поиска",
            "details": str(e) if settings.DEBUG else "Внутренняя ошибка сервера"
        }, status=500)

def startup_detail_by_id(request, startup_id):
    startup = get_object_or_404(Startups, startup_id=startup_id)
    if startup.slug:
        return redirect("startup_detail", slug=startup.slug, permanent=True)
    return startup_detail(request, slug=str(startup_id))

def startup_detail(request, slug):
    try:
        startup = Startups.objects.select_related("owner", "direction", "stage").get(slug=slug)
    except Startups.DoesNotExist:
        # Fallback: try as ID for old links
        try:
            startup = Startups.objects.select_related("owner", "direction", "stage").get(startup_id=slug)
            if startup.slug:
                return redirect("startup_detail", slug=startup.slug, permanent=True)
        except (Startups.DoesNotExist, ValueError):
            raise Http404("Стартап не найден")
    if request.method == "POST":
        if "status" in request.POST:
            if not is_moderator(request.user):
                messages.error(request, "У вас нет прав для этого действия.")
                return redirect("startup_detail", slug=startup.slug or startup.startup_id)
            new_status = (request.POST.get("status", "") or "").strip().lower()
            allowed_statuses = {"approved", "blocked", "closed", "pending", "rejected"}
            if new_status in allowed_statuses:
                startup.status = new_status
                try:
                    status_map = {
                        "approved": "Approved",
                        "blocked": "Blocked",
                        "closed": "Closed",
                        "pending": "Pending",
                        "rejected": "Rejected",
                    }
                    startup.status_id = ReviewStatuses.objects.get(status_name=status_map[new_status])
                except ReviewStatuses.DoesNotExist:
                    pass
                startup.save(update_fields=["status", "status_id"])
                messages.success(request, "Статус стартапа обновлён.")
            else:
                messages.error(request, "Недопустимый статус.")
            return redirect("startup_detail", slug=startup.slug or startup.startup_id)
        if not request.user.is_authenticated:
            return redirect("login")
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.startup_id = startup
            comment.user_id = request.user

            try:
                new_rating = int(form.cleaned_data.get("user_rating") or 0)
            except (TypeError, ValueError):
                new_rating = 0
            from django.db import transaction
            with transaction.atomic():
                user_vote = UserVotes.objects.filter(user=request.user, startup=startup).first()
                if 1 <= new_rating <= 5:
                    comment.user_rating = new_rating
                    if user_vote:
                        if user_vote.rating != new_rating:
                            old_rating = int(user_vote.rating or 0)
                            user_vote.rating = new_rating
                            user_vote.save(update_fields=["rating"])
                            Startups.objects.filter(startup_id=startup.startup_id).update(
                                sum_votes=models.F("sum_votes") + (new_rating - old_rating)
                            )
                    else:
                        UserVotes.objects.create(user=request.user, startup=startup, rating=new_rating)
                        Startups.objects.filter(startup_id=startup.startup_id).update(
                            total_voters=models.F("total_voters") + 1,
                            sum_votes=models.F("sum_votes") + new_rating,
                        )
                else:
                    if user_vote:
                        comment.user_rating = user_vote.rating

                comment.save()
            messages.success(request, "Ваш комментарий был добавлен.")
            return redirect("startup_detail", slug=startup.slug or startup.startup_id)
        else:
            messages.error(request, "Ошибка при добавлении комментария.")
    else:
        form = CommentForm()
    comments_with_rating = (
        Comments.objects.filter(startup_id=startup, parent_comment_id__isnull=True)
        .select_related("user_id")
        .annotate(
            user_vote_rating=models.Subquery(
                UserVotes.objects.filter(
                    startup=startup, user=models.OuterRef("user_id_id")
                ).values("rating")[:1]
            )
        )
        .order_by("-created_at")
    )
    form = CommentForm()
    average_rating = startup.get_average_rating()
    total_votes = startup.total_voters
    total_comments_count = comments_with_rating.count()
    user_has_voted = False
    if request.user.is_authenticated:
        user_has_voted = UserVotes.objects.filter(
            user=request.user, startup=startup
        ).exists()
    rating_distribution_query = (
        UserVotes.objects.filter(startup=startup)
        .values("rating")
        .annotate(count=Count("rating"))
        .order_by("-rating")
    )
    rating_distribution = {
        item["rating"]: item["count"] for item in rating_distribution_query
    }
    for i in range(1, 6):
        rating_distribution.setdefault(i, 0)
    similar_startups = (
        Startups.objects.filter(status="approved")
        .select_related("owner", "direction")
        .exclude(startup_id=startup.startup_id)
        .order_by("?")[:4]
    )
    similar_startups = similar_startups.annotate(
        average_rating_calc=Avg(
            models.ExpressionWrapper(
                models.F("sum_votes") * 1.0 / models.F("total_voters"),
                output_field=FloatField(),
            ),
            filter=models.Q(total_voters__gt=0),
        )
    ).annotate(average_rating=Coalesce("average_rating_calc", 0.0))
    logo_urls = startup.logo_urls if isinstance(startup.logo_urls, list) else []
    creatives_urls = (
        startup.creatives_urls if isinstance(startup.creatives_urls, list) else []
    )
    slider_images = startup.slider_images if isinstance(startup.slider_images, list) else []
    all_creatives = creatives_urls
    video_urls = startup.video_urls if isinstance(startup.video_urls, list) else []

    # Batch-resolve all file URLs (1 S3 call per type instead of N calls per file)
    from .utils import batch_resolve_file_urls
    file_url_map = {}
    file_url_map.update(batch_resolve_file_urls(logo_urls, startup.startup_id, "logo", "startup"))
    file_url_map.update(batch_resolve_file_urls(all_creatives, startup.startup_id, "creative", "startup"))
    file_url_map.update(batch_resolve_file_urls(video_urls, startup.startup_id, "video", "startup"))
    show_moderator_comment = False
    if startup.moderator_comment and (
        request.user == startup.owner
        or (
            request.user.is_authenticated
            and hasattr(request.user, "role")
            and request.user.role.role_name == "moderator"
        )
    ):
        show_moderator_comment = True
    progress_percentage = startup.get_progress_percentage()
    investors_count = startup.get_investors_count()
    timeline_events = StartupTimeline.objects.filter(startup=startup).order_by(
        "step_number"
    )
    # Получаем синхронизированные документы (только те, что есть в редакторе)
    startup_documents = get_synced_files(startup, "proof", "proofs_urls")
    proof_ids = [d.file_url for d in startup_documents] if startup_documents else []
    file_url_map.update(batch_resolve_file_urls(proof_ids, startup.startup_id, "proof", "startup"))
    ai_rating = None
    if startup.customization_data and isinstance(startup.customization_data, dict):
        ai_rating = startup.customization_data.get('ai_rating', 5)
    else:
        ai_rating = 5
    context = {
        "startup": startup,
        "comments": comments_with_rating,
        "form": form,
        "average_rating": average_rating,
        "total_votes_count": total_votes,
        "total_comments_count": total_comments_count,
        "user_has_voted": user_has_voted,
        "rating_distribution": rating_distribution,
        "similar_startups": similar_startups,
        "logo_urls": logo_urls,
        "creatives_urls": creatives_urls,
        "slider_images": slider_images,
        "all_creatives": all_creatives,
        "video_urls": video_urls,
        "show_moderator_comment": show_moderator_comment,
        "progress_percentage": progress_percentage,
        "investors_count": investors_count,
        "timeline_events": timeline_events,
        "startup_documents": startup_documents,
        "ai_rating": ai_rating,
        "canonical_url": request.build_absolute_uri(),
        "file_url_map": file_url_map,
    }
    return render(request, "accounts/startup_detail.html", context)
def load_similar_startups(request, startup_id: int):
    current_startup_id = startup_id
    similar_startups = (
        Startups.objects.filter(status="approved")
        .exclude(startup_id=current_startup_id)
        .order_by("?")[:4]
    )
    similar_startups = similar_startups.annotate(
        average_rating_calc=Avg(
            models.ExpressionWrapper(
                models.F("sum_votes") * 1.0 / models.F("total_voters"),
                output_field=FloatField(),
            ),
            filter=models.Q(total_voters__gt=0),
        )
    ).annotate(average_rating=Coalesce("average_rating_calc", 0.0))
    html = render_to_string(
        "accounts/_similar_startup_cards.html",
        {"similar_startups": similar_startups, "request": request},
    )
    return HttpResponse(html)
@login_required
def investments(request):
    if not hasattr(request.user, "role") or request.user.role.role_name != "investor":
        messages.error(request, "Доступ к этой странице разрешен только инвесторам.")
        return redirect("profile")
    default_month_labels = [
        "Янв",
        "Фев",
        "Мар",
        "Апр",
        "Май",
        "Июн",
        "Июл",
        "Авг",
        "Сен",
        "Окт",
        "Ноя",
        "Дек",
    ]
    safe_context = {
        "startups_count": 0,
        "total_investment": Decimal("0"),
        "max_investment": Decimal("0"),
        "min_investment": Decimal("0"),
        "investment_categories": [],
        "month_labels": default_month_labels,
        "chart_monthly_category_data": [],
        "chart_categories": [],
        "all_directions": [],
        "invested_category_data": {},
        "user_investments": [],
        "user_owned_startups": [],
        "startup_applications": [],
        "current_sort": "newest",
        "planetary_investments": [],
        "planetary_investments_json": [],
        "investor_logo_url": request.user.get_profile_picture_url() or "https://via.placeholder.com/60",
    }
    try:
        base_tx_qs = InvestmentTransactions.objects.filter(
            investor=request.user
        )
        user_investments_qs = (
            base_tx_qs.filter(amount__gt=0)
            .select_related("startup", "startup__direction", "startup__owner")
            .defer("franchise")
        )
        logger.info(
            f"[investments] tx count for {request.user.email}: {user_investments_qs.count()} (base={base_tx_qs.count()})"
        )
        total_investment_data = user_investments_qs.aggregate(
            total_investment=Sum("amount"),
            max_investment=Max("amount"),
            startups_count=Count("startup", distinct=True),
        )
        total_investment = total_investment_data.get("total_investment") or Decimal("0")
        max_investment = total_investment_data.get("max_investment") or Decimal("0")

        investments_with_amount = user_investments_qs.filter(amount__gt=0)
        min_investment_data = investments_with_amount.aggregate(
            min_investment=Min("amount")
        )
        min_investment = min_investment_data.get("min_investment") or Decimal("0")
        startups_count = total_investment_data.get("startups_count", 0)
        logger.info(
            f"[investments] User: {request.user.email}, Total Investment: {total_investment}"
        )
        category_data_raw = (
            user_investments_qs.values("startup__direction__direction_name")
            .annotate(category_total=Sum("amount"))
            .order_by("-category_total")
        )
        investment_categories = []
        invested_category_data_dict = {}
        total_for_category_percentage = (
            total_investment if total_investment > 0 else Decimal("1")
        )
        for cat_data in category_data_raw:
            percentage = 0
            category_sum = cat_data.get("category_total")
            category_name = (
                cat_data.get("startup__direction__direction_name") or "Без категории"
            )
            if category_sum and total_for_category_percentage > 0:
                try:
                    percentage = round(
                        (Decimal(category_sum) / total_for_category_percentage) * 100
                    )
                    percentage = min(percentage, 100)
                except Exception as e:
                    logger.error(
                        f"Ошибка расчета процента для категории '{category_name}': {e}"
                    )
                    percentage = 0
            investment_categories.append(
                {"name": category_name, "percentage": percentage}
            )
            invested_category_data_dict[category_name] = percentage
        end_dt = timezone.now()
        start_dt = (end_dt - relativedelta(months=11)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        logger.info(
            f"[investments] Preparing chart data for user {request.user.email}, range: {start_dt.date()}..{end_dt.date()}"
        )
        monthly_data_direct = (
            user_investments_qs.filter(created_at__date__gte=start_dt.date(), created_at__date__lte=end_dt.date())
            .annotate(month=TruncMonth("created_at"))
            .values("month")
            .annotate(monthly_total=Sum(Coalesce("amount", Decimal(0))))
            .order_by("month")
        )
        month_labels = []
        month_cursor = start_dt
        for _ in range(12):
            month_labels.append(month_cursor.strftime("%b"))
            month_cursor = month_cursor + relativedelta(months=1)
        monthly_totals = [0] * 12
        month_to_index = {}
        month_cursor = start_dt
        for idx in range(12):
            month_to_index[month_cursor.strftime("%Y-%m")] = idx
            month_cursor = month_cursor + relativedelta(months=1)
        for data in monthly_data_direct:
            month_key = data["month"].strftime("%Y-%m")
            idx = month_to_index.get(month_key)
            if idx is not None:
                monthly_total_decimal = data.get("monthly_total", Decimal(0)) or Decimal(0)
                monthly_totals[idx] = float(monthly_total_decimal)
        monthly_category_data_raw = (
            user_investments_qs.filter(
                created_at__date__gte=start_dt.date(),
                created_at__date__lte=end_dt.date(),
                startup__direction__isnull=False,
            )
            .annotate(month=TruncMonth("created_at"))
            .values("month", "startup__direction__direction_name")
            .annotate(monthly_category_total=Sum(Coalesce("amount", Decimal(0))))
            .order_by("month", "startup__direction__direction_name")
        )
        logger.info(
            f"[investments] Raw monthly category data from DB: {list(monthly_category_data_raw)}"
        )
        structured_monthly_data = collections.defaultdict(
            lambda: collections.defaultdict(float)
        )
        unique_categories = set()
        for data in monthly_category_data_raw:
            month_dt = data["month"]
            category_name = data["startup__direction__direction_name"]
            amount = float(data.get("monthly_category_total", 0) or 0)
            month_key = month_dt.strftime("%Y-%m-01")
            structured_monthly_data[month_key][category_name] += amount
            unique_categories.add(category_name)
        sorted_categories = sorted(list(unique_categories))
        logger.info(
            f"[investments] Unique categories found for chart: {sorted_categories}"
        )
        chart_data_list = []
        rolling_start = start_dt.date()
        for i in range(12):
            month_key = (rolling_start + relativedelta(months=i)).strftime("%Y-%m-01")
            month_data = {
                "month_key": month_key,
                "category_data": dict(structured_monthly_data[month_key]),
            }
            chart_data_list.append(month_data)
        logger.info(
            f"[investments] Final structured chart data list: {chart_data_list}"
        )
        try:
            s3_client = client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                region_name=settings.AWS_S3_REGION_NAME,
            )
        except Exception as s3_init_err:
            logger.error(f"[investments] S3 client init failed: {s3_init_err}")
            s3_client = None
        invested_startups_qs = (
            user_investments_qs.select_related("startup").defer("franchise")
            .annotate(
                average_rating=Avg(
                    ExpressionWrapper(
                        Coalesce(F("startup__sum_votes"), 0)
                        * 1.0
                        / Coalesce(F("startup__total_voters"), 1),
                        output_field=FloatField(),
                    ),
                    filter=Q(startup__total_voters__gt=0),
                    default=0.0,
                ),
                comment_count=Count("startup__comments", distinct=True),
                investors_count=Count(
                    "startup__investmenttransactions__investor", distinct=True
                ),
            )
            .order_by("-amount")[:5]
        )
        owned_startups_qs = (
            Startups.objects.filter(owner_id=request.user.user_id, status="approved")
            .select_related("direction")
            .annotate(
                average_rating=Avg(
                    ExpressionWrapper(
                        Coalesce(F("sum_votes"), 0)
                        * 1.0
                        / Coalesce(F("total_voters"), 1),
                        output_field=FloatField(),
                    ),
                    filter=Q(total_voters__gt=0),
                    default=0.0,
                ),
                comment_count=Count("comments", distinct=True),
                investors_count=Count(
                    "investmenttransactions__investor", distinct=True
                ),
            )
            .order_by("-amount_raised")[:5]
        )
        planetary_investments = []
        min_orbit_size = 200
        max_orbit_size = 800
        orbit_step = 50
        available_sizes = list(
            range(min_orbit_size, max_orbit_size + orbit_step, orbit_step)
        )
        shuffle(available_sizes)
        for idx, startup in enumerate(
            list(invested_startups_qs) + list(owned_startups_qs), 1
        ):
            if hasattr(startup, "startup"):
                startup_obj = startup.startup
            else:
                startup_obj = startup
            if (
                not startup_obj.logo_urls
                or not isinstance(startup_obj.logo_urls, list)
                or len(startup_obj.logo_urls) == 0
            ):
                logger.warning(
                    f"Стартап {startup_obj.startup_id} ({startup_obj.title}) не имеет логотипа в logo_urls"
                )
                logo_url = "https://via.placeholder.com/150"
            else:
                if s3_client is not None:
                    try:
                        prefix = f"startups/{startup_obj.startup_id}/logos/"
                        response = s3_client.list_objects_v2(
                            Bucket=settings.AWS_STORAGE_BUCKET_NAME, Prefix=prefix
                        )
                        if "Contents" in response and len(response["Contents"]) > 0:
                            file_key = response["Contents"][0]["Key"]
                            logo_url = f"{settings.AWS_S3_ENDPOINT_URL}/{settings.AWS_STORAGE_BUCKET_NAME}/{file_key}"
                            logger.info(
                                f"Сгенерирован URL для логотипа стартапа {startup_obj.startup_id}: {logo_url}"
                            )
                        else:
                            logger.warning(
                                f"Файл для логотипа стартапа {startup_obj.startup_id} не найден в бакете по префиксу {prefix}"
                            )
                            logo_url = "https://via.placeholder.com/150"
                    except Exception as e:
                        logger.error(
                            f"Ошибка при генерации URL для логотипа стартапа {startup_obj.startup_id}: {str(e)}"
                        )
                        logo_url = "https://via.placeholder.com/150"
                else:
                    logo_url = "https://via.placeholder.com/150"
            orbit_size = (idx * 100) + 100
            orbit_time = (idx * 20) + 60
            planet_size = (idx * 2) + 50
            investment_type = (
                "Инвестирование"
                if startup_obj.only_invest
                else "Выкуп"
                if startup_obj.only_buy
                else "Выкуп+инвестирование"
                if startup_obj.both_mode
                else "Не указано"
            )
            planet_data = {
                "id": str(idx),
                "startup_id": startup_obj.startup_id,
                "name": startup_obj.title or "Без названия",
                "description": strip_tags(startup_obj.description) if startup_obj.description else "Описание отсутствует",
                "rating": f"{(startup.average_rating or 0):.1f}/5 ({startup_obj.total_voters or 0})",
                "comment_count": startup.comment_count or 0,
                "progress": f"{(startup_obj.amount_raised / startup_obj.funding_goal * 100 if startup_obj.funding_goal else 0):.0f}%",
                "direction": startup_obj.direction.direction_name
                if startup_obj.direction
                else "Не указано",
                "investment_type": investment_type,
                "funding": f"{int(startup_obj.amount_raised or 0):,d} ₽".replace(
                    ",", " "
                ),
                "funding_goal": f"{int(startup_obj.funding_goal or 0):,d} ₽".replace(
                    ",", " "
                ),
                "investors": f"Инвесторов: {startup.investors_count or 0}",
                "image": logo_url,
                "planet_image": startup_obj.planet_image,
                "orbit_size": orbit_size,
                "orbit_time": orbit_time,
                "planet_size": planet_size,
            }
            planetary_investments.append(planet_data)
        logger.info(
            f"[investments] Planetary investments for user {request.user.email}: {planetary_investments}"
        )
        user_investments = (
            user_investments_qs.select_related("startup")
            .annotate(
                startup_average_rating=Avg(
                    ExpressionWrapper(
                        F("startup__sum_votes") * 1.0 / F("startup__total_voters"),
                        output_field=FloatField(),
                    ),
                    filter=Q(startup__total_voters__gt=0),
                    default=0.0,
                ),
                startup_comment_count=Count("startup__comments", distinct=True),
            )
            .order_by("-created_at")
        )
        user_owned_startups = (
            Startups.objects.filter(owner_id=request.user.user_id)
            .select_related("direction", "stage", "status_id")
            .annotate(
                average_rating=Avg(
                    ExpressionWrapper(
                        Coalesce(F("sum_votes"), 0)
                        * 1.0
                        / Coalesce(F("total_voters"), 1),
                        output_field=FloatField(),
                    ),
                    filter=Q(total_voters__gt=0),
                    default=0.0,
                ),
                comment_count=Count("comments"),
            )
            .order_by("-created_at")
        )
        
        # Получаем заявки пользователя (стартапы, которые он создал)
        startup_applications = (
            Startups.objects.filter(owner_id=request.user.user_id)
            .select_related("status_id")
            .order_by("-updated_at")
        )
        
        all_directions_qs = Directions.objects.all().order_by("direction_name")
        all_directions_list = list(all_directions_qs.values("pk", "direction_name"))
        context = {
            "startups_count": startups_count,
            "total_investment": total_investment,
            "max_investment": max_investment,
            "min_investment": min_investment,
            "investment_categories": investment_categories[:7],
            "month_labels": month_labels,
            "chart_monthly_category_data": chart_data_list,
            "chart_categories": sorted_categories,
            "all_directions": all_directions_list,
            "invested_category_data": invested_category_data_dict,
            "user_investments": user_investments,
            "user_owned_startups": user_owned_startups,
            "startup_applications": startup_applications,
            "current_sort": "newest",
            "planetary_investments": planetary_investments,
            "planetary_investments_json": planetary_investments,
            "investor_logo_url": request.user.get_profile_picture_url()
            or "https://via.placeholder.com/60",
        }
        return render(request, "accounts/investments.html", context)
    except Exception as e:
        logger.error(f"Произошла ошибка в investments: {str(e)}", exc_info=True)
        try:
            user_investments_qs = InvestmentTransactions.objects.filter(
                investor=request.user, transaction_type__type_name__iexact="investment"
            ).select_related("startup", "startup__direction")
            total_investment_data = user_investments_qs.aggregate(
                total_investment=Sum("amount"),
                max_investment=Max("amount"),
                startups_count=Count("startup", distinct=True),
            )
            total_investment = total_investment_data.get("total_investment") or Decimal("0")
            max_investment = total_investment_data.get("max_investment") or Decimal("0")
            investments_with_amount = user_investments_qs.filter(amount__gt=0)
            min_investment = investments_with_amount.aggregate(m=Min("amount")).get("m") or Decimal("0")
            category_data_raw = (
                user_investments_qs.values("startup__direction__direction_name").annotate(category_total=Sum("amount")).order_by("-category_total")
            )
            investment_categories = []
            invested_category_data_dict = {}
            denom = total_investment if total_investment > 0 else Decimal("1")
            for c in category_data_raw:
                name = c.get("startup__direction__direction_name") or "Без категории"
                s = c.get("category_total") or Decimal("0")
                pct = int(round((Decimal(s) / denom) * 100)) if s else 0
                pct = max(0, min(pct, 100))
                investment_categories.append({"name": name, "percentage": pct})
                invested_category_data_dict[name] = pct
            end_dt = timezone.now()
            start_dt = (end_dt - relativedelta(months=11)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            month_labels = []
            month_cursor = start_dt
            for _ in range(12):
                month_labels.append(month_cursor.strftime("%b"))
                month_cursor = month_cursor + relativedelta(months=1)
            monthly_category_data_raw = (
                user_investments_qs.filter(created_at__date__gte=start_dt.date(), created_at__date__lte=end_dt.date(), amount__gt=0, startup__direction__isnull=False)
                .annotate(month=TruncMonth("created_at"))
                .values("month", "startup__direction__direction_name")
                .annotate(monthly_category_total=Sum(Coalesce("amount", Decimal(0))))
                .order_by("month", "startup__direction__direction_name")
            )
            structured = collections.defaultdict(lambda: collections.defaultdict(float))
            cats = set()
            for row in monthly_category_data_raw:
                mkey = row["month"].strftime("%Y-%m-01") if row.get("month") else None
                if not mkey:
                    continue
                cname = row["startup__direction__direction_name"]
                val = float(row.get("monthly_category_total") or 0)
                structured[mkey][cname] += val
                cats.add(cname)
            chart_data_list = []
            rolling_start = start_dt.date()
            for i in range(12):
                k = (rolling_start + relativedelta(months=i)).strftime("%Y-%m-01")
                chart_data_list.append({"month_key": k, "category_data": dict(structured[k])})
            sorted_categories = sorted(list(cats))
            context = {
                "startups_count": total_investment_data.get("startups_count", 0),
                "total_investment": total_investment,
                "max_investment": max_investment,
                "min_investment": min_investment,
                "investment_categories": investment_categories[:7],
                "month_labels": month_labels,
                "chart_monthly_category_data": chart_data_list,
                "chart_categories": sorted_categories,
                "all_directions": list(Directions.objects.values("pk", "direction_name")),
                "invested_category_data": invested_category_data_dict,
                "user_investments": user_investments_qs.order_by("-created_at")[:12],
                "user_owned_startups": Startups.objects.filter(owner_id=request.user.user_id)[:12],
                "startup_applications": Startups.objects.filter(owner_id=request.user.user_id).order_by("-updated_at"),
                "current_sort": "newest",
                "planetary_investments": [],
                "planetary_investments_json": [],
                "investor_logo_url": request.user.get_profile_picture_url() or "https://via.placeholder.com/60",
            }
            return render(request, "accounts/investments.html", context)
        except Exception as e2:
            logger.error(f"[investments] Fallback building failed: {e2}", exc_info=True)
            return render(request, "accounts/investments.html", safe_context)
def legal(request):
    return render(request, "accounts/legal.html")
@login_required
def profile(request, user_id=None):
    if not user_id:
        user_id_param = request.GET.get("user_id")
        if user_id_param:
            try:
                user_id = int(user_id_param)
            except ValueError:
                user_id = None
    if user_id:
        user = get_object_or_404(Users, user_id=user_id)
        is_own_profile = request.user.user_id == user.user_id
    else:
        user = request.user
        is_own_profile = True
    if request.headers.get("x-requested-with") == "XMLHttpRequest" and request.method == "GET":
        user_data = {
            "user_id": user.user_id,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "role": user.role.role_name if user.role else "",
            "profile_picture_url": user.get_profile_picture_url() if hasattr(user, "get_profile_picture_url") else "",
            "rating": getattr(user, "rating", None),
            "bio": getattr(user, "bio", ""),
        }
        return JsonResponse(user_data)
    # Модальное окно выбора роли больше не нужно - всем назначается роль startuper
    show_role_selection = False
    if request.method == "POST" and is_own_profile:
        if "select_role" in request.POST:
            role_id = request.POST.get("role_id")
            if role_id in ["1", "2"]:
                user.role_id = int(role_id)
                user.save(update_fields=["role"])
                messages.success(request, "Роль успешно выбрана!")
                return redirect("profile")
            else:
                messages.error(request, "Выбрана неверная роль.")
                return redirect("profile")
        elif "edit_profile" in request.POST:
            form = ProfileEditForm(request.POST, instance=user)
            if form.is_valid():
                form.save()
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse(
                        {"success": True, "message": "Профиль успешно обновлен!"}
                    )
                messages.success(request, "Профиль успешно обновлен!")
                return redirect("profile")
            else:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "errors": form.errors})
                messages.error(request, "Пожалуйста, исправьте ошибки.")
        elif "avatar" in request.FILES:
            avatar = request.FILES["avatar"]
            allowed_mimes = ["image/jpeg", "image/png"]
            if avatar.content_type not in allowed_mimes:
                messages.error(request, "Допустимы только файлы PNG или JPEG.")
                return redirect("profile")
            try:
                avatar_id = str(uuid.uuid4())
                file_path = f"users/{user.user_id}/avatar/{avatar_id}_{avatar.name}"
                s3_client = boto3.client(
                    "s3",
                    endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                    region_name=settings.AWS_S3_REGION_NAME,
                )
                bucket_name = settings.AWS_STORAGE_BUCKET_NAME
                # Delete old avatar files from S3
                prefix = f"users/{user.user_id}/avatar/"
                response = s3_client.list_objects_v2(Bucket=bucket_name, Prefix=prefix)
                if "Contents" in response:
                    for obj in response["Contents"]:
                        s3_client.delete_object(Bucket=bucket_name, Key=obj["Key"])
                FileStorage.objects.filter(
                    entity_type__type_name="user",
                    entity_id=user.user_id,
                    file_type__type_name="avatar",
                ).delete()
                default_storage.save(file_path, avatar)
                user.profile_picture_url = avatar_id
                user.save(update_fields=["profile_picture_url"])
                entity_type_obj, _ = EntityTypes.objects.get_or_create(type_name="user")
                file_type_obj, _ = FileTypes.objects.get_or_create(type_name="avatar")
                FileStorage.objects.create(
                    entity_type=entity_type_obj,
                    entity_id=user.user_id,
                    file_url=avatar_id,
                    file_type=file_type_obj,
                    uploaded_at=timezone.now(),
                )
                messages.success(request, "Аватар успешно обновлен!")
            except Exception as e:
                logger.error("Avatar upload failed for user %d: %s", user.user_id, e)
                messages.error(request, "Ошибка при загрузке аватара.")
            return redirect("profile")
    form = ProfileEditForm(instance=user)
    startups_list = Startups.objects.filter(owner=user).order_by("-created_at")
    startups_paginator = Paginator(startups_list, 5)
    startups_page_number = request.GET.get("startups_page")
    startups_page_obj = startups_paginator.get_page(startups_page_number)
    news_list = NewsArticles.objects.filter(author=user).order_by("-published_at")
    news_paginator = Paginator(news_list, 6)
    news_page_number = request.GET.get("news_page")
    news_page_obj = news_paginator.get_page(news_page_number)
    context = {
        "user": user,
        "is_own_profile": is_own_profile,
        "show_role_selection": show_role_selection,
        "form": form,
        "startups_page": startups_page_obj,
        "news_page": news_page_obj,
    }
    return render(request, "accounts/profile.html", context)
@login_required
def delete_avatar(request):
    if request.method == "POST":
        user = request.user
        if "avatar" in request.FILES:
            avatar = request.FILES["avatar"]
            allowed_mimes = ["image/jpeg", "image/png"]
            if avatar.content_type not in allowed_mimes:
                messages.error(request, "Допустимы только файлы PNG или JPEG.")
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse(
                        {
                            "success": False,
                            "error": "Допустимы только файлы PNG или JPEG.",
                        }
                    )
                return render(
                    request,
                    "accounts/profile.html",
                    {
                        "user": user,
                        "is_own_profile": True,
                        "form": form,
                        "startups_page": startups_page,
                        "news_page": news_page,
                        "show_role_selection": show_role_selection,
                    },
                )
            max_size = 5 * 1024 * 1024
            if avatar.size > max_size:
                messages.error(request, "Размер файла не должен превышать 5 МБ.")
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse(
                        {
                            "success": False,
                            "error": "Размер файла не должен превышать 5 МБ.",
                        }
                    )
                return render(
                    request,
                    "accounts/profile.html",
                    {
                        "user": user,
                        "is_own_profile": True,
                        "form": form,
                        "startups_page": startups_page,
                        "news_page": news_page,
                        "show_role_selection": show_role_selection,
                    },
                )
            avatar_id = str(uuid.uuid4())
            file_path = f"users/{request.user.user_id}/avatar/{avatar_id}_{avatar.name}"
            try:
                s3_client = boto3.client(
                    "s3",
                    endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                    region_name=settings.AWS_S3_REGION_NAME,
                )
                bucket_name = settings.AWS_STORAGE_BUCKET_NAME
                prefix = f"users/{request.user.user_id}/avatar/"
                response = s3_client.list_objects_v2(Bucket=bucket_name, Prefix=prefix)
                if "Contents" in response:
                    for obj in response["Contents"]:
                        s3_client.delete_object(Bucket=bucket_name, Key=obj["Key"])
                        logger.info(f"Удалён старый аватар: {obj['Key']}")
                FileStorage.objects.filter(
                    entity_type__type_name="user",
                    entity_id=request.user.user_id,
                    file_type__type_name="avatar",
                ).delete()
                default_storage.save(file_path, avatar)
                request.user.profile_picture_url = avatar_id
                request.user.save()
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="user")
                file_type, _ = FileTypes.objects.get_or_create(type_name="avatar")
                FileStorage.objects.create(
                    entity_type=entity_type,
                    entity_id=request.user.user_id,
                    file_url=avatar_id,
                    file_type=file_type,
                    uploaded_at=timezone.now(),
                )
                logger.info(
                    f"Аватар сохранён для user_id {request.user.user_id} по пути: {file_path}, UUID: {avatar_id}"
                )
                messages.success(request, "Аватарка успешно загружена!")
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse(
                        {"success": True, "message": "Аватарка успешно загружена!"}
                    )
            except Exception as e:
                logger.error(
                    f"Ошибка при сохранении аватара для user_id {request.user.user_id}: {str(e)}"
                )
                messages.error(request, "Ошибка при загрузке аватара.")
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse(
                        {"success": False, "error": "Ошибка при загрузке аватара."}
                    )
            return redirect("profile")
    return render(
        request,
        "accounts/profile.html",
        {
            "user": user,
            "is_own_profile": profile_user == request.user,
            "form": form,
            "startups_page": startups_page,
            "news_page": news_page,
            "show_role_selection": show_role_selection,
        },
    )
@login_required
def start_deal(request, chat_id):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    logger.info(
        f"Starting deal check for chat {chat_id}, participants: {chat.chatparticipants_set.count()}"
    )
    if chat.is_group_chat or chat.is_deal:
        logger.error(f"Chat {chat_id} is group or already a deal")
        return JsonResponse(
            {
                "success": False,
                "error": "Сделку можно начать только в личном чате, который ещё не является сделкой",
            },
            status=400,
        )
    participants = chat.chatparticipants_set.all()
    if (
        participants.count() < 2
    ):
        logger.error(
            f"Chat {chat_id} has {participants.count()} participants, expected at least 2"
        )
        return JsonResponse(
            {"success": False, "error": "В чате должно быть как минимум два участника"},
            status=400,
        )
    roles = {
        p.user.role.role_name.lower() for p in participants if p.user and p.user.role
    }
    if not {"startuper", "investor"}.issubset(roles):
        logger.error(
            f"Chat {chat_id} roles: {roles}, expected 'startuper' and 'investor'"
        )
        return JsonResponse(
            {
                "success": False,
                "error": "Чат должен включать одного стартапера и одного инвестора",
            },
            status=400,
        )
    try:
        data = json.loads(request.body)
        initiator_name = data.get(
            "initiator_name", request.user.get_full_name() or "Пользователь"
        )
    except json.JSONDecodeError:
        initiator_name = request.user.get_full_name() or "Пользователь"
    with transaction.atomic():
        chat.is_deal = True
        chat.deal_status = "pending"
        chat.updated_at = timezone.now()
        chat.save()
        moderators = Users.objects.filter(role__role_name="moderator")
        if not moderators.exists():
            return JsonResponse(
                {"success": False, "error": "Нет доступных модераторов"}, status=500
            )
        moderator = choice(list(moderators))
        moderator_participant, created = ChatParticipants.objects.get_or_create(
            conversation=chat, user=moderator
        )
        if not created and not moderator_participant:
            logger.error(
                f"Failed to create or find moderator {moderator.user_id} for chat {chat_id}"
            )
            return JsonResponse(
                {"success": False, "error": "Ошибка назначения модератора"}, status=500
            )
        logger.info(
            f"Moderator {moderator.user_id} added to chat {chat_id}, created: {created}"
        )
        message = Messages(
            conversation=chat,
            sender=None,
            message_text=f"Сделку начал {initiator_name}. Назначен модератор: {moderator.get_full_name()}",
            status=MessageStatuses.objects.get(status_name="sent"),
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        message.save()
    participants_data = [
        {
            "user_id": p.user.user_id,
            "name": p.user.get_full_name(),
            "role": p.user.role.role_name if p.user.role else "unknown",
        }
        for p in chat.chatparticipants_set.all()
    ]
    logger.info(
        f"Сделка начата в чате {chat_id}, модератор {moderator.user_id} назначен"
    )
    return JsonResponse(
        {
            "success": True,
            "message": "Сделка начата, модератор назначен",
            "moderator": {
                "user_id": moderator.user_id,
                "name": moderator.get_full_name(),
            },
            "participants": participants_data,
        }
    )
@login_required
def deals_view(request):
    if not is_moderator(request.user):
        messages.error(request, "Доступ к этой странице разрешен только модераторам.")
        logger.warning(
            f"Access denied for user {request.user.user_id} - not a moderator"
        )
        return redirect("home")
    status_filter = request.GET.get("status", "pending")
    valid_statuses = ["pending", "approved", "rejected"]
    if status_filter not in valid_statuses:
        status_filter = "pending"
    logger.info(
        f"Processing deals_view for user_id={request.user.user_id}, status_filter={status_filter}"
    )
    try:
        deals_query = (
            ChatConversations.objects.filter(is_deal=True, deal_status=status_filter)
            .prefetch_related("chatparticipants_set__user")
            .order_by("-updated_at")
        )
        logger.info(f"Initial query returned {deals_query.count()} deals")
    except Exception as e:
        logger.error(f"Error in initial query: {str(e)}")
        return JsonResponse({"error": f"Database query failed: {str(e)}"}, status=500)
    deals = deals_query.filter(chatparticipants__user=request.user)
    logger.info(f"Filtered deals for moderator {request.user.user_id}: {deals.count()}")
    for deal in deals:
        try:
            participants = deal.chatparticipants_set.all()
            logger.debug(
                f"Deal {deal.conversation_id}: Participants {[(p.user.user_id, p.user.role.role_name if p.user.role else 'None') for p in participants]}, Status: {deal.deal_status}"
            )
        except Exception as e:
            logger.error(f"Error processing deal {deal.conversation_id}: {str(e)}")
    deal_data = []
    selected_chat = None
    chat_id = request.GET.get("chat_id")
    if chat_id:
        try:
            selected_chat = get_object_or_404(
                ChatConversations, conversation_id=chat_id, is_deal=True
            )
            if not selected_chat.chatparticipants_set.filter(
                user=request.user
            ).exists():
                messages.error(request, "У вас нет доступа к этому чату.")
                logger.warning(
                    f"No access to chat {chat_id} for user {request.user.user_id}"
                )
                selected_chat = None
            else:
                messages = Messages.objects.filter(conversation=selected_chat).order_by(
                    "created_at"
                )
                messages_data = [
                    {
                        "message_id": msg.message_id,
                        "sender_name": msg.sender.get_full_name()
                        if msg.sender
                        else "Система",
                        "message_text": msg.message_text,
                        "created_at": msg.created_at.strftime("%H:%M %d/%m/%Y")
                        if msg.created_at
                        else "",
                        "is_own": msg.sender == request.user if msg.sender else False,
                    }
                    for msg in messages
                ]
                selected_chat_messages = messages_data
                logger.info(f"Loaded {len(messages_data)} messages for chat {chat_id}")
        except Exception as e:
            logger.error(f"Error loading chat {chat_id}: {str(e)}")
            messages.error(request, "Ошибка загрузки чата.")
            selected_chat = None
    for deal in deals:
        try:
            participants = deal.chatparticipants_set.all()
            moderator = next(
                (
                    p.user
                    for p in participants
                    if p.user.role and p.user.role.role_name == "moderator"
                ),
                None,
            )
            other_participants = [
                p.user for p in participants if p.user and p.user != moderator
            ]
            deal_data.append(
                {
                    "conversation_id": deal.conversation_id,
                    "name": deal.name or f"Сделка {deal.conversation_id}",
                    "participants": [
                        f"{p.first_name} {p.last_name}" for p in other_participants
                    ],
                    "moderator": moderator.get_full_name()
                    if moderator
                    else "Не назначен",
                    "last_message": deal.get_last_message().message_text
                    if deal.get_last_message()
                    else "Нет сообщений",
                    "created_at": deal.created_at.strftime("%H:%M")
                    if deal.created_at
                    else "",
                    "date": deal.created_at.strftime("%d/%m/%Y")
                    if deal.created_at
                    else "",
                    "unread_count": Messages.objects.filter(
                        conversation=deal, status__status_name="sent"
                    )
                    .exclude(sender=moderator)
                    .count()
                    if moderator
                    else 0,
                    "deal_status": deal.deal_status,
                }
            )
        except Exception as e:
            logger.error(
                f"Error processing deal data for {deal.conversation_id}: {str(e)}"
            )
    context = {
        "deals": deal_data,
        "current_status": status_filter,
        "selected_chat": selected_chat,
        "chat_messages": selected_chat_messages if selected_chat else [],
    }
    logger.info(f"Rendering deals.html with {len(deal_data)} deals")
    return render(request, "accounts/deals.html", context)
@login_required
def send_message(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    form = MessageForm(request.POST)
    if not form.is_valid():
        return JsonResponse({"success": False, "error": "Неверные данные формы"})
    chat_id = request.POST.get("chat_id")
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}
        )
    if not is_moderator(request.user):
        return JsonResponse(
            {
                "success": False,
                "error": "Только модератор может отправлять сообщения здесь",
            }
        )
    message = Messages(
        conversation=chat,
        sender=request.user,
        message_text=form.cleaned_data["message_text"],
        status=MessageStatuses.objects.get(status_name="sent"),
        created_at=timezone.now(),
        updated_at=timezone.now(),
    )
    message.save()
    chat.updated_at = timezone.now()
    chat.save()
    return JsonResponse(
        {
            "success": True,
            "message": {
                "message_id": message.message_id,
                "sender_name": request.user.get_full_name(),
                "message_text": message.message_text,
                "created_at": message.created_at.strftime("%H:%M %d/%m/%Y"),
                "is_own": True,
            },
        }
    )
@login_required
def approve_deal(request, chat_id):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists() or (
        request.user.role and (request.user.role.role_name or "").lower() != "moderator"
    ):
        return JsonResponse(
            {"success": False, "error": "У вас нет прав для этого действия"}, status=403
        )
    if not chat.is_deal:
        return JsonResponse({"success": False, "error": "Это не сделка"}, status=400)
    with transaction.atomic():
        chat.deal_status = "approved"
        chat.updated_at = timezone.now()
        chat.save()
        message = Messages(
            conversation=chat,
            sender=None,
            message_text=f"Сделка #{chat.conversation_id} одобрена модератором {request.user.get_full_name()}",
            status=MessageStatuses.objects.get(status_name="sent"),
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        message.save()
    logger.info(f"Сделка {chat_id} одобрена модератором {request.user.user_id}")
    return JsonResponse({"success": True, "message": "Сделка одобрена"})
@login_required
def reject_deal(request, chat_id):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists() or (
        request.user.role and (request.user.role.role_name or "").lower() != "moderator"
    ):
        return JsonResponse(
            {"success": False, "error": "У вас нет прав для этого действия"}, status=403
        )
    if not chat.is_deal:
        return JsonResponse({"success": False, "error": "Это не сделка"}, status=400)
    with transaction.atomic():
        chat.deal_status = "rejected"
        chat.updated_at = timezone.now()
        chat.save()
        message = Messages(
            conversation=chat,
            sender=None,
            message_text=f"Сделка #{chat.conversation_id} отклонена модератором {request.user.get_full_name()}",
            status=MessageStatuses.objects.get(status_name="sent"),
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        message.save()
    logger.info(f"Сделка {chat_id} отклонена модератором {request.user.user_id}")
    return JsonResponse({"success": True, "message": "Сделка отклонена"})
@login_required
def create_startup(request):
    allowed_roles = ["startuper", "moderator"]
    if not hasattr(request.user, "role") or request.user.role.role_name.lower() not in allowed_roles:
        messages.error(request, "Доступ к созданию стартапа разрешён только пользователям с ролью 'Стартаппер' или 'Модератор'.")
        return redirect("home")
    if request.method == "POST":
        # Логирование входящих данных
        logger.info(f"=== CREATE_STARTUP START === User: {request.user.user_id} ({request.user.email})")
        logger.info(f"POST keys: {list(request.POST.keys())}")
        logger.info(f"FILES keys: {list(request.FILES.keys())}")
        for key, files in request.FILES.lists():
            logger.info(f"  {key}: {len(files)} файлов, sizes: {[f.size for f in files]}, names: {[f.name for f in files]}")
        total_size = sum(f.size for f in request.FILES.values())
        logger.info(f"Total upload size: {total_size / 1024 / 1024:.2f} MB")
        
        # Защита от дубликатов: проверяем, не создал ли пользователь стартап в последние 5 секунд
        five_seconds_ago = timezone.now() - datetime.timedelta(seconds=5)
        recent_startups = Startups.objects.filter(
            owner=request.user,
            created_at__gte=five_seconds_ago
        ).order_by('-created_at')
        
        if recent_startups.exists():
            latest = recent_startups.first()
            logger.warning(f"=== DUPLICATE STARTUP PREVENTED === User {request.user.user_id} tried to create duplicate, existing startup_id: {latest.startup_id}")
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("startup_creation_success"),
                })
            return redirect("startup_creation_success")
        
        form = StartupForm(request.POST, request.FILES)
        if form.is_valid():
            logger.info("Form is valid, creating startup...")
            startup = form.save(commit=False)
            startup.owner = request.user
            startup.created_at = timezone.now()
            startup.updated_at = timezone.now()
            startup.status = "pending"
            try:
                startup.status_id = ReviewStatuses.objects.get(status_name="Pending")
            except ReviewStatuses.DoesNotExist:
                logger.error("Статус 'Pending' не найден в базе данных.")
                messages.error(request, "Статус 'Pending' не найден в базе данных.")
                return render(
                    request,
                    "accounts/create_startup.html",
                    {"form": form, "timeline_steps": request.POST},
                )
            investment_type = form.cleaned_data.get("investment_type")
            if investment_type == "invest":
                startup.only_invest = True
                startup.only_buy = False
                startup.both_mode = False
            elif investment_type == "buy":
                startup.only_invest = False
                startup.only_buy = True
                startup.both_mode = False
            elif investment_type == "both":
                startup.only_invest = False
                startup.only_buy = False
                startup.both_mode = True
            startup.step_number = int(request.POST.get("step_number", 1))
            startup.planet_image = form.cleaned_data.get("planet_image")
            logger.info("Сохранение стартапа перед обработкой файлов...")
            startup.save()
            logger.info(f"Стартап сохранен, startup_id: {startup.startup_id}")
            if not startup.startup_id:
                logger.error("Ошибка: startup_id не сгенерирован после сохранения!")
                messages.error(
                    request,
                    "Произошла ошибка при создании стартапа: ID не сгенерирован.",
                )
                return render(
                    request,
                    "accounts/create_startup.html",
                    {"form": form, "timeline_steps": request.POST},
                )
            
            # Обработка временных медиа-файлов из localStorage
            temp_media_data = request.POST.get('temp_media_data', '')
            if temp_media_data:
                try:
                    temp_files = json.loads(temp_media_data)
                    logger.info(f"Обработка {len(temp_files)} временных медиа-файлов")
                    
                    uploaded_content_type, _ = FileTypes.objects.get_or_create(type_name='uploaded_content')
                    entity_type, _ = EntityTypes.objects.get_or_create(type_name='startup')
                    
                    for temp_file in temp_files:
                        # Создаем запись в FileStorage для временного файла
                        file_id = str(uuid.uuid4())
                        safe_create_file_storage(
                            entity_type=entity_type,
                            entity_id=startup.startup_id,
                            file_type=uploaded_content_type,
                            file_url=file_id,
                            uploaded_at=timezone.now(),
                            startup=startup,
                            original_file_name=temp_file.get('name', 'temp_file'),
                        )
                        logger.info(f"Создана запись для временного файла: {temp_file.get('name')}")
                        
                except Exception as e:
                    logger.error(f"Ошибка обработки временных медиа-файлов: {e}", exc_info=True)
                    messages.warning(request, "Не удалось обработать временные медиа-файлы.")
            for i in range(1, 6):
                description = request.POST.get(f"step_description_{i}", "").strip()
                if description:
                    StartupTimeline.objects.create(
                        startup=startup,
                        step_number=i,
                        title=f"Этап {i}",
                        description=description,
                    )
            file_save_errors = []
            # Upload files using shared helpers
            logo_ids = upload_entity_logo(form, startup, 'startup')
            upload_entity_catalog_card(form, startup, 'startup')
            creatives_ids = upload_entity_files(request, form, startup, 'startup', 'creatives', 'creative', 'image/jpeg')
            proofs_ids = upload_entity_files(request, form, startup, 'startup', 'proofs', 'proof', 'application/pdf')
            video_ids = []
            # Video handled separately below (async Celery)
            # Old logo/catalog_card/creatives/proofs upload code removed — handled by shared helpers above
            
            videos = request.FILES.getlist("video")
            if videos:
                for video in videos:
                    if not hasattr(video, "name"):
                        logger.warning(f"Пропущено видео: {video}")
                        continue
                    try:
                        unique_filename = get_unique_filename(video.name, startup.startup_id, "video")
                        video_data = video.read()
                        content_type = getattr(video, 'content_type', 'video/mp4')
                        
                        video_data_b64 = base64.b64encode(video_data).decode('utf-8')
                        
                        upload_video_to_s3.delay(
                            video_data=video_data_b64,
                            video_name=video.name,
                            video_content_type=content_type,
                            entity_id=startup.startup_id,
                            original_filename=unique_filename,
                            entity_type_name='startup'
                        )
                        logger.info(f"Видео отправлено в очередь Celery: {video.name}, размер: {len(video_data)} байт")
                        messages.info(request, f"Видео {video.name} загружается в фоновом режиме.")
                    except Exception as e:
                        logger.error(f"Ошибка отправки видео в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось отправить видео {video.name} на загрузку.")
                        file_save_errors.append({"field": "video", "file": getattr(video, "name", ""), "error": str(e)})
            startup.logo_urls = logo_ids
            startup.creatives_urls = creatives_ids
            startup.proofs_urls = proofs_ids
            startup.video_urls = video_ids
            logger.info(f"UUID сохранены: logo={logo_ids}, creatives={creatives_ids}, proofs={proofs_ids}, videos={video_ids}. Файлы загружаются асинхронно через Celery.")
            
            # При создании автоматически добавляем первые 4 креатива в слайдер
            # (при редактировании пользователь может выбрать вручную через чекбоксы)
            slider_images = request.POST.getlist("slider_images")
            if not slider_images and creatives_ids:
                # Если slider_images не был отправлен, используем creatives_ids
                slider_images = creatives_ids[:4]
                logger.info(f"Автоматически установлены slider_images из creatives_ids: {slider_images}")
            elif len(slider_images) > 4:
                slider_images = slider_images[:4]
            startup.slider_images = slider_images
            
            startup.save()
            logger.info(
                f"Стартап создан: ID={startup.startup_id}, Planet={startup.planet_image}"
            )
            
            # Отправляем уведомление в Telegram о новой заявке
            try:
                owner_name = f"{request.user.first_name or ''} {request.user.last_name or ''}".strip() or request.user.email
                send_telegram_new_entity_notification(
                    entity_type='startup',
                    entity_title=startup.title,
                    owner_name=owner_name,
                    owner_email=request.user.email,
                    entity_id=startup.startup_id
                )
            except Exception as e:
                logger.warning(f"Failed to send Telegram notification for startup {startup.startup_id}: {e}")
            
            clear_temp_files(request, 'startupForm')
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("startup_creation_success"),
                    "file_save_errors": file_save_errors,
                })
            messages.success(
                request,
                f'Стартап "{startup.title}" успешно создан и отправлен на модерацию!',
            )
            return redirect("startup_creation_success")
        else:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                logger.warning(f"=== CREATE_STARTUP FORM INVALID (AJAX) === User: {request.user.user_id}")
                logger.warning(f"Form errors: {form.errors.as_json()}")
                return JsonResponse({
                    "success": False,
                    "errors": {f: [x["message"] for x in e.get_json_data()] for f, e in form.errors.items()},
                    "non_field_errors": [str(e) for e in form.non_field_errors()],
                }, status=400)
            logger.warning(f"=== CREATE_STARTUP FORM INVALID === User: {request.user.user_id}")
            logger.warning(f"Form errors: {form.errors.as_json()}")
            messages.error(request, "Форма содержит ошибки.")
            return render(
                request,
                "accounts/create_startup.html",
                {"form": form, "timeline_steps": request.POST},
            )
    else:
        form = StartupForm()
        # Очищаем temp_files из сессии при свежем открытии формы
        clear_temp_files(request, 'startupForm')
    return render(request, "accounts/create_startup.html", {"form": form})

@login_required
def create_franchise(request):
    allowed_roles = ["startuper", "moderator"]
    if not hasattr(request.user, "role") or request.user.role.role_name.lower() not in allowed_roles:
        messages.error(request, "Доступ к созданию франшизы разрешён только пользователям с ролью 'Стартаппер' или 'Модератор'.")
        return redirect("home")
    if request.method == "POST":
      try:
        # Логирование входящих данных
        logger.info(f"=== CREATE_FRANCHISE START === User: {request.user.user_id} ({request.user.email})")
        logger.info(f"POST keys: {list(request.POST.keys())}")
        logger.info(f"FILES keys: {list(request.FILES.keys())}")
        for key, files in request.FILES.lists():
            logger.info(f"  {key}: {len(files)} файлов, sizes: {[f.size for f in files]}, names: {[f.name for f in files]}")
        total_size = sum(f.size for f in request.FILES.values())
        logger.info(f"Total upload size: {total_size / 1024 / 1024:.2f} MB")
        
        # Защита от дубликатов: проверяем, не создал ли пользователь франшизу в последние 5 секунд
        five_seconds_ago = timezone.now() - datetime.timedelta(seconds=5)
        recent_franchises = Franchises.objects.filter(
            owner=request.user,
            created_at__gte=five_seconds_ago
        ).order_by('-created_at')
        
        if recent_franchises.exists():
            latest = recent_franchises.first()
            logger.warning(f"=== DUPLICATE FRANCHISE PREVENTED === User {request.user.user_id} tried to create duplicate, existing franchise_id: {latest.franchise_id}")
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("franchise_creation_success"),
                })
            return redirect("franchise_creation_success")
        
        form = FranchiseForm(request.POST, request.FILES)
        
        # Проверяем наличие изображений через request.FILES.getlist
        creatives_files = request.FILES.getlist("creatives")
        if not creatives_files:
            form.add_error("creatives", "Загрузите хотя бы одно изображение.")
        
        if form.is_valid():
            logger.info("Form is valid, creating franchise...")
            franchise = form.save(commit=False)
            franchise.owner = request.user
            franchise.created_at = timezone.now()
            franchise.updated_at = timezone.now()
            franchise.status = "pending"
            try:
                franchise.status_id = ReviewStatuses.objects.get(status_name="Pending")
            except ReviewStatuses.DoesNotExist:
                messages.error(request, "Статус 'Pending' не найден в базе данных.")
                return render(request, "accounts/create_franchise.html", {"form": form})
            franchise.planet_image = form.cleaned_data.get("planet_image")
            franchise.save()
            
            def try_save_file(file_obj, file_path):
                try:
                    default_storage.save(file_path, file_obj)
                    return True
                except Exception as e:
                    logger.error(f"Ошибка default_storage.save для {file_path}: {e}", exc_info=True)
                    try:
                        s3 = boto3.client(
                            's3',
                            endpoint_url=getattr(settings, 'AWS_S3_ENDPOINT_URL', None),
                            aws_access_key_id=getattr(settings, 'AWS_ACCESS_KEY_ID', None),
                            aws_secret_access_key=getattr(settings, 'AWS_SECRET_ACCESS_KEY', None),
                            region_name=getattr(settings, 'AWS_S3_REGION_NAME', None),
                            config=boto3.session.Config(s3={'addressing_style': getattr(settings, 'AWS_S3_ADDRESSING_STYLE', 'virtual')})
                        )
                        bucket = getattr(settings, 'AWS_STORAGE_BUCKET_NAME', None)
                        content_type = getattr(file_obj, 'content_type', 'application/octet-stream')
                        body_bytes = file_obj.read()
                        try:
                            s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ContentType=content_type, ACL='public-read')
                        except Exception:
                            s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ACL='public-read')
                        return True
                    except Exception as e2:
                        logger.error(f"Ошибка прямой загрузки в S3 для {file_path}: {e2}", exc_info=True)
                        return False
            
            # Обработка временных медиа-файлов из localStorage
            temp_media_data = request.POST.get('temp_media_data', '')
            if temp_media_data:
                try:
                    temp_files = json.loads(temp_media_data)
                    logger.info(f"Обработка {len(temp_files)} временных медиа-файлов для франшизы")
                    
                    uploaded_content_type, _ = FileTypes.objects.get_or_create(type_name='uploaded_content')
                    entity_type, _ = EntityTypes.objects.get_or_create(type_name='franchise')
                    
                    for temp_file in temp_files:
                        # Создаем запись в FileStorage для временного файла
                        file_id = str(uuid.uuid4())
                        safe_create_file_storage(
                            entity_type=entity_type,
                            entity_id=franchise.franchise_id,
                            file_type=uploaded_content_type,
                            file_url=file_id,
                            uploaded_at=timezone.now(),
                            startup=None,
                            original_file_name=temp_file.get('name', 'temp_file'),
                        )
                        logger.info(f"Создана запись для временного файла франшизы: {temp_file.get('name')}")
                        
                except Exception as e:
                    logger.error(f"Ошибка обработки временных медиа-файлов франшизы: {e}", exc_info=True)
                    messages.warning(request, "Не удалось обработать временные медиа-файлы.")

            # Upload all files using shared helpers
            logo_ids = upload_entity_logo(form, franchise, 'franchise')
            upload_entity_catalog_card(form, franchise, 'franchise')
            creatives_ids = upload_entity_files(request, form, franchise, 'franchise', 'creatives', 'creative', 'image/jpeg')
            proofs_ids = upload_entity_files(request, form, franchise, 'franchise', 'proofs', 'proof', 'application/pdf')

            video_ids = []
            videos = request.FILES.getlist("video")
            if videos:
                for video in videos:
                    if not hasattr(video, "name"):
                        logger.warning(f"Пропущено видео: {video}")
                        continue
                    try:
                        unique_filename = get_unique_filename(video.name, franchise.franchise_id, "video")
                        video_data = video.read()
                        content_type = getattr(video, 'content_type', 'video/mp4')
                        
                        video_data_b64 = base64.b64encode(video_data).decode('utf-8')
                        
                        upload_video_to_s3.delay(
                            video_data=video_data_b64,
                            video_name=video.name,
                            video_content_type=content_type,
                            entity_id=franchise.franchise_id,
                            original_filename=unique_filename,
                            entity_type_name='franchise'
                        )
                        logger.info(f"Видео отправлено в очередь Celery: {video.name}, размер: {len(video_data)} байт")
                        messages.info(request, f"Видео {video.name} загружается в фоновом режиме.")
                    except Exception as e:
                        logger.error(f"Ошибка отправки видео в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось отправить видео {video.name} на загрузку.")

            franchise.logo_urls = logo_ids
            franchise.creatives_urls = creatives_ids
            franchise.proofs_urls = proofs_ids
            franchise.video_urls = video_ids
            logger.info(f"UUID сохранены: logo={logo_ids}, creatives={creatives_ids}, proofs={proofs_ids}, videos={video_ids}. Файлы загружаются асинхронно через Celery.")
            
            # При создании автоматически добавляем первые 4 креатива в слайдер
            # (при редактировании пользователь может выбрать вручную через чекбоксы)
            slider_images = request.POST.getlist("slider_images")
            if not slider_images and creatives_ids:
                # Если slider_images не был отправлен, используем creatives_ids
                slider_images = creatives_ids[:4]
                logger.info(f"Автоматически установлены slider_images из creatives_ids: {slider_images}")
            elif len(slider_images) > 4:
                slider_images = slider_images[:4]
            franchise.slider_images = slider_images
            
            franchise.save()
            logger.info(f"=== CREATE_FRANCHISE SUCCESS === franchise_id: {franchise.franchise_id}")
            
            # Отправляем уведомление в Telegram о новой заявке
            try:
                owner_name = f"{request.user.first_name or ''} {request.user.last_name or ''}".strip() or request.user.email
                send_telegram_new_entity_notification(
                    entity_type='franchise',
                    entity_title=franchise.title,
                    owner_name=owner_name,
                    owner_email=request.user.email,
                    entity_id=franchise.franchise_id
                )
            except Exception as e:
                logger.warning(f"Failed to send Telegram notification for franchise {franchise.franchise_id}: {e}")

            clear_temp_files(request, 'franchiseForm')
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("franchise_creation_success"),
                })
            messages.success(request, f'Франшиза "{franchise.title}" успешно создана и отправлена на модерацию!')
            return redirect("franchise_creation_success")
        else:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                logger.warning(f"=== CREATE_FRANCHISE FORM INVALID (AJAX) === User: {request.user.user_id}")
                logger.warning(f"Form errors: {form.errors.as_json()}")
                return JsonResponse({
                    "success": False,
                    "errors": {f: [x["message"] for x in e.get_json_data()] for f, e in form.errors.items()},
                    "non_field_errors": [str(e) for e in form.non_field_errors()],
                }, status=400)
            logger.warning(f"=== CREATE_FRANCHISE FORM INVALID === User: {request.user.user_id}")
            logger.warning(f"Form errors: {form.errors.as_json()}")
            messages.error(request, "Форма содержит ошибки.")
            return render(request, "accounts/create_franchise.html", {"form": form})
      except Exception as exc:
        logger.exception(f"=== CREATE_FRANCHISE CRASHED === User: {request.user.user_id}")
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({
                "success": False,
                "errors": {},
                "non_field_errors": [f"Внутренняя ошибка сервера: {exc}"],
            }, status=500)
        messages.error(request, f"Ошибка при создании франшизы: {exc}")
        form = FranchiseForm(request.POST)
        return render(request, "accounts/create_franchise.html", {"form": form})
    else:
        form = FranchiseForm()
        clear_temp_files(request, 'franchiseForm')
    return render(request, "accounts/create_franchise.html", {"form": form})

@login_required
def create_agency(request):
    allowed_roles = ["startuper", "moderator"]
    if not hasattr(request.user, "role") or request.user.role.role_name.lower() not in allowed_roles:
        messages.error(request, "Доступ к созданию агентства разрешён только пользователям с ролью 'Стартаппер' или 'Модератор'.")
        return redirect("home")
    if request.method == "POST":
        # Логирование входящих данных
        logger.info(f"=== CREATE_AGENCY START === User: {request.user.user_id} ({request.user.email})")
        logger.info(f"POST keys: {list(request.POST.keys())}")
        logger.info(f"FILES keys: {list(request.FILES.keys())}")
        for key, files in request.FILES.lists():
            logger.info(f"  {key}: {len(files)} файлов, sizes: {[f.size for f in files]}, names: {[f.name for f in files]}")
        total_size = sum(f.size for f in request.FILES.values())
        logger.info(f"Total upload size: {total_size / 1024 / 1024:.2f} MB")
        
        # Защита от дубликатов: проверяем, не создал ли пользователь агентство в последние 5 секунд
        five_seconds_ago = timezone.now() - datetime.timedelta(seconds=5)
        recent_agencies = Agencies.objects.filter(
            owner=request.user,
            created_at__gte=five_seconds_ago
        ).order_by('-created_at')
        
        if recent_agencies.exists():
            latest = recent_agencies.first()
            logger.warning(f"=== DUPLICATE AGENCY PREVENTED === User {request.user.user_id} tried to create duplicate, existing agency_id: {latest.agency_id}")
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("agency_creation_success"),
                })
            return redirect("agency_creation_success")
        
        form = AgencyForm(request.POST, request.FILES)
        
        # Проверяем наличие изображений через request.FILES.getlist
        creatives_files = request.FILES.getlist("creatives")
        if not creatives_files:
            form.add_error("creatives", "Загрузите хотя бы одно изображение.")
        
        if form.is_valid():
            logger.info("Form is valid, creating agency...")
            
            # Проверяем сколько агентств уже есть в базе перед созданием
            qs = Agencies.objects.filter(status="pending")
            existing_count = qs.count()
            logger.info(f"=== BEFORE SAVE === Pending agencies in DB: {existing_count}")
            logger.info(f"=== SQL QUERY === {qs.query}")
            
            agency = form.save(commit=False)
            agency.owner = request.user
            agency.created_at = timezone.now()
            agency.updated_at = timezone.now()
            agency.status = "pending"
            agency.planet_image = form.cleaned_data.get("planet_image")
            agency.save()
            
            logger.info(f"=== AFTER FIRST SAVE === Created agency with ID: {agency.agency_id}")
            
            # Сохраняем категорию и дополнительные поля в customization_data
            data = agency.customization_data or {}
            cat = form.cleaned_data.get("agency_category")
            if cat:
                data["agency_category"] = cat
            agency_services = form.cleaned_data.get("agency_services")
            if agency_services:
                data["agency_services"] = agency_services
            successful_projects = form.cleaned_data.get("successful_projects")
            if successful_projects is not None:
                data["successful_projects"] = successful_projects
            if data:
                agency.customization_data = data
                agency.save(update_fields=["customization_data"])
                logger.info(f"=== AFTER CUSTOMIZATION_DATA SAVE === agency_id: {agency.agency_id}")
            
            def try_save_file(file_obj, file_path):
                try:
                    default_storage.save(file_path, file_obj)
                    return True
                except Exception as e:
                    logger.error(f"Ошибка default_storage.save для {file_path}: {e}", exc_info=True)
                    try:
                        s3 = boto3.client(
                            's3',
                            endpoint_url=getattr(settings, 'AWS_S3_ENDPOINT_URL', None),
                            aws_access_key_id=getattr(settings, 'AWS_ACCESS_KEY_ID', None),
                            aws_secret_access_key=getattr(settings, 'AWS_SECRET_ACCESS_KEY', None),
                            region_name=getattr(settings, 'AWS_S3_REGION_NAME', None),
                            config=boto3.session.Config(s3={'addressing_style': getattr(settings, 'AWS_S3_ADDRESSING_STYLE', 'virtual')})
                        )
                        bucket = getattr(settings, 'AWS_STORAGE_BUCKET_NAME', None)
                        content_type = getattr(file_obj, 'content_type', 'application/octet-stream')
                        body_bytes = file_obj.read()
                        try:
                            s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ContentType=content_type, ACL='public-read')
                        except Exception:
                            s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ACL='public-read')
                        return True
                    except Exception as e2:
                        logger.error(f"Ошибка прямой загрузки в S3 для {file_path}: {e2}", exc_info=True)
                        return False
            
            # Обработка временных медиа-файлов из localStorage
            temp_media_data = request.POST.get('temp_media_data', '')
            if temp_media_data:
                try:
                    temp_files = json.loads(temp_media_data)
                    logger.info(f"Обработка {len(temp_files)} временных медиа-файлов для агентства")
                    
                    uploaded_content_type, _ = FileTypes.objects.get_or_create(type_name='uploaded_content')
                    entity_type, _ = EntityTypes.objects.get_or_create(type_name='agency')
                    
                    for temp_file in temp_files:
                        # Создаем запись в FileStorage для временного файла
                        file_id = str(uuid.uuid4())
                        safe_create_file_storage(
                            entity_type=entity_type,
                            entity_id=agency.agency_id,
                            file_type=uploaded_content_type,
                            file_url=file_id,
                            uploaded_at=timezone.now(),
                            startup=None,
                            original_file_name=temp_file.get('name', 'temp_file'),
                        )
                        logger.info(f"Создана запись для временного файла агентства: {temp_file.get('name')}")
                        
                except Exception as e:
                    logger.error(f"Ошибка обработки временных медиа-файлов агентства: {e}", exc_info=True)
                    messages.warning(request, "Не удалось обработать временные медиа-файлы.")

            # Upload files using shared helpers
            logo_ids = upload_entity_logo(form, agency, 'agency')
            upload_entity_catalog_card(form, agency, 'agency')
            creatives_ids = upload_entity_files(request, form, agency, 'agency', 'creatives', 'creative', 'image/jpeg')
            proofs_ids = upload_entity_files(request, form, agency, 'agency', 'proofs', 'proof', 'application/pdf')
            video_ids = []
            # Old catalog_card/creatives/proofs code removed — handled by shared helpers above
            videos = request.FILES.getlist("video")
            if videos:
                for video in videos:
                    if not hasattr(video, "name"):
                        logger.warning(f"Пропущено видео: {video}")
                        continue
                    try:
                        unique_filename = get_unique_filename(video.name, agency.agency_id, "video")
                        video_data = video.read()
                        content_type = getattr(video, 'content_type', 'video/mp4')
                        
                        video_data_b64 = base64.b64encode(video_data).decode('utf-8')
                        
                        upload_video_to_s3.delay(
                            video_data=video_data_b64,
                            video_name=video.name,
                            video_content_type=content_type,
                            entity_id=agency.agency_id,
                            original_filename=unique_filename,
                            entity_type_name='agency'
                        )
                        logger.info(f"Видео агентства отправлено в очередь Celery: {video.name}, размер: {len(video_data)} байт")
                        messages.info(request, f"Видео {video.name} загружается в фоновом режиме.")
                    except Exception as e:
                        logger.error(f"Ошибка отправки видео в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось отправить видео {video.name} на загрузку.")

            agency.logo_urls = logo_ids
            agency.creatives_urls = creatives_ids
            agency.proofs_urls = proofs_ids
            agency.video_urls = video_ids
            logger.info(f"UUID сохранены: logo={logo_ids}, creatives={creatives_ids}, proofs={proofs_ids}, videos={video_ids}. Файлы загружаются асинхронно через Celery.")
            
            # При создании автоматически добавляем первые 4 креатива в слайдер
            # (при редактировании пользователь может выбрать вручную через чекбоксы)
            slider_images = request.POST.getlist("slider_images")
            if not slider_images and creatives_ids:
                # Если slider_images не был отправлен, используем creatives_ids
                slider_images = creatives_ids[:4]
                logger.info(f"Автоматически установлены slider_images из creatives_ids: {slider_images}")
            elif len(slider_images) > 4:
                slider_images = slider_images[:4]
            agency.slider_images = slider_images
            
            agency.save()
            
            # Финальная проверка - сколько агентств с pending статусом после сохранения
            final_count = Agencies.objects.filter(status="pending").count()
            logger.info(f"=== CREATE_AGENCY SUCCESS === agency_id: {agency.agency_id}")
            logger.info(f"=== AFTER FINAL SAVE === Pending agencies in DB: {final_count}")
            
            # Проверяем нет ли дубликатов через ORM
            all_pending = list(Agencies.objects.filter(status="pending").values_list('agency_id', 'title'))
            logger.info(f"=== ALL PENDING AGENCIES (ORM) === {all_pending}")
            
            # Проверяем через distinct
            distinct_pending = list(Agencies.objects.filter(status="pending").distinct().values_list('agency_id', 'title'))
            logger.info(f"=== ALL PENDING AGENCIES (DISTINCT) === {distinct_pending}")
            
            # Проверяем через raw SQL
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute("SELECT agency_id, title FROM agencies WHERE status = 'pending'")
                raw_result = cursor.fetchall()
                logger.info(f"=== ALL PENDING AGENCIES (RAW SQL) === {raw_result}")
            
            # Отправляем уведомление в Telegram о новой заявке
            try:
                owner_name = f"{request.user.first_name or ''} {request.user.last_name or ''}".strip() or request.user.email
                send_telegram_new_entity_notification(
                    entity_type='agency',
                    entity_title=agency.title,
                    owner_name=owner_name,
                    owner_email=request.user.email,
                    entity_id=agency.agency_id
                )
            except Exception as e:
                logger.warning(f"Failed to send Telegram notification for agency {agency.agency_id}: {e}")

            clear_temp_files(request, 'agencyForm')
            messages.success(request, f'Агентство "{agency.title}" успешно создано и отправлено на модерацию!')
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("agency_creation_success"),
                })
            return redirect("agency_creation_success")
        else:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                logger.warning(f"=== CREATE_AGENCY FORM INVALID (AJAX) === User: {request.user.user_id}")
                logger.warning(f"Form errors: {form.errors.as_json()}")
                return JsonResponse({
                    "success": False,
                    "errors": {f: [x["message"] for x in e.get_json_data()] for f, e in form.errors.items()},
                    "non_field_errors": [str(e) for e in form.non_field_errors()],
                }, status=400)
            logger.warning(f"=== CREATE_AGENCY FORM INVALID === User: {request.user.user_id}")
            logger.warning(f"Form errors: {form.errors.as_json()}")
            messages.error(request, "Форма содержит ошибки.")
            return render(request, "accounts/create_agency.html", {"form": form})
    else:
        form = AgencyForm()
        clear_temp_files(request, 'agencyForm')
    return render(request, "accounts/create_agency.html", {"form": form})

@login_required
def create_specialist(request):
    allowed_roles = ["startuper", "moderator"]
    if not hasattr(request.user, "role") or request.user.role.role_name.lower() not in allowed_roles:
        messages.error(request, "Доступ к созданию профиля специалиста разрешён только пользователям с ролью 'Стартаппер' или 'Модератор'.")
        return redirect("home")
    if request.method == "POST":
        # Логирование входящих данных
        logger.info(f"=== CREATE_SPECIALIST START === User: {request.user.user_id} ({request.user.email})")
        logger.info(f"POST keys: {list(request.POST.keys())}")
        logger.info(f"FILES keys: {list(request.FILES.keys())}")
        for key, files in request.FILES.lists():
            logger.info(f"  {key}: {len(files)} файлов, sizes: {[f.size for f in files]}, names: {[f.name for f in files]}")
        total_size = sum(f.size for f in request.FILES.values())
        logger.info(f"Total upload size: {total_size / 1024 / 1024:.2f} MB")
        
        # Защита от дубликатов: проверяем, не создал ли пользователь специалиста в последние 5 секунд
        five_seconds_ago = timezone.now() - datetime.timedelta(seconds=5)
        recent_specialists = Specialists.objects.filter(
            owner=request.user,
            created_at__gte=five_seconds_ago
        ).order_by('-created_at')
        
        if recent_specialists.exists():
            latest = recent_specialists.first()
            logger.warning(f"=== DUPLICATE SPECIALIST PREVENTED === User {request.user.user_id} tried to create duplicate, existing specialist_id: {latest.specialist_id}")
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("specialist_creation_success"),
                })
            return redirect("specialist_creation_success")
        
        form = SpecialistForm(request.POST, request.FILES)
        
        # Проверяем наличие изображений через request.FILES.getlist
        creatives_files = request.FILES.getlist("creatives")
        if not creatives_files:
            form.add_error("creatives", "Загрузите хотя бы одно изображение.")
        
        if form.is_valid():
            logger.info("Form is valid, creating specialist...")
            spec = form.save(commit=False)
            spec.owner = request.user
            spec.created_at = timezone.now()
            spec.updated_at = timezone.now()
            spec.status = "pending"
            spec.planet_image = form.cleaned_data.get("planet_image")
            spec.save()
            
            # Сохраняем данные кастомизации
            cat = form.cleaned_data.get("specialist_category")
            
            data = spec.customization_data or {}
            if cat:
                data["specialist_category"] = cat
            
            if data:
                spec.customization_data = data
                spec.save(update_fields=["customization_data"])
            
            def try_save_file(file_obj, file_path):
                try:
                    default_storage.save(file_path, file_obj)
                    return True
                except Exception as e:
                    logger.error(f"Ошибка default_storage.save для {file_path}: {e}", exc_info=True)
                    try:
                        s3 = boto3.client(
                            's3',
                            endpoint_url=getattr(settings, 'AWS_S3_ENDPOINT_URL', None),
                            aws_access_key_id=getattr(settings, 'AWS_ACCESS_KEY_ID', None),
                            aws_secret_access_key=getattr(settings, 'AWS_SECRET_ACCESS_KEY', None),
                            region_name=getattr(settings, 'AWS_S3_REGION_NAME', None),
                            config=boto3.session.Config(s3={'addressing_style': getattr(settings, 'AWS_S3_ADDRESSING_STYLE', 'virtual')})
                        )
                        bucket = getattr(settings, 'AWS_STORAGE_BUCKET_NAME', None)
                        content_type = getattr(file_obj, 'content_type', 'application/octet-stream')
                        body_bytes = file_obj.read()
                        try:
                            s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ContentType=content_type, ACL='public-read')
                        except Exception:
                            s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ACL='public-read')
                        return True
                    except Exception as e2:
                        logger.error(f"Ошибка прямой загрузки в S3 для {file_path}: {e2}", exc_info=True)
                        return False
            
            # Обработка временных медиа-файлов из localStorage
            temp_media_data = request.POST.get('temp_media_data', '')
            if temp_media_data:
                try:
                    temp_files = json.loads(temp_media_data)
                    logger.info(f"Обработка {len(temp_files)} временных медиа-файлов для специалиста")
                    
                    uploaded_content_type, _ = FileTypes.objects.get_or_create(type_name='uploaded_content')
                    entity_type, _ = EntityTypes.objects.get_or_create(type_name='specialist')
                    
                    for temp_file in temp_files:
                        # Создаем запись в FileStorage для временного файла
                        file_id = str(uuid.uuid4())
                        safe_create_file_storage(
                            entity_type=entity_type,
                            entity_id=spec.specialist_id,
                            file_type=uploaded_content_type,
                            file_url=file_id,
                            uploaded_at=timezone.now(),
                            startup=None,
                            original_file_name=temp_file.get('name', 'temp_file'),
                        )
                        logger.info(f"Создана запись для временного файла специалиста: {temp_file.get('name')}")
                        
                except Exception as e:
                    logger.error(f"Ошибка обработки временных медиа-файлов специалиста: {e}", exc_info=True)
                    messages.warning(request, "Не удалось обработать временные медиа-файлы.")

            # Upload files using shared helpers
            logo_ids = upload_entity_logo(form, spec, 'specialist')
            upload_entity_catalog_card(form, spec, 'specialist')
            creatives_ids = upload_entity_files(request, form, spec, 'specialist', 'creatives', 'creative', 'image/jpeg')
            proofs_ids = upload_entity_files(request, form, spec, 'specialist', 'proofs', 'proof', 'application/pdf')
            video_ids = []

            videos = request.FILES.getlist("video")
            if videos:
                for video in videos:
                    if not hasattr(video, "name"):
                        logger.warning(f"Пропущено видео: {video}")
                        continue
                    try:
                        unique_filename = get_unique_filename(video.name, spec.specialist_id, "video")
                        video_data = video.read()
                        content_type = getattr(video, 'content_type', 'video/mp4')
                        
                        video_data_b64 = base64.b64encode(video_data).decode('utf-8')
                        
                        upload_video_to_s3.delay(
                            video_data=video_data_b64,
                            video_name=video.name,
                            video_content_type=content_type,
                            entity_id=spec.specialist_id,
                            original_filename=unique_filename,
                            entity_type_name='specialist'
                        )
                        logger.info(f"Видео специалиста отправлено в очередь Celery: {video.name}, размер: {len(video_data)} байт")
                        messages.info(request, f"Видео {video.name} загружается в фоновом режиме.")
                    except Exception as e:
                        logger.error(f"Ошибка отправки видео в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось отправить видео {video.name} на загрузку.")

            spec.logo_urls = logo_ids
            spec.creatives_urls = creatives_ids
            spec.proofs_urls = proofs_ids
            spec.video_urls = video_ids
            logger.info(f"UUID сохранены: logo={logo_ids}, creatives={creatives_ids}, proofs={proofs_ids}, videos={video_ids}. Файлы загружаются асинхронно через Celery.")
            
            # При создании автоматически добавляем первые 4 креатива в слайдер
            # (при редактировании пользователь может выбрать вручную через чекбоксы)
            slider_images = request.POST.getlist("slider_images")
            if not slider_images and creatives_ids:
                # Если slider_images не был отправлен, используем creatives_ids
                slider_images = creatives_ids[:4]
                logger.info(f"Автоматически установлены slider_images из creatives_ids: {slider_images}")
            elif len(slider_images) > 4:
                slider_images = slider_images[:4]
            spec.slider_images = slider_images
            
            spec.save()
            logger.info(f"=== CREATE_SPECIALIST SUCCESS === specialist_id: {spec.specialist_id}")
            
            # Отправляем уведомление в Telegram о новой заявке
            try:
                owner_name = f"{request.user.first_name or ''} {request.user.last_name or ''}".strip() or request.user.email
                send_telegram_new_entity_notification(
                    entity_type='specialist',
                    entity_title=spec.title,
                    owner_name=owner_name,
                    owner_email=request.user.email,
                    entity_id=spec.specialist_id
                )
            except Exception as e:
                logger.warning(f"Failed to send Telegram notification for specialist {spec.specialist_id}: {e}")

            clear_temp_files(request, 'specialistForm')
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("specialist_creation_success"),
                })
            messages.success(request, f'Профиль специалиста "{spec.title}" успешно создан и отправлен на модерацию!')
            return redirect("specialist_creation_success")
        else:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                logger.warning(f"=== CREATE_SPECIALIST FORM INVALID (AJAX) === User: {request.user.user_id}")
                logger.warning(f"Form errors: {form.errors.as_json()}")
                return JsonResponse({
                    "success": False,
                    "errors": {f: [x["message"] for x in e.get_json_data()] for f, e in form.errors.items()},
                    "non_field_errors": [str(e) for e in form.non_field_errors()],
                }, status=400)
            logger.warning(f"=== CREATE_SPECIALIST FORM INVALID === User: {request.user.user_id}")
            logger.warning(f"Form errors: {form.errors.as_json()}")
            messages.error(request, "Форма содержит ошибки.")
            return render(request, "accounts/create_specialist.html", {"form": form})
    else:
        form = SpecialistForm()
        clear_temp_files(request, 'specialistForm')
    return render(request, "accounts/create_specialist.html", {"form": form})

@login_required
def startup_creation_success(request):
    return render(request, "accounts/startup_creation_success.html")

@login_required
def agency_creation_success(request):
    return render(request, "accounts/agency_creation_success.html")

@login_required
def franchise_creation_success(request):
    return render(request, "accounts/franchise_creation_success.html")

@login_required
def specialist_creation_success(request):
    return render(request, "accounts/specialist_creation_success.html")

@login_required
def delete_message(request, message_id):
    message = get_object_or_404(Messages, message_id=message_id)
    chat = message.conversation
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    if request.user.role and request.user.role.role_name.lower() == "moderator":
        message.is_deleted = True
        message.save()
        return JsonResponse({"success": True})
    return JsonResponse(
        {"success": False, "error": "Только модератор может удалить сообщение"},
        status=403,
    )
@login_required
def reorder_files(request, entity_type, entity_id):
    """
    AJAX view для изменения порядка файлов (креативов)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method allowed'}, status=405)
    
    try:
        # Получаем данные из запроса
        data = json.loads(request.body)
        file_type = data.get('file_type')  # 'creative', 'proof', 'video'
        new_order = data.get('new_order')  # список file_id в новом порядке
        
        if not file_type or not new_order:
            return JsonResponse({'success': False, 'error': 'Missing file_type or new_order'}, status=400)
        
        # Получаем сущность
        if entity_type == 'startup':
            entity = get_object_or_404(Startups, startup_id=entity_id)
            if not (request.user == entity.owner or request.user.role.role_name == 'moderator'):
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        elif entity_type == 'franchise':
            entity = get_object_or_404(Franchises, franchise_id=entity_id)
            if not (request.user == entity.owner or request.user.role.role_name == 'moderator'):
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        elif entity_type == 'agency':
            entity = get_object_or_404(Agencies, agency_id=entity_id)
            if not (request.user == entity.owner or request.user.role.role_name == 'moderator'):
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        elif entity_type == 'specialist':
            entity = get_object_or_404(Specialists, specialist_id=entity_id)
            if not (request.user == entity.owner or request.user.role.role_name == 'moderator'):
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        else:
            return JsonResponse({'success': False, 'error': 'Invalid entity type'}, status=400)
        
        # Обновляем порядок файлов
        if file_type == 'creative':
            entity.creatives_urls = new_order
        elif file_type == 'proof':
            entity.proofs_urls = new_order
        elif file_type == 'video':
            entity.video_urls = new_order
        else:
            return JsonResponse({'success': False, 'error': 'Invalid file type'}, status=400)
        
        entity.save()
        
        return JsonResponse({'success': True, 'message': 'File order updated successfully'})
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error reordering files: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Internal server error'}, status=500)


@login_required
def delete_file(request, entity_type, entity_id):
    """
    AJAX view для удаления отдельного файла
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method allowed'}, status=405)
    
    try:
        # Получаем данные из запроса
        data = json.loads(request.body)
        file_id = data.get('file_id')
        file_type = data.get('file_type')  # 'creative', 'proof', 'video'
        
        if not file_id or not file_type:
            return JsonResponse({'success': False, 'error': 'Missing file_id or file_type'}, status=400)
        
        # Получаем сущность
        if entity_type == 'startup':
            entity = get_object_or_404(Startups, startup_id=entity_id)
            if not (request.user == entity.owner or request.user.role.role_name == 'moderator'):
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        elif entity_type == 'franchise':
            entity = get_object_or_404(Franchises, franchise_id=entity_id)
            if not (request.user == entity.owner or request.user.role.role_name == 'moderator'):
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        elif entity_type == 'agency':
            entity = get_object_or_404(Agencies, agency_id=entity_id)
            if not (request.user == entity.owner or request.user.role.role_name == 'moderator'):
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        elif entity_type == 'specialist':
            entity = get_object_or_404(Specialists, specialist_id=entity_id)
            if not (request.user == entity.owner or request.user.role.role_name == 'moderator'):
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        else:
            return JsonResponse({'success': False, 'error': 'Invalid entity type'}, status=400)
        
        # Получаем информацию о файле
        entity_type_obj, _ = EntityTypes.objects.get_or_create(type_name=entity_type)
        file_storage = FileStorage.objects.filter(
            entity_type=entity_type_obj,
            entity_id=entity_id,
            file_url=file_id
        ).first()
        
        if not file_storage:
            return JsonResponse({'success': False, 'error': 'File not found'}, status=404)
        
        # Удаляем файл из S3
        if file_type == 'creative':
            file_path = f"{entity_type}s/{entity_id}/creatives/{file_id}_{file_storage.original_file_name or 'unknown'}"
        elif file_type == 'proof':
            file_path = f"{entity_type}s/{entity_id}/proofs/{file_id}_{file_storage.original_file_name or 'unknown'}"
        elif file_type == 'video':
            file_path = f"{entity_type}s/{entity_id}/videos/{file_id}_{file_storage.original_file_name or 'unknown'}"
        else:
            return JsonResponse({'success': False, 'error': 'Invalid file type'}, status=400)
        
        try:
            delete_file_from_s3(file_path)
        except Exception as e:
            logger.warning(f"Failed to delete file from S3: {e}")
        
        # Удаляем запись из базы данных
        file_storage.delete()
        
        # Удаляем из списка URL
        if file_type == 'creative' and entity.creatives_urls:
            entity.creatives_urls = [url for url in entity.creatives_urls if url != file_id]
        elif file_type == 'proof' and entity.proofs_urls:
            entity.proofs_urls = [url for url in entity.proofs_urls if url != file_id]
        elif file_type == 'video' and entity.video_urls:
            entity.video_urls = [url for url in entity.video_urls if url != file_id]
        
        entity.save()
        
        return JsonResponse({'success': True, 'message': 'File deleted successfully'})
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error deleting file: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Internal server error'}, status=500)


@login_required
def remove_participant(request, chat_id):
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    user_id = request.POST.get("user_id")
    if not user_id:
        return JsonResponse(
            {"success": False, "error": "Не указан пользователь"}, status=400
        )
    try:
        user_to_remove = Users.objects.get(user_id=user_id)
    except Users.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "Пользователь не найден"}, status=404
        )
    if (
        request.user.role
        and request.user.role.role_name.lower() == "moderator"
        and chat.is_group_chat
    ):
        participant = chat.chatparticipants_set.filter(user=user_to_remove).first()
        if participant:
            participant.delete()
            chat.updated_at = timezone.now()
            chat.save()
            return JsonResponse({"success": True})
    return JsonResponse(
        {
            "success": False,
            "error": "Только модератор может исключить участника из группового чата",
        },
        status=403,
    )
@login_required
def edit_startup(request, startup_id):
    logger.debug(f"Request method: {request.method}")
    logger.debug(f"Request POST: {request.POST}")
    logger.debug(f"Request FILES: {dict(request.FILES)}")
    startup = get_object_or_404(Startups, startup_id=startup_id)
    
    # Очищаем дубликаты в базе данных
    if startup.creatives_urls:
        # Удаляем дубликаты из creatives_urls
        original_count = len(startup.creatives_urls)
        startup.creatives_urls = list(dict.fromkeys(startup.creatives_urls))  # Удаляет дубликаты, сохраняя порядок
        if len(startup.creatives_urls) != original_count:
            logger.info(f"Очищено {original_count - len(startup.creatives_urls)} дубликатов из creatives_urls")
            startup.save(update_fields=['creatives_urls'])
    
    if startup.proofs_urls:
        # Удаляем дубликаты из proofs_urls
        original_count = len(startup.proofs_urls)
        startup.proofs_urls = list(dict.fromkeys(startup.proofs_urls))
        if len(startup.proofs_urls) != original_count:
            logger.info(f"Очищено {original_count - len(startup.proofs_urls)} дубликатов из proofs_urls")
            startup.save(update_fields=['proofs_urls'])
    
    if startup.video_urls:
        # Удаляем дубликаты из video_urls
        original_count = len(startup.video_urls)
        startup.video_urls = list(dict.fromkeys(startup.video_urls))
        if len(startup.video_urls) != original_count:
            logger.info(f"Очищено {original_count - len(startup.video_urls)} дубликатов из video_urls")
            startup.save(update_fields=['video_urls'])
    
    if not (
        request.user == startup.owner
        or (
            hasattr(request.user, "role")
            and request.user.role
            and request.user.role.role_name == "moderator"
        )
    ):
        messages.error(request, "У вас нет прав для редактирования этого стартапа.")
        return redirect("startup_detail", slug=startup.slug or startup_id)
    timeline = StartupTimeline.objects.filter(startup=startup)
    timeline_steps = timeline
    if request.method == "POST":
        # Сохраняем оригинальные данные для сравнения
        original_data = {
            'title': startup.title,
            'short_description': startup.short_description,
            'description': startup.description,
            'terms': startup.terms,
            'funding_goal': startup.funding_goal,
            'amount_raised': startup.amount_raised,
            'valuation': startup.valuation,
            'pitch_deck_url': startup.pitch_deck_url,
            'planet_image': startup.planet_image,
        }
        
        form = StartupEditForm(request.POST, request.FILES, instance=startup)
        
        if form.is_valid():
            logger.info("=== ФОРМА ВАЛИДНА ===")
            logger.info(f"Form data: {form.cleaned_data}")
            startup = form.save(commit=False)
            logger.info(f"Стартап загружен: {startup.startup_id}")
            logger.info(f"Planet image value: {startup.planet_image}")
            
            # Проверяем были ли изменения
            has_changes = False
            
            # Проверяем основные поля
            if (startup.title != original_data['title'] or 
                startup.short_description != original_data['short_description'] or
                startup.description != original_data['description'] or
                startup.terms != original_data['terms'] or
                startup.funding_goal != original_data['funding_goal'] or
                startup.amount_raised != original_data['amount_raised'] or
                startup.valuation != original_data['valuation'] or
                startup.pitch_deck_url != original_data['pitch_deck_url'] or
                startup.planet_image != original_data['planet_image']):
                has_changes = True
            
            # Проверяем загружены ли новые файлы
            if not has_changes:
                # Проверяем файлы в request.FILES
                if request.FILES:
                    has_changes = True
                    logger.info("Найдены файлы в request.FILES - установлен has_changes = True")
                    logger.info(f"Файлы: {list(request.FILES.keys())}")
                    for key, file_list in request.FILES.lists():
                        logger.info(f"  {key}: {len(file_list)} файлов")
                        for i, file_obj in enumerate(file_list):
                            logger.info(f"    Файл {i}: {file_obj.name} ({file_obj.size} байт)")
                
                # Также проверяем cleaned_data для файлов
                if not has_changes:
                    file_fields = ['logo', 'creatives', 'proofs', 'video']
                    for field_name in file_fields:
                        field_data = form.cleaned_data.get(field_name)
                        if field_data:
                            has_changes = True
                            logger.info(f"Найдены файлы в form.cleaned_data['{field_name}'] - установлен has_changes = True")
                            if isinstance(field_data, list):
                                logger.info(f"  {field_name}: {len(field_data)} файлов")
                            else:
                                logger.info(f"  {field_name}: 1 файл")
                            break
                
                # Дополнительная проверка - если есть файлы в request.FILES, но нет в cleaned_data
                if not has_changes and request.FILES:
                    # Проверяем, есть ли файлы с реальным содержимым
                    for key, file_list in request.FILES.lists():
                        for file_obj in file_list:
                            if file_obj.size > 0:  # Файл не пустой
                                has_changes = True
                                logger.info(f"Найден непустой файл {key}: {file_obj.name} ({file_obj.size} байт) - установлен has_changes = True")
                                break
                        if has_changes:
                            break
            
            # Проверяем удалены ли файлы
            if not has_changes:
                deleted_files_json = request.POST.get('deleted_files', '[]')
                try:
                    deleted_files = json.loads(deleted_files_json)
                    if deleted_files:
                        has_changes = True
                except json.JSONDecodeError:
                    pass
            
            # Проверяем изменения в этапах
            if not has_changes:
                for i in range(1, 6):
                    description = request.POST.get(f"step_description_{i}", "").strip()
                    if description:
                        try:
                            timeline_entry = StartupTimeline.objects.get(
                                startup=startup,
                                step_number=i
                            )
                            if timeline_entry.description != description:
                                has_changes = True
                                break
                        except StartupTimeline.DoesNotExist:
                            # Новый этап - это изменение
                            has_changes = True
                            break
            
            # Принудительная проверка файлов - если есть файлы в request.FILES, устанавливаем has_changes = True
            if request.FILES:
                has_changes = True
                logger.info("ПРИНУДИТЕЛЬНО установлен has_changes = True из-за наличия файлов в request.FILES")
            
            # Устанавливаем статус в зависимости от наличия изменений
            logger.info(f"has_changes: {has_changes}")
            
            # Логика изменения статуса: только approved -> pending
            if has_changes and startup.status == "approved":
                startup.status = "pending"
                startup.is_edited = True
                startup.save(update_fields=['status', 'is_edited'])
                logger.info("Статус изменен: approved -> pending")
            elif not has_changes and startup.status == "pending":
                # Если нет изменений, но статус pending - оставляем как есть
                logger.info("Статус остается: pending (изменений нет)")
            else:
                # Для всех остальных случаев (approved без изменений, другие статусы)
                logger.info(f"Статус остается: {startup.status}")
            
            startup.updated_at = timezone.now()
            if "step_number" in request.POST:
                new_step = int(request.POST.get("step_number"))
                startup.step_number = new_step
            investment_type = form.cleaned_data.get("investment_type")
            if investment_type == "invest":
                startup.only_invest = True
                startup.only_buy = False
                startup.both_mode = False
            elif investment_type == "buy":
                startup.only_invest = False
                startup.only_buy = True
                startup.both_mode = False
            elif investment_type == "both":
                startup.only_invest = False
                startup.only_buy = False
                startup.both_mode = True
            logger.info("=== СОХРАНЕНИЕ СТАРТАПА ===")
            startup.save()
            logger.info("Стартап сохранен")
            
            logo_ids = startup.logo_urls or []
            creative_ids = startup.creatives_urls or []  # Существующие креативы
            proofs_ids = startup.proofs_urls or []  # Существующие документы
            video_ids = startup.video_urls or []  # Существующие видео
            logger.info("Переменные инициализированы")
            # Обработка логотипа — синхронная загрузка в S3
            logo = request.FILES.get("logo")
            if logo and logo.size > 0:
                logo_id = str(uuid.uuid4())
                try:
                    logo.seek(0)
                    file_data = logo.read()
                    content_type = getattr(logo, 'content_type', 'image/jpeg')
                    unique_filename = get_unique_filename(logo.name, startup.startup_id, "logo")

                    result = upload_file_to_s3_sync(
                        file_data=file_data,
                        file_name=logo.name,
                        content_type=content_type,
                        entity_type_name='startup',
                        entity_id=startup.startup_id,
                        file_type_name='logo',
                        original_filename=unique_filename,
                        file_id=logo_id
                    )
                    if result:
                        logo_ids = [logo_id]
                        logger.info(f"Логотип стартапа загружен синхронно: {logo.name}, размер: {len(file_data)} байт")
                    else:
                        logger.error(f"Синхронная загрузка логотипа стартапа не удалась: {logo.name}")
                        messages.warning(request, "Не удалось сохранить логотип")
                        logo_ids = startup.logo_urls or []
                except Exception as e:
                    logger.error(f"Ошибка загрузки логотипа стартапа: {e}", exc_info=True)
                    messages.warning(request, f"Не удалось сохранить логотип: {e}")
                    logo_ids = startup.logo_urls or []
            else:
                logo_ids = startup.logo_urls or []
            # Обработка креативов
            creatives = request.FILES.getlist("creatives")
            proofs = request.FILES.getlist("proofs")
            videos = request.FILES.getlist("video")
            
            # Проверка лимитов файлов
            if len(creatives) > 10:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 10 изображений"}, status=400)
                messages.error(request, "Максимально 10 изображений")
                return render(request, "accounts/edit_startup.html", {"form": form, "startup": startup, "timeline_steps": timeline_steps})
            
            if len(proofs) > 15:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 15 документов"}, status=400)
                messages.error(request, "Максимально 15 документов")
                return render(request, "accounts/edit_startup.html", {"form": form, "startup": startup, "timeline_steps": timeline_steps})
            
            if len(videos) > 1:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 1 видео"}, status=400)
                messages.error(request, "Максимально 1 видео")
                return render(request, "accounts/edit_startup.html", {"form": form, "startup": startup, "timeline_steps": timeline_steps})
            
            if creatives:
                creative_type = FileTypes.objects.get(type_name="creative")
                entity_type = EntityTypes.objects.get(type_name="startup")
                creative_ids = []  # Инициализируем список для новых файлов
                # Не очищаем существующие файлы, Celery добавит новые через atomic append
                for creative_file in creatives:
                    if not hasattr(creative_file, "name"):
                        logger.warning(
                            f"Пропущен креатив, так как это не файл: {creative_file}"
                        )
                        continue

                    # Проверяем, не существует ли уже файл с таким именем
                    existing_file = FileStorage.objects.filter(
                        entity_type=entity_type,
                        entity_id=startup.startup_id,
                        file_type=creative_type,
                        original_file_name=creative_file.name
                    ).first()

                    if existing_file:
                        logger.warning(f"Креатив {creative_file.name} уже существует, пропускаем создание")
                        continue

                    try:
                        unique_filename = get_unique_filename(creative_file.name, startup.startup_id, "creative")
                        creative_id = str(uuid.uuid4())
                        file_data = creative_file.read()
                        content_type = getattr(creative_file, 'content_type', 'image/jpeg')
                        result = upload_file_to_s3_sync(
                            file_data=file_data,
                            file_name=creative_file.name,
                            content_type=content_type,
                            entity_type_name='startup',
                            entity_id=startup.startup_id,
                            file_type_name='creative',
                            original_filename=unique_filename,
                            file_id=creative_id
                        )
                        if result:
                            creative_ids.append(creative_id)
                        logger.info(f"Креатив стартапа загружен: {creative_file.name}, размер: {len(file_data)} байт")
                    except Exception as e:
                        logger.error(f"Ошибка отправки креатива стартапа в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось сохранить изображение {creative_file.name}")
            else:
                creative_ids = startup.creatives_urls or []
            if proofs:
                proof_type = FileTypes.objects.get(type_name="proof")
                entity_type = EntityTypes.objects.get(type_name="startup")
                proofs_ids = []  # Инициализируем список для новых файлов
                # Не очищаем существующие файлы, Celery добавит через atomic append
                for proof_file in proofs:
                    if not hasattr(proof_file, "name"):
                        logger.warning(
                            f"Пропущен пруф, так как это не файл: {proof_file}"
                        )
                        continue

                    # Проверяем, не существует ли уже файл с таким именем
                    existing_file = FileStorage.objects.filter(
                        entity_type=entity_type,
                        entity_id=startup.startup_id,
                        file_type=proof_type,
                        original_file_name=proof_file.name
                    ).first()

                    if existing_file:
                        logger.warning(f"Файл {proof_file.name} уже существует, пропускаем создание")
                        continue

                    try:
                        unique_filename = get_unique_filename(proof_file.name, startup.startup_id, "proof")
                        proof_id = str(uuid.uuid4())
                        file_data = proof_file.read()
                        content_type = getattr(proof_file, 'content_type', 'application/pdf')
                        result = upload_file_to_s3_sync(
                            file_data=file_data,
                            file_name=proof_file.name,
                            content_type=content_type,
                            entity_type_name='startup',
                            entity_id=startup.startup_id,
                            file_type_name='proof',
                            original_filename=unique_filename,
                            file_id=proof_id
                        )
                        if result:
                            proofs_ids.append(proof_id)
                        logger.info(f"Пруф стартапа загружен: {proof_file.name}, размер: {len(file_data)} байт")
                    except Exception as e:
                        logger.error(f"Ошибка отправки пруфа стартапа в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось сохранить документ {proof_file.name}")
            videos = request.FILES.getlist("video")
            if videos:
                video_type, _ = FileTypes.objects.get_or_create(type_name="video")
                entity_type = EntityTypes.objects.get(type_name="startup")
                for video in videos:
                    if not hasattr(video, "name"):
                        logger.warning(f"Пропущено видео, так как это не файл: {video}")
                        continue
                    
                    try:
                        unique_filename = get_unique_filename(video.name, startup.startup_id, "video")
                        video_data = video.read()
                        content_type = getattr(video, 'content_type', 'video/mp4')
                        
                        video_data_b64 = base64.b64encode(video_data).decode('utf-8')
                        
                        upload_video_to_s3.delay(
                            video_data=video_data_b64,
                            video_name=video.name,
                            video_content_type=content_type,
                            entity_id=startup.startup_id,
                            original_filename=unique_filename,
                            entity_type_name='startup'
                        )
                        logger.info(f"Видео отправлено в очередь Celery: {video.name}, размер: {len(video_data)} байт")
                        messages.info(request, f"Видео {video.name} загружается в фоновом режиме.")
                    except Exception as e:
                        logger.error(f"Ошибка отправки видео в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось отправить видео {video.name} на загрузку.")
                logger.info("Видео загружается асинхронно через Celery, video_urls обновится автоматически")
                if startup.video_urls is None:
                    startup.video_urls = []
            startup.logo_urls = logo_ids
            if 'creative_ids' in locals() and creative_ids:
                existing_creatives = startup.creatives_urls or []
                new_creatives = [url for url in creative_ids if url not in existing_creatives]
                startup.creatives_urls = existing_creatives + new_creatives
                logger.info(f"Добавлено {len(new_creatives)} новых креативов (было {len(existing_creatives)}, дубликатов пропущено: {len(creative_ids) - len(new_creatives)})")
            if 'proofs_ids' in locals() and proofs_ids:
                existing_proofs = startup.proofs_urls or []
                new_proofs = [url for url in proofs_ids if url not in existing_proofs]
                startup.proofs_urls = existing_proofs + new_proofs
                logger.info(f"Добавлено {len(new_proofs)} новых документов (было {len(existing_proofs)}, дубликатов пропущено: {len(proofs_ids) - len(new_proofs)})")
            # Обработка удаленных файлов
            deleted_files_json = request.POST.get('deleted_files', '[]')
            try:
                deleted_files = json.loads(deleted_files_json)
                for deleted_file in deleted_files:
                    file_id = deleted_file.get('id')
                    file_type = deleted_file.get('type')
                    if file_id and file_type:
                        from django.db.models import Q
                        entity_type = EntityTypes.objects.get(type_name="startup")
                        
                        # Получаем информацию о файле перед удалением
                        file_storage = FileStorage.objects.filter(
                            Q(entity_type=entity_type, entity_id=startup.startup_id) | Q(startup=startup),
                            file_url=file_id
                        ).first()
                        
                        if file_storage:
                            # Удаляем файл из S3 - используем правильный путь
                            if file_type == 'creative':
                                # Используем оригинальное имя файла как при загрузке
                                file_path = f"startups/{startup.startup_id}/creatives/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            elif file_type == 'proof':
                                # Используем оригинальное имя файла как при загрузке
                                file_path = f"startups/{startup.startup_id}/proofs/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            elif file_type == 'video':
                                file_path = f"startups/{startup.startup_id}/videos/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            else:
                                file_path = f"startups/{startup.startup_id}/{file_type}s/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            
                            logger.info(f"Попытка удаления файла из S3: {file_path}")
                            delete_file_from_s3(file_path)
                        
                        # Удаляем запись из базы данных
                        FileStorage.objects.filter(
                            Q(entity_type=entity_type, entity_id=startup.startup_id) | Q(startup=startup),
                            file_url=file_id
                        ).delete()
                        
                        # Удаляем из списка URL
                        if file_type == 'creative' and startup.creatives_urls:
                            startup.creatives_urls = [url for url in startup.creatives_urls if url != file_id]
                        elif file_type == 'proof' and startup.proofs_urls:
                            startup.proofs_urls = [url for url in startup.proofs_urls if url != file_id]
                        elif file_type == 'video' and startup.video_urls:
                            startup.video_urls = [url for url in startup.video_urls if url != file_id]
                        logger.info(f"Удален файл {file_type}: {file_id}")
            except json.JSONDecodeError:
                logger.warning("Ошибка при разборе deleted_files JSON")
            
            # Обработка этапов
            for i in range(1, 6):
                description = request.POST.get(f"step_description_{i}", "").strip()
                if description:
                    timeline_entry, created = StartupTimeline.objects.get_or_create(
                        startup=startup,
                        step_number=i,
                        defaults={"title": f"Этап {i}", "description": description},
                    )
                    if not created and timeline_entry.description != description:
                        timeline_entry.description = description
                        timeline_entry.save()
            
            slider_images = request.POST.getlist("slider_images")
            if len(slider_images) > 4:
                slider_images = slider_images[:4]
            startup.slider_images = slider_images
            
            # Обработка catalog_card_image
            catalog_card_image = form.cleaned_data.get("catalog_card_image") or request.FILES.get("catalog_card_image")
            if catalog_card_image and hasattr(catalog_card_image, 'read'):
                catalog_card_id = str(uuid.uuid4())
                # Конвертируем в WebP
                processed_catalog_image, processed_catalog_name, _ = process_uploaded_image(catalog_card_image, quality=85)
                base_name = os.path.splitext(processed_catalog_name)[0]
                ext = os.path.splitext(processed_catalog_name)[1]
                safe_base_name = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
                safe_name = slugify(safe_base_name) + ext
                file_path = f"catalog_cards/{catalog_card_id}_{safe_name}"
                try:
                    logger.info(f"Попытка сохранить изображение карточки по пути: {file_path}")
                    # Используем boto3 напрямую с ACL='public-read'
                    s3 = boto3.client(
                        's3',
                        endpoint_url=getattr(settings, 'AWS_S3_ENDPOINT_URL', None),
                        aws_access_key_id=getattr(settings, 'AWS_ACCESS_KEY_ID', None),
                        aws_secret_access_key=getattr(settings, 'AWS_SECRET_ACCESS_KEY', None),
                        region_name=getattr(settings, 'AWS_S3_REGION_NAME', None),
                        config=boto3.session.Config(s3={'addressing_style': getattr(settings, 'AWS_S3_ADDRESSING_STYLE', 'virtual')})
                    )
                    bucket = getattr(settings, 'AWS_STORAGE_BUCKET_NAME', None)
                    content_type = 'image/webp' if processed_catalog_name.endswith('.webp') else getattr(catalog_card_image, 'content_type', 'application/octet-stream')
                    if hasattr(processed_catalog_image, 'read'):
                        body_bytes = processed_catalog_image.read()
                    else:
                        body_bytes = processed_catalog_image.getvalue() if hasattr(processed_catalog_image, 'getvalue') else processed_catalog_image
                    try:
                        s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ContentType=content_type, ACL='public-read')
                    except Exception:
                        s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ACL='public-read')
                    logger.info(f"Изображение карточки успешно сохранено по пути: {file_path}")
                    startup.catalog_card_image = f"{catalog_card_id}_{safe_name}"
                    logger.info(f"Изображение карточки сохранено с ID: {catalog_card_id}")
                except Exception as e:
                    logger.error(f"Ошибка сохранения изображения карточки: {e}", exc_info=True)
                    messages.warning(request, "Не удалось сохранить изображение для карточки.")
            
            logger.info("=== ФИНАЛЬНОЕ СОХРАНЕНИЕ ===")
            startup.save()
            logger.info("Стартап финально сохранен")
            logger.info("=== Обновление стартапа ===")
            logger.info(f"Стартап ID: {startup.startup_id}")
            if logo:
                logger.info(f"Логотип: {logo.name}, размер: {logo.size} байт")
                logger.info(
                    f"ID логотипа: {logo_ids[0] if logo_ids else 'Не сохранён'}"
                )
            else:
                logger.info("Логотип не загружен")
            if creatives:
                logger.info(f"Креативы: {len(creatives)} файлов")
                for i, creative_file in enumerate(creatives, 1):
                    if hasattr(creative_file, "name"):
                        logger.info(
                            f"Креатив {i}: {creative_file.name}, размер: {creative_file.size} байт"
                        )
                    else:
                        logger.info(
                            f"Креатив {i}: Неверный формат (не файл): {creative_file}"
                        )
            else:
                logger.info("Креативы не загружены")
            if proofs:
                logger.info(f"Пруфы: {len(proofs)} файлов")
                for i, proof_file in enumerate(proofs, 1):
                    if hasattr(proof_file, "name"):
                        logger.info(
                            f"Пруф {i}: {proof_file.name}, размер: {proof_file.size} байт"
                        )
                    else:
                        logger.info(
                            f"Пруф {i}: Неверный формат (не файл): {proof_file}"
                        )
            else:
                logger.info("Пруфы не загружены")
            if videos:
                logger.info(f"Видео: {len(videos)} файлов")
                for i, video_file in enumerate(videos, 1):
                    if hasattr(video_file, "name"):
                        logger.info(
                            f"Видео {i}: {video_file.name}, размер: {video_file.size} байт"
                        )
                    else:
                        logger.info(
                            f"Видео {i}: Неверный формат (не файл): {video_file}"
                        )
            else:
                logger.info("Видео не загружено")
            logger.info("=== S3 storage check ===")
            logger.info(f"Storage backend: {default_storage.__class__.__name__}")
            logger.info(f"S3 endpoint configured: {'yes' if getattr(settings, 'AWS_S3_ENDPOINT_URL', None) else 'no'}")
            logger.info(f"S3 bucket configured: {'yes' if getattr(settings, 'AWS_STORAGE_BUCKET_NAME', None) else 'no'}")
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                # Собираем информацию о файлах
                files_info = {
                    "logo_saved": bool(logo and hasattr(logo, 'size') and logo.size > 0),
                    "creatives_saved": len(creatives) if creatives else 0,
                    "proofs_saved": len(proofs) if proofs else 0,
                    "videos_saved": len(videos) if videos else 0,
                }
                
                # Собираем информацию о статусе
                previous_status = startup.status_id.status_name if startup.status_id else "неизвестно"
                new_status = startup.status_id.status_name if startup.status_id else "неизвестно"
                status_changed = 'has_changes' in locals() and has_changes
                
                status_info = {
                    "previous_status": previous_status,
                    "new_status": new_status,
                    "status_changed": status_changed,
                    "reason": "Файлы загружены" if files_info["creatives_saved"] > 0 or files_info["proofs_saved"] > 0 or files_info["videos_saved"] > 0 or files_info["logo_saved"] else "Текстовые изменения" if 'has_changes' in locals() and has_changes else "Изменений нет"
                }
                
                # Собираем информацию об изменениях
                changes_info = {
                    "has_changes": 'has_changes' in locals() and has_changes,
                    "text_changed": 'has_changes' in locals() and has_changes and not (files_info["creatives_saved"] > 0 or files_info["proofs_saved"] > 0 or files_info["videos_saved"] > 0 or files_info["logo_saved"]),
                    "files_changed": files_info["creatives_saved"] > 0 or files_info["proofs_saved"] > 0 or files_info["videos_saved"] > 0 or files_info["logo_saved"],
                    "files_details": {
                        "creatives": [f.name for f in creatives] if creatives else [],
                        "proofs": [f.name for f in proofs] if proofs else [],
                        "videos": [f.name for f in videos] if videos else [],
                        "logo": logo.name if logo and hasattr(logo, 'name') else None
                    }
                }
                
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("profile"),
                    "files_info": files_info,
                    "status_info": status_info,
                    "changes_info": changes_info,
                })
            
            # Разные сообщения в зависимости от статуса
            if 'has_changes' in locals() and has_changes:
                messages.success(
                    request,
                    f'Стартап "{startup.title}" успешно отредактирован и отправлен на модерацию!',
                )
            else:
                messages.success(
                    request,
                    f'Стартап "{startup.title}" сохранен без изменений и остался в каталоге!',
                )
            return redirect("profile")
        else:
            logger.info("=== ФОРМА НЕ ВАЛИДНА ===")
            logger.info(f"Form errors: {form.errors}")
            logger.info(f"Form data: {request.POST}")
            for field, errors in form.errors.items():
                logger.error(f"Поле {field}: {errors}")
            logger.info("=== ФОРМА НЕ ВАЛИДНА ===")
            
            # Инициализируем has_changes для невалидной формы
            has_changes = False
            
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": False,
                    "errors": {f: [x["message"] for x in e.get_json_data()] for f, e in form.errors.items()},
                    "non_field_errors": [str(e) for e in form.non_field_errors()],
                }, status=400)
            messages.error(request, "Форма содержит ошибки.")
            return render(
                request,
                "accounts/edit_startup.html",
                {
                    "form": form,
                    "startup": startup,
                    "timeline_steps": timeline_steps,
                },
            )
    else:
        form = StartupEditForm(instance=startup)
    return render(
        request,
        "accounts/edit_startup.html",
        {
            "form": form,
            "startup": startup,
            "timeline_steps": timeline_steps,
        },
    )
def get_startup_updates(startup):
    """
    Получает обновления для стартапа (инвестиции и комментарии)
    """
    updates = []
    
    try:
        # Получаем последние инвестиции
        recent_investments = InvestmentTransactions.objects.filter(
            startup=startup
        ).select_related("investor").defer("franchise").order_by("-created_at")[:3]
        
        for investment in recent_investments:
            if investment.investor:
                updates.append({
                    "text": f"{investment.investor.get_full_name()} инвестировал {investment.amount:,.0f} ₽".replace(",", " "),
                    "timestamp": investment.created_at,
                    "type": "investment"
                })
        
        # Получаем последние комментарии
        recent_comments = Comments.objects.filter(
            startup_id=startup
        ).select_related("user_id").order_by("-created_at")[:3]
        
        for comment in recent_comments:
            if comment.user_id:
                updates.append({
                    "text": f"{comment.user_id.get_full_name()} оставил комментарий",
                    "timestamp": comment.created_at,
                    "type": "comment"
                })
        
        # Сортируем по времени (новые сверху)
        updates.sort(key=lambda x: x["timestamp"] if x["timestamp"] else datetime.min, reverse=True)
        
        # Берем только первые 3 обновления
        updates = updates[:3]
        
        # Возвращаем только тексты для простоты
        return [update["text"] for update in updates]
        
    except Exception as e:
        # В случае ошибки возвращаем пустой массив
        return []


@login_required
def main_page_moderator(request):
    """
    Отображает главную страницу для модератора.
    """
    if not is_moderator(request.user):
        return redirect("home")
    
    # Получаем реальные данные для карусели
    carousel_data = []
    
    try:
        # Сначала получаем стартапы с инвестициями
        startups_with_investments = (
            Startups.objects.filter(status="approved")
            .annotate(
                total_amount=Sum(
                    "investmenttransactions__amount",
                    filter=Q(
                        investmenttransactions__amount__gt=0,
                        investmenttransactions__transaction_status="completed",
                    ),
                ),
                total_investors=Count(
                    "investmenttransactions__investor",
                    filter=Q(
                        investmenttransactions__amount__gt=0,
                        investmenttransactions__transaction_status="completed",
                    ),
                    distinct=True,
                ),
            )
            .filter(total_amount__gt=0)
            .order_by("-startup_id")[:6]
        )
        
        added_startup_ids = set()
        
        for startup in startups_with_investments:
            updates = get_startup_updates(startup)
            chat_url = f"/cosmochat/?start_chat_with={startup.owner.user_id}" if startup.owner else "/cosmochat/"
            
            carousel_data.append({
                "startup_id": startup.startup_id,
                "name": startup.title,
                "logo_url": startup.get_logo_url() or "/static/accounts/images/main_page_moderator/planet_logo_carusel.webp",
                "total_investors": startup.total_investors or 0,
                "total_amount": float(startup.total_amount or 0),
                "updates": updates,
                "chat_url": chat_url,
                "startup_url": f"/startups/{startup.startup_id}/"
            })
            added_startup_ids.add(startup.startup_id)
        
        # Добавляем остальные одобренные стартапы, чтобы в сумме было до 6
        remaining_count = 6 - len(carousel_data)
        if remaining_count > 0:
            approved_startups = (
                Startups.objects.filter(status="approved")
                .exclude(startup_id__in=added_startup_ids)
                .order_by("-startup_id")[:remaining_count]
            )
            
            for startup in approved_startups:
                updates = get_startup_updates(startup)
                chat_url = f"/cosmochat/?start_chat_with={startup.owner.user_id}" if startup.owner else "/cosmochat/"
                
                carousel_data.append({
                    "startup_id": startup.startup_id,
                    "name": startup.title,
                    "logo_url": startup.get_logo_url() or "/static/accounts/images/main_page_moderator/planet_logo_carusel.webp",
                    "total_investors": 0,
                    "total_amount": 0,
                    "updates": updates,
                    "chat_url": chat_url,
                    "startup_url": f"/startups/{startup.startup_id}/"
                })
            
    except Exception as e:
        # В случае ошибки оставляем пустой массив
        carousel_data = []
    
    context = {
        "carousel_data": json.dumps(carousel_data, ensure_ascii=False)
    }
    
    return render(request, "accounts/moderator_main.html", context)
@login_required
def investor_main(request):
    """
    Отображает главную страницу инвестора с планетарной системой стартапов.
    """
    directions_data_json = FIXED_CATEGORIES.copy()
    selected_direction_name = request.GET.get("direction", "All")
    startups_query = Startups.objects.filter(status="approved").annotate(
        rating_avg=Coalesce(Avg("uservotes__rating"), 0.0, output_field=FloatField()),
        voters_count=Count("uservotes", distinct=True),
        total_investors=Count("investmenttransactions", distinct=True),
        current_funding=Coalesce(
            Sum("investmenttransactions__amount"), 0, output_field=DecimalField()
        ),
        comment_count=Count("comments", distinct=True),
    )
    if selected_direction_name != "All" and selected_direction_name != "Все":
        from django.db.models import Q
        direction_filter = Q()
        for category in FIXED_CATEGORIES:
            if category['original_name'] == selected_direction_name or category['direction_name'] == selected_direction_name:
                names = category.get('match_names') or [category['original_name']]
                direction_filter |= Q(direction__direction_name__in=names)
        if direction_filter:
            startups_query = startups_query.filter(direction_filter)
    startups_filtered = startups_query.annotate(
        progress=Case(
            When(funding_goal__gt=0, then=(F("amount_raised") * 100.0 / F("funding_goal"))),
            default=Value(0),
            output_field=FloatField(),
        )
    )[:6]
    planets_data_for_template = []
    fixed_orbit_sizes = [200, 300, 400, 500, 600, 700]
    orbit_times = [80, 95, 110, 125, 140, 160]
    planet_sizes = [60, 70, 56, 64, 50, 60]
    import random
    for idx, startup in enumerate(startups_filtered):
        if startup.planet_image:
            image_path = get_planet_image_url(startup.planet_image)
        else:
            image_path = get_fallback_planet_url(startup.startup_id)
        planets_data_for_template.append(
            {
                "id": startup.startup_id,
                "image": image_path,
                "orbit_size": fixed_orbit_sizes[idx % len(fixed_orbit_sizes)],
                "orbit_time": orbit_times[idx % len(orbit_times)],
                "planet_size": planet_sizes[idx % len(planet_sizes)],
            }
        )
    planets_data_json = []
    for startup in startups_filtered:
        investment_type = (
            "Инвестирование"
            if startup.only_invest
            else "Выкуп"
            if startup.only_buy
            else "Выкуп+инвестирование"
            if startup.both_mode
            else "Не указано"
        )
        if startup.planet_image:
            planet_image_url = get_planet_image_url(startup.planet_image)
        else:
            planet_image_url = get_fallback_planet_url(startup.startup_id)
        planets_data_json.append({
            "id": startup.startup_id,
            "name": startup.title,
            "image": planet_image_url,
            "rating": round(startup.rating_avg, 2),
            "progress": f"{startup.progress:.2f}%" if startup.progress is not None else "0%",
            "direction": startup.direction.direction_name if startup.direction else "Не указано",
            "investors": startup.total_investors,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не определена",
            "comment_count": startup.comment_count,
            "startup_id": startup.startup_id,
            "description": startup.short_description,
            "investment_type": investment_type,
            "logo": startup.get_logo_url(),
        })
    is_authenticated = request.user.is_authenticated
    is_startuper = is_authenticated and hasattr(request.user, 'role') and request.user.role and request.user.role.role_name == 'startuper'
    logo_data = {"image": static("accounts/images/planetary_system/gi.svg")}
    all_startups_query = Startups.objects.filter(status="approved").annotate(
        rating_avg=Coalesce(Avg("uservotes__rating"), 0.0, output_field=FloatField()),
        voters_count=Count("uservotes", distinct=True),
        total_investors=Count("investmenttransactions", distinct=True),
        current_funding=Coalesce(
            Sum("investmenttransactions__amount"), 0, output_field=DecimalField()
        ),
        comment_count=Count("comments", distinct=True),
        progress=Case(
            When(funding_goal__gt=0, then=(F("amount_raised") * 100.0 / F("funding_goal"))),
            default=Value(0),
            output_field=FloatField(),
        )
    )
    all_startups_data = []
    for startup in all_startups_query:
        investment_type = (
            "Инвестирование"
            if startup.only_invest
            else "Выкуп"
            if startup.only_buy
            else "Выкуп+инвестирование"
            if startup.both_mode
            else "Не указано"
        )
        if startup.planet_image:
            planet_image_url = get_planet_image_url(startup.planet_image)
        else:
            planet_image_url = get_fallback_planet_url(startup.startup_id)
        direction_name = startup.direction.direction_name if startup.direction else "Не указано"
        # Нормализуем здоровье в одну категорию 'Health'
        if direction_name in ['Healthcare', 'Medicine']:
            original_direction = 'Health'
        else:
            original_direction = None
            for category in FIXED_CATEGORIES:
                if category['direction_name'] == direction_name:
                    original_direction = category['original_name']
                    break
            if not original_direction:
                original_direction = direction_name
        all_startups_data.append({
            "id": startup.startup_id,
            "name": startup.title,
            "image": planet_image_url,
            "rating": round(startup.rating_avg, 2),
            "voters_count": startup.voters_count,
            "progress": round(startup.progress, 2) if startup.progress is not None else 0,
            "direction": original_direction,
            "investors": startup.total_investors,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не определена",
            "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указана",
            "comment_count": startup.comment_count,
            "startup_id": startup.startup_id,
            "description": startup.short_description,
            "investment_type": investment_type,
            "logo": startup.get_logo_url(),
        })
    try:
        latest_news = list(NewsArticles.objects.filter(
            status="published"
        ).select_related("author").order_by("-published_at")[:5])
    except Exception:
        latest_news = []

    context = {
        "planets_data": planets_data_for_template,
        "logo_data": logo_data,
        "directions": directions_data_json,
        "selected_galaxy": selected_direction_name,
        "planets_data_json": json.dumps(planets_data_json, cls=DjangoJSONEncoder),
        "directions_data_json": json.dumps(directions_data_json, cls=DjangoJSONEncoder),
        "all_startups_data_json": json.dumps(all_startups_data, cls=DjangoJSONEncoder),
        "is_startuper": is_startuper,
        "latest_news": latest_news,
    }
    return render(request, "accounts/investor_main.html", context)
@login_required
def startuper_main(request):
    """
    Отображает главную страницу стартаппера с планетарной системой стартапов.
    """
    directions_data_json = FIXED_CATEGORIES.copy()
    selected_direction_name = request.GET.get("direction", "All")
    startups_query = Startups.objects.filter(status="approved").annotate(
        rating_avg=Coalesce(Avg("uservotes__rating"), 0.0, output_field=FloatField()),
        voters_count=Count("uservotes", distinct=True),
        total_investors=Count("investmenttransactions", distinct=True),
        current_funding=Coalesce(
            Sum("investmenttransactions__amount"), 0, output_field=DecimalField()
        ),
        comment_count=Count("comments", distinct=True),
    )
    from accounts.models import Directions
    all_directions = Directions.objects.all()

    if selected_direction_name != "All" and selected_direction_name != "Все":
        from django.db.models import Q
        direction_filter = Q()
        for category in FIXED_CATEGORIES:
            if category['original_name'] == selected_direction_name or category['direction_name'] == selected_direction_name:
                names = category.get('match_names') or [category['original_name']]
                direction_filter |= Q(direction__direction_name__in=names)
        if direction_filter:
            startups_query = startups_query.filter(direction_filter)

    startups_filtered = startups_query.annotate(
        progress=Case(
            When(funding_goal__gt=0, then=(F("amount_raised") * 100.0 / F("funding_goal"))),
            default=Value(0),
            output_field=FloatField(),
        )
    )[:6]
    planets_data_for_template = []
    fixed_orbit_sizes = [200, 300, 400, 500, 600, 700]
    orbit_times = [80, 95, 110, 125, 140, 160]
    planet_sizes = [60, 70, 56, 64, 50, 60]
    import random
    for idx, startup in enumerate(startups_filtered):
        if startup.planet_image:
            image_path = get_planet_image_url(startup.planet_image)
        else:
            image_path = get_fallback_planet_url(startup.startup_id)
        planets_data_for_template.append(
            {
                "id": startup.startup_id,
                "image": image_path,
                "orbit_size": fixed_orbit_sizes[idx % len(fixed_orbit_sizes)],
                "orbit_time": orbit_times[idx % len(orbit_times)],
                "planet_size": planet_sizes[idx % len(planet_sizes)],
            }
        )
    planets_data_json = []
    for startup in startups_filtered:
        investment_type = (
            "Инвестирование"
            if startup.only_invest
            else "Выкуп"
            if startup.only_buy
            else "Выкуп+инвестирование"
            if startup.both_mode
            else "Не указано"
        )
        if startup.planet_image:
            planet_image_url = get_planet_image_url(startup.planet_image)
        else:
            planet_image_url = get_fallback_planet_url(startup.startup_id)
        planets_data_json.append({
            "id": startup.startup_id,
            "name": startup.title,
            "image": planet_image_url,
            "rating": round(startup.rating_avg, 2),
            "progress": f"{startup.progress:.2f}%" if startup.progress is not None else "0%",
            "direction": startup.direction.direction_name if startup.direction else "Не указано",
            "investors": startup.total_investors,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не определена",
            "comment_count": startup.comment_count,
            "startup_id": startup.startup_id,
            "description": startup.short_description,
            "investment_type": investment_type,
            "logo": startup.get_logo_url(),
        })
    is_authenticated = request.user.is_authenticated
    is_startuper = is_authenticated and hasattr(request.user, 'role') and request.user.role and request.user.role.role_name == 'startuper'
    logo_data = {"image": static("accounts/images/planetary_system/gi.svg")}
    all_startups_query = Startups.objects.filter(status="approved").annotate(
        rating_avg=Coalesce(Avg("uservotes__rating"), 0.0, output_field=FloatField()),
        voters_count=Count("uservotes", distinct=True),
        total_investors=Count("investmenttransactions", distinct=True),
        current_funding=Coalesce(
            Sum("investmenttransactions__amount"), 0, output_field=DecimalField()
        ),
        comment_count=Count("comments", distinct=True),
        progress=Case(
            When(funding_goal__gt=0, then=(F("amount_raised") * 100.0 / F("funding_goal"))),
            default=Value(0),
            output_field=FloatField(),
        )
    )
    all_startups_data = []
    for startup in all_startups_query:
        investment_type = (
            "Инвестирование"
            if startup.only_invest
            else "Выкуп"
            if startup.only_buy
            else "Выкуп+инвестирование"
            if startup.both_mode
            else "Не указано"
        )
        if startup.planet_image:
            planet_image_url = get_planet_image_url(startup.planet_image)
        else:
            planet_image_url = get_fallback_planet_url(startup.startup_id)
        direction_name = startup.direction.direction_name if startup.direction else "Не указано"
        russian_direction = DIRECTION_TRANSLATIONS.get(direction_name, direction_name)
        original_direction = None
        for category in FIXED_CATEGORIES:
            if category['direction_name'] == direction_name:
                original_direction = category['original_name']
                break
        if not original_direction:
            original_direction = direction_name
        all_startups_data.append({
            "id": startup.startup_id,
            "name": startup.title,
            "image": planet_image_url,
            "rating": round(startup.rating_avg, 2),
            "voters_count": startup.voters_count,
            "progress": round(startup.progress, 2) if startup.progress is not None else 0,
            "direction": original_direction,
            "investors": startup.total_investors,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не определена",
            "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указана",
            "comment_count": startup.comment_count,
            "startup_id": startup.startup_id,
            "description": startup.short_description,
            "investment_type": investment_type,
            "logo": startup.get_logo_url(),
        })

    try:
        random_startups = Startups.objects.filter(status="approved").order_by('?')[:3]
        random_startups_data = []

        for startup in random_startups:

            # Получаем логотип стартапа
            startup_logo = None
            if hasattr(startup, 'get_logo_url') and startup.get_logo_url():
                startup_logo = startup.get_logo_url()

            # Получаем планету
            if startup.planet_image:
                planet_image = get_planet_image_url(startup.planet_image)
            else:
                planet_image = get_fallback_planet_url(startup.startup_id)
            
            # Основное изображение - логотип если есть, иначе планета
            startup_image = startup_logo if startup_logo else planet_image


            if hasattr(startup, 'owner') and startup.owner and hasattr(startup.owner, 'get_profile_picture_url'):
                owner_avatar = startup.owner.get_profile_picture_url() or static('accounts/images/default_icon.svg')
            else:
                owner_avatar = static('accounts/images/default_icon.svg')


            try:
                rating = startup.get_average_rating() or 0
            except Exception:
                rating = 0
            rating_formatted = str(round(rating))


            description = getattr(startup, 'short_description', '') or getattr(startup, 'description', '')
            if description:
                description = strip_tags(description)
                if len(description) > 100:
                    description = description[:100] + "..."
            else:
                description = "Описание не указано"


            startup_id = getattr(startup, 'startup_id', None)
            if startup_id and str(startup_id).isdigit():
                startup_url = f"/startups/{startup_id}/"
            else:
                startup_url = "/startups_list/"

            startup_data = {
                'id': startup_id or 'Unknown',
                'name': getattr(startup, 'title', 'Unknown'),
                'rating': rating_formatted,
                'description': description,
                'image': startup_image,
                'planet_image': planet_image,
                'has_logo': bool(startup_logo),
                'owner_avatar': owner_avatar,
                'url': startup_url
            }
            random_startups_data.append(startup_data)

    except Exception as e:
        logger.error(f"Error getting random startups for startuper_main: {e}")

        random_startups_data = [
            {
                'id': 'fallback1',
                'name': 'VoltForge Dynamics',
                'rating': '4',
                'description': 'VoltForge разрабатывает твердотельные батареи с графеновыми наноструктурами, которые заряжаются...',
                'image': static('accounts/images/main_page/volt_forge.webp'),
                'planet_image': static('accounts/images/planetary_system/textures/planet_1.webp'),
                'has_logo': False,
                'owner_avatar': static('accounts/images/default_icon.svg'),
                'url': '/startups_list/'
            },
            {
                'id': 'fallback2',
                'name': 'NeuroBloom',
                'rating': '5',
                'description': 'NeuroBloom предлагает носимый гаджет с ИИ, который анализирует нейронные паттерны для раннего выявления тревоги, депрессии и выгорания.',
                'image': static('accounts/images/main_page/neuro_bloom.webp'),
                'planet_image': static('accounts/images/planetary_system/textures/planet_2.webp'),
                'has_logo': False,
                'owner_avatar': static('accounts/images/default_icon.svg'),
                'url': '/startups_list/'
            },
            {
                'id': 'fallback3',
                'name': 'BioCrop Nexus',
                'rating': '4',
                'description': 'BioCrop Nexus создает генетически оптимизированные семена, устойчивые к экстремальным климатическим условиям и вредителям.',
                'image': static('accounts/images/main_page/biocrop_nexus.webp'),
                'planet_image': static('accounts/images/planetary_system/textures/planet_3.webp'),
                'has_logo': False,
                'owner_avatar': static('accounts/images/default_icon.svg'),
                'url': '/startups_list/'
            }
        ]


    try:
        random_startupers = Users.objects.filter(role__role_name='startuper').order_by('?')[:3]
        random_startupers_data = []

        for startuper in random_startupers:

            if hasattr(startuper, 'get_profile_picture_url'):
                avatar_url = startuper.get_profile_picture_url() or static('accounts/images/avatars/default_avatar_ufo.png')
            else:
                avatar_url = static('accounts/images/avatars/default_avatar_ufo.png')


            rating = getattr(startuper, 'rating_avg', 0.0)
            rating_formatted = str(round(rating)) if rating else "0"


            first_name = getattr(startuper, 'first_name', '') or ''
            last_name = getattr(startuper, 'last_name', '') or ''
            if first_name and last_name:
                full_name = f"{first_name} {last_name}"
            elif first_name:
                full_name = first_name
            elif last_name:
                full_name = last_name
            else:
                full_name = "Стартапер"

            startuper_data = {
                'id': getattr(startuper, 'user_id', 'Unknown'),
                'name': full_name,
                'rating': rating_formatted,
                'avatar': avatar_url
            }
            random_startupers_data.append(startuper_data)

    except Exception as e:
        logger.error(f"Error getting random startupers for startuper_main: {e}")

        random_startupers_data = [
            {
                'id': 'fallback1',
                'name': 'Виктор Смирнов',
                'rating': '5',
                'avatar': static('accounts/images/avatars/default_avatar_ufo.png')
            },
            {
                'id': 'fallback2',
                'name': 'Анна Кузнецова',
                'rating': '5',
                'avatar': static('accounts/images/avatars/default_avatar_ufo.png')
            },
            {
                'id': 'fallback3',
                'name': 'Дмитрий Иванов',
                'rating': '4',
                'avatar': static('accounts/images/avatars/default_avatar_ufo.png')
            }
        ]

    try:
        latest_news = list(NewsArticles.objects.filter(
            status="published"
        ).select_related("author").order_by("-published_at")[:5])
    except Exception:
        latest_news = []

    context = {
        "planets_data": planets_data_for_template,
        "logo_data": logo_data,
        "directions": directions_data_json,
        "selected_galaxy": selected_direction_name,
        "planets_data_json": json.dumps(planets_data_json, cls=DjangoJSONEncoder),
        "directions_data_json": json.dumps(directions_data_json, cls=DjangoJSONEncoder),
        "all_startups_data_json": json.dumps(all_startups_data, cls=DjangoJSONEncoder),
        "is_startuper": is_startuper,
        "random_startups": random_startups_data,
        "random_startupers": random_startupers_data,
        "latest_news": latest_news,
    }

    return render(request, "accounts/startuper_main.html", context)


@login_required
def moderator_dashboard(request):
    if not is_moderator(request.user):
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    pending_startups_list = Startups.objects.filter(status="pending").select_related("owner", "direction", "stage")
    pending_franchises_list = Franchises.objects.filter(status="pending").select_related("owner", "direction", "stage")
    pending_agencies_list = Agencies.objects.filter(status="pending").select_related("owner", "direction").distinct()
    pending_specialists_list = Specialists.objects.filter(status="pending").select_related("owner", "direction")
    
    # Логирование для отладки дублирования
    logger.info(f"=== MODERATOR_DASHBOARD === pending_agencies count: {pending_agencies_list.count()}")
    for agency in pending_agencies_list:
        logger.info(f"  Agency ID: {agency.agency_id}, title: {agency.title}, status: {agency.status}")
    
    logger.info(f"=== MODERATOR_DASHBOARD === pending_specialists count: {pending_specialists_list.count()}")
    for specialist in pending_specialists_list:
        logger.info(f"  Specialist ID: {specialist.specialist_id}, title: {specialist.title}, status: {specialist.status}")
    
    # Получаем категории как в каталогах
    # Стартапы - только определенные категории
    startup_categories = Directions.objects.filter(
        direction_name__in=[
            'Technology', 'Healthcare', 'Finance', 'Education', 'Entertainment',
            'Fashion', 'Food', 'Gaming', 'Real Estate', 'Travel', 'Agriculture',
            'Energy', 'Environment', 'Social', 'Medicine', 'Auto', 'Delivery',
            'Cafe', 'Fastfood', 'Health', 'Beauty', 'Transport', 'Sport',
            'Psychology', 'AI', 'IT', 'Retail'
        ]
    ).order_by('direction_name')
    
    # Франшизы - только существующие направления
    existing_dir_ids = (
        Franchises.objects.filter(status="approved", direction__isnull=False)
        .values_list("direction_id", flat=True)
        .distinct()
    )
    franchise_categories = Directions.objects.filter(direction_id__in=existing_dir_ids).order_by("direction_name")
    
    # Категории для агентств и специалистов (как в каталогах)
    agency_categories = [
        "Веб-разработка", "Мобильная разработка", "Дизайн", "Маркетинг", 
        "ИИ", "Брендинг", "Видео и мультимедиа", "Перевод"
    ]
    specialist_categories = [
        "Веб-разработка", "Мобильная разработка", "Дизайн", "Маркетинг", 
        "ИИ", "Брендинг", "Видео и мультимедиа", "Перевод"
    ]
    
    # Объединяем все категории для фильтра
    all_categories = []
    for cat in startup_categories:
        all_categories.append({"name": cat.direction_name, "type": "startup"})
    for cat in franchise_categories:
        all_categories.append({"name": cat.direction_name, "type": "franchise"})
    for cat in agency_categories:
        all_categories.append({"name": cat, "type": "agency"})
    for cat in specialist_categories:
        all_categories.append({"name": cat, "type": "specialist"})
    
    selected_category_name = request.GET.get("category")
    selected_category_type = request.GET.get("category_type", "startup")
    sort_order = request.GET.get("sort")
    filter_type = request.GET.get("filter")
    search_query = request.GET.get("search", "").strip()
    
    if filter_type == "all":
        selected_category_name = None
        selected_category_type = None
        sort_order = None
        search_query = ""
    
    # Применяем поиск
    if search_query:
        pending_startups_list = pending_startups_list.filter(
            Q(title__icontains=search_query) | Q(description__icontains=search_query)
        )
        pending_franchises_list = pending_franchises_list.filter(
            Q(title__icontains=search_query) | Q(description__icontains=search_query)
        )
        pending_agencies_list = pending_agencies_list.filter(
            Q(title__icontains=search_query) | Q(description__icontains=search_query)
        )
        pending_specialists_list = pending_specialists_list.filter(
            Q(title__icontains=search_query) | Q(description__icontains=search_query)
        )
    
    # Применяем фильтры по категориям
    if selected_category_name:
        if selected_category_type == "startup":
            pending_startups_list = pending_startups_list.filter(
                direction__direction_name__iexact=selected_category_name
            )
            # Скрываем все остальные типы заявок
            pending_franchises_list = pending_franchises_list.none()
            pending_agencies_list = pending_agencies_list.none()
            pending_specialists_list = pending_specialists_list.none()
        elif selected_category_type == "franchise":
            pending_franchises_list = pending_franchises_list.filter(
                direction__direction_name__iexact=selected_category_name
            )
            # Скрываем все остальные типы заявок
            pending_startups_list = pending_startups_list.none()
            pending_agencies_list = pending_agencies_list.none()
            pending_specialists_list = pending_specialists_list.none()
        elif selected_category_type == "agency":
            pending_agencies_list = pending_agencies_list.filter(
                customization_data__agency_category__iexact=selected_category_name
            )
            # Скрываем все остальные типы заявок
            pending_startups_list = pending_startups_list.none()
            pending_franchises_list = pending_franchises_list.none()
            pending_specialists_list = pending_specialists_list.none()
        elif selected_category_type == "specialist":
            pending_specialists_list = pending_specialists_list.filter(
                customization_data__specialist_category__iexact=selected_category_name
            )
            # Скрываем все остальные типы заявок
            pending_startups_list = pending_startups_list.none()
            pending_franchises_list = pending_franchises_list.none()
            pending_agencies_list = pending_agencies_list.none()
    
    # Применяем сортировку
    if sort_order == "newest":
        if hasattr(Startups, "created_at"):
            pending_startups_list = pending_startups_list.order_by("-created_at")
            pending_franchises_list = pending_franchises_list.order_by("-created_at")
            pending_agencies_list = pending_agencies_list.order_by("-created_at")
            pending_specialists_list = pending_specialists_list.order_by("-created_at")
        else:
            pending_startups_list = pending_startups_list.order_by("-startup_id")
            pending_franchises_list = pending_franchises_list.order_by("-franchise_id")
            pending_agencies_list = pending_agencies_list.order_by("-agency_id")
            pending_specialists_list = pending_specialists_list.order_by("-specialist_id")
    else:
        # Дефолтная сортировка - по ID (старые сверху)
        pending_startups_list = pending_startups_list.order_by("startup_id")
        pending_franchises_list = pending_franchises_list.order_by("franchise_id")
        pending_agencies_list = pending_agencies_list.order_by("agency_id")
        pending_specialists_list = pending_specialists_list.order_by("specialist_id")
    
    context = {
        "pending_startups": pending_startups_list,
        "pending_franchises": pending_franchises_list,
        "pending_agencies": pending_agencies_list,
        "pending_specialists": pending_specialists_list,
        "all_categories": all_categories,
        "selected_category_name": selected_category_name,
        "selected_category_type": selected_category_type,
        "current_sort_order": sort_order,
        "filter_type": filter_type,
        "search_query": search_query,
    }
    return render(request, "accounts/moderator_dashboard.html", context)
@login_required
def approve_startup(request, startup_id):
    if not is_moderator(request.user):
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    startup = get_object_or_404(Startups, startup_id=startup_id)
    if request.method == "POST":
        from accounts.moderation import approve_entity
        moderator_comment = request.POST.get("moderator_comment", "")
        approve_entity(startup, request.user, "startup", moderator_comment)
        messages.success(request, "Стартап одобрен.")
    return redirect("moderator_dashboard")
@login_required
def reject_startup(request, startup_id):
    if not is_moderator(request.user):
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    startup = get_object_or_404(Startups, startup_id=startup_id)
    if request.method == "POST":
        from accounts.moderation import reject_entity
        moderator_comment = request.POST.get("moderator_comment", "")
        reject_entity(startup, request.user, "startup", moderator_comment)
        messages.success(request, "Стартап отклонен.")
    return redirect("moderator_dashboard")
@login_required
def vote_startup(request, startup_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    startup = get_object_or_404(Startups, startup_id=startup_id)
    try:
        rating = int(request.POST.get("rating", 0))
    except (TypeError, ValueError):
        return JsonResponse({"success": False, "error": "Некорректное значение рейтинга"})
    if not 1 <= rating <= 5:
        return JsonResponse(
            {"success": False, "error": "Недопустимое значение рейтинга"}
        )
    from django.db import transaction
    try:
        with transaction.atomic():
            # Атомарная проверка и создание голоса (предотвращает race condition)
            vote, created = UserVotes.objects.get_or_create(
                user=request.user,
                startup=startup,
                defaults={"rating": rating, "created_at": timezone.now()},
            )
            if not created:
                return JsonResponse(
                    {"success": False, "error": "Вы уже голосовали за этот стартап"}
                )
            # Атомарное обновление счётчиков через F() (предотвращает потерю данных)
            Startups.objects.filter(startup_id=startup_id).update(
                total_voters=models.F("total_voters") + 1,
                sum_votes=models.F("sum_votes") + rating,
            )
    except Exception:
        return JsonResponse({"success": False, "error": "Ошибка при сохранении голоса"})
    # Перечитываем актуальные значения
    startup.refresh_from_db(fields=["total_voters", "sum_votes"])
    average_rating = (
        startup.sum_votes / startup.total_voters if startup.total_voters > 0 else 0
    )
    return JsonResponse({"success": True, "average_rating": average_rating})
@login_required
def invest(request, startup_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    startup = get_object_or_404(Startups, startup_id=startup_id)
    if not request.user.is_authenticated:
        return JsonResponse(
            {"success": False, "error": "Требуется авторизация"}
        )
    if startup.status in ["blocked", "closed"]:
        return JsonResponse(
            {
                "success": False,
                "error": f"Инвестирование запрещено: стартап {startup.status}",
            }
        )
    
    user_role = request.user.role.role_name if request.user.role else None
    
    if user_role == "moderator":
        try:
            amount = Decimal(request.POST.get("amount", "0"))
            if amount <= 0:
                return JsonResponse(
                    {"success": False, "error": "Сумма должна быть больше 0"}
                )
            
            transaction_type, _ = TransactionTypes.objects.get_or_create(
                type_name="investment",
                defaults={"type_name": "investment"}
            )
            payment_method, _ = PaymentMethods.objects.get_or_create(
                method_name="default",
                defaults={"method_name": "default"}
            )
            
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO investment_transactions 
                    (startup_id, investor_id, amount, is_micro, transaction_type_id, transaction_status, payment_method_id, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    startup.startup_id,
                    request.user.user_id,
                    amount,
                    startup.micro_investment_available if hasattr(startup, 'micro_investment_available') else False,
                    transaction_type.type_id,
                    "completed",
                    payment_method.method_id,
                    timezone.now(),
                    timezone.now(),
                ])
            startup.amount_raised = (startup.amount_raised or Decimal("0")) + amount
            startup.total_invested = (startup.total_invested or Decimal("0")) + amount
            startup.save()
            investors_count = startup.get_investors_count()
            progress_percentage = startup.get_progress_percentage()
            return JsonResponse(
                {
                    "success": True,
                    "amount_raised": float(startup.amount_raised),
                    "investors_count": investors_count,
                    "progress_percentage": float(progress_percentage),
                }
            )
        except Exception as e:
            logger.error(f"Ошибка при инвестировании: {str(e)}", exc_info=True)
            return JsonResponse(
                {"success": False, "error": f"Произошла ошибка при инвестировании: {str(e)}"}
            )
    elif user_role in ["investor", "startuper"]:
        if not startup.owner:
            return JsonResponse(
                {"success": False, "error": "У стартапа нет владельца"}
            )
        if startup.owner.user_id == request.user.user_id:
            return JsonResponse(
                {"success": False, "error": "Нельзя инвестировать в свой стартап"}
            )
        
        try:
            from django.db import transaction as db_transaction
            
            existing_chats = ChatConversations.objects.filter(
                is_group_chat=False,
                chatparticipants__user=request.user
            ).filter(
                chatparticipants__user=startup.owner
            ).annotate(
                num_participants=Count("chatparticipants")
            ).filter(num_participants=2).distinct()
            
            existing_chat = existing_chats.first()
            
            if existing_chat:
                chat = existing_chat
                chat_existed = True
            else:
                chat = ChatConversations.objects.create(
                    name=f"Чат {request.user.first_name} и {startup.owner.first_name}",
                    is_group_chat=False,
                    created_at=timezone.now(),
                    updated_at=timezone.now(),
                )
                ChatParticipants.objects.create(conversation=chat, user=request.user)
                ChatParticipants.objects.create(conversation=chat, user=startup.owner)
                chat_existed = False
            
            if not chat.is_deal:
                participants = chat.chatparticipants_set.all()
                participant_roles = [p.user.role.role_name.lower() if p.user and p.user.role else None for p in participants]
                participant_roles_set = {r for r in participant_roles if r}
                
                if len(participants) >= 2 and {"startuper", "investor"}.issubset(participant_roles_set):
                    with db_transaction.atomic():
                        chat.is_deal = True
                        chat.deal_status = "pending"
                        chat.updated_at = timezone.now()
                        chat.save()
                        moderators = Users.objects.filter(role__role_name="moderator")
                        if moderators.exists():
                            moderator = choice(list(moderators))
                            moderator_participant, created = ChatParticipants.objects.get_or_create(
                                conversation=chat, user=moderator
                            )
                            if created:
                                sent_status, _ = MessageStatuses.objects.get_or_create(
                                    status_name="sent",
                                    defaults={"status_name": "sent"}
                                )
                                message = Messages(
                                    conversation=chat,
                                    sender=None,
                                    message_text=f"Сделку начал {request.user.get_full_name()}. Назначен модератор: {moderator.get_full_name()}",
                                    status=sent_status,
                                    created_at=timezone.now(),
                                    updated_at=timezone.now(),
                                )
                                message.save()
            
            return JsonResponse(
                {
                    "success": True,
                    "redirect": True,
                    "chat_id": chat.conversation_id,
                    "chat_existed": chat_existed,
                }
            )
        except Exception as e:
            logger.error(f"Ошибка при создании чата: {str(e)}", exc_info=True)
            return JsonResponse(
                {"success": False, "error": f"Произошла ошибка при создании чата: {str(e)}"}
            )
    else:
        return JsonResponse(
            {"success": False, "error": "Недостаточно прав для инвестирования"}
        )

@login_required
def edit_ai_rating(request, entity_type, entity_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    
    if not request.user.is_authenticated or not request.user.role or request.user.role.role_name != "moderator":
        return JsonResponse({"success": False, "error": "Недостаточно прав"})
    
    try:
        rating = Decimal(request.POST.get("ai_rating", "0"))
        if rating < 1 or rating > 10:
            return JsonResponse({"success": False, "error": "Оценка должна быть от 1 до 10"})
        
        if entity_type == "startup":
            entity = get_object_or_404(Startups, startup_id=entity_id)
        elif entity_type == "franchise":
            entity = get_object_or_404(Franchises, franchise_id=entity_id)
        elif entity_type == "agency":
            entity = get_object_or_404(Agencies, agency_id=entity_id)
        elif entity_type == "specialist":
            entity = get_object_or_404(Specialists, specialist_id=entity_id)
        else:
            return JsonResponse({"success": False, "error": "Неверный тип сущности"})
        
        if not hasattr(entity, 'customization_data') or not entity.customization_data:
            entity.customization_data = {}
        
        if not isinstance(entity.customization_data, dict):
            entity.customization_data = {}
        
        entity.customization_data['ai_rating'] = float(rating)
        entity.save()
        
        return JsonResponse({"success": True, "ai_rating": float(rating)})
    except Exception as e:
        logger.error(f"Ошибка при сохранении AI рейтинга: {str(e)}", exc_info=True)
        return JsonResponse({"success": False, "error": f"Произошла ошибка: {str(e)}"})

@login_required
def invest_franchise(request, franchise_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
    if not request.user.is_authenticated or request.user.role.role_name != "investor":
        return JsonResponse(
            {"success": False, "error": "Только инвесторы могут инвестировать"}
        )
    if franchise.status in ["blocked", "closed"]:
        return JsonResponse(
            {
                "success": False,
                "error": f"Инвестирование запрещено: франшиза {franchise.status}",
            }
        )
    try:
        amount = Decimal(request.POST.get("amount", "0"))
        if amount <= 0:
            return JsonResponse(
                {"success": False, "error": "Сумма должна быть больше 0"}
            )
        transaction_type, _ = TransactionTypes.objects.get_or_create(
            type_name="investment",
            defaults={"type_name": "investment"}
        )
        payment_method, _ = PaymentMethods.objects.get_or_create(
            method_name="default",
            defaults={"method_name": "default"}
        )
        
        transaction = InvestmentTransactions(
            franchise=franchise,
            investor=request.user,
            amount=amount,
            is_micro=False,
            transaction_type=transaction_type,
            transaction_status="completed",
            payment_method=payment_method,
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        transaction.save()
        franchise.total_invested = (franchise.total_invested or Decimal("0")) + amount
        franchise.save()
        investors_count = franchise.get_investors_count()
        return JsonResponse(
            {
                "success": True,
                "amount_raised": float(franchise.total_invested),
                "investors_count": investors_count,
            }
        )
    except Exception as e:
        logger.error(f"Ошибка при инвестировании франшизы: {str(e)}")
        return JsonResponse(
            {"success": False, "error": "Произошла ошибка при инвестировании"}
        )

@login_required
def suggest_news(request):
    """Page for authenticated users to suggest a news article for moderation."""
    user = request.user
    user_startups = Startups.objects.filter(owner=user, status_id__status_name="Одобрено")
    user_franchises = Franchises.objects.filter(owner=user, status_id__status_name="Одобрено")
    all_categories = NewsCategories.objects.all().order_by("sort_order")

    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        content = request.POST.get("content", "").strip()
        category_id = request.POST.get("category") or None
        tags = request.POST.get("tags", "").strip()
        entity_val = request.POST.get("entity_type", "personal")
        image = request.FILES.get("image")

        if not title or not content:
            messages.error(request, "Заголовок и текст обязательны.")
            return redirect("suggest_news")

        article = NewsArticles(
            title=title,
            content=content,
            author=user,
            status="pending",
            tags=tags or None,
            updated_at=timezone.now(),
        )
        if category_id:
            try:
                article.category = NewsCategories.objects.get(pk=category_id)
            except NewsCategories.DoesNotExist:
                pass

        # Parse entity type
        if entity_val.startswith("startup_"):
            sid = entity_val.replace("startup_", "")
            if user_startups.filter(startup_id=sid).exists():
                article.entity_type = "startup"
                article.linked_startup_id = int(sid)
        elif entity_val.startswith("franchise_"):
            fid = entity_val.replace("franchise_", "")
            if user_franchises.filter(franchise_id=fid).exists():
                article.entity_type = "franchise"
                article.linked_franchise_id = int(fid)
        else:
            article.entity_type = "personal"

        article.save()  # triggers slug generation

        if image:
            image_id = str(uuid.uuid4())
            file_path = f"news/{article.article_id}/{image_id}_{image.name}"
            default_storage.save(file_path, image)
            article.image_url = file_path
            article.save()

        messages.success(request, "Новость отправлена на модерацию! Мы рассмотрим её в ближайшее время.")
        return redirect("news")

    context = {
        "user_startups": user_startups,
        "user_franchises": user_franchises,
        "all_categories": all_categories,
    }
    return render(request, "accounts/suggest_news.html", context)


def news(request):
    from .forms import NewsForm

    # Backfill/fix slugs: regenerate non-ASCII or empty slugs
    import re as _re
    for art in NewsArticles.objects.all():
        if not art.slug or _re.search(r'[^\x00-\x7F]', art.slug):
            art.slug = None  # reset to trigger save() transliteration
            art.save()

    articles = NewsArticles.objects.filter(
        status="published"
    ).select_related("author", "category")

    articles = articles.annotate(
        likes_count_agg=Count("newslikes", distinct=True),
        comments_count_agg=Count("comments", distinct=True),
    )

    sort_order = request.GET.get("sort", "new")
    if sort_order == "old":
        articles = articles.order_by("published_at")
    elif sort_order == "rating":
        articles = articles.order_by("-likes_count_agg", "-published_at")
    else:
        articles = articles.order_by("-published_at")

    selected_categories = request.GET.getlist("category")
    if selected_categories:
        articles = articles.filter(category__slug__in=selected_categories)

    # Content type filter (news / article)
    content_type_filter = request.GET.get("content_type", "")
    if content_type_filter in ("news", "article"):
        articles = articles.filter(content_type=content_type_filter)

    # Entity focus filter (franchise / startup / etc)
    entity_focus_filter = request.GET.get("entity_focus", "")
    if entity_focus_filter:
        articles = articles.filter(entity_focus=entity_focus_filter)

    search_query = request.GET.get("search")
    if search_query:
        articles = articles.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(tags__icontains=search_query)
        )

    try:
        page_number = int(request.GET.get("page", 1))
        if page_number < 1:
            page_number = 1
    except (ValueError, TypeError):
        page_number = 1

    paginator = Paginator(articles, 12)

    if page_number > paginator.num_pages and paginator.num_pages > 0:
        page_number = paginator.num_pages

    page_obj = paginator.get_page(page_number)

    # Assign card_format for dynamic grid layout (3-column grid, no gaps)
    # Rules: featured/text-full = 3 cols, medium = 2 cols, standard = 1 col
    # Track columns per row so every row sums to exactly 3
    articles_list = list(page_obj)
    col = 0  # columns used in current row (0, 1, or 2)
    row_count = 0  # count image-rows for pattern variety

    for idx, article in enumerate(articles_list):
        if not article.image_url:
            if col == 0:
                # Row start: text-full takes full row (3 cols)
                article.card_format = 'text-full'
            else:
                # Mid-row: use standard (1 col) to avoid gap, even without image
                article.card_format = 'standard'
                col = (col + 1) % 3
                continue
        elif idx == 0 and col == 0:
            article.card_format = 'featured'
            col = 0
        elif col == 0:
            # Start of row: alternate between medium+standard and 3x standard
            if row_count % 3 == 2:
                # Every 3rd image-row: use 3 standards for variety
                article.card_format = 'standard'
                col = 1
            else:
                article.card_format = 'medium'
                col = 2
            row_count += 1
        elif col == 2:
            # 1 col left: standard to complete the row
            article.card_format = 'standard'
            col = 0
        elif col == 1:
            # 2 cols left: medium to complete, or standard+standard
            if row_count % 3 == 0:
                article.card_format = 'medium'
                col = 0
            else:
                article.card_format = 'standard'
                col = 2
        else:
            article.card_format = 'standard'
            col = (col + 1) % 3

    all_categories = NewsCategories.objects.all().order_by("sort_order")

    from accounts.promotions import get_active_ads
    context = {
        "articles": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "sort_order": sort_order,
        "selected_categories": selected_categories,
        "all_categories": all_categories,
        "search_query": search_query,
        "news_ads": get_active_ads("news_sidebar"),
        "content_type_filter": content_type_filter,
        "entity_focus_filter": entity_focus_filter,
    }

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string("accounts/partials/_news_cards.html", context, request=request)
        return JsonResponse({
            "html": html,
            "has_next": page_obj.has_next(),
            "page_number": page_obj.number,
            "num_pages": paginator.num_pages,
            "count": paginator.count,
        })

    return render(request, "accounts/news.html", context)
def news_detail(request, slug):
    from .forms import NewsCommentForm

    article = get_object_or_404(NewsArticles, slug=slug)

    # Не показывать черновики/архив обычным пользователям
    if article.status != "published":
        is_mod = request.user.is_authenticated and (request.user.role.role_name or "").lower() == "moderator"
        is_author = request.user.is_authenticated and article.author_id == request.user.user_id
        if not is_mod and not is_author:
            return redirect("news")

    user = request.user if request.user.is_authenticated else None

    # Трекинг просмотров (только авторизованные)
    if user:
        if not NewsViews.objects.filter(article=article, user=user).exists():
            NewsViews.objects.create(article=article, user=user)

    # Счётчики
    views_count = NewsViews.objects.filter(article=article).count()
    likes_count = NewsLikes.objects.filter(article=article).count()
    dislikes_count = NewsDislikes.objects.filter(article=article).count()
    user_liked = NewsLikes.objects.filter(article=article, user=user).exists() if user else False
    user_disliked = NewsDislikes.objects.filter(article=article, user=user).exists() if user else False

    # POST: лайк / дизлайк / комментарий
    if request.method == "POST" and user:
        if "like" in request.POST:
            if user_liked:
                NewsLikes.objects.filter(article=article, user=user).delete()
            else:
                NewsLikes.objects.get_or_create(article=article, user=user)
                NewsDislikes.objects.filter(article=article, user=user).delete()
            return redirect("news_detail", slug=article.slug)

        elif "dislike" in request.POST:
            if user_disliked:
                NewsDislikes.objects.filter(article=article, user=user).delete()
            else:
                NewsDislikes.objects.get_or_create(article=article, user=user)
                NewsLikes.objects.filter(article=article, user=user).delete()
            return redirect("news_detail", slug=article.slug)

        elif "comment" in request.POST:
            comment_form = NewsCommentForm(request.POST)
            if comment_form.is_valid():
                comment = comment_form.save(commit=False)
                comment.article = article
                comment.user = user
                parent_id = request.POST.get("parent_comment_id")
                if parent_id:
                    try:
                        comment.parent_comment_id = int(parent_id)
                    except (ValueError, TypeError):
                        pass
                comment.save()
                return redirect("news_detail", slug=article.slug)

    # Комментарии (top-level)
    comments = NewsComments.objects.filter(
        article=article, parent_comment__isnull=True
    ).select_related("user").order_by("-created_at")

    # Ответы на комментарии
    all_replies = NewsComments.objects.filter(
        article=article, parent_comment__isnull=False
    ).select_related("user").order_by("created_at")
    replies_by_parent = {}
    for reply in all_replies:
        replies_by_parent.setdefault(reply.parent_comment_id, []).append(reply)

    comments_count = NewsComments.objects.filter(article=article).count()

    # Похожие статьи (по категории, потом рандом)
    similar = NewsArticles.objects.filter(
        status="published"
    ).exclude(article_id=article.article_id).select_related("author")

    if article.category_id:
        similar_by_cat = list(similar.filter(category_id=article.category_id).order_by("-published_at")[:6])
    else:
        similar_by_cat = []

    if len(similar_by_cat) < 6:
        exclude_ids = [article.article_id] + [a.article_id for a in similar_by_cat]
        extra = list(similar.exclude(article_id__in=exclude_ids).order_by("-published_at")[:6 - len(similar_by_cat)])
        similar_by_cat.extend(extra)

    comment_form = NewsCommentForm()

    return render(
        request,
        "accounts/news_detail.html",
        {
            "article": article,
            "views_count": views_count,
            "likes_count": likes_count,
            "dislikes_count": dislikes_count,
            "user_liked": user_liked,
            "user_disliked": user_disliked,
            "comments": comments,
            "replies_by_parent": replies_by_parent,
            "comments_count": comments_count,
            "similar_articles": similar_by_cat,
            "comment_form": comment_form,
            "canonical_url": request.build_absolute_uri(),
            "reading_time": max(1, len((article.content or "").split()) // 200),
        },
    )


def api_similar_news(request):
    """API for lazy-loading similar articles."""
    from django.http import JsonResponse
    from django.template.defaultfilters import truncatewords, striptags

    article_id = request.GET.get("article_id")
    offset = int(request.GET.get("offset", 0))
    limit = int(request.GET.get("limit", 6))

    if not article_id:
        return JsonResponse({"articles": [], "has_more": False})

    try:
        article = NewsArticles.objects.get(article_id=article_id)
    except NewsArticles.DoesNotExist:
        return JsonResponse({"articles": [], "has_more": False})

    qs = NewsArticles.objects.filter(status="published").exclude(
        article_id=article.article_id
    ).order_by("-published_at")

    if article.category_id:
        cat_qs = qs.filter(category_id=article.category_id)
        other_qs = qs.exclude(category_id=article.category_id)
        combined = list(cat_qs) + list(other_qs)
    else:
        combined = list(qs)

    page = combined[offset : offset + limit]
    has_more = len(combined) > offset + limit

    articles_data = []
    for a in page:
        articles_data.append({
            "slug": a.slug,
            "title": a.title,
            "excerpt": truncatewords(striptags(a.content or ""), 15),
            "image_url": a.get_image_url() if a.get_image_url() else None,
            "date": a.published_at.strftime("%d %b %Y") if a.published_at else "",
        })

    return JsonResponse({"articles": articles_data, "has_more": has_more})


@login_required
def edit_news(request, slug):
    from .forms import NewsEditForm

    article = get_object_or_404(NewsArticles, slug=slug)
    is_mod = (request.user.role.role_name or "").lower() == "moderator"
    is_author = article.author_id == request.user.user_id
    if not is_mod and not is_author:
        return JsonResponse({"success": False, "error": "Нет прав для редактирования."})

    if request.method == "POST":
        form = NewsEditForm(request.POST, request.FILES, instance=article)
        if form.is_valid():
            article = form.save(commit=False)
            article.updated_at = timezone.now()
            image = form.cleaned_data.get("image")
            if image:
                # Удаляем старую картинку
                if article.image_url:
                    try:
                        default_storage.delete(article.image_url)
                    except Exception:
                        pass
                image_id = str(uuid.uuid4())
                file_path = f"news/{article.article_id}/{image_id}_{image.name}"
                default_storage.save(file_path, image)
                article.image_url = file_path
            article.save()
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": True})
            messages.success(request, "Новость обновлена!")
            return redirect("news_detail", slug=article.slug)
        else:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "errors": form.errors})
    else:
        form = NewsEditForm(instance=article)

    return render(request, "accounts/edit_news.html", {"form": form, "article": article})


@login_required
def delete_news_comment(request, comment_id):
    comment = get_object_or_404(NewsComments, comment_id=comment_id)
    is_mod = (request.user.role.role_name or "").lower() == "moderator"
    is_comment_author = comment.user_id == request.user.user_id
    if not is_mod and not is_comment_author:
        return JsonResponse({"success": False, "error": "Нет прав для удаления комментария."})

    if request.method == "POST":
        article_id = comment.article_id
        comment.delete()
        if is_mod:
            ModerationLog.objects.create(
                moderator=request.user,
                action="delete_comment",
                entity_type="news_article",
                entity_id=article_id,
                entity_title=f"Комментарий #{comment_id}",
                comment=f"Удалён комментарий к новости #{article_id}",
            )
        return JsonResponse({"success": True})
    return JsonResponse({"success": False, "error": "Требуется метод POST."})


@login_required
def delete_news(request, slug):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        return JsonResponse(
            {"success": False, "error": "У вас нет прав для этого действия."}
        )
    article = get_object_or_404(NewsArticles, slug=slug)
    if article.image_url:
        try:
            default_storage.delete(article.image_url)
        except Exception as e:
            logger.error(f"Ошибка при удалении картинки новости {article_id}: {str(e)}")
    article.delete()
    return JsonResponse({"success": True})
@login_required
def cosmochat(request):
    if not request.user.is_authenticated:
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse(
                {"success": False, "error": "Требуется авторизация"}, status=401
            )
        messages.error(
            request, "Пожалуйста, войдите в систему, чтобы получить доступ к чату."
        )
        return redirect("login")
    chats = (
        ChatConversations.objects.filter(chatparticipants__user=request.user)
        .prefetch_related(
            "chatparticipants_set__user"
        )
        .annotate(
            latest_message_time=Max("messages__created_at")
        )
        .order_by(F("latest_message_time").desc(nulls_last=True), "-updated_at")
    )
    for chat in chats:
        if chat.is_group_chat:
            chat.display_name = chat.name
            chat.display_avatar = None
        else:
            other_participant = None
            for p in chat.chatparticipants_set.all():
                if p.user_id != request.user.user_id:
                    other_participant = p
                    break
            if other_participant and other_participant.user:
                user_profile = other_participant.user
                # Используем переименованное имя чата, если оно есть, иначе имя пользователя
                if chat.name:
                    chat.display_name = chat.name
                else:
                    chat.display_name = f"{user_profile.first_name or ''} {user_profile.last_name or ''}".strip()
                chat.display_avatar = user_profile.get_profile_picture_url()
            else:
                chat.display_name = "Удаленный чат"
                chat.display_avatar = None
    search_form = UserSearchForm(request.GET)
    users = (
        Users.objects.all()
        .select_related("role")
        .annotate(
            invested_startups_count=Count(
                "investmenttransactions__startup",
                filter=(
                    Q(investmenttransactions__amount__gt=0)
                    & (
                        Q(investmenttransactions__transaction_status="completed")
                        | Q(investmenttransactions__transaction_status__isnull=True)
                    )
                ),
                distinct=True,
            ),
            owned_startups_count=Count("startups", distinct=True),
        )
    )
    if search_form.is_valid():
        query = search_form.cleaned_data.get("query", "")
        roles = search_form.cleaned_data.get("roles", [])
        if query:
            users = users.filter(
                Q(email__icontains=query)
                | Q(first_name__icontains=query)
                | Q(last_name__icontains=query)
            )
        if roles:
            users = users.filter(role__role_name__in=roles)
    users = users.exclude(user_id=request.user.user_id)
    chat_id = request.GET.get("chat_id")
    if chat_id:
        chat = ChatConversations.objects.filter(conversation_id=chat_id).first()
        if chat:
            participant_ids = chat.chatparticipants_set.values_list(
                "user_id", flat=True
            )
            users = users.exclude(user_id__in=participant_ids)
        else:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse(
                    {"success": False, "error": "Чат не найден"}, status=404
                )
    for user in users[:5]:
        profile_url = (
            user.get_profile_picture_url() if user.profile_picture_url else "None"
        )
        logger.info(
            f"Cosmochat User ID: {user.user_id}, Profile Picture URL: {user.profile_picture_url}, Generated URL: {profile_url}"
        )
    for chat in chats[:5]:
        participants = chat.chatparticipants_set.all()
        participant_info = [
            f"ID: {p.user.user_id}, Picture: {p.user.get_profile_picture_url() or 'None'}"
            for p in participants
            if p.user and p.user != request.user
        ]
        logger.info(
            f"Chat ID: {chat.conversation_id}, Participants (excluding self): {participant_info}"
        )
    message_form = MessageForm()
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        users_data = [
            {
                "user_id": user.user_id,
                "name": f"{user.first_name} {user.last_name}",
                "role": user.role.role_name if user.role else "Система",
            }
            for user in users
        ]
        return JsonResponse({"users": users_data})
    from accounts.promotions import get_active_ads
    return render(
        request,
        "accounts/cosmochat.html",
        {
            "search_form": search_form,
            "users": users,
            "chats": chats,
            "message_form": message_form,
            "cosmochat_ads": get_active_ads("cosmochat_banner"),
        },
    )

def chat_list(request):
    """API endpoint для получения списка чатов пользователя"""
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "error": "Требуется авторизация"}, status=401)

    try:
        chats = (
            ChatConversations.objects.filter(chatparticipants__user=request.user)
            .prefetch_related(
                "chatparticipants_set__user",
                "messages_set"
            )
            .annotate(
                latest_message_time=Max("messages__created_at"),
                has_messages=Count("messages"),
                chat_created_at=F("created_at")
            )
            .order_by(
                "-has_messages",
                F("latest_message_time").desc(nulls_last=True),
                "-updated_at",
                "-chat_created_at"
            )
        )

        chats_data = []
        for chat in chats:
            if chat.is_group_chat:
                try:
                    chat_name = chat.name or "Групповой чат"
                except:
                    chat_name = "Групповой чат"

                try:
                    conversation_id = chat.conversation_id or 0
                except:
                    conversation_id = 0

                try:
                    created_at = chat.created_at.isoformat() if chat.created_at else None
                except:
                    created_at = None

                try:
                    updated_at = chat.updated_at.isoformat() if chat.updated_at else None
                except:
                    updated_at = None

                try:
                    has_messages = chat.has_messages or 0
                except:
                    has_messages = 0

                try:
                    latest_message_time = chat.latest_message_time.isoformat() if chat.latest_message_time else None
                except:
                    latest_message_time = None

                chat_data = {
                    "conversation_id": conversation_id,
                    "name": chat_name,
                    "is_group_chat": True,
                    "is_deleted": False,
                    "has_left": False,
                    "is_deal": False,
                    "latest_message": None,
                    "unread_count": 0,
                    "created_at": created_at,
                    "updated_at": updated_at,
                    "has_messages": has_messages,
                    "latest_message_time": latest_message_time
                }
            else:
                other_participant = None
                for p in chat.chatparticipants_set.all():
                    if p.user_id != request.user.user_id:
                        other_participant = p
                        break

                if other_participant and other_participant.user:
                    user_profile = other_participant.user
                    # Используем переименованное имя чата, если оно есть, иначе имя пользователя
                    if chat.name:
                        user_name = chat.name
                        logger.info(f"Используем переименованное имя чата {chat.conversation_id}: {chat.name}")
                    else:
                        try:
                            user_name = f"{user_profile.first_name or ''} {user_profile.last_name or ''}".strip() or user_profile.email
                        except:
                            try:
                                user_id = user_profile.user_id or 0
                                user_name = user_profile.email or f"Пользователь {user_id}"
                            except:
                                try:
                                    user_id = user_profile.user_id or 0
                                except:
                                    user_id = 0
                                user_name = f"Пользователь {user_id}"

                    try:
                        conversation_id = chat.conversation_id or 0
                    except:
                        conversation_id = 0

                    try:
                        created_at = chat.created_at.isoformat() if chat.created_at else None
                    except:
                        created_at = None

                    try:
                        updated_at = chat.updated_at.isoformat() if chat.updated_at else None
                    except:
                        updated_at = None

                    try:
                        has_messages = chat.has_messages or 0
                    except:
                        has_messages = 0

                    try:
                        latest_message_time = chat.latest_message_time.isoformat() if chat.latest_message_time else None
                    except:
                        latest_message_time = None

                    chat_data = {
                        "conversation_id": conversation_id,
                        "name": user_name,
                        "is_group_chat": False,
                        "is_deleted": False,
                        "has_left": False,
                        "is_deal": False,
                        "latest_message": None,
                        "unread_count": 0,
                        "created_at": created_at,
                        "updated_at": updated_at,
                        "has_messages": has_messages,
                        "latest_message_time": latest_message_time
                    }
                else:
                    try:
                        conversation_id = chat.conversation_id or 0
                    except:
                        conversation_id = 0

                    try:
                        created_at = chat.created_at.isoformat() if chat.created_at else None
                    except:
                        created_at = None

                    try:
                        updated_at = chat.updated_at.isoformat() if chat.updated_at else None
                    except:
                        updated_at = None

                    try:
                        has_messages = chat.has_messages or 0
                    except:
                        has_messages = 0

                    try:
                        latest_message_time = chat.latest_message_time.isoformat() if chat.latest_message_time else None
                    except:
                        latest_message_time = None

                    chat_data = {
                        "conversation_id": conversation_id,
                        "name": "Удаленный чат",
                        "is_group_chat": False,
                        "is_deleted": True,
                        "has_left": True,
                        "is_deal": False,
                        "latest_message": None,
                        "unread_count": 0,
                        "created_at": created_at,
                        "updated_at": updated_at,
                        "has_messages": has_messages,
                        "latest_message_time": latest_message_time
                    }


            latest_message = chat.messages_set.order_by('-created_at').first()
            if latest_message:
                try:
                    is_read = latest_message.is_read()
                except:
                    is_read = False

                try:
                    sender_name = f"{latest_message.sender.first_name} {latest_message.sender.last_name}" if latest_message.sender else "Неизвестно"
                except:
                    sender_name = "Неизвестно"

                try:
                    created_at = latest_message.created_at.strftime("%d.%m.%Y %H:%M") if latest_message.created_at else ""
                    created_at_time = latest_message.created_at.strftime("%H:%M") if latest_message.created_at else ""
                    created_at_date = latest_message.created_at.strftime("%d.%m") if latest_message.created_at else ""
                except:
                    created_at = ""
                    created_at_time = ""
                    created_at_date = ""

                try:
                    message_text = latest_message.message_text or ""
                except:
                    message_text = ""

                try:
                    message_id = latest_message.message_id or 0
                except:
                    message_id = 0

                try:
                    sender_id = latest_message.sender.user_id if latest_message.sender else None
                except:
                    sender_id = None

                chat_data["last_message"] = {
                    "message_id": message_id,
                    "message_text": message_text,
                    "sender_id": sender_id,
                    "sender_name": sender_name,
                    "created_at": created_at,
                    "created_at_time": created_at_time,
                    "created_at_date": created_at_date,
                    "is_read": is_read
                }


            try:
                unread_count = chat.messages_set.filter(
                    ~Q(sender=request.user),
                    is_read=False
                ).count()
            except:
                unread_count = 0
            chat_data["unread_count"] = unread_count


            if not chat.is_group_chat:
                other_participant = None
                for p in chat.chatparticipants_set.all():
                    if p.user_id != request.user.user_id:
                        other_participant = p
                        break

                if other_participant and other_participant.user:
                    try:
                        profile_picture_url = other_participant.user.get_profile_picture_url()
                    except:
                        profile_picture_url = None

                    try:
                        first_name = other_participant.user.first_name or ""
                        last_name = other_participant.user.last_name or ""
                    except:
                        first_name = ""
                        last_name = ""

                    try:
                        user_id = other_participant.user.user_id or 0
                    except:
                        user_id = 0

                    chat_data["participant"] = {
                        "user_id": user_id,
                        "first_name": first_name,
                        "last_name": last_name,
                        "profile_picture_url": profile_picture_url
                    }

            chats_data.append(chat_data)

        return JsonResponse({
            "success": True,
            "chats": chats_data
        })

    except Exception as e:
        import traceback
        logger.error(f"Ошибка при получении списка чатов: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({
            "success": False,
            "error": "Внутренняя ошибка сервера"
        }, status=500)

def get_chat_messages(request, chat_id):
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "error": "Требуется авторизация"})
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}
        )
    since = request.GET.get("since")
    messages = chat.messages_set.all()
    if since:
        try:
            since_dt = dt.fromisoformat(since.replace("Z", "+00:00"))
            messages = messages.filter(created_at__gt=since_dt)
        except ValueError:
            return JsonResponse(
                {"success": False, "error": "Неверный формат параметра since"}
            )
    messages = messages.order_by("created_at")
    messages_data = [
        {
            "message_id": msg.message_id,
            "sender_id": msg.sender.user_id if msg.sender else None,
            "sender_name": (
                f"{msg.sender.first_name} {msg.sender.last_name}"
                if msg.sender
                else "Неизвестно"
            ),
            "message_text": msg.message_text,
            "created_at": (
                msg.created_at.strftime("%d.%m.%Y %H:%M") if msg.created_at else ""
            ),
            "created_at_iso": msg.created_at.isoformat() if msg.created_at else "",
            "is_read": msg.is_read(),
            "is_own": msg.sender == request.user if msg.sender else False,
        }
        for msg in messages
    ]
    participants = chat.get_participants()
    participants_data = [
        {
            "user_id": p.user.user_id,
            "name": f"{p.user.first_name} {p.user.last_name}",
            "role": p.user.role.role_name if p.user.role else "Неизвестно",
        }
        for p in participants
    ]
    # Определяем имя чата
    if chat.is_group_chat:
        chat_name = chat.name or "Групповой чат"
    else:
        # Для личных чатов используем переименованное имя чата, если оно есть
        if chat.name:
            chat_name = chat.name
        else:
            other_participant = None
            for p in participants:
                if p.user.user_id != request.user.user_id:
                    other_participant = p
                    break
            if other_participant and other_participant.user:
                chat_name = f"{other_participant.user.first_name or ''} {other_participant.user.last_name or ''}".strip()
            else:
                chat_name = "Удаленный чат"
    
    logger.info(f"Возвращаем имя чата {chat_id}: {chat_name}")
    return JsonResponse(
        {"success": True, "messages": messages_data, "participants": participants_data, "chat_name": chat_name}
    )


@login_required
def send_message(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    form = MessageForm(request.POST)
    if not form.is_valid():
        return JsonResponse({"success": False, "error": "Неверные данные формы"})
    chat_id = request.POST.get("chat_id")
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}
        )
    message = Messages(
        conversation=chat,
        sender=request.user,
        message_text=form.cleaned_data["message_text"],
        status=MessageStatuses.objects.get(status_name="sent"),
        created_at=timezone.now(),
        updated_at=timezone.now(),
    )
    message.save()
    chat.updated_at = timezone.now()
    chat.save()

    msg_data = {
        "message_id": message.message_id,
        "sender_id": request.user.user_id,
        "sender_name": f"{request.user.first_name} {request.user.last_name}",
        "message_text": message.message_text,
        "created_at": message.created_at.strftime("%d.%m.%Y %H:%M"),
        "created_at_iso": message.created_at.isoformat(),
        "is_read": message.is_read(),
        "is_own": True,
    }

    # Уведомляем WebSocket-клиентов о новом сообщении
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                f"chat_{chat_id}",
                {
                    "type": "chat_new_message",
                    "message": msg_data,
                    "sender_id": request.user.user_id,
                },
            )
    except Exception:
        pass  # WebSocket недоступен — не блокируем HTTP-ответ

    return JsonResponse({"success": True, "message": msg_data})


@login_required
def mark_messages_read(request, chat_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}
        )
    read_status = MessageStatuses.objects.get(status_name="read")
    messages = chat.messages_set.filter(status__status_name="sent").exclude(
        sender=request.user
    )
    messages.update(status=read_status, updated_at=timezone.now())
    return JsonResponse({"success": True})


@login_required
def start_chat(request, user_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    target_user = get_object_or_404(Users, user_id=user_id)
    if target_user == request.user:
        return JsonResponse(
            {"success": False, "error": "Нельзя создать чат с самим собой"}
        )
    existing_chat = (
        ChatConversations.objects.annotate(num_participants=Count("chatparticipants"))
        .filter(
            is_group_chat=False, num_participants=2, chatparticipants__user=request.user
        )
        .filter(chatparticipants__user=target_user)
        .first()
    )
    if existing_chat:
        return JsonResponse(
            {"success": True, "chat_id": existing_chat.conversation_id, "existed": True}
        )
    chat = ChatConversations.objects.create(
        name=f"Чат {request.user.first_name} и {target_user.first_name}",
        is_group_chat=False,
        created_at=timezone.now(),
        updated_at=timezone.now(),
    )
    ChatParticipants.objects.create(conversation=chat, user=request.user)
    ChatParticipants.objects.create(conversation=chat, user=target_user)
    chat_data = {
        "conversation_id": chat.conversation_id,
        "name": chat.name,
        "is_group_chat": chat.is_group_chat,
        "participant": {
            "user_id": target_user.user_id,
            "first_name": target_user.first_name,
            "profile_picture_url": target_user.get_profile_picture_url(),
        },
        "last_message": None,
        "unread_count": 0,
    }
    return JsonResponse({"success": True, "chat": chat_data, "existed": False})


@login_required
def add_participant(request, chat_id):
    logger.debug(
        f"Adding participant to chat {chat_id}, user_id: {request.POST.get('user_id')}"
    )
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    user_id = request.POST.get("user_id")
    if not user_id:
        return JsonResponse(
            {"success": False, "error": "Не указан пользователь"}, status=400
        )
    try:
        new_user = Users.objects.get(user_id=user_id)
    except Users.DoesNotExist:
        logger.error(f"User {user_id} not found")
        return JsonResponse(
            {"success": False, "error": "Пользователь не найден"}, status=404
        )
    if chat.chatparticipants_set.filter(user=new_user).exists():
        return JsonResponse(
            {"success": False, "error": "Пользователь уже в чате"}, status=400
        )
    if not chat.is_group_chat:
        participants = chat.get_participants()
        if participants.count() >= 3:
            return JsonResponse(
                {"success": False, "error": "В личном чате максимум 3 участника"},
                status=400,
            )
        current_roles = {
            p.user.role.role_name.lower()
            for p in participants
            if p.user and p.user.role
        }
        if new_user.role and new_user.role.role_name.lower() in current_roles:
            return JsonResponse(
                {"success": False, "error": "Пользователь с такой ролью уже в чате"},
                status=400,
            )
    ChatParticipants.objects.create(conversation=chat, user=new_user)
    chat.updated_at = timezone.now()
    chat.save()
    logger.info(f"Добавлен участник {new_user.user_id} в чат {chat.conversation_id}")
    return JsonResponse(
        {
            "success": True,
            "new_participant": {
                "user_id": new_user.user_id,
                "name": f"{new_user.first_name} {new_user.last_name}",
                "role": new_user.role.role_name if new_user.role else "Система",
            },
        }
    )


@login_required
def available_users_for_chat(request, chat_id):
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    current_participant_ids = chat.chatparticipants_set.values_list(
        "user_id", flat=True
    )
    users = Users.objects.exclude(user_id__in=current_participant_ids).exclude(
        user_id=request.user.user_id
    )
    if chat.is_group_chat:
        users = users.exclude(role__role_name="moderator")
    else:
        current_roles = chat.chatparticipants_set.exclude(
            user=request.user
        ).values_list("user__role__role_name", flat=True)
        users = users.filter(
            role__role_name__in=["startuper", "investor", "moderator"]
        ).exclude(role__role_name__in=current_roles)
    users_data = [
        {
            "user_id": user.user_id,
            "name": f"{user.first_name} {user.last_name}",
            "role": user.role.role_name if user.role else "Неизвестно",
        }
        for user in users
    ]
    return JsonResponse({"success": True, "users": users_data})


@login_required
def leave_chat(request, chat_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}
        )
    ChatParticipants.objects.filter(conversation=chat, user=request.user).delete()
    remaining_participants = chat.chatparticipants_set.all()
    if remaining_participants.exists():
        message = Messages(
            conversation=chat,
            sender=None,
            message_text=f"{request.user.first_name} {request.user.last_name} покинул чат",
            status=MessageStatuses.objects.get(status_name="sent"),
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        message.save()
        chat.updated_at = timezone.now()
        chat.save()
    else:
        chat.delete()
        return JsonResponse({"success": True, "deleted": True})
    return JsonResponse({"success": True, "deleted": False})


def planetary_system(request):
    """
    Планетарная система - отображает стартапы как планеты на орбитах
    """
    directions_data = FIXED_CATEGORIES.copy()
    selected_direction_name = request.GET.get("direction", "All")
    logger.info(f"🪐 Планетарная система: выбрано направление '{selected_direction_name}'")
    startups_query = Startups.objects.filter(
        status="approved"
    ).select_related("direction", "owner").order_by("-created_at")
    if selected_direction_name != "All" and selected_direction_name != "Все":
        from django.db.models import Q
        direction_filter = Q()
        for category in FIXED_CATEGORIES:
            if category['original_name'] == selected_direction_name or category['direction_name'] == selected_direction_name:
                direction_filter |= Q(direction__direction_name=category['direction_name'])
        if direction_filter:
            startups_query = startups_query.filter(direction_filter)
    startups_list = list(startups_query)
    logger.info(f"🪐 Загружено стартапов: {len(startups_list)}")
    selected_startups = []
    if len(startups_list) > 0:
        selected_startups = startups_list[:6]
    else:
        selected_startups = []
    planets_data = []
    for i, startup in enumerate(selected_startups):
        planet_image_url = None

        if startup.planet_image:
            planet_image_url = get_planet_image_url(startup.planet_image)

        if not planet_image_url:
            planet_image_url = get_fallback_planet_url(startup.startup_id)

        direction_original = 'Не указано'
        if startup.direction:
            for cat in directions_data:
                if cat['direction_name'] == startup.direction.direction_name or cat['original_name'] == getattr(startup.direction, 'original_name', None):
                    direction_original = cat['original_name']
                    break
        planets_data.append({
            "id": startup.startup_id,
            "startup_id": startup.startup_id,
            "name": startup.title,
            "description": strip_tags(startup.short_description or startup.description[:200]) if (startup.short_description or startup.description) else "",
            "image": planet_image_url,
            "rating": startup.get_average_rating(),
            "voters_count": startup.total_voters,
            "comment_count": startup.comments.count(),
            "direction": direction_original,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не указано",
            "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указано",
            "investors": startup.get_investors_count(),
            "progress": startup.get_progress_percentage(),
            "investment_type": "Выкуп+инвестирование" if startup.both_mode else ("Только выкуп" if startup.only_buy else "Только инвестирование"),
            "logo": startup.get_logo_url(),
        })
    all_approved_startups = list(Startups.objects.filter(status="approved").select_related("direction", "owner").order_by("-created_at"))
    all_startups_data = []
    for idx, startup in enumerate(all_approved_startups):
        planet_image_url = None

        if startup.planet_image:
            planet_image_url = get_planet_image_url(startup.planet_image)

        if not planet_image_url:
            planet_image_url = get_fallback_planet_url(startup.startup_id)

        direction_original = 'Не указано'
        if startup.direction:
            for cat in directions_data:
                if cat['direction_name'] == startup.direction.direction_name or cat['original_name'] == getattr(startup.direction, 'original_name', None):
                    direction_original = cat['original_name']
                    break
        all_startups_data.append({
            "id": startup.startup_id,
            "startup_id": startup.startup_id,
            "name": startup.title,
            "description": strip_tags(startup.short_description or startup.description[:200]) if (startup.short_description or startup.description) else "",
            "image": planet_image_url,
            "rating": startup.get_average_rating(),
            "voters_count": startup.total_voters,
            "comment_count": startup.comments.count(),
            "direction": direction_original,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не указано",
            "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указано",
            "investors": startup.get_investors_count(),
            "progress": startup.get_progress_percentage(),
            "investment_type": "Выкуп+инвестирование" if startup.both_mode else ("Только выкуп" if startup.only_buy else "Только инвестирование"),
            "logo": startup.get_logo_url(),
        })
    logo_data = {
        "image": "/static/accounts/images/logo.png"
    }
    context = {
        "planets_data_json": json.dumps(planets_data, ensure_ascii=False),
        "directions_data_json": json.dumps(directions_data, ensure_ascii=False),
        "all_startups_data_json": json.dumps(all_startups_data, ensure_ascii=False),
        "logo_data": logo_data,
        "directions": directions_data,
        "selected_galaxy": selected_direction_name,
    }
    return render(request, "accounts/planetary_system.html", context)


@login_required
def my_startups(request):
    try:
        if request.user.role and request.user.role.role_name == 'startuper':
            user_startups_qs = (
                Startups.objects.filter(owner=request.user)
                .select_related("direction", "stage", "status_id")
                .prefetch_related("comments")
            )
        else:
            user_startups_qs = (
                Startups.objects.all()
                .select_related("direction", "stage", "status_id")
                .prefetch_related("comments")
            )
        total_user_startups_count = user_startups_qs.count()
        approved_startups_qs = user_startups_qs.filter(status="approved")
        financial_analytics_data = approved_startups_qs.aggregate(
            total_raised=Sum("amount_raised"),
            max_raised=Max("amount_raised"),
            approved_startups_count=Count("startup_id"),
        )
        approved_startups_count = financial_analytics_data.get(
            "approved_startups_count", 0
        )
        total_amount_raised = financial_analytics_data.get("total_raised") or Decimal(
            "0"
        )
        max_raised = financial_analytics_data.get("max_raised") or Decimal("0")

        startups_with_funding = approved_startups_qs.filter(amount_raised__gt=0)
        min_raised_data = startups_with_funding.aggregate(
            min_raised=Min("amount_raised")
        )
        min_raised = min_raised_data.get("min_raised") or Decimal("0")
        category_data_raw = (
            user_startups_qs.values("direction__direction_name")
            .annotate(category_count=Count("startup_id"))
            .order_by("-category_count")
        )
        investment_categories = []
        invested_category_data_dict = {}
        total_for_category_percentage = (
            total_user_startups_count if total_user_startups_count > 0 else 1
        )
        for cat_data in category_data_raw:
            percentage = 0
            category_count = cat_data.get("category_count")
            category_name = cat_data.get("direction__direction_name") or "Без категории"
            if category_count and total_for_category_percentage > 0:
                try:
                    percentage = round(
                        (int(category_count) / total_for_category_percentage) * 100
                    )
                    percentage = min(percentage, 100)
                except Exception as e:
                    logger.error(
                        f"Ошибка расчета процента (по количеству) для категории '{category_name}': {e}"
                    )
                    percentage = 0
            investment_categories.append(
                {
                    "name": category_name,
                    "percentage": percentage,
                }
            )
            invested_category_data_dict[category_name] = percentage
        current_year = timezone.now().year
        logger.info(
            f"[my_startups] Preparing chart data for user {request.user.email}, year: {current_year}"
        )
        monthly_data_direct = (
            approved_startups_qs.filter(
                updated_at__year=current_year, amount_raised__gt=0
            )
            .annotate(month=TruncMonth("updated_at"))
            .values("month")
            .annotate(monthly_total=Sum(Coalesce("amount_raised", Decimal(0))))
            .order_by("month")
        )
        month_labels = [
            "Янв",
            "Фев",
            "Мар",
            "Апр",
            "Май",
            "Июн",
            "Июл",
            "Авг",
            "Сен",
            "Окт",
            "Ноя",
            "Дек",
        ]
        monthly_totals = [0] * 12
        for data in monthly_data_direct:
            month_index = data["month"].month - 1
            if 0 <= month_index < 12:
                monthly_total_decimal = data.get(
                    "monthly_total", Decimal(0)
                ) or Decimal(0)
                monthly_totals[month_index] = float(monthly_total_decimal)
        logger.info(
            f"[my_startups] Preparing chart data for user {request.user.email}, year: {current_year}"
        )
        monthly_category_data_raw = (
            approved_startups_qs.filter(
                updated_at__year=current_year,
                amount_raised__gt=0,
                direction__isnull=False,
            )
            .annotate(month=TruncMonth("updated_at"))
            .values("month", "direction__direction_name")
            .annotate(monthly_category_total=Sum(Coalesce("amount_raised", Decimal(0))))
            .order_by("month", "direction__direction_name")
        )
        logger.info(
            f"[my_startups] Raw monthly category data from DB: {list(monthly_category_data_raw)}"
        )
        structured_monthly_data = collections.defaultdict(
            lambda: collections.defaultdict(float)
        )
        unique_categories = set()
        for data in monthly_category_data_raw:
            month_dt = data["month"]
            category_name = data["direction__direction_name"]
            amount = float(data.get("monthly_category_total", 0) or 0)
            month_key = month_dt.strftime("%Y-%m-01")
            structured_monthly_data[month_key][category_name] += amount
            unique_categories.add(category_name)
        sorted_categories = sorted(list(unique_categories))
        logger.info(
            f"[my_startups] Unique categories found for chart: {sorted_categories}"
        )
        chart_data_list = []
        start_date = datetime.date(current_year, 1, 1)
        for i in range(12):
            current_month_key = (start_date + relativedelta(months=i)).strftime(
                "%Y-%m-01"
            )
            month_data = {
                "month_key": current_month_key,
                "category_data": dict(structured_monthly_data[current_month_key]),
            }
            chart_data_list.append(month_data)
        logger.info(
            f"[my_startups] Final structured chart data list: {chart_data_list}"
        )
        try:
            all_directions_qs = Directions.objects.all().order_by("direction_name")
            all_directions_list = [
                {"direction_name": d.direction_name} for d in all_directions_qs
            ]
        except Exception as e:
            logger.error(f"Ошибка при получении всех направлений: {str(e)}")
            all_directions_list = []
        try:
            approved_startups_annotated = (
                approved_startups_qs.annotate(
                    average_rating=Avg(
                        models.ExpressionWrapper(
                            Coalesce(models.F("sum_votes"), 0)
                            * 1.0
                            / Coalesce(models.F("total_voters"), 1),
                            output_field=FloatField(),
                        ),
                        filter=models.Q(total_voters__gt=0),
                        default=0.0,
                    ),
                    comment_count=Count("comments"),
                )
                .annotate(average_rating=Coalesce("average_rating", 0.0))
                .order_by("-created_at")
            )
        except Exception as e:
            logger.error(f"Ошибка при получении одобренных стартапов: {str(e)}")
            approved_startups_annotated = []
        s3_client = boto3.client(
            "s3",
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            endpoint_url=settings.AWS_S3_ENDPOINT_URL,
            region_name=settings.AWS_S3_REGION_NAME,
        )
        planetary_startups = []
        for idx, startup in enumerate(approved_startups_annotated, start=1):
            orbit_size = (idx * 100) + 150
            orbit_time = (idx * 10) + 40
            planet_size = 60
            planet_data = {
                "startup_id": startup.startup_id,
                "title": startup.title or "Без названия",
                "planet_image": startup.planet_image,
                "logo_urls": startup.logo_urls,
                "average_rating": startup.average_rating or 0,
                "rating": startup.average_rating or 0,
                "total_voters": startup.total_voters or 0,
                "comment_count": startup.comment_count or 0,
                "description": strip_tags(startup.description) if startup.description else "Описание отсутствует.",
                "short_description": strip_tags(startup.description) if startup.description else "Описание отсутствует.",
                "progress": startup.get_progress_percentage() or 0,
                "funding_goal": startup.funding_goal or 0,
                "amount_raised": startup.amount_raised or 0,
                "get_investors_count": startup.get_investors_count(),
                "direction": startup.direction.direction_name if startup.direction else "Не указано",
                "investment_type": "Не указано",
                "orbit_size": orbit_size,
                "orbit_time": orbit_time,
                "planet_size": planet_size,
            }
            planetary_startups.append(planet_data)
    except Exception as e:
        logger.error(f"Критическая ошибка в my_startups view: {e}", exc_info=True)
        messages.error(
            request, "Произошла ошибка при загрузке страницы ваших стартапов."
        )
        return redirect("profile")
    context = {
        "user_startups": approved_startups_annotated,
        "planetary_startups": planetary_startups,
        "total_investment": total_amount_raised,
        "startups_count": approved_startups_count,
        "max_investment": max_raised,
        "min_investment": min_raised,
        "investment_categories": investment_categories[
            :7
        ],
        "invested_category_data": invested_category_data_dict,
        "all_directions": all_directions_list,
        "month_labels": month_labels,
        "chart_monthly_category_data": chart_data_list,
        "chart_categories": sorted_categories,
        "startup_applications": user_startups_qs.order_by("-updated_at"),
    }
    context["planetary_startups_json"] = json.dumps(
        planetary_startups, cls=DjangoJSONEncoder, ensure_ascii=False
    )
    return render(request, "accounts/my_startups.html", context)


@login_required
def notifications_view(request):
    return render(request, "accounts/notifications.html")


@login_required
def create_group_chat(request):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Метод не разрешен."}, status=405
        )
    try:
        data = json.loads(request.body)
        chat_name = data.get("name", "").strip()
        user_ids = data.get("user_ids", [])
        if not chat_name:
            return JsonResponse(
                {"success": False, "error": "Название чата не может быть пустым."},
                status=400,
            )
        if not user_ids:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Необходимо выбрать хотя бы одного участника.",
                },
                status=400,
            )
        try:
            participant_ids = list(set(int(uid) for uid in user_ids))
        except (ValueError, TypeError):
            return JsonResponse(
                {"success": False, "error": "Неверный формат ID пользователей."},
                status=400,
            )
        if request.user.user_id in participant_ids:
            participant_ids.remove(request.user.user_id)
        if not participant_ids:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Нельзя создать групповой чат только с самим собой.",
                },
                status=400,
            )
        if Users.objects.filter(
            user_id__in=participant_ids, role__role_name="moderator"
        ).exists():
            return JsonResponse(
                {
                    "success": False,
                    "error": "Модераторы не могут быть добавлены в групповой чат.",
                },
                status=400,
            )
        with transaction.atomic():
            conversation = ChatConversations.objects.create(
                name=chat_name,
                is_group_chat=True,
                created_at=timezone.now(),
                updated_at=timezone.now(),
            )
            all_participant_users = [request.user]
            users_to_add = Users.objects.filter(user_id__in=participant_ids)
            all_participant_users.extend(list(users_to_add))
            if len(all_participant_users) != len(participant_ids) + 1:
                logger.error(
                    f"Не все пользователи найдены для создания чата. Передано ID: {participant_ids}"
                )
                raise Exception("Один или несколько пользователей не найдены.")
            participants_to_create = [
                ChatParticipants(conversation=conversation, user=user)
                for user in all_participant_users
            ]
            ChatParticipants.objects.bulk_create(participants_to_create)
        chat_data = {
            "conversation_id": conversation.conversation_id,
            "name": conversation.name,
            "is_group_chat": conversation.is_group_chat,
            "participant": None,
            "last_message": None,
            "unread_count": 0,
        }
        logger.info(
            f"Групповой чат создан: ID={conversation.conversation_id}, Название={chat_name}, Участников={len(all_participant_users)}"
        )
        return JsonResponse({"success": True, "chat": chat_data})
    except json.JSONDecodeError:
        logger.error("Неверный формат JSON в create_group_chat")
        return JsonResponse(
            {"success": False, "error": "Неверный формат данных (JSON)."}, status=400
        )
    except Exception as e:
        logger.error(f"Ошибка при создании группового чата: {str(e)}", exc_info=True)
        return JsonResponse(
            {"success": False, "error": "Внутренняя ошибка сервера."}, status=500
        )


def support_page_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    return redirect('support_orders')


@login_required
def change_owner(request, startup_id):
    logger.info(f"Change owner request for startup {startup_id} by user {request.user.user_id}")

    if request.method != "POST":
        logger.warning(f"Invalid method {request.method} for change_owner")
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})

    if not is_moderator(request.user):
        logger.warning(f"User {request.user.user_id} does not have moderator role")
        return JsonResponse(
            {"success": False, "error": "У вас нет прав для этого действия"}
        )

    try:
        startup = get_object_or_404(Startups, startup_id=startup_id)
        new_owner_id = request.POST.get("new_owner_id")

        if not new_owner_id:
            logger.error("No new_owner_id provided")
            return JsonResponse({"success": False, "error": "Не указан новый владелец"})

        new_owner = get_object_or_404(Users, user_id=new_owner_id)
        startup.owner = new_owner
        startup.save()

        logger.info(f"Successfully changed owner of startup {startup_id} to user {new_owner_id}")
        return JsonResponse({"success": True})

    except Exception as e:
        logger.error(f"Error changing owner for startup {startup_id}: {str(e)}")
        return JsonResponse({"success": False, "error": f"Ошибка при смене владельца: {str(e)}"})


@login_required
def get_investors(request, startup_id):
    logger.info(f"Get investors request for startup {startup_id} by user {request.user.user_id}")

    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        logger.warning(f"User {request.user.user_id} does not have moderator role for get_investors")
        return JsonResponse({"error": "Доступ запрещен"}, status=403)

    try:
        startup = get_object_or_404(Startups, startup_id=startup_id)
        investors = InvestmentTransactions.objects.filter(startup=startup).select_related(
            "investor"
        ).defer("franchise")

        logger.info(f"Found {investors.count()} investment transactions for startup {startup_id}")

        investor_list = []
        for tx in investors:
            if tx.investor:
                investor_list.append(
                    {
                        "transaction_id": tx.transaction_id,
                        "user_id": tx.investor.user_id,
                        "name": tx.investor.get_full_name() or tx.investor.email,
                        "amount": float(tx.amount),
                    }
                )

        html = render_to_string(
            "accounts/partials/_investors_list.html",
            {"investors": investor_list, "startup": startup, "user": request.user},
        )

        logger.info(f"Generated HTML for {len(investor_list)} investors")
        return JsonResponse({"html": html})

    except Exception as e:
        logger.error(f"Error getting investors for startup {startup_id}: {str(e)}")
        return JsonResponse({"error": f"Ошибка при получении списка инвесторов: {str(e)}"}, status=500)


@login_required
def add_investor(request, startup_id):
    logger.info(f"Add investor request for startup {startup_id} by user {request.user.user_id}")

    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        logger.warning(f"User {request.user.user_id} does not have moderator role for add_investor")
        return JsonResponse({"error": "Доступ запрещен"}, status=403)

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            user_id = data.get("user_id")
            amount = Decimal(data.get("amount"))

            logger.info(f"Adding investor {user_id} with amount {amount} to startup {startup_id}")

            startup = get_object_or_404(Startups, startup_id=startup_id)
            user_to_invest = get_object_or_404(Users, user_id=user_id)

            if amount <= 0:
                logger.warning(f"Invalid amount {amount} for startup {startup_id}")
                return JsonResponse(
                    {"success": False, "error": "Сумма должна быть положительной."}
                )

            logger.info(f"Creating new investment for user {user_id} in startup {startup_id}")
            try:
                investment_type_obj = TransactionTypes.objects.get(
                    type_name="investment"
                )
                payment_method, _ = PaymentMethods.objects.get_or_create(
                    method_name="default",
                    defaults={"method_name": "default"}
                )
                from django.db import connection
                with connection.cursor() as cursor:
                    cursor.execute("""
                        INSERT INTO investment_transactions 
                        (startup_id, investor_id, amount, transaction_type_id, payment_method_id, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, [
                        startup.startup_id,
                        user_to_invest.user_id,
                        amount,
                        investment_type_obj.type_id,
                        payment_method.method_id,
                        timezone.now(),
                        timezone.now(),
                    ])
                startup.amount_raised = (startup.amount_raised or Decimal("0")) + amount
            except TransactionTypes.DoesNotExist:
                logger.error("Transaction type 'investment' not found")
                return JsonResponse(
                    {"error": "Тип транзакции 'investment' не найден в системе."},
                    status=500,
                )

            startup.save(update_fields=["amount_raised"])
            new_investor_count = startup.get_investors_count()

            logger.info(f"Successfully added investor to startup {startup_id}. New amount: {startup.amount_raised}, investors: {new_investor_count}")

            return JsonResponse(
                {
                    "success": True,
                    "new_amount_raised": float(startup.amount_raised),
                    "new_investor_count": new_investor_count,
                }
            )
        except (json.JSONDecodeError, TypeError, ValueError) as e:
            logger.error(f"Data format error in add_investor: {str(e)}")
            return JsonResponse(
                {"error": f"Неверный формат данных: {str(e)}"}, status=400
            )
        except Exception as e:
            logger.error(f"Unexpected error in add_investor: {str(e)}")
            return JsonResponse(
                {"error": f"Внутренняя ошибка сервера: {str(e)}"}, status=500
            )

    logger.warning(f"Invalid method {request.method} for add_investor")
    return JsonResponse({"error": "Метод не поддерживается"}, status=405)


@login_required
def edit_investment(request, startup_id, user_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    if not is_moderator(request.user):
        return JsonResponse(
            {"success": False, "error": "У вас нет прав для этого действия"}
        )
    startup = get_object_or_404(Startups, startup_id=startup_id)
    investor = get_object_or_404(Users, user_id=user_id)
    new_amount = Decimal(request.POST.get("amount", "0"))
    if new_amount <= 0:
        return JsonResponse({"success": False, "error": "Сумма должна быть больше 0"})
    transaction = get_object_or_404(
        InvestmentTransactions,
        startup=startup,
        investor=investor,
        transaction_status="completed",
    )
    old_amount = transaction.amount
    transaction.amount = new_amount
    transaction.updated_at = timezone.now()
    transaction.save()
    startup.amount_raised = (
        (startup.amount_raised or Decimal("0")) - old_amount + new_amount
    )
    startup.save()
    return JsonResponse({"success": True})


@login_required
def delete_investment(request, startup_id, user_id):
    logger.info(f"Delete investment request for startup {startup_id}, user {user_id} by user {request.user.user_id}")

    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        logger.warning(f"User {request.user.user_id} does not have moderator role for delete_investment")
        return JsonResponse({"error": "Доступ запрещен"}, status=403)

    if request.method == "POST":
        with transaction.atomic():
            try:
                data = {}
                if request.body:
                    try:
                        data = json.loads(request.body)
                    except (json.JSONDecodeError, ValueError):
                        pass
                
                transaction_id = data.get("transaction_id")
                
                if transaction_id:
                    tx_qs = InvestmentTransactions.objects.filter(
                        transaction_id=transaction_id,
                        startup_id=startup_id,
                    ).defer("franchise")
                    if not tx_qs.exists():
                        return JsonResponse({"success": False, "error": "Инвестиция не найдена"}, status=404)
                    tx = tx_qs.first()
                else:
                    user_to_delete = get_object_or_404(Users, pk=user_id)
                    tx_qs = InvestmentTransactions.objects.filter(
                        startup_id=startup_id,
                        investor=user_to_delete,
                    ).defer("franchise")
                    if not tx_qs.exists():
                        return JsonResponse({"success": False, "error": "Инвестиция не найдена"}, status=404)
                    tx = tx_qs.first()

                logger.info(f"Found investment transaction {tx.transaction_id} for deletion")

                deleted_amount = tx.amount
                startup = get_object_or_404(Startups, startup_id=startup_id)
                
                from django.db import connection
                with connection.cursor() as cursor:
                    cursor.execute(
                        "DELETE FROM investment_transactions WHERE transaction_id = %s",
                        [tx.transaction_id]
                    )

                startup.amount_raised = (startup.amount_raised or Decimal("0")) - deleted_amount
                if startup.amount_raised < 0:
                    startup.amount_raised = Decimal("0")
                startup.save(update_fields=["amount_raised"])
                new_investor_count = startup.get_investors_count()

                logger.info(f"Successfully deleted investment. New total: {startup.amount_raised}, investors: {new_investor_count}")

                return JsonResponse(
                    {
                        "success": True,
                        "new_amount_raised": float(startup.amount_raised),
                        "new_investor_count": new_investor_count,
                    }
                )
            except InvestmentTransactions.DoesNotExist:
                logger.warning(f"Investment transaction not found for startup {startup_id}, user {user_id}")
                return JsonResponse({"success": False, "error": "Инвестиция не найдена"}, status=404)
            except Exception as e:
                logger.error(f"Error deleting investment: {str(e)}", exc_info=True)
                return JsonResponse({"success": False, "error": f"Ошибка при удалении: {str(e)}"}, status=500)

    logger.warning(f"Invalid method {request.method} for delete_investment")
    return JsonResponse({"error": "Неверный метод запроса"}, status=405)


@login_required
def support_orders_view(request):
    user_is_mod = is_moderator(request.user)
    if user_is_mod:
        orders = SupportTicket.objects.all().order_by("-created_at")
    else:
        orders = SupportTicket.objects.filter(user=request.user).order_by("-created_at")


    page_number = request.GET.get('page', 1)
    paginator = Paginator(orders, 10)
    page_obj = paginator.get_page(page_number)

    context = {
        "orders": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "is_moderator": user_is_mod
    }
    return render(request, "accounts/support_orders.html", context)


@login_required
def support_ticket_detail(request, ticket_id):
    ticket = get_object_or_404(SupportTicket, pk=ticket_id)
    user = request.user
    user_is_mod = is_moderator(user)
    if not (user == ticket.user or user_is_mod):
        return HttpResponseForbidden("У вас нет доступа к этой заявке.")

    if user_is_mod:
        all_tickets = SupportTicket.objects.all().order_by("-created_at")
    else:
        all_tickets = SupportTicket.objects.filter(user=user).order_by("-created_at")

    form = None
    if user_is_mod:
        if request.method == "POST":
            form = ModeratorTicketForm(request.POST, instance=ticket)
            if form.is_valid():
                form.save()
                ticket.refresh_from_db()
                messages.success(request, "Заявка успешно обновлена.")
                return redirect("support_ticket_detail", ticket_id=ticket.ticket_id)
        else:
            form = ModeratorTicketForm(instance=ticket)

    context = {
        "ticket": ticket,
        "form": form,
        "is_moderator": user_is_mod,
        "all_tickets": all_tickets,
    }
    return render(request, "accounts/support_ticket_detail.html", context)


@login_required
def close_support_ticket(request, ticket_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"}, status=405)

    ticket = get_object_or_404(SupportTicket, pk=ticket_id)
    user = request.user
    user_is_mod = is_moderator(user)

    if not (user == ticket.user or user_is_mod):
        return JsonResponse({"success": False, "error": "У вас нет доступа к этой заявке"}, status=403)

    try:
        ticket.status = 'closed'
        ticket.save()
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
def update_ticket_status(request, ticket_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"}, status=405)

    ticket = get_object_or_404(SupportTicket, pk=ticket_id)
    user = request.user
    user_is_mod = is_moderator(user)

    if not user_is_mod:
        return JsonResponse({"success": False, "error": "У вас нет прав для изменения статуса"}, status=403)

    try:
        data = json.loads(request.body)
        new_status = data.get('status')

        if new_status not in ['new', 'in_progress', 'closed']:
            return JsonResponse({"success": False, "error": "Неверный статус"}, status=400)

        ticket.status = new_status
        ticket.save()
        return JsonResponse({"success": True})
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Неверный формат данных"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
def support_contact_view(request):
    if request.method == "POST":
        form = SupportTicketForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.user = request.user
            ticket.save()
            try:
                logger.info(f"Dispatching Telegram for support ticket {ticket.ticket_id}")
                sent_ok = send_telegram_support_message(ticket)
                logger.info(f"Telegram dispatch result for ticket {ticket.ticket_id}: {sent_ok}")
            except Exception as e:
                logger.error(f"Unexpected error during Telegram dispatch for ticket {ticket.ticket_id}: {e}", exc_info=True)
            messages.success(
                request, "Ваше обращение успешно отправлено! Мы скоро с вами свяжемся."
            )
            return redirect("support_contact")
    else:
        form = SupportTicketForm()
    context = {"form": form}
    return render(request, "accounts/support_contact.html", context)


@login_required
def rename_chat(request, chat_id):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    try:
        data = json.loads(request.body)
        new_name = data.get("name", "").strip()
        if not new_name:
            return JsonResponse(
                {"success": False, "error": "Название не может быть пустым"}, status=400
            )
        with transaction.atomic():
            chat.name = new_name
            chat.updated_at = timezone.now()
            chat.save()
        logger.info(f"Чат {chat.conversation_id} переименован в {new_name}")
        return JsonResponse({"success": True, "chat_name": new_name})
    except json.JSONDecodeError:
        logger.error("Неверный формат JSON в rename_chat")
        return JsonResponse(
            {"success": False, "error": "Неверный формат данных"}, status=400
        )
    except Exception as e:
        logger.error(f"Ошибка при переименовании чата {chat_id}: {str(e)}")
        return JsonResponse(
            {"success": False, "error": f"Ошибка: {str(e)}"}, status=500
        )


@login_required
def available_users(request):
    users = Users.objects.exclude(user_id=request.user.user_id).exclude(
        role__role_name="moderator"
    )
    users_data = [
        {
            "user_id": user.user_id,
            "name": f"{user.first_name} {user.last_name}",
            "role": user.role.role_name if user.role else "unknown",
            "profile_picture_url": user.get_profile_picture_url() or "",
        }
        for user in users
    ]
    return JsonResponse({"success": True, "users": users_data})


@login_required
def find_or_create_chat(request, recipient_id):
    if request.method == "POST":
        recipient = get_object_or_404(Users, user_id=recipient_id)
        if request.user.user_id == recipient.user_id:
            return JsonResponse(
                {"error": "You cannot start a chat with yourself."}, status=400
            )
        user_chats = ChatConversations.objects.filter(
            is_group_chat=False, chatparticipants__user=request.user
        ).annotate(num_participants=Count("chatparticipants"))
        personal_chats = user_chats.filter(num_participants=2)
        chat = personal_chats.filter(chatparticipants__user=recipient).first()
        if not chat:
            chat = ChatConversations.objects.create(
                is_group_chat=False,
                created_at=timezone.now(),
                updated_at=timezone.now(),
            )
            ChatParticipants.objects.create(conversation=chat, user=request.user)
            ChatParticipants.objects.create(conversation=chat, user=recipient)
        chat_url = reverse("cosmochat") + f"?chat_id={chat.conversation_id}"
        return JsonResponse({"chat_url": chat_url})
    return JsonResponse({"error": "Invalid request method."}, status=405)
def get_user_rating_for_startup(user_id, startup_id):
    """
    // ... existing code ...
    """
    pass
def custom_404(request, exception):
    return render(request, "accounts/404.html", status=404)


def robots_txt(request):
    lines = [
        "User-agent: *",
        "Allow: /startups/",
        "Allow: /franchises/",
        "Allow: /agencies/",
        "Allow: /specialists/",
        "Allow: /blog/",
        "",
        "Disallow: /profile/",
        "Disallow: /admin/",
        "Disallow: /edit-",
        "Disallow: /create-",
        "Disallow: /api/",
        "Disallow: /vote-",
        "Disallow: /silk/",
        "",
        "Sitemap: https://www.greatideas.ru/sitemap.xml",
    ]
    return HttpResponse("\n".join(lines), content_type="text/plain")


def _handle_moderation_callback(bot_token, callback_data, chat_id, message_id, original_text, callback_query):
    """Обрабатывает нажатие кнопок Одобрить/Отклонить из Telegram уведомления."""
    from accounts.moderation import approve_entity, reject_entity

    parts = callback_data.split("_")
    # mod_approve_startup_123 or mod_reject_franchise_456
    if len(parts) < 4:
        return

    action = parts[1]  # approve / reject
    entity_type = parts[2]  # startup / franchise / agency / specialist
    try:
        entity_id = int(parts[3])
    except (ValueError, IndexError):
        return

    model_map = {
        'startup': (Startups, 'startup_id'),
        'franchise': (Franchises, 'franchise_id'),
        'agency': (Agencies, 'agency_id'),
        'specialist': (Specialists, 'specialist_id'),
    }

    if entity_type not in model_map:
        return

    Model, pk_field = model_map[entity_type]
    entity = Model.objects.filter(**{pk_field: entity_id}).first()
    if not entity:
        _tg_edit_message(bot_token, chat_id, message_id, escape(original_text) + "\n\n⚠️ Объект не найден в базе данных.")
        return

    if entity.status == "approved" and action == "approve":
        _tg_edit_message(bot_token, chat_id, message_id, escape(original_text) + "\n\nℹ️ Уже одобрено ранее.")
        return
    if entity.status == "rejected" and action == "reject":
        _tg_edit_message(bot_token, chat_id, message_id, escape(original_text) + "\n\nℹ️ Уже отклонено ранее.")
        return

    # Найти модератора по telegram_id отправителя callback
    from_user_id = str(callback_query.get("from", {}).get("id", ""))
    moderator = Users.objects.filter(telegram_id=from_user_id).first()
    if not moderator:
        # Fallback: любой модератор
        moderator = Users.objects.filter(role__role_name="moderator").first()
    if not moderator:
        _tg_edit_message(bot_token, chat_id, message_id, escape(original_text) + "\n\n⚠️ Модератор не найден.")
        return

    entity_names = {'startup': 'Стартап', 'franchise': 'Франшиза', 'agency': 'Агентство', 'specialist': 'Специалист'}
    entity_name_ru = entity_names.get(entity_type, 'Заявка')

    if action == "approve":
        approve_entity(entity, moderator, entity_type, moderator_comment="Одобрено через Telegram")
        status_text = "✅ ОДОБРЕНО"
    else:
        reject_entity(entity, moderator, entity_type, moderator_comment="Отклонено через Telegram")
        status_text = "❌ ОТКЛОНЕНО"

    mod_name = f"{moderator.first_name or ''} {moderator.last_name or ''}".strip() or moderator.email
    safe_mod_name = escape(mod_name)
    now_str = timezone.now().strftime("%d.%m.%Y %H:%M")

    updated_text = escape(original_text) + f"\n\n<b>{status_text}</b>\n👤 Модератор: {safe_mod_name}\n⏰ {now_str}"

    # Убираем кнопки действий, оставляем только ссылки
    entity_url_paths = {'startup': 'startups', 'franchise': 'franchises', 'agency': 'agencies', 'specialist': 'specialists'}
    url_path = entity_url_paths.get(entity_type, 'startups')
    view_url = f"https://greatideas.ru/{url_path}/{entity_id}/"

    new_keyboard = {
        "inline_keyboard": [[
            {"text": f"{'✅' if action == 'approve' else '❌'} {status_text}", "callback_data": "noop"},
            {"text": "👁 Посмотреть", "url": view_url},
        ]]
    }

    _tg_edit_message(bot_token, chat_id, message_id, updated_text, new_keyboard)
    logger.info(f"Entity {entity_type} ID={entity_id} {action}d via Telegram by {moderator.email}")


def _tg_edit_message(bot_token, chat_id, message_id, text, reply_markup=None):
    """Редактирует сообщение в Telegram."""
    payload = {
        "chat_id": chat_id,
        "message_id": message_id,
        "text": text,
        "parse_mode": "HTML",
    }
    if reply_markup:
        payload["reply_markup"] = reply_markup
    try:
        requests.post(f"https://api.telegram.org/bot{bot_token}/editMessageText", json=payload, timeout=10)
    except Exception as e:
        logger.error(f"Failed to edit Telegram message: {e}")


@csrf_exempt
@require_POST
def telegram_webhook(request, token):
    from django.conf import settings
    bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
    if not bot_token:
        logger.error("TELEGRAM_BOT_TOKEN is not configured")
        return HttpResponse(status=500)
    if token != bot_token:
        logger.warning("Invalid token in webhook URL.")
        return HttpResponseForbidden("Invalid token")
    try:
        data = json.loads(request.body)
        logger.info(f"Webhook received data: {data}")
        if "callback_query" not in data:
            return HttpResponse(status=200)
        callback_query = data["callback_query"]
        callback_data = callback_query["data"]
        message = callback_query["message"]
        chat_id = message["chat"]["id"]
        message_id = message["message_id"]
        requests.post(
            f"https://api.telegram.org/bot{bot_token}/answerCallbackQuery",
            json={"callback_query_id": callback_query["id"]},
        )
        new_text = message.get("text", "")
        new_keyboard = None
        ticket = None
        if callback_data == "noop":
            return HttpResponse(status=200)
        elif callback_data.startswith("mod_approve_") or callback_data.startswith("mod_reject_"):
            _handle_moderation_callback(bot_token, callback_data, chat_id, message_id, new_text, callback_query)
            return HttpResponse(status=200)
        elif callback_data.startswith("close_ticket_"):
            ticket_id = int(callback_data.split("_")[2])
            ticket = SupportTicket.objects.filter(pk=ticket_id).first()
            if ticket:
                ticket.status = "closed"
                ticket.save(update_fields=["status"])
                status_line = "\n\n<b>✅ ЗАЯВКА ЗАКРЫТА</b>"
                if status_line not in new_text:
                    new_text += status_line
                new_keyboard = {
                    "inline_keyboard": [
                        [
                            {
                                "text": "↩️ Вернуть в работу",
                                "callback_data": f"reopen_ticket_{ticket.ticket_id}",
                            }
                        ]
                    ]
                }
                logger.info(f"Ticket {ticket_id} was closed via Telegram.")
        elif callback_data.startswith("reopen_ticket_"):
            ticket_id = int(callback_data.split("_")[2])
            ticket = SupportTicket.objects.filter(pk=ticket_id).first()
            if ticket:
                ticket.status = "new"
                ticket.save(update_fields=["status"])
                status_line = "\n\n<b>✅ ЗАЯВКА ЗАКРЫТА</b>"
                if new_text.endswith(status_line):
                    new_text = new_text[: -len(status_line)]
                new_keyboard = {
                    "inline_keyboard": [
                        [
                            {
                                "text": "✅ Исполнено",
                                "callback_data": f"close_ticket_{ticket.ticket_id}",
                            }
                        ]
                    ]
                }
                logger.info(f"Ticket {ticket_id} was reopened via Telegram.")
        if new_keyboard:
            payload = {
                "chat_id": chat_id,
                "message_id": message_id,
                "text": new_text,
                "parse_mode": "HTML",
                "reply_markup": new_keyboard,
            }
            requests.post(
                f"https://api.telegram.org/bot{bot_token}/editMessageText", json=payload
            )
        return HttpResponse(status=200)
    except json.JSONDecodeError:
        logger.error("Error decoding JSON from Telegram webhook.")
        return HttpResponse(status=400)
    except Exception as e:
        logger.error(f"Error processing Telegram webhook: {e}", exc_info=True)
        return HttpResponse(status=500)

@login_required
def download_startups_report(request):
    try:

        wb = Workbook()
        ws = wb.active
        ws.title = "Стартапы"

        ws.merge_cells('A1:K1')
        title_cell = ws.cell(row=1, column=1, value="Отчет по стартапам")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=1, value="Владелец").font = Font(bold=True)
        owner_name = ""
        if request.user.role and request.user.role.role_name == 'startuper':
            first_name = request.user.first_name or ""
            last_name = request.user.last_name or ""
            owner_name = f"{first_name} {last_name}".strip()
        else:
            owner_name = "Все стартапы"
        ws.cell(row=2, column=2, value=owner_name)

        headers = [
            "ID", "Название", "Статус", "Категория", "Стадия",
            "Цель финансирования", "Собрано", "Рейтинг", "Количество инвесторов", "Список инвесторов", "Дата создания"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        if request.user.role and request.user.role.role_name == 'startuper':
            startups = Startups.objects.select_related('owner', 'direction', 'stage').filter(owner=request.user)
        else:
            startups = Startups.objects.select_related('owner', 'direction', 'stage').all()

        for row, startup in enumerate(startups, 5):
            try:
                try:
                    startup_id = startup.startup_id
                except Exception:
                    startup_id = row - 1
                ws.cell(row=row, column=1, value=startup_id)

                try:
                    title = startup.title or ""
                except Exception:
                    title = ""
                ws.cell(row=row, column=2, value=title)

                try:
                    status_display = startup.get_status_display()
                except Exception:
                    status_display = startup.status or "Неизвестен"
                ws.cell(row=row, column=3, value=status_display)

                try:
                    direction_name = startup.direction.direction_name if startup.direction else "Не указана"
                except Exception:
                    direction_name = "Не указана"
                ws.cell(row=row, column=4, value=direction_name)

                try:
                    stage_name = startup.stage.stage_name if startup.stage else "Не указана"
                except Exception:
                    stage_name = "Не указана"
                ws.cell(row=row, column=5, value=stage_name)

                try:
                    funding_goal = startup.funding_goal or 0
                except Exception:
                    funding_goal = 0
                ws.cell(row=row, column=6, value=funding_goal)

                try:
                    amount_raised = startup.amount_raised or 0
                except Exception:
                    amount_raised = 0
                ws.cell(row=row, column=7, value=amount_raised)

                try:
                    avg_rating = UserVotes.objects.filter(startup=startup).aggregate(Avg('rating'))['rating__avg']
                    ws.cell(row=row, column=8, value=round(avg_rating, 2) if avg_rating else 0)
                except Exception:
                    ws.cell(row=row, column=8, value=0)

                try:
                    investors_count = startup.get_investors_count()
                except Exception:
                    investors_count = 0
                ws.cell(row=row, column=9, value=investors_count)

                try:
                    investors = (
                        InvestmentTransactions.objects
                        .filter(startup=startup)
                        .select_related('investor')
                        .values_list('investor__first_name', 'investor__last_name')
                        .distinct()
                    )
                    investors_list = [
                        f"{first or ''} {last or ''}".strip() for first, last in investors if first or last
                    ]
                    cell = ws.cell(row=row, column=10, value="\n".join(investors_list))
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                except Exception:
                    cell = ws.cell(row=row, column=10, value="")
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

                try:
                    created_date = startup.created_at.strftime("%d.%m.%Y") if startup.created_at else ""
                except Exception:
                    created_date = ""
                ws.cell(row=row, column=11, value=created_date)
            except Exception as e:
                logger.error(f"Ошибка при обработке стартапа {startup.startup_id}: {e}")
                continue


        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


        for row_num in range(5, len(startups) + 5):
            ws.row_dimensions[row_num].height = 50

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        report_date = dt.now().strftime('%Y-%m-%d')
        filename = f'Отчет по стартапам_{report_date}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response
    except Exception as e:
        logger.error(f"Ошибка при генерации отчета: {e}", exc_info=True)
        return HttpResponse("Ошибка при генерации отчета", status=500)

@login_required
def approve_franchise(request, franchise_id):
    if not is_moderator(request.user):
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
    if request.method == "POST":
        from accounts.moderation import approve_entity
        moderator_comment = request.POST.get("moderator_comment", "")
        approve_entity(franchise, request.user, "franchise", moderator_comment)
        messages.success(request, "Франшиза одобрена.")
    return redirect("moderator_dashboard")


@login_required
def reject_franchise(request, franchise_id):
    if not is_moderator(request.user):
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
    if request.method == "POST":
        from accounts.moderation import reject_entity
        moderator_comment = request.POST.get("moderator_comment", "")
        reject_entity(franchise, request.user, "franchise", moderator_comment)
        messages.success(request, "Франшиза отклонена.")
    return redirect("moderator_dashboard")


@login_required
def approve_agency(request, agency_id):
    if not is_moderator(request.user):
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    agency = Agencies.objects.filter(agency_id=agency_id).first()
    if not agency:
        messages.error(request, "Агентство не найдено.")
        return redirect("moderator_dashboard")
    if request.method == "POST":
        from accounts.moderation import approve_entity
        moderator_comment = request.POST.get("moderator_comment", "")
        approve_entity(agency, request.user, "agency", moderator_comment)
        messages.success(request, "Агентство одобрено.")
    return redirect("moderator_dashboard")


@login_required
def reject_agency(request, agency_id):
    if not is_moderator(request.user):
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    agency = Agencies.objects.filter(agency_id=agency_id).first()
    if not agency:
        messages.error(request, "Агентство не найдено.")
        return redirect("moderator_dashboard")
    if request.method == "POST":
        from accounts.moderation import reject_entity
        moderator_comment = request.POST.get("moderator_comment", "")
        reject_entity(agency, request.user, "agency", moderator_comment)
        messages.success(request, "Агентство отклонено.")
    return redirect("moderator_dashboard")


@login_required
def approve_specialist(request, specialist_id):
    if not is_moderator(request.user):
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    spec = Specialists.objects.filter(specialist_id=specialist_id).first()
    if not spec:
        messages.error(request, "Специалист не найден.")
        return redirect("moderator_dashboard")
    if request.method == "POST":
        from accounts.moderation import approve_entity
        moderator_comment = request.POST.get("moderator_comment", "")
        approve_entity(spec, request.user, "specialist", moderator_comment)
        messages.success(request, "Специалист одобрен.")
    return redirect("moderator_dashboard")


@login_required
def reject_specialist(request, specialist_id):
    if not is_moderator(request.user):
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    spec = Specialists.objects.filter(specialist_id=specialist_id).first()
    if not spec:
        messages.error(request, "Специалист не найден.")
        return redirect("moderator_dashboard")
    if request.method == "POST":
        from accounts.moderation import reject_entity
        moderator_comment = request.POST.get("moderator_comment", "")
        reject_entity(spec, request.user, "specialist", moderator_comment)
        messages.success(request, "Специалист отклонен.")
    return redirect("moderator_dashboard")


@login_required
def vote_franchise(request, franchise_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
    try:
        rating = int(request.POST.get("rating", 0))
    except (TypeError, ValueError):
        return JsonResponse({"success": False, "error": "Некорректное значение рейтинга"})
    if not 1 <= rating <= 5:
        return JsonResponse(
            {"success": False, "error": "Недопустимое значение рейтинга"}
        )
    from django.db import transaction
    try:
        with transaction.atomic():
            vote, created = FranchiseVotes.objects.get_or_create(
                user=request.user,
                franchise=franchise,
                defaults={"rating": rating, "created_at": timezone.now()},
            )
            if not created:
                return JsonResponse(
                    {"success": False, "error": "Вы уже голосовали за эту франшизу"}
                )
            Franchises.objects.filter(franchise_id=franchise_id).update(
                total_voters=models.F("total_voters") + 1,
                sum_votes=models.F("sum_votes") + rating,
            )
    except Exception:
        return JsonResponse({"success": False, "error": "Ошибка при сохранении голоса"})
    franchise.refresh_from_db(fields=["total_voters", "sum_votes"])
    average_rating = (
        franchise.sum_votes / franchise.total_voters if franchise.total_voters > 0 else 0
    )
    return JsonResponse({"success": True, "average_rating": average_rating})


@login_required
def vote_agency(request, agency_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    agency = Agencies.objects.filter(agency_id=agency_id).first()
    if not agency:
        return JsonResponse({"success": False, "error": "Агентство не найдено"})
    try:
        rating = int(request.POST.get("rating", 0))
    except (TypeError, ValueError):
        return JsonResponse({"success": False, "error": "Некорректное значение рейтинга"})
    if not 1 <= rating <= 5:
        return JsonResponse({"success": False, "error": "Недопустимое значение рейтинга"})
    from django.db import transaction
    try:
        with transaction.atomic():
            vote, created = AgencyVotes.objects.get_or_create(
                user=request.user,
                agency=agency,
                defaults={"rating": rating, "created_at": timezone.now()},
            )
            if not created:
                return JsonResponse({"success": False, "error": "Вы уже голосовали за это агентство"})
            Agencies.objects.filter(agency_id=agency_id).update(
                total_voters=models.F("total_voters") + 1,
                sum_votes=models.F("sum_votes") + rating,
            )
    except Exception:
        return JsonResponse({"success": False, "error": "Ошибка при сохранении голоса"})
    agency.refresh_from_db(fields=["total_voters", "sum_votes"])
    average_rating = agency.sum_votes / agency.total_voters if agency.total_voters > 0 else 0
    return JsonResponse({"success": True, "average_rating": average_rating})

@login_required
def vote_specialist(request, specialist_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    specialist = get_object_or_404(Specialists, specialist_id=specialist_id)
    try:
        rating = int(request.POST.get("rating", 0))
    except (TypeError, ValueError):
        return JsonResponse({"success": False, "error": "Некорректное значение рейтинга"})
    if not 1 <= rating <= 5:
        return JsonResponse({"success": False, "error": "Недопустимое значение рейтинга"})
    from django.db import transaction
    try:
        with transaction.atomic():
            vote, created = SpecialistVotes.objects.get_or_create(
                user=request.user,
                specialist=specialist,
                defaults={"rating": rating, "created_at": timezone.now()},
            )
            if not created:
                return JsonResponse({"success": False, "error": "Вы уже голосовали за этого специалиста"})
            Specialists.objects.filter(specialist_id=specialist_id).update(
                total_voters=models.F("total_voters") + 1,
                sum_votes=models.F("sum_votes") + rating,
            )
    except Exception:
        return JsonResponse({"success": False, "error": "Ошибка при сохранении голоса"})
    specialist.refresh_from_db(fields=["total_voters", "sum_votes"])
    average_rating = specialist.sum_votes / specialist.total_voters if specialist.total_voters > 0 else 0
    return JsonResponse({"success": True, "average_rating": average_rating})

def load_similar_franchises(request, franchise_id: int):
    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        similar_franchises = (
            Franchises.objects.filter(
                direction=franchise.direction,
                status="approved",
            )
            .exclude(franchise_id=franchise.franchise_id)
            .order_by("?")[:4]
        )

        context = {
            "similar_franchises": similar_franchises,
        }
        if similar_franchises.count() < 4:
            return HttpResponse("")
        return render(request, "accounts/partials/_similar_franchise_cards.html", context)
    except Exception as e:
        logger.error(f"Ошибка при загрузке похожих франшиз: {e}")
        return JsonResponse({"error": "Ошибка при загрузке похожих франшиз"}, status=500)

@login_required
def load_similar_agencies(request, agency_id: int):
    try:
        agency = Agencies.objects.filter(agency_id=agency_id).first()
        if not agency:
            return JsonResponse({"error": "Агентство не найдено"}, status=404)
        if agency.customization_data and "agency_category" in agency.customization_data:
            similar_qs = (
                Agencies.objects.filter(
                    customization_data__agency_category=agency.customization_data.get("agency_category"),
                    status="approved",
                )
                .exclude(agency_id=agency.agency_id)
                .order_by("?")[:4]
            )
        else:
            similar_qs = Agencies.objects.filter(status="approved").exclude(agency_id=agency.agency_id).order_by("?")[:4]
        html = render_to_string(
            "accounts/partials/_similar_agency_cards.html",
            {"similar_franchises": similar_qs, "request": request},
        )
        return HttpResponse(html)
    except Exception as e:
        logger.error(f"Ошибка при загрузке похожих агентств: {e}")
        return JsonResponse({"error": "Ошибка при загрузке похожих агентств"}, status=500)

@login_required
def load_similar_specialists(request, specialist_id: int):
    try:
        specialist = get_object_or_404(Specialists, specialist_id=specialist_id)
        if specialist.customization_data and "specialist_category" in specialist.customization_data:
            similar_qs = (
                Specialists.objects.filter(
                    customization_data__specialist_category=specialist.customization_data.get("specialist_category"),
                    status="approved",
                )
                .exclude(specialist_id=specialist.specialist_id)
                .order_by("?")[:4]
            )
        else:
            similar_qs = Specialists.objects.filter(status="approved").exclude(specialist_id=specialist.specialist_id).order_by("?")[:4]
        html = render_to_string(
            "accounts/partials/_similar_specialist_cards.html",
            {"similar_specialists": similar_qs, "request": request},
        )
        return HttpResponse(html)
    except Exception as e:
        logger.error(f"Ошибка при загрузке похожих специалистов: {e}")
        return JsonResponse({"error": "Ошибка при загрузке похожих специалистов"}, status=500)

@login_required
def delete_startup_comment(request, comment_id: int):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Недопустимый метод"}, status=405)
    if not hasattr(request.user, "role") or (request.user.role.role_name or "") != "moderator":
        return JsonResponse({"success": False, "error": "Нет прав"}, status=403)
    comment = get_object_or_404(Comments, pk=comment_id)
    comment.delete()
    return JsonResponse({"success": True})

@login_required
def delete_franchise_comment(request, comment_id: int):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Недопустимый метод"}, status=405)
    if not hasattr(request.user, "role") or (request.user.role.role_name or "") != "moderator":
        return JsonResponse({"success": False, "error": "Нет прав"}, status=403)
    from .models import FranchiseComments
    comment = get_object_or_404(FranchiseComments, pk=comment_id)
    comment.delete()
    return JsonResponse({"success": True})

@login_required
def delete_agency_comment(request, comment_id: int):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Недопустимый метод"}, status=405)
    if not hasattr(request.user, "role") or (request.user.role.role_name or "") != "moderator":
        return JsonResponse({"success": False, "error": "Нет прав"}, status=403)
    comment = get_object_or_404(AgencyComments, pk=comment_id)
    comment.delete()
    return JsonResponse({"success": True})

@login_required
def delete_specialist_comment(request, comment_id: int):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Недопустимый метод"}, status=405)
    if not hasattr(request.user, "role") or (request.user.role.role_name or "") != "moderator":
        return JsonResponse({"success": False, "error": "Нет прав"}, status=403)
    comment = get_object_or_404(SpecialistComments, pk=comment_id)
    comment.delete()
    return JsonResponse({"success": True})


@login_required
def edit_franchise(request, franchise_id):
    franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
    if request.user != franchise.owner and request.user.role.role_name != 'moderator':
        messages.error(request, "У вас нет прав для редактирования этой франшизы.")
        return redirect("franchise_detail", slug=franchise.slug or franchise_id)

    if request.method == "POST":
        # Сохраняем оригинальные данные для сравнения
        original_data = {
            'title': franchise.title,
            'short_description': franchise.short_description,
            'description': franchise.description,
            'terms': franchise.terms,
            'investment_size': franchise.investment_size,
            'franchise_cost': franchise.franchise_cost,
            'profit_calculation': franchise.profit_calculation,
            'pitch_deck_url': franchise.pitch_deck_url,
            'planet_image': franchise.planet_image,
        }
        
        form = FranchiseEditForm(request.POST, request.FILES, instance=franchise)
        
        if form.is_valid():
            franchise = form.save(commit=False)
            
            # Проверяем были ли изменения
            has_changes = False
            
            # Проверяем основные поля
            if (franchise.title != original_data['title'] or 
                franchise.short_description != original_data['short_description'] or
                franchise.description != original_data['description'] or
                franchise.terms != original_data['terms'] or
                franchise.investment_size != original_data['investment_size'] or
                franchise.franchise_cost != original_data['franchise_cost'] or
                franchise.profit_calculation != original_data['profit_calculation'] or
                franchise.pitch_deck_url != original_data['pitch_deck_url'] or
                franchise.planet_image != original_data['planet_image']):
                has_changes = True
            
            # Проверяем загружены ли новые файлы
            if not has_changes:
                if request.FILES:
                    has_changes = True
            
            # Проверяем удалены ли файлы
            if not has_changes:
                deleted_files_json = request.POST.get('deleted_files', '[]')
                try:
                    deleted_files = json.loads(deleted_files_json)
                    if deleted_files:
                        has_changes = True
                except json.JSONDecodeError:
                    pass
            
            # Принудительная проверка файлов
            if request.FILES:
                has_changes = True
            
            # Устанавливаем статус в зависимости от наличия изменений
            if has_changes and franchise.status == "approved":
                franchise.status = "pending"
                franchise.is_edited = True
                franchise.save(update_fields=['status', 'is_edited'])
            
            franchise.updated_at = timezone.now()
            franchise.save()
            
            logo_ids = franchise.logo_urls or []
            creative_ids = franchise.creatives_urls or []  # Существующие креативы
            proofs_ids = franchise.proofs_urls or []
            video_ids = franchise.video_urls or []
            
            # Обработка логотипа — синхронная загрузка в S3
            logo = request.FILES.get("logo")
            if logo and logo.size > 0:
                logo_id = str(uuid.uuid4())
                try:
                    logo.seek(0)
                    file_data = logo.read()
                    content_type = getattr(logo, 'content_type', 'image/jpeg')
                    unique_filename = get_unique_filename(logo.name, franchise.franchise_id, "logo")

                    result = upload_file_to_s3_sync(
                        file_data=file_data,
                        file_name=logo.name,
                        content_type=content_type,
                        entity_type_name='franchise',
                        entity_id=franchise.franchise_id,
                        file_type_name='logo',
                        original_filename=unique_filename,
                        file_id=logo_id
                    )
                    if result:
                        logo_ids = [logo_id]
                        logger.info(f"Логотип франшизы загружен синхронно: {logo.name}, размер: {len(file_data)} байт")
                    else:
                        logger.error(f"Синхронная загрузка логотипа франшизы не удалась: {logo.name}")
                        messages.warning(request, "Не удалось сохранить логотип")
                except Exception as e:
                    logger.error(f"Ошибка загрузки логотипа франшизы: {e}", exc_info=True)
                    messages.warning(request, f"Не удалось сохранить логотип: {e}")

            # Обработка креативов
            creatives = request.FILES.getlist("creatives")
            proofs = request.FILES.getlist("proofs")
            videos = request.FILES.getlist("video")

            # Проверка лимитов файлов
            if len(creatives) > 10:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 10 изображений"}, status=400)
                messages.error(request, "Максимально 10 изображений")
                return render(request, "accounts/edit_franchise.html", {"form": form, "franchise": franchise})

            if len(proofs) > 15:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 15 документов"}, status=400)
                messages.error(request, "Максимально 15 документов")
                return render(request, "accounts/edit_franchise.html", {"form": form, "franchise": franchise})

            if len(videos) > 1:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 1 видео"}, status=400)
                messages.error(request, "Максимально 1 видео")
                return render(request, "accounts/edit_franchise.html", {"form": form, "franchise": franchise})

            if creatives:
                for creative_file in creatives:
                    if not hasattr(creative_file, "name"):
                        continue
                    try:
                        creative_id = str(uuid.uuid4())
                        file_data = creative_file.read()
                        content_type = getattr(creative_file, 'content_type', 'image/jpeg')
                        unique_filename = get_unique_filename(creative_file.name, franchise.franchise_id, "creative")

                        result = upload_file_to_s3_sync(
                            file_data=file_data,
                            file_name=creative_file.name,
                            content_type=content_type,
                            entity_type_name='franchise',
                            entity_id=franchise.franchise_id,
                            file_type_name='creative',
                            original_filename=unique_filename,
                            file_id=creative_id
                        )
                        if result:
                            creative_ids.append(creative_id)
                        logger.info(f"Креатив франшизы загружен: {creative_file.name}, размер: {len(file_data)} байт")
                    except Exception as e:
                        logger.error(f"Ошибка отправки креатива франшизы в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось сохранить креатив: {e}")

            if proofs:
                for proof_file in proofs:
                    if not hasattr(proof_file, "name"):
                        continue
                    try:
                        proof_id = str(uuid.uuid4())
                        file_data = proof_file.read()
                        content_type = getattr(proof_file, 'content_type', 'application/pdf')
                        unique_filename = get_unique_filename(proof_file.name, franchise.franchise_id, "proof")

                        result = upload_file_to_s3_sync(
                            file_data=file_data,
                            file_name=proof_file.name,
                            content_type=content_type,
                            entity_type_name='franchise',
                            entity_id=franchise.franchise_id,
                            file_type_name='proof',
                            original_filename=unique_filename,
                            file_id=proof_id
                        )
                        if result:
                            proofs_ids.append(proof_id)
                        logger.info(f"Пруф франшизы загружен: {proof_file.name}, размер: {len(file_data)} байт")
                    except Exception as e:
                        logger.error(f"Ошибка отправки пруфа франшизы в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось сохранить документ: {e}")
            
            if videos:
                for video in videos:
                    if not hasattr(video, "name"):
                        logger.warning(f"Пропущено видео франшизы: {video}")
                        continue
                    try:
                        unique_filename = get_unique_filename(video.name, franchise.franchise_id, "video")
                        video_data = video.read()
                        content_type = getattr(video, 'content_type', 'video/mp4')
                        
                        video_data_b64 = base64.b64encode(video_data).decode('utf-8')
                        
                        upload_video_to_s3.delay(
                            video_data=video_data_b64,
                            video_name=video.name,
                            video_content_type=content_type,
                            entity_id=franchise.franchise_id,
                            original_filename=unique_filename,
                            entity_type_name='franchise'
                        )
                        logger.info(f"Видео франшизы отправлено в очередь Celery: {video.name}, размер: {len(video_data)} байт")
                        messages.info(request, f"Видео {video.name} загружается в фоновом режиме.")
                    except Exception as e:
                        logger.error(f"Ошибка отправки видео франшизы в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось отправить видео {video.name} на загрузку.")
                logger.info("Видео франшизы загружается асинхронно через Celery")
                if franchise.video_urls is None:
                    franchise.video_urls = []
            
            # Обновляем URL файлов
            franchise.logo_urls = logo_ids
            franchise.creatives_urls = creative_ids
            franchise.proofs_urls = proofs_ids
            
            # Обработка удаленных файлов
            deleted_files_json = request.POST.get('deleted_files', '[]')
            try:
                deleted_files = json.loads(deleted_files_json)
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="franchise")
                for deleted_file in deleted_files:
                    file_id = deleted_file.get('id')
                    file_type = deleted_file.get('type')
                    if file_id and file_type:
                        # Получаем информацию о файле перед удалением
                        file_storage = FileStorage.objects.filter(
                            entity_type=entity_type,
                            entity_id=franchise.franchise_id,
                            file_url=file_id
                        ).first()
                        
                        if file_storage:
                            # Удаляем файл из S3
                            if file_type == 'creative':
                                file_path = f"franchises/{franchise.franchise_id}/creatives/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            elif file_type == 'proof':
                                file_path = f"franchises/{franchise.franchise_id}/proofs/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            elif file_type == 'video':
                                file_path = f"franchises/{franchise.franchise_id}/videos/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            else:
                                file_path = f"franchises/{franchise.franchise_id}/{file_type}s/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            
                            delete_file_from_s3(file_path)
                        
                        # Удаляем запись из базы данных
                        FileStorage.objects.filter(
                            entity_type=entity_type,
                            entity_id=franchise.franchise_id,
                            file_url=file_id
                        ).delete()
                        
                        # Удаляем из списка URL
                        if file_type == 'creative' and franchise.creatives_urls:
                            franchise.creatives_urls = [url for url in franchise.creatives_urls if url != file_id]
                        elif file_type == 'proof' and franchise.proofs_urls:
                            franchise.proofs_urls = [url for url in franchise.proofs_urls if url != file_id]
                        elif file_type == 'video' and franchise.video_urls:
                            franchise.video_urls = [url for url in franchise.video_urls if url != file_id]
                        logger.info(f"Удален файл {file_type}: {file_id}")
            except json.JSONDecodeError:
                pass
            
            slider_images = request.POST.getlist("slider_images")
            if len(slider_images) > 4:
                slider_images = slider_images[:4]
            franchise.slider_images = slider_images
            
            # Обработка catalog_card_image
            catalog_card_image = form.cleaned_data.get("catalog_card_image") or request.FILES.get("catalog_card_image")
            if catalog_card_image and hasattr(catalog_card_image, 'read'):
                catalog_card_id = str(uuid.uuid4())
                # Конвертируем в WebP
                processed_catalog_image, processed_catalog_name, _ = process_uploaded_image(catalog_card_image, quality=85)
                base_name = os.path.splitext(processed_catalog_name)[0]
                ext = os.path.splitext(processed_catalog_name)[1]
                safe_base_name = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
                safe_name = slugify(safe_base_name) + ext
                file_path = f"catalog_cards/{catalog_card_id}_{safe_name}"
                try:
                    s3 = boto3.client(
                        's3',
                        endpoint_url=getattr(settings, 'AWS_S3_ENDPOINT_URL', None),
                        aws_access_key_id=getattr(settings, 'AWS_ACCESS_KEY_ID', None),
                        aws_secret_access_key=getattr(settings, 'AWS_SECRET_ACCESS_KEY', None),
                        region_name=getattr(settings, 'AWS_S3_REGION_NAME', None),
                        config=boto3.session.Config(s3={'addressing_style': getattr(settings, 'AWS_S3_ADDRESSING_STYLE', 'virtual')})
                    )
                    bucket = getattr(settings, 'AWS_STORAGE_BUCKET_NAME', None)
                    content_type = 'image/webp' if processed_catalog_name.endswith('.webp') else getattr(catalog_card_image, 'content_type', 'application/octet-stream')
                    if hasattr(processed_catalog_image, 'read'):
                        body_bytes = processed_catalog_image.read()
                    else:
                        body_bytes = processed_catalog_image.getvalue() if hasattr(processed_catalog_image, 'getvalue') else processed_catalog_image
                    try:
                        s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ContentType=content_type, ACL='public-read')
                    except Exception:
                        s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ACL='public-read')
                    franchise.catalog_card_image = f"{catalog_card_id}_{safe_name}"
                except Exception as e:
                    logger.error(f"Ошибка сохранения изображения карточки: {e}", exc_info=True)
                    messages.warning(request, "Не удалось сохранить изображение для карточки.")
            
            franchise.save()
            
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("franchise_detail", kwargs={"slug": franchise.slug or franchise.franchise_id})
                })
            messages.success(request, "Франшиза успешно обновлена.")
            return redirect("franchise_detail", slug=franchise.slug or franchise.franchise_id)
        else:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": False,
                    "errors": {f: [x["message"] for x in e.get_json_data()] for f, e in form.errors.items()},
                    "non_field_errors": [str(e) for e in form.non_field_errors()],
                }, status=400)
            messages.error(request, "Форма содержит ошибки.")
    else:
        form = FranchiseEditForm(instance=franchise)
    
    context = {
        'form': form,
        'franchise': franchise,
    }
    return render(request, 'accounts/edit_franchise.html', context)


@login_required
def edit_agency(request, agency_id):
    agency = Agencies.objects.filter(agency_id=agency_id).first()
    if not agency:
        messages.error(request, "Агентство не найдено.")
        return redirect("agencies_list")
    if request.user != agency.owner and request.user.role.role_name != 'moderator':
        messages.error(request, "У вас нет прав для редактирования этого агентства.")
        return redirect("agency_detail", slug=agency.slug or agency_id)

    if request.method == "POST":
        # Сохраняем оригинальные данные для сравнения
        original_data = {
            'title': agency.title,
            'short_description': agency.short_description,
            'description': agency.description,
            'terms': agency.terms,
            'pitch_deck_url': agency.pitch_deck_url,
            'planet_image': agency.planet_image,
        }
        
        form = AgencyEditForm(request.POST, request.FILES, instance=agency)
        
        if form.is_valid():
            agency = form.save(commit=False)
            
            # Проверяем были ли изменения
            has_changes = False
            
            # Проверяем основные поля
            if (agency.title != original_data['title'] or 
                agency.short_description != original_data['short_description'] or
                agency.description != original_data['description'] or
                agency.terms != original_data['terms'] or
                agency.pitch_deck_url != original_data['pitch_deck_url'] or
                agency.planet_image != original_data['planet_image']):
                has_changes = True
            
            # Обрабатываем successful_projects и другие поля customization_data
            successful_projects = form.cleaned_data.get('successful_projects')
            if agency.customization_data is None:
                agency.customization_data = {}
            if successful_projects is not None:
                agency.customization_data['successful_projects'] = successful_projects
            
            # Сохраняем новые поля customization_data
            agency.customization_data['agency_category'] = form.cleaned_data.get('agency_category', '')
            agency.customization_data['agency_services'] = form.cleaned_data.get('agency_services', '')
            
            # Проверяем загружены ли новые файлы
            if not has_changes:
                if request.FILES:
                    has_changes = True
            
            # Проверяем удалены ли файлы
            if not has_changes:
                deleted_files_json = request.POST.get('deleted_files', '[]')
                try:
                    deleted_files = json.loads(deleted_files_json)
                    if deleted_files:
                        has_changes = True
                except json.JSONDecodeError:
                    pass
            
            # Принудительная проверка файлов
            if request.FILES:
                has_changes = True
            
            # Устанавливаем статус в зависимости от наличия изменений
            if has_changes and agency.status == "approved":
                agency.status = "pending"
                agency.save(update_fields=['status'])
            
            agency.updated_at = timezone.now()
            agency.save()
            
            logo_ids = agency.logo_urls or []
            creative_ids = agency.creatives_urls or []  # Существующие креативы
            proofs_ids = agency.proofs_urls or []
            video_ids = agency.video_urls or []
            
            # Обработка логотипа — синхронная загрузка в S3
            logo = request.FILES.get("logo")
            if logo and logo.size > 0:
                logo_id = str(uuid.uuid4())
                try:
                    logo.seek(0)
                    file_data = logo.read()
                    content_type = getattr(logo, 'content_type', 'image/jpeg')
                    unique_filename = get_unique_filename(logo.name, agency.agency_id, "logo")

                    result = upload_file_to_s3_sync(
                        file_data=file_data,
                        file_name=logo.name,
                        content_type=content_type,
                        entity_type_name='agency',
                        entity_id=agency.agency_id,
                        file_type_name='logo',
                        original_filename=unique_filename,
                        file_id=logo_id
                    )
                    if result:
                        logo_ids = [logo_id]
                        logger.info(f"Логотип агентства загружен синхронно: {logo.name}, размер: {len(file_data)} байт")
                    else:
                        logger.error(f"Синхронная загрузка логотипа агентства не удалась: {logo.name}")
                        messages.warning(request, "Не удалось сохранить логотип")
                except Exception as e:
                    logger.error(f"Ошибка загрузки логотипа агентства: {e}", exc_info=True)
                    messages.warning(request, f"Не удалось сохранить логотип: {e}")

            # Обработка креативов
            creatives = request.FILES.getlist("creatives")
            proofs = request.FILES.getlist("proofs")
            videos = request.FILES.getlist("video")

            # Проверка лимитов файлов
            if len(creatives) > 10:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 10 изображений"}, status=400)
                messages.error(request, "Максимально 10 изображений")
                return render(request, "accounts/edit_agency.html", {"form": form, "agency": agency})

            if len(proofs) > 15:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 15 документов"}, status=400)
                messages.error(request, "Максимально 15 документов")
                return render(request, "accounts/edit_agency.html", {"form": form, "agency": agency})

            if len(videos) > 1:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 1 видео"}, status=400)
                messages.error(request, "Максимально 1 видео")
                return render(request, "accounts/edit_agency.html", {"form": form, "agency": agency})

            if creatives:
                for creative_file in creatives:
                    if not hasattr(creative_file, "name"):
                        continue
                    try:
                        creative_id = str(uuid.uuid4())
                        file_data = creative_file.read()
                        content_type = getattr(creative_file, 'content_type', 'image/jpeg')
                        unique_filename = get_unique_filename(creative_file.name, agency.agency_id, "creative")

                        result = upload_file_to_s3_sync(
                            file_data=file_data,
                            file_name=creative_file.name,
                            content_type=content_type,
                            entity_type_name='agency',
                            entity_id=agency.agency_id,
                            file_type_name='creative',
                            original_filename=unique_filename,
                            file_id=creative_id
                        )
                        if result:
                            creative_ids.append(creative_id)
                        logger.info(f"Креатив агентства загружен: {creative_file.name}, размер: {len(file_data)} байт")
                    except Exception as e:
                        logger.error(f"Ошибка отправки креатива агентства в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось сохранить креатив: {e}")

            if proofs:
                for proof_file in proofs:
                    if not hasattr(proof_file, "name"):
                        continue
                    try:
                        proof_id = str(uuid.uuid4())
                        file_data = proof_file.read()
                        content_type = getattr(proof_file, 'content_type', 'application/pdf')
                        unique_filename = get_unique_filename(proof_file.name, agency.agency_id, "proof")

                        result = upload_file_to_s3_sync(
                            file_data=file_data,
                            file_name=proof_file.name,
                            content_type=content_type,
                            entity_type_name='agency',
                            entity_id=agency.agency_id,
                            file_type_name='proof',
                            original_filename=unique_filename,
                            file_id=proof_id
                        )
                        if result:
                            proofs_ids.append(proof_id)
                        logger.info(f"Пруф агентства загружен: {proof_file.name}, размер: {len(file_data)} байт")
                    except Exception as e:
                        logger.error(f"Ошибка отправки пруфа агентства в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось сохранить документ: {e}")
            
            if videos:
                for video in videos:
                    if not hasattr(video, "name"):
                        logger.warning(f"Пропущено видео агентства: {video}")
                        continue
                    try:
                        unique_filename = get_unique_filename(video.name, agency.agency_id, "video")
                        video_data = video.read()
                        content_type = getattr(video, 'content_type', 'video/mp4')
                        
                        video_data_b64 = base64.b64encode(video_data).decode('utf-8')
                        
                        upload_video_to_s3.delay(
                            video_data=video_data_b64,
                            video_name=video.name,
                            video_content_type=content_type,
                            entity_id=agency.agency_id,
                            original_filename=unique_filename,
                            entity_type_name='agency'
                        )
                        logger.info(f"Видео агентства отправлено в очередь Celery: {video.name}, размер: {len(video_data)} байт")
                        messages.info(request, f"Видео {video.name} загружается в фоновом режиме.")
                    except Exception as e:
                        logger.error(f"Ошибка отправки видео агентства в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось отправить видео {video.name} на загрузку.")
                logger.info("Видео агентства загружается асинхронно через Celery")
                if agency.video_urls is None:
                    agency.video_urls = []
            
            # Обработка удаленных файлов
            deleted_files_json = request.POST.get('deleted_files', '[]')
            try:
                deleted_files = json.loads(deleted_files_json)
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="agency")
                for deleted_file in deleted_files:
                    file_id = deleted_file.get('id')
                    file_type = deleted_file.get('type')
                    if file_id and file_type:
                        # Получаем информацию о файле перед удалением
                        file_storage = FileStorage.objects.filter(
                            entity_type=entity_type,
                            entity_id=agency.agency_id,
                            file_url=file_id
                        ).first()
                        
                        if file_storage:
                            # Удаляем файл из S3
                            if file_type == 'creative':
                                file_path = f"agencies/{agency.agency_id}/creatives/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            elif file_type == 'proof':
                                file_path = f"agencies/{agency.agency_id}/proofs/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            elif file_type == 'video':
                                file_path = f"agencies/{agency.agency_id}/videos/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            else:
                                file_path = f"agencies/{agency.agency_id}/{file_type}s/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            
                            delete_file_from_s3(file_path)
                        
                        # Удаляем запись из базы данных
                        FileStorage.objects.filter(
                            entity_type=entity_type,
                            entity_id=agency.agency_id,
                            file_url=file_id
                        ).delete()
                        
                        # Удаляем из списка URL
                        if file_type == 'creative' and creative_ids:
                            creative_ids = [url for url in creative_ids if url != file_id]
                        elif file_type == 'proof' and proofs_ids:
                            proofs_ids = [url for url in proofs_ids if url != file_id]
                        elif file_type == 'video' and video_ids:
                            video_ids = [url for url in video_ids if url != file_id]
                        logger.info(f"Удален файл {file_type}: {file_id}")
            except json.JSONDecodeError:
                pass
            
            # Обновляем URL файлов
            agency.logo_urls = logo_ids
            agency.creatives_urls = creative_ids
            agency.proofs_urls = proofs_ids
            
            slider_images = request.POST.getlist("slider_images")
            if len(slider_images) > 4:
                slider_images = slider_images[:4]
            agency.slider_images = slider_images
            
            # Обработка catalog_card_image
            catalog_card_image = form.cleaned_data.get("catalog_card_image") or request.FILES.get("catalog_card_image")
            if catalog_card_image and hasattr(catalog_card_image, 'read'):
                catalog_card_id = str(uuid.uuid4())
                # Конвертируем в WebP
                processed_catalog_image, processed_catalog_name, _ = process_uploaded_image(catalog_card_image, quality=85)
                base_name = os.path.splitext(processed_catalog_name)[0]
                ext = os.path.splitext(processed_catalog_name)[1]
                safe_base_name = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
                safe_name = slugify(safe_base_name) + ext
                file_path = f"catalog_cards/{catalog_card_id}_{safe_name}"
                try:
                    s3 = boto3.client(
                        's3',
                        endpoint_url=getattr(settings, 'AWS_S3_ENDPOINT_URL', None),
                        aws_access_key_id=getattr(settings, 'AWS_ACCESS_KEY_ID', None),
                        aws_secret_access_key=getattr(settings, 'AWS_SECRET_ACCESS_KEY', None),
                        region_name=getattr(settings, 'AWS_S3_REGION_NAME', None),
                        config=boto3.session.Config(s3={'addressing_style': getattr(settings, 'AWS_S3_ADDRESSING_STYLE', 'virtual')})
                    )
                    bucket = getattr(settings, 'AWS_STORAGE_BUCKET_NAME', None)
                    content_type = 'image/webp' if processed_catalog_name.endswith('.webp') else getattr(catalog_card_image, 'content_type', 'application/octet-stream')
                    if hasattr(processed_catalog_image, 'read'):
                        body_bytes = processed_catalog_image.read()
                    else:
                        body_bytes = processed_catalog_image.getvalue() if hasattr(processed_catalog_image, 'getvalue') else processed_catalog_image
                    try:
                        s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ContentType=content_type, ACL='public-read')
                    except Exception:
                        s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ACL='public-read')
                    agency.catalog_card_image = f"{catalog_card_id}_{safe_name}"
                except Exception as e:
                    logger.error(f"Ошибка сохранения изображения карточки: {e}", exc_info=True)
                    messages.warning(request, "Не удалось сохранить изображение для карточки.")
            
            agency.save()
            
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("agency_detail", kwargs={"slug": agency.slug or agency.agency_id})
                })
            messages.success(request, "Агентство успешно обновлено.")
            return redirect("agency_detail", slug=agency.slug or agency.agency_id)
        else:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": False,
                    "errors": {f: [x["message"] for x in e.get_json_data()] for f, e in form.errors.items()},
                    "non_field_errors": [str(e) for e in form.non_field_errors()],
                }, status=400)
            messages.error(request, "Форма содержит ошибки.")
    else:
        form = AgencyEditForm(instance=agency)
    
    context = {
        'form': form,
        'agency': agency,
    }
    return render(request, 'accounts/edit_agency.html', context)


@login_required
def edit_specialist(request, specialist_id):
    specialist = get_object_or_404(Specialists, specialist_id=specialist_id)
    if request.user != specialist.owner and request.user.role.role_name != 'moderator':
        messages.error(request, "У вас нет прав для редактирования этого специалиста.")
        return redirect("specialist_detail", specialist_id=specialist_id)

    if request.method == "POST":
        # Сохраняем оригинальные данные для сравнения
        original_data = {
            'title': specialist.title,
            'short_description': specialist.short_description,
            'description': specialist.description,
            'terms': specialist.terms,
            'additional_info': specialist.additional_info,
            'pitch_deck_url': specialist.pitch_deck_url,
            'planet_image': specialist.planet_image,
        }
        
        form = SpecialistEditForm(request.POST, request.FILES, instance=specialist)
        
        if form.is_valid():
            specialist = form.save(commit=False)
            
            # Проверяем были ли изменения
            has_changes = False
            
            # Проверяем основные поля
            if (specialist.title != original_data['title'] or 
                specialist.short_description != original_data['short_description'] or
                specialist.description != original_data['description'] or
                specialist.terms != original_data['terms'] or
                specialist.additional_info != original_data['additional_info'] or
                specialist.pitch_deck_url != original_data['pitch_deck_url'] or
                specialist.planet_image != original_data['planet_image']):
                has_changes = True
            
            # Обрабатываем successful_projects и другие поля customization_data
            successful_projects = form.cleaned_data.get('successful_projects', 12)
            if specialist.customization_data is None:
                specialist.customization_data = {}
            specialist.customization_data['successful_projects'] = successful_projects
            
            # Сохраняем новые поля customization_data
            specialist.customization_data['specialist_category'] = form.cleaned_data.get('specialist_category', '')
            
            # Проверяем загружены ли новые файлы
            if not has_changes:
                if request.FILES:
                    has_changes = True
            
            # Проверяем удалены ли файлы
            if not has_changes:
                deleted_files_json = request.POST.get('deleted_files', '[]')
                try:
                    deleted_files = json.loads(deleted_files_json)
                    if deleted_files:
                        has_changes = True
                except json.JSONDecodeError:
                    pass
            
            # Принудительная проверка файлов
            if request.FILES:
                has_changes = True
            
            # Устанавливаем статус в зависимости от наличия изменений
            if has_changes and specialist.status == "approved":
                specialist.status = "pending"
                specialist.save(update_fields=['status'])
            
            specialist.updated_at = timezone.now()
            specialist.save()
            
            logo_ids = specialist.logo_urls or []
            creative_ids = specialist.creatives_urls or []  # Существующие креативы
            proofs_ids = specialist.proofs_urls or []
            video_ids = specialist.video_urls or []
            
            # Обработка логотипа — синхронная загрузка в S3
            logo = request.FILES.get("logo")
            if logo and logo.size > 0:
                logo_id = str(uuid.uuid4())
                try:
                    logo.seek(0)
                    file_data = logo.read()
                    content_type = getattr(logo, 'content_type', 'image/jpeg')
                    unique_filename = get_unique_filename(logo.name, specialist.specialist_id, "logo")

                    result = upload_file_to_s3_sync(
                        file_data=file_data,
                        file_name=logo.name,
                        content_type=content_type,
                        entity_type_name='specialist',
                        entity_id=specialist.specialist_id,
                        file_type_name='logo',
                        original_filename=unique_filename,
                        file_id=logo_id
                    )
                    if result:
                        logo_ids = [logo_id]
                        logger.info(f"Логотип специалиста загружен синхронно: {logo.name}, размер: {len(file_data)} байт")
                    else:
                        logger.error(f"Синхронная загрузка логотипа специалиста не удалась: {logo.name}")
                        messages.warning(request, "Не удалось сохранить логотип")
                except Exception as e:
                    logger.error(f"Ошибка загрузки логотипа специалиста: {e}", exc_info=True)
                    messages.warning(request, f"Не удалось сохранить логотип: {e}")

            # Обработка креативов
            creatives = request.FILES.getlist("creatives")
            proofs = request.FILES.getlist("proofs")
            videos = request.FILES.getlist("video")

            # Проверка лимитов файлов
            if len(creatives) > 10:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 10 изображений"}, status=400)
                messages.error(request, "Максимально 10 изображений")
                return render(request, "accounts/edit_specialist.html", {"form": form, "specialist": specialist})

            if len(proofs) > 15:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 15 документов"}, status=400)
                messages.error(request, "Максимально 15 документов")
                return render(request, "accounts/edit_specialist.html", {"form": form, "specialist": specialist})

            if len(videos) > 1:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Максимально 1 видео"}, status=400)
                messages.error(request, "Максимально 1 видео")
                return render(request, "accounts/edit_specialist.html", {"form": form, "specialist": specialist})

            if creatives:
                for creative_file in creatives:
                    if not hasattr(creative_file, "name"):
                        continue
                    try:
                        creative_id = str(uuid.uuid4())
                        file_data = creative_file.read()
                        content_type = getattr(creative_file, 'content_type', 'image/jpeg')
                        unique_filename = get_unique_filename(creative_file.name, specialist.specialist_id, "creative")

                        result = upload_file_to_s3_sync(
                            file_data=file_data,
                            file_name=creative_file.name,
                            content_type=content_type,
                            entity_type_name='specialist',
                            entity_id=specialist.specialist_id,
                            file_type_name='creative',
                            original_filename=unique_filename,
                            file_id=creative_id
                        )
                        if result:
                            creative_ids.append(creative_id)
                        logger.info(f"Креатив специалиста загружен: {creative_file.name}, размер: {len(file_data)} байт")
                    except Exception as e:
                        logger.error(f"Ошибка отправки креатива специалиста в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось сохранить креатив: {e}")

            if proofs:
                for proof_file in proofs:
                    if not hasattr(proof_file, "name"):
                        continue
                    try:
                        proof_id = str(uuid.uuid4())
                        file_data = proof_file.read()
                        content_type = getattr(proof_file, 'content_type', 'application/pdf')
                        unique_filename = get_unique_filename(proof_file.name, specialist.specialist_id, "proof")

                        result = upload_file_to_s3_sync(
                            file_data=file_data,
                            file_name=proof_file.name,
                            content_type=content_type,
                            entity_type_name='specialist',
                            entity_id=specialist.specialist_id,
                            file_type_name='proof',
                            original_filename=unique_filename,
                            file_id=proof_id
                        )
                        if result:
                            proofs_ids.append(proof_id)
                        logger.info(f"Пруф специалиста загружен: {proof_file.name}, размер: {len(file_data)} байт")
                    except Exception as e:
                        logger.error(f"Ошибка отправки пруфа специалиста в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось сохранить документ: {e}")
            
            if videos:
                for video in videos:
                    if not hasattr(video, "name"):
                        logger.warning(f"Пропущено видео специалиста: {video}")
                        continue
                    try:
                        unique_filename = get_unique_filename(video.name, specialist.specialist_id, "video")
                        video_data = video.read()
                        content_type = getattr(video, 'content_type', 'video/mp4')
                        
                        video_data_b64 = base64.b64encode(video_data).decode('utf-8')
                        
                        upload_video_to_s3.delay(
                            video_data=video_data_b64,
                            video_name=video.name,
                            video_content_type=content_type,
                            entity_id=specialist.specialist_id,
                            original_filename=unique_filename,
                            entity_type_name='specialist'
                        )
                        logger.info(f"Видео специалиста отправлено в очередь Celery: {video.name}, размер: {len(video_data)} байт")
                        messages.info(request, f"Видео {video.name} загружается в фоновом режиме.")
                    except Exception as e:
                        logger.error(f"Ошибка отправки видео специалиста в очередь: {e}", exc_info=True)
                        messages.warning(request, f"Не удалось отправить видео {video.name} на загрузку.")
                logger.info("Видео специалиста загружается асинхронно через Celery")
                if specialist.video_urls is None:
                    specialist.video_urls = []
            
            # Обновляем URL файлов
            specialist.logo_urls = logo_ids
            specialist.creatives_urls = creative_ids
            specialist.proofs_urls = proofs_ids
            
            # Обработка удаленных файлов
            deleted_files_json = request.POST.get('deleted_files', '[]')
            try:
                deleted_files = json.loads(deleted_files_json)
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="specialist")
                for deleted_file in deleted_files:
                    file_id = deleted_file.get('id')
                    file_type = deleted_file.get('type')
                    if file_id and file_type:
                        # Получаем информацию о файле перед удалением
                        file_storage = FileStorage.objects.filter(
                            entity_type=entity_type,
                            entity_id=specialist.specialist_id,
                            file_url=file_id
                        ).first()
                        
                        if file_storage:
                            # Удаляем файл из S3
                            if file_type == 'creative':
                                file_path = f"specialists/{specialist.specialist_id}/creatives/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            elif file_type == 'proof':
                                file_path = f"specialists/{specialist.specialist_id}/proofs/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            elif file_type == 'video':
                                file_path = f"specialists/{specialist.specialist_id}/videos/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            else:
                                file_path = f"specialists/{specialist.specialist_id}/{file_type}s/{file_id}_{file_storage.original_file_name or 'unknown'}"
                            
                            delete_file_from_s3(file_path)
                        
                        # Удаляем запись из базы данных
                        FileStorage.objects.filter(
                            entity_type=entity_type,
                            entity_id=specialist.specialist_id,
                            file_url=file_id
                        ).delete()
                        
                        # Удаляем из списка URL
                        if file_type == 'creative' and specialist.creatives_urls:
                            specialist.creatives_urls = [url for url in specialist.creatives_urls if url != file_id]
                        elif file_type == 'proof' and specialist.proofs_urls:
                            specialist.proofs_urls = [url for url in specialist.proofs_urls if url != file_id]
                        elif file_type == 'video' and specialist.video_urls:
                            specialist.video_urls = [url for url in specialist.video_urls if url != file_id]
                        logger.info(f"Удален файл {file_type}: {file_id}")
            except json.JSONDecodeError:
                pass
            
            slider_images = request.POST.getlist("slider_images")
            if len(slider_images) > 4:
                slider_images = slider_images[:4]
            specialist.slider_images = slider_images
            
            # Обработка catalog_card_image
            catalog_card_image = form.cleaned_data.get("catalog_card_image") or request.FILES.get("catalog_card_image")
            if catalog_card_image and hasattr(catalog_card_image, 'read'):
                catalog_card_id = str(uuid.uuid4())
                # Конвертируем в WebP
                processed_catalog_image, processed_catalog_name, _ = process_uploaded_image(catalog_card_image, quality=85)
                base_name = os.path.splitext(processed_catalog_name)[0]
                ext = os.path.splitext(processed_catalog_name)[1]
                safe_base_name = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
                safe_name = slugify(safe_base_name) + ext
                file_path = f"catalog_cards/{catalog_card_id}_{safe_name}"
                try:
                    s3 = boto3.client(
                        's3',
                        endpoint_url=getattr(settings, 'AWS_S3_ENDPOINT_URL', None),
                        aws_access_key_id=getattr(settings, 'AWS_ACCESS_KEY_ID', None),
                        aws_secret_access_key=getattr(settings, 'AWS_SECRET_ACCESS_KEY', None),
                        region_name=getattr(settings, 'AWS_S3_REGION_NAME', None),
                        config=boto3.session.Config(s3={'addressing_style': getattr(settings, 'AWS_S3_ADDRESSING_STYLE', 'virtual')})
                    )
                    bucket = getattr(settings, 'AWS_STORAGE_BUCKET_NAME', None)
                    content_type = 'image/webp' if processed_catalog_name.endswith('.webp') else getattr(catalog_card_image, 'content_type', 'application/octet-stream')
                    if hasattr(processed_catalog_image, 'read'):
                        body_bytes = processed_catalog_image.read()
                    else:
                        body_bytes = processed_catalog_image.getvalue() if hasattr(processed_catalog_image, 'getvalue') else processed_catalog_image
                    try:
                        s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ContentType=content_type, ACL='public-read')
                    except Exception:
                        s3.put_object(Bucket=bucket, Key=file_path, Body=body_bytes, ACL='public-read')
                    specialist.catalog_card_image = f"{catalog_card_id}_{safe_name}"
                except Exception as e:
                    logger.error(f"Ошибка сохранения изображения карточки: {e}", exc_info=True)
                    messages.warning(request, "Не удалось сохранить изображение для карточки.")
            
            specialist.save()
            
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("specialist_detail", kwargs={"slug": specialist.slug or specialist.specialist_id})
                })
            messages.success(request, "Специалист успешно обновлен.")
            return redirect("specialist_detail", slug=specialist.slug or specialist.specialist_id)
        else:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": False,
                    "errors": {f: [x["message"] for x in e.get_json_data()] for f, e in form.errors.items()},
                    "non_field_errors": [str(e) for e in form.non_field_errors()],
                }, status=400)
            messages.error(request, "Форма содержит ошибки.")
    else:
        form = SpecialistEditForm(instance=specialist)
    
    context = {
        'form': form,
        'specialist': specialist,
    }
    return render(request, 'accounts/edit_specialist.html', context)


@login_required
def change_owner_franchise(request, franchise_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})

    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        new_owner_id = request.POST.get('new_owner_id')

        if not new_owner_id:
            return JsonResponse({'success': False, 'error': 'ID нового владельца не указан'})

        new_owner = get_object_or_404(Users, user_id=new_owner_id)
        franchise.owner = new_owner
        franchise.save()

        return JsonResponse({'success': True, 'message': 'Владелец франшизы изменен'})
    except Exception as e:
        logger.error(f"Ошибка при смене владельца франшизы: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при смене владельца'})


@login_required
def change_owner_agency(request, agency_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})

    try:
        agency = get_object_or_404(Agencies, agency_id=agency_id)
        new_owner_id = request.POST.get('new_owner_id')

        if not new_owner_id:
            return JsonResponse({'success': False, 'error': 'ID нового владельца не указан'})

        new_owner = get_object_or_404(Users, user_id=new_owner_id)
        agency.owner = new_owner
        agency.save()

        return JsonResponse({'success': True, 'message': 'Владелец агентства изменен'})
    except Exception as e:
        logger.error(f"Ошибка при смене владельца агентства: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при смене владельца'})


@login_required
def change_owner_specialist(request, specialist_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})

    try:
        specialist = get_object_or_404(Specialists, specialist_id=specialist_id)
        new_owner_id = request.POST.get('new_owner_id')

        if not new_owner_id:
            return JsonResponse({'success': False, 'error': 'ID нового владельца не указан'})

        new_owner = get_object_or_404(Users, user_id=new_owner_id)
        specialist.owner = new_owner
        specialist.save()

        return JsonResponse({'success': True, 'message': 'Владелец профиля специалиста изменен'})
    except Exception as e:
        logger.error(f"Ошибка при смене владельца специалиста: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при смене владельца'})


@login_required
def get_investors_franchise(request, franchise_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})

    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        investors = InvestmentTransactions.objects.filter(franchise=franchise).select_related('investor')

        investors_data = []
        for transaction in investors:
            investors_data.append({
                'user_id': transaction.investor.user_id,
                'name': f"{transaction.investor.first_name or ''} {transaction.investor.last_name or ''}".strip(),
                'amount': float(transaction.amount),
                'date': transaction.created_at.strftime('%d.%m.%Y')
            })

        return JsonResponse({'success': True, 'investors': investors_data})
    except Exception as e:
        logger.error(f"Ошибка при получении инвесторов франшизы: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при получении данных'})


@login_required
def add_investor_franchise(request, franchise_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})

    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        investor_id = request.POST.get('investor_id')
        amount = request.POST.get('amount')

        if not investor_id or not amount:
            return JsonResponse({'success': False, 'error': 'Не указаны ID инвестора или сумма'})

        investor = get_object_or_404(Users, user_id=investor_id)
        amount_decimal = Decimal(amount)

        transaction = InvestmentTransactions(
            franchise=franchise,
            investor=investor,
            amount=amount_decimal,
            transaction_type=TransactionTypes.objects.get(type_name="investment"),
            transaction_status="completed",
            payment_method=PaymentMethods.objects.get(method_name="default"),
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        transaction.save()

        return JsonResponse({'success': True, 'message': 'Инвестор добавлен'})
    except Exception as e:
        logger.error(f"Ошибка при добавлении инвестора франшизы: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при добавлении инвестора'})


@login_required
def edit_investment_franchise(request, franchise_id, user_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})

    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        investor = get_object_or_404(Users, user_id=user_id)
        new_amount = request.POST.get('amount')

        if not new_amount:
            return JsonResponse({'success': False, 'error': 'Не указана сумма'})

        transaction = InvestmentTransactions.objects.get(franchise=franchise, investor=investor)
        transaction.amount = Decimal(new_amount)
        transaction.save()

        return JsonResponse({'success': True, 'message': 'Инвестиция обновлена'})
    except InvestmentTransactions.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Инвестиция не найдена'})
    except Exception as e:
        logger.error(f"Ошибка при редактировании инвестиции франшизы: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при обновлении'})


@login_required
def delete_investment_franchise(request, franchise_id, user_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})

    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        investor = get_object_or_404(Users, user_id=user_id)

        transaction = InvestmentTransactions.objects.get(franchise=franchise, investor=investor)
        transaction.delete()

        return JsonResponse({'success': True, 'message': 'Инвестиция удалена'})
    except InvestmentTransactions.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Инвестиция не найдена'})
    except Exception as e:
        logger.error(f"Ошибка при удалении инвестиции франшизы: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при удалении'})


@login_required
def ckeditor_upload(request):
    """CKEditor 5 SimpleUploadAdapter endpoint — загрузка изображений в S3."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST allowed'}, status=405)

    uploaded = request.FILES.get('upload')
    if not uploaded:
        return JsonResponse({'error': {'message': 'Файл не предоставлен'}}, status=400)

    # Валидация: только изображения, макс 5 MB
    allowed_types = ('image/jpeg', 'image/png', 'image/gif', 'image/webp')
    if uploaded.content_type not in allowed_types:
        return JsonResponse({'error': {'message': 'Допустимы только изображения (JPEG, PNG, GIF, WebP)'}}, status=400)
    if uploaded.size > 5 * 1024 * 1024:
        return JsonResponse({'error': {'message': 'Максимальный размер файла — 5 МБ'}}, status=400)

    # Уникальное имя файла
    ext = os.path.splitext(uploaded.name)[1].lower() or '.jpg'
    filename = f'ckeditor/{uuid.uuid4().hex}{ext}'
    saved_path = default_storage.save(filename, uploaded)

    # Публичный URL (без подписи) — не истекает, т.к. бакет public-read
    file_url = f"{settings.S3_PUBLIC_BASE_URL}/{saved_path}"

    return JsonResponse({'url': file_url})


@login_required
def upload_description_media(request, entity_type, entity_id):
    """
    API endpoint для загрузки медиа-контента в описание
    """
    print(f"[UPLOAD] Request received: user={request.user.username}, entity_type={entity_type}, entity_id={entity_id}, method={request.method}")
    from accounts.utils import get_file_url as utils_get_file_url
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method allowed'}, status=405)
    
    try:
        # Проверяем тип сущности
        valid_entity_types = ['startup', 'franchise', 'agency', 'specialist']
        if entity_type not in valid_entity_types:
            return JsonResponse({'success': False, 'error': 'Invalid entity type'}, status=400)
        
        # Получаем сущность
        if entity_type == 'startup':
            entity = get_object_or_404(Startups, startup_id=entity_id)
        elif entity_type == 'franchise':
            entity = get_object_or_404(Franchises, franchise_id=entity_id)
        elif entity_type == 'agency':
            entity = get_object_or_404(Agencies, agency_id=entity_id)
        elif entity_type == 'specialist':
            entity = get_object_or_404(Specialists, specialist_id=entity_id)
        
        # Проверяем права доступа
        is_owner = entity.owner and request.user == entity.owner
        user_is_mod = is_moderator(request.user)

        owner_id = getattr(entity.owner, 'user_id', 'None') if entity.owner else 'None'
        print(f"[UPLOAD] user={request.user.user_id} ({request.user.username}), owner={owner_id}, is_owner={is_owner}, is_moderator={user_is_mod}")
        logger.info(f"Upload access check: user={request.user.user_id}, owner={owner_id}, is_owner={is_owner}, is_moderator={user_is_mod}")

        if not (is_owner or user_is_mod):
            print(f"[UPLOAD] PERMISSION DENIED for user {request.user.user_id} on {entity_type} {entity_id}")
            logger.warning(f"Upload permission denied for user {request.user.user_id} on {entity_type} {entity_id}")
            return JsonResponse({
                'success': False, 
                'error': f'Permission denied. You are not the owner of this {entity_type}.'
            }, status=403)
        
        # Получаем файлы
        files = request.FILES.getlist('files')
        if not files:
            return JsonResponse({'success': False, 'error': 'No files provided'}, status=400)
        
        # Валидация файлов
        allowed_image_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp']
        allowed_video_types = ['video/mp4', 'video/quicktime', 'video/x-msvideo', 'video/webm']
        max_image_size = 5 * 1024 * 1024  # 5MB
        max_video_size = 50 * 1024 * 1024  # 50MB
        max_images = 10
        max_videos = 2
        
        # Проверяем количество файлов по типам
        entity_type_obj, _ = EntityTypes.objects.get_or_create(type_name=entity_type)
        file_type_obj, _ = FileTypes.objects.get_or_create(type_name='uploaded_content')
        
        uploaded_files_qs = FileStorage.objects.filter(
            entity_type=entity_type_obj,
            entity_id=entity_id,
            file_type=file_type_obj
        )
        
        print(f"[UPLOAD] Found {uploaded_files_qs.count()} existing uploaded_content files for {entity_type} {entity_id}")
        
        existing_images = sum(1 for f in uploaded_files_qs if f.original_file_name and any(f.original_file_name.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp']))
        existing_videos = sum(1 for f in uploaded_files_qs if f.original_file_name and any(f.original_file_name.lower().endswith(ext) for ext in ['.mp4', '.mov', '.avi', '.webm']))
        
        print(f"[UPLOAD] Existing: {existing_images} images, {existing_videos} videos")
        
        # Подсчитываем новые файлы
        new_images = sum(1 for f in files if f.content_type in allowed_image_types)
        new_videos = sum(1 for f in files if f.content_type in allowed_video_types)
        
        print(f"[UPLOAD] New files to upload: {new_images} images, {new_videos} videos")
        
        if existing_images + new_images > max_images:
            return JsonResponse({
                'success': False, 
                'error': f'Максимум {max_images} изображений. Уже загружено: {existing_images}'
            }, status=400)
            
        if existing_videos + new_videos > max_videos:
            return JsonResponse({
                'success': False, 
                'error': f'Максимум {max_videos} видео. Уже загружено: {existing_videos}'
            }, status=400)
        
        uploaded_files = []
        
        for file in files:
            # Валидация типа файла
            file_type = file.content_type
            is_image = file_type in allowed_image_types
            is_video = file_type in allowed_video_types
            
            if not (is_image or is_video):
                return JsonResponse({
                    'success': False, 
                    'error': f'File {file.name} has unsupported type. Allowed: images (jpg, jpeg, png, gif, webp) and videos (mp4, mov, avi)'
                }, status=400)
            
            # Валидация размера
            max_size = max_image_size if is_image else max_video_size
            if file.size > max_size:
                size_mb = max_size / (1024 * 1024)
                return JsonResponse({
                    'success': False, 
                    'error': f'File {file.name} is too large. Maximum size: {size_mb}MB'
                }, status=400)
            
            # Генерируем UUID для файла
            file_id = str(uuid.uuid4())
            
            # Создаем путь для файла  
            base_name = os.path.splitext(file.name)[0]
            ext = os.path.splitext(file.name)[1]
            safe_base_name = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
            safe_name = slugify(safe_base_name) + ext
            file_path = f"{entity_type}s/{entity_id}/modal_download/{file_id}_{safe_name}"
            
            # Сохраняем файл
            try:
                default_storage.save(file_path, file)
                
                # Создаем запись в FileStorage
                file_type_obj, _ = FileTypes.objects.get_or_create(type_name='uploaded_content')
                entity_type_obj, _ = EntityTypes.objects.get_or_create(type_name=entity_type)
                
                safe_create_file_storage(
                    entity_type=entity_type_obj,
                    entity_id=entity_id,
                    file_type=file_type_obj,
                    file_url=file_id,
                    uploaded_at=timezone.now(),
                    startup=entity if entity_type == 'startup' else None,
                    original_file_name=file.name,
                )
                
                print(f"[UPLOAD] File saved to S3: {file_path}, FileStorage created with ID: {file_id}")
                
                # Получаем URL файла через get_file_url (правильно обрабатывает все имена файлов)
                file_url = utils_get_file_url(
                    file_id=file_id,
                    entity_id=entity_id,
                    file_type='uploaded_content',
                    entity_type=entity_type
                )
                
                print(f"[UPLOAD] Generated URL for file {file_id}: {file_url}")
                
                if not file_url:
                    file_url = f"{settings.S3_PUBLIC_BASE_URL}/{file_path}"
                    print(f"[UPLOAD] Fallback URL used: {file_url}")
                
                uploaded_files.append({
                    'file_id': file_id,
                    'file_url': file_url,
                    'file_type': 'image' if is_image else 'video',
                    'file_name': file.name
                })
                
            except Exception as e:
                logger.error(f"Error saving file {file.name}: {e}")
                return JsonResponse({
                    'success': False, 
                    'error': f'Error saving file {file.name}'
                }, status=500)
        
        return JsonResponse({
            'success': True,
            'files': uploaded_files
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"[UPLOAD ERROR] {error_details}")
        logger.error(f"Error in upload_description_media: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required  
def get_description_media(request, entity_type, entity_id):
    """
    API endpoint для получения списка загруженных медиа-файлов
    """
    from accounts.utils import get_file_url as utils_get_file_url
    
    print(f"[GET_MEDIA] Request received: entity_type={entity_type}, entity_id={entity_id}, method={request.method}")
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Only GET method allowed'}, status=405)
    
    logger.info(f"get_description_media called for {entity_type} {entity_id}")
    
    try:
        # Проверяем тип сущности
        valid_entity_types = ['startup', 'franchise', 'agency', 'specialist']
        if entity_type not in valid_entity_types:
            return JsonResponse({'success': False, 'error': 'Invalid entity type'}, status=400)
        
        # Получаем сущность
        if entity_type == 'startup':
            entity = get_object_or_404(Startups, startup_id=entity_id)
        elif entity_type == 'franchise':
            entity = get_object_or_404(Franchises, franchise_id=entity_id)
        elif entity_type == 'agency':
            entity = get_object_or_404(Agencies, agency_id=entity_id)
        elif entity_type == 'specialist':
            entity = get_object_or_404(Specialists, specialist_id=entity_id)
        
        # Проверяем права доступа
        is_owner = entity.owner and request.user == entity.owner
        user_is_mod = is_moderator(request.user)

        if not (is_owner or user_is_mod):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

        entity_type_obj, _ = EntityTypes.objects.get_or_create(type_name=entity_type)

        creative_storages = FileStorage.objects.filter(
            entity_type=entity_type_obj,
            entity_id=entity_id,
            file_type__type_name='creative'
        ).order_by('-uploaded_at')
        
        video_storages = FileStorage.objects.filter(
            entity_type=entity_type_obj,
            entity_id=entity_id,
            file_type__type_name='video'
        ).order_by('-uploaded_at')
        
        uploaded_storages = FileStorage.objects.filter(
            entity_type=entity_type_obj,
            entity_id=entity_id,
            file_type__type_name='uploaded_content'
        ).order_by('-uploaded_at')
        
        print(f"[GET_MEDIA] Found {creative_storages.count()} creative files, {video_storages.count()} video files, {uploaded_storages.count()} uploaded files")
        
        files = []
        
        for file_storage in creative_storages:
            original_name = getattr(file_storage, 'original_file_name', '')
            file_ext = os.path.splitext(original_name)[1].lower()
            
            # Определяем тип файла
            if file_ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.svg']:
                file_type = 'image'
            elif file_ext in ['.mp4', '.mov', '.avi', '.webm']:
                file_type = 'video'
            else:
                file_type = 'unknown'
            
            # Пропускаем неизвестные типы
            if file_type == 'unknown':
                continue
            
            file_url = utils_get_file_url(
                file_id=file_storage.file_url,
                entity_id=entity_id,
                file_type='creative',
                entity_type=entity_type
            )
            
            if not file_url:
                logger.warning(f"Could not get URL for file {file_storage.file_url}")
                continue
            
            files.append({
                'id': file_storage.file_url,
                'url': file_url,
                'type': file_type,
                'name': original_name,
                'uploaded_at': file_storage.uploaded_at.isoformat() if file_storage.uploaded_at else None,
                'source': 'gallery'
            })
        
        # Обрабатываем видео из галереи
        for file_storage in video_storages:
            original_name = getattr(file_storage, 'original_file_name', '')
            
            file_url = utils_get_file_url(
                file_id=file_storage.file_url,
                entity_id=entity_id,
                file_type='video',
                entity_type=entity_type
            )
            
            if not file_url:
                logger.warning(f"Could not get URL for video file {file_storage.file_url}")
                continue
            
            files.append({
                'id': file_storage.file_url,
                'url': file_url,
                'type': 'video',
                'name': original_name,
                'uploaded_at': file_storage.uploaded_at.isoformat() if file_storage.uploaded_at else None,
                'source': 'gallery'
            })
        
        for file_storage in uploaded_storages:
            original_name = getattr(file_storage, 'original_file_name', '')
            file_ext = os.path.splitext(original_name)[1].lower()
            
            print(f"[GET_MEDIA] Processing uploaded file: {original_name}, id: {file_storage.file_url}")
            
            file_url = utils_get_file_url(
                file_id=file_storage.file_url,
                entity_id=entity_id,
                file_type='uploaded_content',
                entity_type=entity_type
            )
            
            print(f"[GET_MEDIA] Generated URL: {file_url}")
            
            if not file_url:
                logger.warning(f"Could not get URL for file {file_storage.file_url}")
                continue
            
            if file_ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.svg']:
                file_type = 'image'
            elif file_ext in ['.mp4', '.mov', '.avi', '.webm']:
                file_type = 'video'
            else:
                file_type = 'unknown'
            
            files.append({
                'id': file_storage.file_url,
                'url': file_url,
                'type': file_type,
                'name': original_name,
                'uploaded_at': file_storage.uploaded_at.isoformat() if file_storage.uploaded_at else None,
                'source': 'uploaded'
            })
        
        return JsonResponse({
            'success': True,
            'files': files
        })
        
    except Exception as e:
        import traceback
        print(f"[ERROR] get_description_media failed: {e}")
        print(traceback.format_exc())
        logger.error(f"Error in get_description_media: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def delete_description_media(request, entity_type, entity_id, file_id):
    print(f"[DELETE] Request received: user={request.user.username}, entity_type={entity_type}, entity_id={entity_id}, file_id={file_id}")
    
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'Only DELETE method allowed'}, status=405)
    
    try:
        valid_entity_types = ['startup', 'franchise', 'agency', 'specialist']
        if entity_type not in valid_entity_types:
            return JsonResponse({'success': False, 'error': 'Invalid entity type'}, status=400)
        
        if entity_type == 'startup':
            entity = get_object_or_404(Startups, startup_id=entity_id)
        elif entity_type == 'franchise':
            entity = get_object_or_404(Franchises, franchise_id=entity_id)
        elif entity_type == 'agency':
            entity = get_object_or_404(Agencies, agency_id=entity_id)
        elif entity_type == 'specialist':
            entity = get_object_or_404(Specialists, specialist_id=entity_id)
        
        # Проверяем права доступа
        is_owner = entity.owner and request.user == entity.owner
        user_is_mod = is_moderator(request.user)

        if not (is_owner or user_is_mod):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

        entity_type_obj, _ = EntityTypes.objects.get_or_create(type_name=entity_type)
        file_storage = get_object_or_404(
            FileStorage,
            entity_type=entity_type_obj,
            entity_id=entity_id,
            file_type__type_name='uploaded_content',
            file_url=file_id
        )
        
        base_name = os.path.splitext(file_storage.original_file_name)[0]
        ext = os.path.splitext(file_storage.original_file_name)[1]
        safe_base_name = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
        safe_name_slugified = slugify(safe_base_name)
        
        safe_name = safe_name_slugified + ext
        file_path = f"{entity_type}s/{entity_id}/modal_download/{file_id}_{safe_name}"
        
        print(f"[DELETE] Attempting to delete file from S3: {file_path}")
        
        try:
            if default_storage.exists(file_path):
                default_storage.delete(file_path)
                print(f"[DELETE] File deleted from S3: {file_path}")
            else:
                print(f"[DELETE] File not found in S3: {file_path}")
        except Exception as e:
            logger.warning(f"Could not delete file from storage: {e}")
            print(f"[DELETE] Error deleting from S3: {e}")
        
        file_storage.delete()
        print(f"[DELETE] FileStorage record deleted for file_id: {file_id}")
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        logger.error(f"Error in delete_description_media: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Internal server error'}, status=500)


# ============================================================================
# Временная загрузка файлов при создании сущности
# ============================================================================

import tempfile
import os
import base64

@login_required
@require_http_methods(["POST"])
def temp_file_upload(request):
    """
    Загружает файлы во временное хранилище (сессию) при создании сущности.
    Файлы сохраняются как base64 в сессии и возвращаются при ошибке валидации.
    """
    try:
        if not request.FILES:
            return JsonResponse({'success': False, 'error': 'No files provided'}, status=400)
        
        field_name = request.POST.get('field_name', 'files')
        form_id = request.POST.get('form_id', 'default')
        
        # Инициализируем хранилище в сессии
        if 'temp_files' not in request.session:
            request.session['temp_files'] = {}
        if form_id not in request.session['temp_files']:
            request.session['temp_files'][form_id] = {}
        if field_name not in request.session['temp_files'][form_id]:
            request.session['temp_files'][form_id][field_name] = []
        
        uploaded = []
        for uploaded_file in request.FILES.getlist('file'):
            # Читаем файл и кодируем в base64
            file_content = uploaded_file.read()
            file_b64 = base64.b64encode(file_content).decode('utf-8')
            
            temp_id = str(uuid.uuid4())
            file_data = {
                'temp_id': temp_id,
                'name': uploaded_file.name,
                'size': uploaded_file.size,
                'content_type': uploaded_file.content_type,
                'data': file_b64,
            }
            
            request.session['temp_files'][form_id][field_name].append(file_data)
            uploaded.append({
                'temp_id': temp_id,
                'name': uploaded_file.name,
                'size': uploaded_file.size,
                'content_type': uploaded_file.content_type,
            })
        
        request.session.modified = True
        
        return JsonResponse({
            'success': True,
            'files': uploaded
        })
        
    except Exception as e:
        logger.error(f"Error in temp_file_upload: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_temp_files(request):
    """
    Возвращает список временных файлов для формы.
    """
    try:
        form_id = request.GET.get('form_id', 'default')
        field_name = request.GET.get('field_name')
        
        temp_files = request.session.get('temp_files', {})
        form_files = temp_files.get(form_id, {})
        
        if field_name:
            files = form_files.get(field_name, [])
            # Не возвращаем данные файла, только метаданные
            result = [{
                'temp_id': f['temp_id'],
                'name': f['name'],
                'size': f['size'],
                'content_type': f['content_type'],
            } for f in files]
            return JsonResponse({'success': True, 'files': result})
        else:
            # Возвращаем все поля
            result = {}
            for fn, files in form_files.items():
                result[fn] = [{
                    'temp_id': f['temp_id'],
                    'name': f['name'],
                    'size': f['size'],
                    'content_type': f['content_type'],
                } for f in files]
            return JsonResponse({'success': True, 'fields': result})
        
    except Exception as e:
        logger.error(f"Error in get_temp_files: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["DELETE", "POST"])
def delete_temp_file(request, temp_id):
    """
    Удаляет временный файл из сессии.
    """
    try:
        form_id = request.GET.get('form_id', request.POST.get('form_id', 'default'))
        
        temp_files = request.session.get('temp_files', {})
        form_files = temp_files.get(form_id, {})
        
        deleted = False
        for field_name, files in form_files.items():
            for i, f in enumerate(files):
                if f['temp_id'] == temp_id:
                    del files[i]
                    deleted = True
                    break
            if deleted:
                break
        
        if deleted:
            request.session.modified = True
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': 'File not found'}, status=404)
        
    except Exception as e:
        logger.error(f"Error in delete_temp_file: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def get_temp_file_content(request, form_id, field_name, temp_id):
    """
    Вспомогательная функция для получения содержимого временного файла.
    Используется при сохранении формы.
    """
    temp_files = request.session.get('temp_files', {})
    form_files = temp_files.get(form_id, {})
    field_files = form_files.get(field_name, [])
    
    for f in field_files:
        if f['temp_id'] == temp_id:
            content = base64.b64decode(f['data'])
            return {
                'name': f['name'],
                'content_type': f['content_type'],
                'content': content,
            }
    return None


def clear_temp_files(request, form_id):
    """
    Очищает временные файлы после успешного сохранения формы.
    """
    temp_files = request.session.get('temp_files', {})
    if form_id in temp_files:
        del temp_files[form_id]
        request.session.modified = True


@login_required
def my_analytics(request):
    from accounts.models import Startups, Franchises, Agencies, Specialists
    from accounts.analytics import get_entity_stats
    import json

    user = request.user
    days = 90  # Always use 90 days

    # Status filter from GET params (default: approved)
    status_filter = request.GET.get('status', 'approved')
    if status_filter not in ('approved', 'all', 'pending', 'rejected'):
        status_filter = 'approved'

    # Get ALL user's entities (for summary totals we always use approved)
    all_startups = list(Startups.objects.filter(owner=user).values('startup_id', 'title', 'slug', 'status'))
    all_franchises = list(Franchises.objects.filter(owner=user).values('franchise_id', 'title', 'slug', 'status'))
    all_agencies = list(Agencies.objects.filter(owner=user).values('agency_id', 'title', 'slug', 'status'))
    all_specialists = list(Specialists.objects.filter(owner=user).values('specialist_id', 'title', 'slug', 'status'))

    # Build full entity list with stats
    def build_entities(items, entity_type, type_label, id_field):
        result = []
        for item in items:
            stats = get_entity_stats(entity_type, item[id_field], days=days)
            result.append({
                'type': entity_type, 'type_label': type_label,
                'id': item[id_field], 'title': item['title'], 'slug': item.get('slug', ''),
                'status': item.get('status', ''), 'stats': stats
            })
        return result

    all_entities = []
    all_entities += build_entities(all_startups, 'startup', '\u0421\u0442\u0430\u0440\u0442\u0430\u043f', 'startup_id')
    all_entities += build_entities(all_franchises, 'franchise', '\u0424\u0440\u0430\u043d\u0448\u0438\u0437\u0430', 'franchise_id')
    all_entities += build_entities(all_agencies, 'agency', '\u0410\u0433\u0435\u043d\u0442\u0441\u0442\u0432\u043e', 'agency_id')
    all_entities += build_entities(all_specialists, 'specialist', '\u0421\u043f\u0435\u0446\u0438\u0430\u043b\u0438\u0441\u0442', 'specialist_id')

    # Summary totals from approved entities only
    approved_entities = [e for e in all_entities if e['status'] == 'approved']
    total_views = sum(e['stats']['total_views'] for e in approved_entities)
    total_unique = sum(e['stats']['unique_views'] for e in approved_entities)
    total_clicks = sum(
        sum(e['stats']['total_clicks'].values()) for e in approved_entities
    )
    total_impressions = sum(e['stats']['total_impressions'] for e in approved_entities)
    total_ctr = round(total_clicks / total_views * 100, 1) if total_views > 0 else 0

    # Aggregate engagement across approved entities
    agg_time_sum = 0
    agg_eng_count = 0
    for e in approved_entities:
        s = e['stats']
        agg_time_sum += s['avg_time_on_page'] * max(s.get('engagement_count', 1), 1) if s['avg_time_on_page'] > 0 else 0
        agg_eng_count += 1 if s['avg_time_on_page'] > 0 else 0
    avg_time_on_page = int(agg_time_sum / agg_eng_count) if agg_eng_count > 0 else 0
    avg_scroll_depth = 0
    if approved_entities:
        scroll_vals = [e['stats']['avg_scroll_depth'] for e in approved_entities if e['stats']['avg_scroll_depth'] > 0]
        avg_scroll_depth = int(sum(scroll_vals) / len(scroll_vals)) if scroll_vals else 0

    # Aggregate sources
    agg_sources = {"direct": 0, "search": 0, "social": 0, "internal": 0, "other": 0}
    for e in approved_entities:
        for k, v in e['stats']['sources'].items():
            agg_sources[k] += v

    # Aggregate period comparison
    comp_views = comp_clicks = comp_impressions = 0
    comp_count = 0
    for e in approved_entities:
        c = e['stats']['comparison']
        comp_views += c['views']
        comp_clicks += c['clicks']
        comp_impressions += c['impressions']
        comp_count += 1
    if comp_count > 0:
        comp_views = round(comp_views / comp_count, 1)
        comp_clicks = round(comp_clicks / comp_count, 1)
        comp_impressions = round(comp_impressions / comp_count, 1)
    agg_comparison = {"views": comp_views, "clicks": comp_clicks, "impressions": comp_impressions}

    # Filter entities for display based on status_filter
    if status_filter == 'all':
        filtered_entities = all_entities
    else:
        filtered_entities = [e for e in all_entities if e['status'] == status_filter]

    # Selected entity from query params
    selected_type = request.GET.get('type', '')
    selected_id = request.GET.get('id', '')

    # If specific entity selected, get its stats
    selected_stats = None
    selected_entity = None
    if selected_type and selected_id:
        selected_stats = get_entity_stats(selected_type, int(selected_id), days=days)
        # Find entity in all_entities (not just filtered)
        for e in all_entities:
            if e['type'] == selected_type and str(e['id']) == selected_id:
                selected_entity = e
                break

    # ── Lead analytics ──────────────────────────────────────────
    from .models import Lead
    from django.db.models.functions import TruncDate

    leads_qs = Lead.objects.filter(entity_owner=user).order_by('-created_at')
    lead_stats = {
        'total': leads_qs.count(),
        'new': leads_qs.filter(status='new').count(),
        'viewed': leads_qs.filter(status='viewed').count(),
        'responded': leads_qs.filter(status='responded').count(),
        'converted': leads_qs.filter(status='converted').count(),
    }

    # Lead status filter
    lead_status_filter = request.GET.get('lead_status', '')
    lead_entity_filter = request.GET.get('lead_entity', '')
    filtered_leads = leads_qs
    if lead_status_filter:
        filtered_leads = filtered_leads.filter(status=lead_status_filter)
    if lead_entity_filter:
        filtered_leads = filtered_leads.filter(entity_type=lead_entity_filter)

    # Mark new leads as viewed when page loads
    leads_qs.filter(status='new').update(status='viewed', viewed_at=timezone.now())

    # Leads per day for chart (last 30 days)
    from datetime import timedelta
    thirty_days_ago = timezone.now() - timedelta(days=30)
    leads_by_day = list(
        leads_qs.filter(created_at__gte=thirty_days_ago)
        .annotate(date=TruncDate('created_at'))
        .values('date')
        .annotate(count=Count('lead_id'))
        .order_by('date')
    )
    leads_chart_data = json.dumps([
        {'date': d['date'].strftime('%Y-%m-%d'), 'count': d['count']}
        for d in leads_by_day
    ], cls=DjangoJSONEncoder)

    # Paginate leads list
    leads_paginator = Paginator(filtered_leads, 15)
    leads_page = leads_paginator.get_page(request.GET.get('leads_page', 1))

    context = {
        'all_entities': filtered_entities,
        'total_views': total_views,
        'total_unique': total_unique,
        'total_clicks': total_clicks,
        'total_impressions': total_impressions,
        'total_ctr': total_ctr,
        'avg_time_on_page': avg_time_on_page,
        'avg_scroll_depth': avg_scroll_depth,
        'agg_sources': json.dumps(agg_sources),
        'agg_comparison': agg_comparison,
        'selected_type': selected_type,
        'selected_id': selected_id,
        'selected_stats': selected_stats,
        'selected_entity': selected_entity,
        'selected_stats_json': json.dumps(selected_stats) if selected_stats else '{}',
        'status_filter': status_filter,
        # Lead data
        'lead_stats': lead_stats,
        'leads_page': leads_page,
        'leads_chart_data': leads_chart_data,
        'lead_status_filter': lead_status_filter,
        'lead_entity_filter': lead_entity_filter,
    }
    return render(request, 'accounts/my_analytics.html', context)


@csrf_exempt
@require_POST
def track_page_view(request):
    """Beacon endpoint for page view tracking."""
    try:
        import json
        from accounts.analytics import record_page_view
        data = json.loads(request.body)
        entity_type = data.get('entity_type', '')
        entity_id = data.get('entity_id', 0)
        if entity_type not in ('startup', 'franchise', 'agency', 'specialist'):
            return JsonResponse({'status': 'error'}, status=400)
        record_page_view(entity_type, int(entity_id), request)
        return JsonResponse({'status': 'ok'})
    except Exception:
        return JsonResponse({'status': 'error'}, status=500)


@csrf_exempt
@require_POST
def track_click(request):
    """Beacon endpoint for click tracking."""
    try:
        import json
        from accounts.analytics import record_click
        data = json.loads(request.body)
        entity_type = data.get('entity_type', '')
        entity_id = data.get('entity_id', 0)
        button_type = data.get('button_type', '')
        valid_types = ('startup', 'franchise', 'agency', 'specialist', 'footer')
        valid_buttons = ('contact', 'website', 'pitch_deck', 'telegram', 'whatsapp', 'subscribe_telegram')
        if entity_type not in valid_types or button_type not in valid_buttons:
            return JsonResponse({'status': 'error'}, status=400)
        record_click(entity_type, int(entity_id), button_type, request)
        return JsonResponse({'status': 'ok'})
    except Exception:
        return JsonResponse({'status': 'error'}, status=500)


@csrf_exempt
@require_POST
def track_impression(request):
    """Beacon endpoint for catalog impression tracking (batch)."""
    try:
        import json
        from accounts.analytics import record_impression
        data = json.loads(request.body)
        items = data.get('items', [])
        valid_types = ('startup', 'franchise', 'agency', 'specialist')
        for item in items[:50]:
            et = item.get('entity_type', '')
            eid = item.get('entity_id', 0)
            if et in valid_types and eid:
                record_impression(et, int(eid), request)
        return JsonResponse({'status': 'ok'})
    except Exception:
        return JsonResponse({'status': 'error'}, status=500)


@csrf_exempt
@require_POST
def track_engagement(request):
    """Beacon endpoint for engagement tracking (time on page + scroll depth)."""
    try:
        import json
        from accounts.analytics import record_engagement
        data = json.loads(request.body)
        entity_type = data.get('entity_type', '')
        entity_id = data.get('entity_id', 0)
        time_on_page = data.get('time_on_page', 0)
        scroll_depth = data.get('scroll_depth', 0)
        valid_types = ('startup', 'franchise', 'agency', 'specialist')
        if entity_type not in valid_types:
            return JsonResponse({'status': 'error'}, status=400)
        record_engagement(entity_type, int(entity_id), time_on_page, scroll_depth, request)
        return JsonResponse({'status': 'ok'})
    except Exception:
        return JsonResponse({'status': 'error'}, status=500)




# ══════════════════════════════════════════════════════════════════
# Lead Generation System
# ══════════════════════════════════════════════════════════════════

@csrf_exempt
@require_POST
def create_lead(request):
    """Create a new lead (inquiry) for an entity."""
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"success": False, "error": "Invalid data"}, status=400)

    entity_type = data.get("entity_type", "")
    entity_id = data.get("entity_id")
    lead_type = data.get("lead_type", "")
    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    phone = (data.get("phone") or "").strip()
    budget_range = data.get("budget_range", "")
    message_text = (data.get("message") or "").strip()
    target_city_id = data.get("target_city") or None
    business_experience = data.get("business_experience", "")
    timeline = data.get("timeline", "")

    valid_entity_types = ("startup", "franchise", "agency", "specialist")
    valid_lead_types = ("invest", "franchise_info", "quote", "consultation")
    if entity_type not in valid_entity_types:
        return JsonResponse({"success": False, "error": "Неверный тип сущности"}, status=400)
    if lead_type not in valid_lead_types:
        return JsonResponse({"success": False, "error": "Неверный тип заявки"}, status=400)
    if not name or not email:
        return JsonResponse({"success": False, "error": "Имя и email обязательны"}, status=400)
    if not entity_id:
        return JsonResponse({"success": False, "error": "Не указан ID объекта"}, status=400)

    model_map = {
        "startup": (Startups, "startup_id"),
        "franchise": (Franchises, "franchise_id"),
        "agency": (Agencies, "agency_id"),
        "specialist": (Specialists, "specialist_id"),
    }
    Model, pk_field = model_map[entity_type]
    entity = Model.objects.filter(**{pk_field: entity_id}).select_related("owner").first()
    if not entity:
        return JsonResponse({"success": False, "error": "Объект не найден"}, status=404)

    entity_owner = getattr(entity, "owner", None)

    if request.user.is_authenticated and entity_owner and request.user.pk == entity_owner.pk:
        return JsonResponse({"success": False, "error": "Нельзя оставить заявку на свой объект"}, status=400)

    from .models import Lead, City
    target_city = None
    if target_city_id:
        target_city = City.objects.filter(city_id=target_city_id).first()

    lead = Lead.objects.create(
        entity_type=entity_type,
        entity_id=entity_id,
        user=request.user if request.user.is_authenticated else None,
        entity_owner=entity_owner,
        lead_type=lead_type,
        name=name,
        email=email,
        phone=phone,
        budget_range=budget_range,
        message=message_text,
        target_city=target_city,
        business_experience=business_experience,
        timeline=timeline,
    )

    try:
        _send_lead_telegram_notification(lead, entity)
    except Exception as e:
        logger.error("Failed to send lead Telegram notification: %s", e)

    if entity_owner:
        try:
            _create_lead_notification(lead, entity_owner, entity)
        except Exception as e:
            logger.error("Failed to create lead notification: %s", e)
        # Send Telegram DM to entity owner if they have telegram_id
        try:
            if entity_owner.telegram_id:
                _send_lead_owner_telegram(lead, entity, entity_owner)
        except Exception as e:
            logger.error("Failed to send owner Telegram DM: %s", e)

    # Send lead to CRM integrations (async via Celery)
    try:
        from .tasks import send_lead_to_crm_task
        send_lead_to_crm_task.delay(lead.lead_id)
    except Exception as e:
        logger.error("Failed to queue CRM sync for lead %s: %s", lead.lead_id, e)

    return JsonResponse({"success": True, "lead_id": lead.lead_id})


def _send_lead_telegram_notification(lead, entity):
    """Send Telegram notification about new lead to admin."""
    from django.conf import settings as django_settings
    bot_token = getattr(django_settings, "TELEGRAM_BOT_TOKEN", None)
    chat_id = getattr(django_settings, "TELEGRAM_OWNER_CHAT_ID", "911873673")
    if not bot_token:
        return

    entity_emojis = {"startup": "\U0001F680", "franchise": "\U0001F3EA", "agency": "\U0001F3E2", "specialist": "\U0001F468\u200D\U0001F4BC"}
    lead_type_names = {"invest": "\u0418\u043D\u0432\u0435\u0441\u0442\u0438\u0446\u0438\u044F", "franchise_info": "\u0424\u0440\u0430\u043D\u0448\u0438\u0437\u0430", "quote": "\u0420\u0430\u0441\u0447\u0451\u0442", "consultation": "\u041A\u043E\u043D\u0441\u0443\u043B\u044C\u0442\u0430\u0446\u0438\u044F"}
    emoji = entity_emojis.get(lead.entity_type, "\U0001F4DD")

    entity_url_paths = {"startup": "startups", "franchise": "franchises", "agency": "agencies", "specialist": "specialists"}
    url_path = entity_url_paths.get(lead.entity_type, "startups")
    view_url = f"https://greatideas.ru/{url_path}/{lead.entity_id}/"

    text = (
        f"\U0001F4E9 <b>\u041D\u043E\u0432\u0430\u044F \u0437\u0430\u044F\u0432\u043A\u0430!</b>\n\n"
        f"{emoji} <b>{entity.title or 'N/A'}</b>\n"
        f"\U0001F4CB \u0422\u0438\u043F: {lead_type_names.get(lead.lead_type, lead.lead_type)}\n\n"
        f"\U0001F464 {lead.name}\n"
        f"\u2709\uFE0F {lead.email}\n"
    )
    if lead.phone:
        text += f"\U0001F4DE {lead.phone}\n"
    if lead.budget_range:
        text += f"\U0001F4B0 \u0411\u044E\u0434\u0436\u0435\u0442: {lead.budget_range}\n"
    if lead.message:
        msg_short = lead.message[:200] + ("..." if len(lead.message) > 200 else "")
        text += f"\n\U0001F4AC {msg_short}\n"

    inline_keyboard = {"inline_keyboard": [[
        {"text": "\U0001F441 \u041F\u043E\u0441\u043C\u043E\u0442\u0440\u0435\u0442\u044C", "url": view_url},
        {"text": "\U0001F4CB \u0410\u0434\u043C\u0438\u043D\u043A\u0430", "url": "https://greatideas.ru/admin/accounts/lead/"},
    ]]}

    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    try:
        requests.post(url, json={"chat_id": chat_id, "text": text, "parse_mode": "HTML", "reply_markup": inline_keyboard}, timeout=10)
    except Exception as e:
        logger.error("Telegram lead notification failed: %s", e)


def _create_lead_notification(lead, entity_owner, entity):
    """Create in-app notification for entity owner."""
    from django.db import connection
    lead_type_names = {"invest": "\u0438\u043D\u0432\u0435\u0441\u0442\u0438\u0446\u0438\u0438", "franchise_info": "\u0444\u0440\u0430\u043D\u0448\u0438\u0437\u0435", "quote": "\u0440\u0430\u0441\u0447\u0451\u0442\u0435", "consultation": "\u043A\u043E\u043D\u0441\u0443\u043B\u044C\u0442\u0430\u0446\u0438\u0438"}
    type_text = lead_type_names.get(lead.lead_type, "\u0432\u0430\u0448\u0435\u043C \u043E\u0431\u044A\u0435\u043A\u0442\u0435")
    msg = f"\u041D\u043E\u0432\u0430\u044F \u0437\u0430\u044F\u0432\u043A\u0430 \u043E {type_text}: {entity.title or 'N/A'} \u043E\u0442 {lead.name}"
    try:
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO notifications (user_id, notification_type_id, message, is_read, created_at) VALUES (%s, (SELECT notification_type_id FROM notification_types WHERE type_name = 'system' LIMIT 1), %s, false, NOW())",
                [entity_owner.pk, msg[:500]],
            )
    except Exception as e:
        logger.warning("Failed to insert notification: %s", e)


def _send_lead_owner_telegram(lead, entity, owner):
    """Send Telegram DM to entity owner about new lead."""
    from django.conf import settings as django_settings
    bot_token = getattr(django_settings, "TELEGRAM_BOT_TOKEN", None)
    if not bot_token or not owner.telegram_id:
        return

    lead_type_names = {"invest": "инвестиции", "franchise_info": "франшизе", "quote": "расчёте", "consultation": "консультации"}
    type_text = lead_type_names.get(lead.lead_type, "вашем объекте")

    entity_url_paths = {"startup": "startups", "franchise": "franchises", "agency": "agencies", "specialist": "specialists"}
    url_path = entity_url_paths.get(lead.entity_type, "startups")
    view_url = f"https://greatideas.ru/{url_path}/{lead.entity_id}/"

    text = (
        f"📩 <b>Новая заявка о {type_text}!</b>\n\n"
        f"📋 <b>{entity.title or 'Ваш объект'}</b>\n\n"
        f"👤 {lead.name}\n"
        f"✉️ {lead.email}\n"
    )
    if lead.phone:
        text += f"📞 {lead.phone}\n"
    if lead.budget_range:
        text += f"💰 Бюджет: {lead.budget_range}\n"
    if lead.message:
        msg_short = lead.message[:150] + ("..." if len(lead.message) > 150 else "")
        text += f"\n💬 {msg_short}\n"
    text += f"\n👉 Посмотреть все заявки: https://greatideas.ru/my-analytics/"

    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    try:
        requests.post(url, json={
            "chat_id": owner.telegram_id,
            "text": text,
            "parse_mode": "HTML",
            "reply_markup": {"inline_keyboard": [[
                {"text": "📋 Мои заявки", "url": "https://greatideas.ru/my-analytics/"},
                {"text": "👁 Объект", "url": view_url},
            ]]},
        }, timeout=10)
    except Exception as e:
        logger.error("Telegram DM to owner %s failed: %s", owner.pk, e)


@login_required
def my_leads(request):
    """Dashboard for entity owners to view their leads."""
    from .models import Lead
    leads_qs = Lead.objects.filter(entity_owner=request.user).order_by("-created_at")

    status_filter = request.GET.get("status", "")
    entity_type_filter = request.GET.get("entity_type", "")
    if status_filter:
        leads_qs = leads_qs.filter(status=status_filter)
    if entity_type_filter:
        leads_qs = leads_qs.filter(entity_type=entity_type_filter)

    # Mark new leads as viewed
    leads_qs.filter(status="new").update(status="viewed", viewed_at=timezone.now())

    all_leads = Lead.objects.filter(entity_owner=request.user)
    stats = {
        "total": all_leads.count(),
        "new": all_leads.filter(status="new").count(),
        "viewed": all_leads.filter(status="viewed").count(),
        "responded": all_leads.filter(status="responded").count(),
        "converted": all_leads.filter(status="converted").count(),
    }

    paginator = Paginator(leads_qs, 20)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(request, "accounts/my_leads.html", {
        "page_obj": page_obj,
        "stats": stats,
        "status_filter": status_filter,
        "entity_type_filter": entity_type_filter,
    })


@login_required
@require_POST
def update_lead_status(request, lead_id):
    """Update lead status (entity owner only)."""
    from .models import Lead

    lead = Lead.objects.filter(lead_id=lead_id, entity_owner=request.user).first()
    if not lead:
        return JsonResponse({"success": False, "error": "Not found"}, status=404)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"success": False, "error": "Invalid data"}, status=400)

    new_status = data.get("status", "")
    valid_transitions = {
        "new": ["viewed", "responded", "converted"],
        "viewed": ["responded", "converted"],
        "responded": ["converted"],
        "converted": [],
    }
    if new_status not in valid_transitions.get(lead.status, []):
        return JsonResponse({"success": False, "error": "Invalid status transition"}, status=400)

    lead.status = new_status
    if new_status == "viewed" and not lead.viewed_at:
        lead.viewed_at = timezone.now()
    elif new_status == "responded" and not lead.responded_at:
        lead.responded_at = timezone.now()
    lead.save()

    return JsonResponse({"success": True, "status": new_status})


@login_required
def export_leads(request):
    """Export leads as Excel file for entity owners."""
    from .models import Lead

    leads_qs = Lead.objects.filter(entity_owner=request.user).order_by("-created_at")

    status_filter = request.GET.get("status", "")
    if status_filter and status_filter in ("new", "viewed", "responded", "converted"):
        leads_qs = leads_qs.filter(status=status_filter)

    entity_type_filter = request.GET.get("entity_type", "")
    if entity_type_filter:
        leads_qs = leads_qs.filter(entity_type=entity_type_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Лиды"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")

    headers = [
        "Дата", "Тип", "Объект", "Имя", "Email", "Телефон",
        "Бюджет", "Город", "Опыт", "Сроки", "Сообщение",
        "Статус", "Дата просмотра", "Дата ответа",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_display = dict(Lead.STATUS_CHOICES)
    experience_display = dict(Lead.EXPERIENCE_CHOICES)
    timeline_display = dict(Lead.TIMELINE_CHOICES)

    for row_num, lead in enumerate(leads_qs[:5000], 2):
        ws.cell(row=row_num, column=1, value=lead.created_at.strftime("%d.%m.%Y %H:%M") if lead.created_at else "")
        ws.cell(row=row_num, column=2, value=lead.get_entity_type_display())
        ws.cell(row=row_num, column=3, value=lead.get_entity_title())
        ws.cell(row=row_num, column=4, value=lead.name)
        ws.cell(row=row_num, column=5, value=lead.email)
        ws.cell(row=row_num, column=6, value=lead.phone)
        ws.cell(row=row_num, column=7, value=lead.budget_range)
        ws.cell(row=row_num, column=8, value=lead.target_city.name if lead.target_city else "")
        ws.cell(row=row_num, column=9, value=experience_display.get(lead.business_experience, ""))
        ws.cell(row=row_num, column=10, value=timeline_display.get(lead.timeline, ""))
        ws.cell(row=row_num, column=11, value=lead.message[:500] if lead.message else "")
        ws.cell(row=row_num, column=12, value=status_display.get(lead.status, lead.status))
        ws.cell(row=row_num, column=13, value=lead.viewed_at.strftime("%d.%m.%Y %H:%M") if lead.viewed_at else "")
        ws.cell(row=row_num, column=14, value=lead.responded_at.strftime("%d.%m.%Y %H:%M") if lead.responded_at else "")

    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    report_date = timezone.now().strftime("%Y-%m-%d")
    response["Content-Disposition"] = f'attachment; filename="leads_{report_date}.xlsx"'
    wb.save(response)
    return response


@login_required
def crm_settings(request):
    """CRM integration settings page."""
    from .models import CRMIntegration

    integrations = CRMIntegration.objects.filter(user=request.user)
    integrations_dict = {i.crm_type: i for i in integrations}

    if request.method == "POST":
        action = request.POST.get("action", "")
        crm_type = request.POST.get("crm_type", "")

        if action == "delete" and crm_type:
            CRMIntegration.objects.filter(user=request.user, crm_type=crm_type).delete()
            messages.success(request, "Интеграция удалена.")
            return redirect("crm_settings")

        if action == "test" and crm_type:
            integration = integrations_dict.get(crm_type)
            if integration:
                from .crm_integration import get_adapter
                adapter = get_adapter(integration)
                if adapter.test_connection():
                    messages.success(request, f"Подключение к {integration.get_crm_type_display()} успешно!")
                else:
                    messages.error(request, f"Не удалось подключиться к {integration.get_crm_type_display()}. Проверьте настройки.")
            return redirect("crm_settings")

        if crm_type in ("bitrix24", "amocrm", "webhook"):
            webhook_url = request.POST.get("webhook_url", "").strip()
            api_key = request.POST.get("api_key", "").strip()
            subdomain = request.POST.get("subdomain", "").strip()
            is_active = request.POST.get("is_active") == "on"

            if not webhook_url:
                messages.error(request, "URL обязателен.")
                return redirect("crm_settings")

            integration, created = CRMIntegration.objects.update_or_create(
                user=request.user,
                crm_type=crm_type,
                defaults={
                    "webhook_url": webhook_url,
                    "api_key": api_key,
                    "subdomain": subdomain,
                    "is_active": is_active,
                    "last_error": "",
                },
            )
            messages.success(request, f"{'Добавлена' if created else 'Обновлена'} интеграция {integration.get_crm_type_display()}")
            return redirect("crm_settings")

    context = {
        "integrations": integrations,
        "integrations_dict": integrations_dict,
    }
    return render(request, "accounts/crm_settings.html", context)


# ── Franchisee Discovery (moderator) ────────────────────────

@login_required
def franchisee_discovery_list(request):
    """Moderator page: list franchises for franchisee contact discovery."""
    if not is_moderator(request.user):
        messages.error(request, "Доступ запрещён.")
        return redirect("home")

    from .models import FranchiseAnalysisLog, FranchiseeContact

    q = request.GET.get("q", "").strip()
    franchises_qs = Franchises.objects.filter(status="approved").select_related("direction")

    if q:
        franchises_qs = franchises_qs.filter(title__icontains=q)

    try:
        franchises_qs = franchises_qs.annotate(
            contacts_count=Count("franchisee_contacts"),
            locations_count=Count("locations", filter=models.Q(locations__status="active")),
            last_analyzed=Max(
                "analysis_logs__created_at",
                filter=models.Q(analysis_logs__status="completed"),
            ),
            is_analyzing=Exists(
                FranchiseAnalysisLog.objects.filter(
                    franchise=OuterRef("franchise_id"),
                    status__in=["pending", "running"],
                )
            ),
        ).order_by("-contacts_count", "-created_at")
    except Exception:
        # Таблицы ещё не созданы — показать без аннотаций
        franchises_qs = franchises_qs.annotate(
            locations_count=Count("locations", filter=models.Q(locations__status="active")),
        ).order_by("-created_at")

    paginator = Paginator(franchises_qs, 25)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    context = {
        "page_obj": page_obj,
        "query": q,
    }
    return render(request, "accounts/franchisee_discovery.html", context)


@login_required
def analyze_franchise_contacts(request, franchise_id):
    """Start async franchisee contact analysis (POST, AJAX)."""
    if not is_moderator(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    from .models import FranchiseAnalysisLog

    franchise = get_object_or_404(Franchises, franchise_id=franchise_id, status="approved")

    # Check not already running
    running = FranchiseAnalysisLog.objects.filter(
        franchise=franchise, status__in=["pending", "running"]
    ).exists()
    if running:
        return JsonResponse({"error": "Анализ уже выполняется"}, status=409)

    log = FranchiseAnalysisLog.objects.create(
        franchise=franchise,
        initiated_by=request.user,
        status="pending",
    )

    from .tasks import analyze_franchise_contacts_task
    task = analyze_franchise_contacts_task.delay(log.log_id)
    log.celery_task_id = task.id
    log.save(update_fields=["celery_task_id"])

    return JsonResponse({"status": "started", "log_id": log.log_id})


@login_required
def franchisee_analysis_status(request, log_id):
    """Poll analysis status (GET, AJAX)."""
    if not is_moderator(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)

    from .models import FranchiseAnalysisLog
    log = get_object_or_404(FranchiseAnalysisLog, log_id=log_id)

    return JsonResponse({
        "status": log.status,
        "contacts_found": log.contacts_found,
        "error": log.error_message,
        "sources_count": len(log.sources_scraped) if log.sources_scraped else 0,
    })


@login_required
def franchisee_contacts_detail(request, franchise_id):
    """View all discovered contacts for a franchise."""
    if not is_moderator(request.user):
        messages.error(request, "Доступ запрещён.")
        return redirect("home")

    from .models import FranchiseAnalysisLog, FranchiseeContact

    franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
    status_filter = request.GET.get("status", "")

    contacts_qs = FranchiseeContact.objects.filter(franchise=franchise)
    if status_filter:
        contacts_qs = contacts_qs.filter(outreach_status=status_filter)

    analysis_logs = FranchiseAnalysisLog.objects.filter(franchise=franchise).order_by("-created_at")[:10]

    # Status summary
    status_counts = {}
    for choice_val, choice_label in FranchiseeContact.OUTREACH_STATUS_CHOICES:
        cnt = FranchiseeContact.objects.filter(franchise=franchise, outreach_status=choice_val).count()
        status_counts[choice_val] = {"label": choice_label, "count": cnt}

    context = {
        "franchise": franchise,
        "contacts": contacts_qs,
        "analysis_logs": analysis_logs,
        "status_filter": status_filter,
        "status_counts": status_counts,
        "outreach_choices": FranchiseeContact.OUTREACH_STATUS_CHOICES,
    }
    return render(request, "accounts/franchisee_contacts.html", context)


@login_required
def update_franchisee_contact(request, contact_id):
    """Update contact outreach status and notes (POST, AJAX)."""
    if not is_moderator(request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    from .models import FranchiseeContact

    contact = get_object_or_404(FranchiseeContact, contact_id=contact_id)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    new_status = data.get("outreach_status", "")
    valid_statuses = [c[0] for c in FranchiseeContact.OUTREACH_STATUS_CHOICES]
    if new_status and new_status in valid_statuses:
        contact.outreach_status = new_status
        if new_status == "contacted" and not contact.contacted_at:
            contact.contacted_at = timezone.now()

    notes = data.get("moderator_notes")
    if notes is not None:
        contact.moderator_notes = notes

    contact.save()
    return JsonResponse({"ok": True, "status": contact.outreach_status})


def franchises_by_city(request, city_slug):
    """SEO landing page for franchises in a specific city."""
    from .models import City, FranchiseLocation

    city = get_object_or_404(City, slug=city_slug)

    franchise_ids = FranchiseLocation.objects.filter(
        city=city, status="active"
    ).values_list("franchise_id", flat=True)

    franchises_qs = Franchises.objects.filter(
        franchise_id__in=franchise_ids, status="approved"
    ).select_related("owner", "direction", "stage").order_by("-created_at")

    paginator = Paginator(franchises_qs, 12)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    # Get other cities for navigation
    other_cities = City.objects.filter(
        franchise_locations__status="active"
    ).exclude(city_id=city.city_id).distinct().order_by("name")[:20]

    # Категории франшиз в этом городе (для перелинковки)
    city_direction_ids = franchises_qs.filter(
        direction__isnull=False
    ).values_list("direction_id", flat=True).distinct()
    city_directions = Directions.objects.filter(
        direction_id__in=city_direction_ids
    ).order_by("direction_name")

    context = {
        "city": city,
        "page_obj": page_obj,
        "paginator": paginator,
        "franchises_count": franchises_qs.count(),
        "other_cities": other_cities,
        "city_directions": city_directions,
    }

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string("accounts/partials/_franchise_cards.html", context, request=request)
        return JsonResponse({
            "html": html,
            "page_number": page_obj.number,
            "num_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
        })

    return render(request, "accounts/franchises_by_city.html", context)


def franchises_by_direction(request, direction_id):
    """SEO landing page for franchises in a specific category/direction."""
    from .models import Directions

    direction = get_object_or_404(Directions, direction_id=direction_id)

    franchises_qs = Franchises.objects.filter(
        direction=direction, status="approved"
    ).select_related("owner", "direction", "stage").order_by("-created_at")

    paginator = Paginator(franchises_qs, 12)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    other_directions = Directions.objects.filter(
        franchises__status="approved"
    ).exclude(direction_id=direction.direction_id).distinct().order_by("direction_name")

    context = {
        "direction": direction,
        "page_obj": page_obj,
        "paginator": paginator,
        "franchises_count": franchises_qs.count(),
        "other_directions": other_directions,
    }

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string("accounts/partials/_franchise_cards.html", context, request=request)
        return JsonResponse({
            "html": html,
            "page_number": page_obj.number,
            "num_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
        })

    return render(request, "accounts/franchises_by_direction.html", context)


# ── Franchise AI Importer ────────────────────────────────────

@login_required
def franchise_import(request):
    """AI-powered franchise import from files and URLs (moderator only)."""
    if not (request.user.is_authenticated and hasattr(request.user, 'role') and request.user.role and request.user.role.role_name == 'moderator'):
        messages.error(request, "Доступ только для модераторов")
        return redirect("home")

    result = None
    warnings = []
    franchise = None

    if request.method == "POST":
        from accounts.franchise_importer import parse_file, parse_url, is_image_file, extract_franchise_data, create_franchise_from_data

        # Collect all text from all sources
        all_texts = []
        image_files = []

        # Parse URL if provided
        url = request.POST.get("import_url", "").strip()
        if url:
            url_text = parse_url(url)
            if url_text:
                all_texts.append(f"=== Данные с сайта {url} ===\n{url_text}")

        # Parse uploaded files
        files = request.FILES.getlist("import_files")
        for f in files:
            if is_image_file(f.name):
                image_files.append(f)
            else:
                text = parse_file(f, f.name)
                if text:
                    all_texts.append(f"=== Файл: {f.name} ===\n{text}")

        if not all_texts:
            messages.warning(request, "Не удалось извлечь текст из загруженных файлов. Добавьте PDF, DOCX, Excel или ссылку на сайт.")
            return render(request, "accounts/franchise_import.html", {})

        combined_text = "\n\n".join(all_texts)

        # AI extraction
        extracted = extract_franchise_data(combined_text)
        if not extracted:
            messages.error(request, "AI не смог обработать данные. Попробуйте загрузить другие файлы.")
            return render(request, "accounts/franchise_import.html", {})

        # Create franchise
        franchise, warnings = create_franchise_from_data(
            extracted, image_files=image_files, user=request.user
        )

        if franchise:
            result = "success"
            messages.success(request, f"Франшиза \"{franchise.title}\" создана! Проверьте и дозаполните данные.")

    return render(request, "accounts/franchise_import.html", {
        "result": result,
        "warnings": warnings,
        "franchise": franchise,
    })
